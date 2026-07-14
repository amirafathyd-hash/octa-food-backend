import os
import shutil
import subprocess
import tempfile
from collections import defaultdict, deque

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DESSERT_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "data", "Tokyo_Dessert_Ordering.xlsm")

MEAL_HEADER_KEYS = {"meal name", "meal", "اسم الوجبة", "الصنف"}
COUNT_HEADER_KEYS = {"total count", "count", "عدد", "العدد", "إجمالي العدد"}
DAY_NAME_TO_NO = {
    "السبت": 1,
    "الأحد": 2,
    "الاحد": 2,
    "الاثنين": 3,
    "الثلاثاء": 4,
    "الأربعاء": 5,
    "الاربعاء": 5,
    "الخميس": 6,
    "الجمعة": 7,
}


def _norm(value):
    return " ".join(str(value or "").strip().lower().split())


def _as_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _find_meal_count_columns(ws):
    max_row = min(ws.max_row, 60)
    max_col = min(ws.max_column, 80)
    for row in range(1, max_row + 1):
        meal_col = None
        count_col = None
        for col in range(1, max_col + 1):
            text = _norm(ws.cell(row=row, column=col).value)
            if text in MEAL_HEADER_KEYS:
                meal_col = col
            if text in COUNT_HEADER_KEYS:
                count_col = col
        if meal_col and count_col:
            return row, meal_col, count_col
    return None


def _detect_day_no(ws):
    for row in range(1, min(ws.max_row, 25) + 1):
        for col in range(1, min(ws.max_column, 5) + 1):
            text = str(ws.cell(row=row, column=col).value or "").strip()
            if text in DAY_NAME_TO_NO:
                return DAY_NAME_TO_NO[text], text
    return None, None


def _find_update_table(ws):
    for row in range(1, min(ws.max_row, 40) + 1):
        label = _norm(ws.cell(row=row, column=1).value)
        total_count = _norm(ws.cell(row=row - 1, column=12).value) if row > 1 else ""
        if label == "row labels" and total_count == "total count":
            return row + 1, 1, 12
    return 10, 1, 12


def read_uploaded_meal_counts(file_storage):
    file_storage.seek(0)
    wb = load_workbook(file_storage, data_only=True, read_only=True)
    try:
        day_no, day_name = None, None
        for ws in wb.worksheets:
            if day_no is None:
                day_no, day_name = _detect_day_no(ws)
            found = _find_meal_count_columns(ws)
            if not found:
                continue
            header_row, meal_col, count_col = found
            rows = []
            for row in range(header_row + 1, ws.max_row + 1):
                meal = ws.cell(row=row, column=meal_col).value
                count = _as_number(ws.cell(row=row, column=count_col).value)
                if not meal or count is None:
                    continue
                rows.append({"meal_name": str(meal).strip(), "count": count})
            if rows:
                return {"rows": rows, "day_no": day_no, "day_name": day_name}

        if "Update" in wb.sheetnames:
            ws = wb["Update"]
            day_no, day_name = _detect_day_no(ws)
            start_row, meal_col, count_col = _find_update_table(ws)
            rows = []
            for row in range(start_row, ws.max_row + 1):
                meal = ws.cell(row=row, column=meal_col).value
                count = _as_number(ws.cell(row=row, column=count_col).value)
                if meal and count is not None:
                    rows.append({"meal_name": str(meal).strip(), "count": count})
            if rows:
                return {"rows": rows, "day_no": day_no, "day_name": day_name}
    finally:
        wb.close()
        file_storage.seek(0)

    raise ValueError("مش لاقي أعمدة Meal name / Total Count في ملف الرفع")


def _target_meal_slots(ws):
    slots = []
    for row in range(2, ws.max_row + 1):
        meal = ws[f"AF{row}"].value
        if meal:
            slots.append({"row": row, "meal_name": str(meal).strip()})
    return slots


def _write_counts(template_path, uploaded_rows):
    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    ws = wb["Ordering"]
    slots = _target_meal_slots(ws)

    matched = []
    unmatched = []

    if len(uploaded_rows) == len(slots):
        for slot, item in zip(slots, uploaded_rows):
            ws[f"AG{slot['row']}"] = item["count"]
            matched.append({"row": slot["row"], "meal_name": slot["meal_name"], "count": item["count"]})
    else:
        by_name = defaultdict(deque)
        for item in uploaded_rows:
            by_name[_norm(item["meal_name"])].append(item["count"])
        for slot in slots:
            queue = by_name.get(_norm(slot["meal_name"]))
            if queue:
                count = queue.popleft()
                ws[f"AG{slot['row']}"] = count
                matched.append({"row": slot["row"], "meal_name": slot["meal_name"], "count": count})
            else:
                unmatched.append(slot["meal_name"])

    out_path = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
    _sync_ordering_counts_to_recipe_sheets(wb)
    wb.save(out_path)
    wb.close()
    return out_path, {"matched": matched, "unmatched": unmatched, "uploaded_count": len(uploaded_rows)}


def _extract_ag_reference(value):
    text = str(value or "").strip().upper()
    if text.startswith("="):
        text = text[1:]
    return text if text.startswith("AG") else None


def _sync_ordering_counts_to_recipe_sheets(wb):
    if "Ordering" not in wb.sheetnames:
        return
    ws = wb["Ordering"]
    for row in range(3, ws.max_row + 1):
        sheet_name = ws[f"AA{row}"].value
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        count_ref = _extract_ag_reference(ws[f"AC{row}"].value)
        count = ws[count_ref].value if count_ref else ws[f"AC{row}"].value
        if count not in (None, ""):
            wb[sheet_name]["V1"] = count


def _apply_edits_to_workbook(wb, edits):
    for edit in edits or []:
        sheet = edit.get("sheet")
        address = edit.get("address")
        value = edit.get("value")
        if not sheet or not address or sheet not in wb.sheetnames:
            continue
        cell = wb[sheet][address]
        if isinstance(cell.value, str) and cell.value.startswith("="):
            continue
        number = _as_number(value)
        cell.value = number if number is not None and str(value).strip() != "" else value


def _build_ordering_aggregates(calculated_workbook_path):
    wb = load_workbook(calculated_workbook_path, data_only=True)
    ws = wb["Ordering"]
    selected_day = _as_number(ws["R1"].value)
    selected_day = int(selected_day) if selected_day is not None else None
    daily = defaultdict(float)
    weekly = defaultdict(float)

    for row in range(3, ws.max_row + 1):
        sheet_name = ws[f"AA{row}"].value
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        row_day = _as_number(ws[f"AB{row}"].value)
        row_day = int(row_day) if row_day is not None else None
        recipe = wb[sheet_name]
        for recipe_row in range(5, recipe.max_row + 1):
            ingredient = recipe[f"B{recipe_row}"].value
            quantity = _as_number(recipe[f"H{recipe_row}"].value)
            if not ingredient or quantity is None:
                continue
            weekly[ingredient] += quantity
            if selected_day is None or row_day == selected_day:
                daily[ingredient] += quantity

    wb.close()
    return daily, weekly


def _write_ordering_aggregates(xlsm_path, daily, weekly):
    wb = load_workbook(xlsm_path, data_only=False, keep_vba=True)
    ws = wb["Ordering"]
    for row in range(3, ws.max_row + 1):
        ingredient = ws[f"A{row}"].value
        if not ingredient:
            continue
        ws[f"D{row}"] = round(daily.get(ingredient, 0), 6)
        ws[f"E{row}"] = round(weekly.get(ingredient, 0), 6)
    out_path = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
    wb.save(out_path)
    wb.close()
    return out_path


def recalc_with_ordering_aggregates(xlsm_path):
    first_xlsx = recalc_workbook_to_xlsx(xlsm_path)
    daily, weekly = _build_ordering_aggregates(first_xlsx)
    aggregated_xlsm = _write_ordering_aggregates(xlsm_path, daily, weekly)
    return recalc_workbook_to_xlsx(aggregated_xlsm)


def _soffice_bin():
    return os.environ.get("SOFFICE_BIN") or shutil.which("soffice") or "soffice"


def recalc_workbook_to_xlsx(xlsm_path):
    out_dir = tempfile.mkdtemp(prefix="dessert_recalc_")
    profile_dir = tempfile.mkdtemp(prefix="dessert_lo_profile_")
    cmd = [
        _soffice_bin(),
        f"-env:UserInstallation=file://{profile_dir}",
        "--headless",
        "--convert-to",
        "xlsx",
        "--outdir",
        out_dir,
        xlsm_path,
    ]
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=120)
    if proc.returncode != 0:
        raise RuntimeError((proc.stderr or proc.stdout or "LibreOffice failed").strip())
    base = os.path.splitext(os.path.basename(xlsm_path))[0] + ".xlsx"
    return os.path.join(out_dir, base)


def _export_workbook_to_pdf(workbook_path):
    out_dir = tempfile.mkdtemp(prefix="dessert_pdf_")
    profile_dir = tempfile.mkdtemp(prefix="dessert_lo_pdf_profile_")
    cmd = [
        _soffice_bin(),
        f"-env:UserInstallation=file://{profile_dir}",
        "--headless",
        "--convert-to",
        "pdf",
        "--outdir",
        out_dir,
        workbook_path,
    ]
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=120)
    if proc.returncode != 0:
        raise RuntimeError((proc.stderr or proc.stdout or "LibreOffice PDF export failed").strip())
    base = os.path.splitext(os.path.basename(workbook_path))[0] + ".pdf"
    return os.path.join(out_dir, base)


def _selected_recipe_sheets(wb, day_no):
    ws = wb["Ordering"]
    sheets = []
    for row in range(3, ws.max_row + 1):
        sheet_name = ws[f"AA{row}"].value
        row_day = _as_number(ws[f"AB{row}"].value)
        if not sheet_name or sheet_name not in wb.sheetnames or row_day is None:
            continue
        if int(row_day) == int(day_no):
            sheets.append(sheet_name)
    return sheets


def _last_recipe_row(ws):
    last = 4
    for row in range(5, ws.max_row + 1):
        if any(ws.cell(row=row, column=col).value not in (None, "") for col in range(1, 9)):
            last = row
    return last


def _prepare_pdf_workbook(calculated_workbook_path, day_no):
    wb = load_workbook(calculated_workbook_path)
    recipe_sheets = _selected_recipe_sheets(wb, day_no)
    if not recipe_sheets:
        wb.close()
        raise ValueError(f"مفيش شيتات وصفات لليوم {day_no}")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.sheet_state = "visible" if sheet_name in recipe_sheets else "hidden"
        if sheet_name not in recipe_sheets:
            continue

        last_row = _last_recipe_row(ws)
        ws.print_area = f"A2:H{last_row}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25
        ws.page_margins.top = 0.95
        ws.page_margins.bottom = 0.75
        ws.oddHeader.center.text = "&A"
        ws.oddHeader.center.size = 14
        ws.oddHeader.center.font = "Arial,Bold"
        ws.oddHeader.right.text = f"Day {int(day_no)}"
        ws.oddHeader.right.size = 12
        ws.oddHeader.right.font = "Arial,Bold"
        ws.oddFooter.center.text = "Page &P of &N"

    wb.active = wb.sheetnames.index(recipe_sheets[0])
    out_path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
    wb.save(out_path)
    wb.close()
    return out_path, recipe_sheets


def _style_report_sheet(ws, freeze="A2"):
    purple = "70306F"
    light = "F6F2EC"
    ws.freeze_panes = freeze
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=purple)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col in range(1, ws.max_column + 1):
        max_len = 10
        for row in range(1, min(ws.max_row, 80) + 1):
            value = ws.cell(row=row, column=col).value
            max_len = max(max_len, len(str(value or "")))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 42)
    for row in range(2, ws.max_row + 1, 2):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=light)


def _money(value):
    number = _as_number(value)
    return round(number or 0, 3)


def _build_cost_report(calculated_workbook_path, day_no):
    source = load_workbook(calculated_workbook_path, data_only=True)
    ordering = source["Ordering"]
    recipe_sheets = _selected_recipe_sheets(source, day_no)

    meals = []
    for sheet_name in recipe_sheets:
        ws = source[sheet_name]
        english_name = ws["A2"].value or sheet_name
        arabic_name = ws["A3"].value or ""
        required = _money(ws["S13"].value or ws["V1"].value)
        safety = _money(ws["S14"].value)
        total_qty = _money(ws["S15"].value)
        total_cost = _money(ws["S16"].value)
        unit_cost = _money(ws["Y8"].value or ws["X8"].value)
        meals.append({
            "sheet": sheet_name,
            "english_name": english_name,
            "arabic_name": arabic_name,
            "required": required,
            "safety": safety,
            "total_qty": total_qty,
            "unit_cost": unit_cost,
            "total_cost": total_cost,
        })

    ingredients = []
    category_totals = defaultdict(float)
    for row in range(3, ordering.max_row + 1):
        item = ordering[f"A{row}"].value
        category = ordering[f"B{row}"].value
        if not item:
            continue
        daily_weight = _money(ordering[f"D{row}"].value)
        daily_order = _money(ordering[f"L{row}"].value)
        daily_price = _money(ordering[f"O{row}"].value)
        if daily_weight == 0 and daily_order == 0 and daily_price == 0:
            continue
        category_totals[category or "Uncategorized"] += daily_price
        ingredients.append({
            "item": item,
            "category": category,
            "unit": ordering[f"C{row}"].value,
            "daily_weight": daily_weight,
            "daily_order": daily_order,
            "final_unit": ordering[f"M{row}"].value,
            "unit_price": _money(ordering[f"J{row}"].value),
            "daily_price": daily_price,
        })

    report = Workbook()
    summary = report.active
    summary.title = "Summary"
    summary.append(["Metric", "Value"])
    summary.append(["Day", int(day_no)])
    summary.append(["Meals Count", len(meals)])
    summary.append(["Total Meal Cost", round(sum(m["total_cost"] for m in meals), 3)])
    summary.append(["Total Ingredient Spend", round(sum(i["daily_price"] for i in ingredients), 3)])
    summary.append(["Total Items Ordered", round(sum(i["daily_order"] for i in ingredients), 3)])

    meal_ws = report.create_sheet("Meal Costs")
    meal_ws.append(["Sheet", "Meal English", "Meal Arabic", "Required", "Safety", "Total Qty", "Unit Cost", "Total Cost"])
    for meal in meals:
        meal_ws.append([
            meal["sheet"], meal["english_name"], meal["arabic_name"], meal["required"],
            meal["safety"], meal["total_qty"], meal["unit_cost"], meal["total_cost"],
        ])

    ingredient_ws = report.create_sheet("Ingredient Map")
    ingredient_ws.append(["Ingredient", "Category", "Unit", "Daily Weight", "Daily Order", "Final Unit", "Unit Price", "Daily Spend"])
    for item in ingredients:
        ingredient_ws.append([
            item["item"], item["category"], item["unit"], item["daily_weight"],
            item["daily_order"], item["final_unit"], item["unit_price"], item["daily_price"],
        ])

    category_ws = report.create_sheet("Category Spend")
    category_ws.append(["Category", "Daily Spend"])
    for category, total in sorted(category_totals.items(), key=lambda pair: pair[1], reverse=True):
        category_ws.append([category, round(total, 3)])

    for ws in [summary, meal_ws, ingredient_ws, category_ws]:
        _style_report_sheet(ws)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.000'

    if len(meals) >= 1:
        chart = BarChart()
        chart.title = "Total Cost by Meal"
        chart.y_axis.title = "Cost"
        chart.x_axis.title = "Meal"
        data = Reference(meal_ws, min_col=8, min_row=1, max_row=meal_ws.max_row)
        cats = Reference(meal_ws, min_col=2, min_row=2, max_row=meal_ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        summary.add_chart(chart, "D2")

    if category_ws.max_row > 1:
        pie = PieChart()
        pie.title = "Spend by Category"
        data = Reference(category_ws, min_col=2, min_row=2, max_row=category_ws.max_row)
        labels = Reference(category_ws, min_col=1, min_row=2, max_row=category_ws.max_row)
        pie.add_data(data)
        pie.set_categories(labels)
        pie.height = 8
        pie.width = 12
        summary.add_chart(pie, "D18")

        cat_bar = BarChart()
        cat_bar.title = "Category Spend"
        cat_bar.y_axis.title = "Cost"
        data = Reference(category_ws, min_col=2, min_row=1, max_row=category_ws.max_row)
        cats = Reference(category_ws, min_col=1, min_row=2, max_row=category_ws.max_row)
        cat_bar.add_data(data, titles_from_data=True)
        cat_bar.set_categories(cats)
        cat_bar.height = 8
        cat_bar.width = 16
        category_ws.add_chart(cat_bar, "D2")

    out_path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
    report.save(out_path)
    source.close()
    return out_path, {"day_no": int(day_no), "meals_count": len(meals), "ingredients_count": len(ingredients)}


def _fill_rgb(cell):
    color = cell.fill.fgColor
    if color and color.type == "rgb" and color.rgb:
        return "#" + color.rgb[-6:]
    return None


def _meal_group_fill(row):
    if row > 38:
        return "#FFFFFF"
    if 3 <= row <= 9:
        return "#E2F0D9"
    if 11 <= row <= 15:
        return "#DDEBF7"
    if 17 <= row <= 23:
        return "#FFF2CC"
    if 25 <= row <= 28:
        return "#EDEDED"
    if 30 <= row <= 33:
        return "#FCE4D6"
    if 35 <= row <= 38:
        return "#D9E2F3"
    return "#FFFFFF"


def _rounded(value):
    if isinstance(value, float):
        return round(value, 3)
    return value


def _cell_payload(ws_formula, ws_values, row, col):
    formula_cell = ws_formula.cell(row=row, column=col)
    value_cell = ws_values.cell(row=row, column=col)
    formula = formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else None
    return {
        "address": formula_cell.coordinate,
        "row": row,
        "col": col,
        "value": _rounded(value_cell.value if formula else formula_cell.value),
        "formula": formula,
        "editable": formula is None,
        "fill": _fill_rgb(formula_cell),
        "bold": bool(formula_cell.font.bold),
        "align": formula_cell.alignment.horizontal,
        "number_format": formula_cell.number_format,
    }


def extract_workbook_state(workbook_path):
    wb_formula = load_workbook(workbook_path, data_only=False, keep_vba=True)
    wb_values = load_workbook(workbook_path, data_only=True, keep_vba=True)
    sheets = []
    for sheet_name in wb_formula.sheetnames:
        ws_formula = wb_formula[sheet_name]
        ws_values = wb_values[sheet_name]
        rows = []
        max_row = ws_formula.max_row
        max_col = ws_formula.max_column
        for row in range(1, max_row + 1):
            rows.append([_cell_payload(ws_formula, ws_values, row, col) for col in range(1, max_col + 1)])
        sheets.append({
            "name": sheet_name,
            "max_row": max_row,
            "max_col": max_col,
            "columns": [get_column_letter(col) for col in range(1, max_col + 1)],
            "rows": rows,
        })
    wb_formula.close()
    wb_values.close()
    return {"sheets": sheets}


def extract_dashboard_state(workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb["Ordering"]

    meals = []
    for row in range(2, ws.max_row + 1):
        meal = ws[f"AF{row}"].value
        if meal:
            meals.append({
                "row": row,
                "meal_name": meal,
                "count": _rounded(ws[f"AG{row}"].value),
                "fill": _meal_group_fill(row) if row > 38 else (_fill_rgb(ws[f"AF{row}"]) or _meal_group_fill(row)),
            })

    ingredients = []
    for row in range(1, ws.max_row + 1):
        item = ws.cell(row=row, column=1).value
        category = ws.cell(row=row, column=2).value
        if row != 1 and not item and not category:
            continue
        ingredients.append({
            "row": row,
            "item": item,
            "category": category,
            "unit": ws.cell(row=row, column=3).value,
            "daily_weight": _rounded(ws.cell(row=row, column=4).value),
            "weekly_weight": _rounded(ws.cell(row=row, column=5).value),
            "yield": _rounded(ws.cell(row=row, column=6).value),
            "order_base": _rounded(ws.cell(row=row, column=7).value),
            "order_unit": ws.cell(row=row, column=8).value,
            "price": _rounded(ws.cell(row=row, column=9).value),
            "cost_per_unit": _rounded(ws.cell(row=row, column=10).value),
            "weekly_order": _rounded(ws.cell(row=row, column=11).value),
            "daily_order": _rounded(ws.cell(row=row, column=12).value),
            "final_unit": ws.cell(row=row, column=13).value,
            "weekly_price": _rounded(ws.cell(row=row, column=14).value),
            "daily_price": _rounded(ws.cell(row=row, column=15).value),
            "fill": _fill_rgb(ws.cell(row=row, column=1)),
        })

    wb.close()
    return {"meals": meals, "ingredients": ingredients}


def update_dessert_ordering_from_upload(file_storage, template_path=DESSERT_TEMPLATE_PATH):
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Tokyo_Dessert_Ordering.xlsm غير موجود في data")
    upload_data = read_uploaded_meal_counts(file_storage)
    uploaded_rows = upload_data["rows"]
    updated_xlsm, report = _write_counts(template_path, uploaded_rows)
    if upload_data.get("day_no"):
        wb = load_workbook(updated_xlsm, data_only=False, keep_vba=True)
        wb["Ordering"]["R1"] = upload_data["day_no"]
        wb.save(updated_xlsm)
        wb.close()
        report["day_no"] = upload_data["day_no"]
        report["day_name"] = upload_data.get("day_name")
    recalculated_xlsx = recalc_with_ordering_aggregates(updated_xlsm)
    state = extract_dashboard_state(recalculated_xlsx)
    state.update(extract_workbook_state(recalculated_xlsx))
    report["matched_count"] = len(report["matched"])
    report["unmatched_count"] = len(report["unmatched"])
    return state, report


def export_dessert_pdf_with_edits(edits, template_path=DESSERT_TEMPLATE_PATH):
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Tokyo_Dessert_Ordering.xlsm غير موجود في data")
    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    _apply_edits_to_workbook(wb, edits)
    day_no = _as_number(wb["Ordering"]["R1"].value) or 1
    _sync_ordering_counts_to_recipe_sheets(wb)
    out_xlsm = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
    wb.save(out_xlsm)
    wb.close()

    recalculated_xlsx = recalc_with_ordering_aggregates(out_xlsm)
    pdf_workbook, recipe_sheets = _prepare_pdf_workbook(recalculated_xlsx, day_no)
    pdf_path = _export_workbook_to_pdf(pdf_workbook)
    return pdf_path, {"day_no": int(day_no), "sheets": recipe_sheets}


def export_dessert_excel_with_edits(edits, template_path=DESSERT_TEMPLATE_PATH):
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Tokyo_Dessert_Ordering.xlsm غير موجود في data")
    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    _apply_edits_to_workbook(wb, edits)
    day_no = _as_number(wb["Ordering"]["R1"].value) or 1
    _sync_ordering_counts_to_recipe_sheets(wb)
    out_xlsm = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
    wb.save(out_xlsm)
    wb.close()

    recalculated_xlsx = recalc_with_ordering_aggregates(out_xlsm)
    return recalculated_xlsx, {"day_no": int(day_no)}


def export_dessert_cost_report_with_edits(edits, template_path=DESSERT_TEMPLATE_PATH):
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Tokyo_Dessert_Ordering.xlsm غير موجود في data")
    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    _apply_edits_to_workbook(wb, edits)
    day_no = _as_number(wb["Ordering"]["R1"].value) or 1
    _sync_ordering_counts_to_recipe_sheets(wb)
    out_xlsm = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
    wb.save(out_xlsm)
    wb.close()

    recalculated_xlsx = recalc_with_ordering_aggregates(out_xlsm)
    return _build_cost_report(recalculated_xlsx, day_no)


def get_dessert_template_state(template_path=DESSERT_TEMPLATE_PATH):
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Tokyo_Dessert_Ordering.xlsm غير موجود في data")
    state = extract_dashboard_state(template_path)
    state.update(extract_workbook_state(template_path))
    return state


def recalculate_dessert_with_edits(edits, template_path=DESSERT_TEMPLATE_PATH):
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Tokyo_Dessert_Ordering.xlsm غير موجود في data")
    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    _apply_edits_to_workbook(wb, edits)
    _sync_ordering_counts_to_recipe_sheets(wb)
    out_path = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
    wb.save(out_path)
    wb.close()
    recalculated_xlsx = recalc_with_ordering_aggregates(out_path)
    state = extract_dashboard_state(recalculated_xlsx)
    state.update(extract_workbook_state(recalculated_xlsx))
    return state
