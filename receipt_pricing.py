import io
import re
from datetime import datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from price_center import normalize_price_item_name, price_item_key


def _to_decimal(value):
    if value is None or value == '':
        return None
    try:
        text = str(value).replace(',', '').strip()
        if not text or text in {'-', '—'}:
            return None
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _clean_header(value):
    return re.sub(r'\s+', '', str(value or '').strip().casefold())


def _name_candidates(name):
    clean = normalize_price_item_name(name)
    if not clean:
        return []
    parts = [clean]
    for sep in (' - ', '-', '–', '—'):
        if sep in clean:
            parts.extend(p.strip() for p in clean.split(sep) if p and p.strip())
    seen = []
    for part in parts:
        key = price_item_key(part)
        if key and key not in seen:
            seen.append(key)
    return seen


def build_price_lookup(price_rows):
    lookup = {}
    for row in price_rows or []:
        name = normalize_price_item_name(row.get('item_name'))
        price = _to_decimal(row.get('order_unit_price'))
        if not name or price is None:
            continue
        item = {
            'item_name': name,
            'order_unit_price': float(price),
        }
        for key in _name_candidates(name):
            lookup.setdefault(key, item)
    return lookup


def price_receipt_rows(receipt_rows, price_rows):
    lookup = build_price_lookup(price_rows)
    priced = []
    missing = []
    for row in receipt_rows or []:
        item_name = normalize_price_item_name(
            row.get('items') or row.get('item') or row.get('name') or row.get('item_name')
        )
        if not item_name:
            continue
        qty = (
            _to_decimal(row.get('received'))
            or _to_decimal(row.get('qty_received'))
            or _to_decimal(row.get('received_qty'))
        )
        if qty is None:
            qty = Decimal('0')
        required_qty = (
            _to_decimal(row.get('daily_order'))
            or _to_decimal(row.get('required'))
            or _to_decimal(row.get('required_qty'))
            or _to_decimal(row.get('qty_required'))
        )
        match = None
        for key in _name_candidates(item_name):
            match = lookup.get(key)
            if match:
                break
        unit_price = _to_decimal(match.get('order_unit_price')) if match else None
        total = (qty * unit_price) if unit_price is not None else None
        out = {
            'item_name': item_name,
            'category': row.get('category') or '',
            'unit': row.get('order_unit') or row.get('unit') or row.get('rec_unit') or '',
            'required': float(required_qty) if required_qty is not None else None,
            'received': float(qty),
            'price_item_name': match.get('item_name') if match else '',
            'order_unit_price': float(unit_price) if unit_price is not None else None,
            'total_cost': float(total) if total is not None else None,
        }
        if not match:
            missing.append(item_name)
        priced.append(out)
    return priced, list(dict.fromkeys(missing))


def _find_external_columns(ws):
    item_col = None
    required_col = None
    qty_col = None
    category_col = None
    unit_col = None
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True), 1):
        for col_idx, value in enumerate(row, 1):
            h = _clean_header(value)
            if h in {'الصنف', 'اسمالصنف', 'items', 'item', 'itemname', 'name'}:
                item_col = col_idx
            if h in {'الاستلام', 'المستلم', 'كميةالمستلم', 'كميةالاستلام', 'received', 'receivedqty', 'qtyreceived'}:
                qty_col = col_idx
            if h in {'المطلوب', 'كميةالمطلوب', 'الكميةالمطلوبة', 'dailyorder', 'required', 'requiredqty', 'qtyrequired'}:
                required_col = col_idx
            if h in {'التصنيف', 'category'}:
                category_col = col_idx
            if h in {'الوحدة', 'unit', 'orderunit'}:
                unit_col = col_idx
        if item_col and qty_col:
            header_row = row_idx
            break
    if not item_col:
        item_col = 1
    if not qty_col:
        raise ValueError('لم يتم العثور على عمود الاستلام/المستلم داخل الشيت')
    return header_row or 1, item_col, qty_col, required_col, category_col, unit_col


def parse_external_receipt_workbook(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True, read_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            header_row, item_col, qty_col, required_col, category_col, unit_col = _find_external_columns(ws)
        except ValueError:
            continue
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            item_name = normalize_price_item_name(row[item_col - 1] if len(row) >= item_col else '')
            if not item_name:
                continue
            qty = row[qty_col - 1] if len(row) >= qty_col else None
            if _to_decimal(qty) is None:
                continue
            rows.append({
                'items': item_name,
                'category': row[category_col - 1] if category_col and len(row) >= category_col else '',
                'order_unit': row[unit_col - 1] if unit_col and len(row) >= unit_col else '',
                'daily_order': row[required_col - 1] if required_col and len(row) >= required_col else '',
                'received': qty,
                'sheet': sheet_name,
            })
    if not rows:
        raise ValueError('لم يتم العثور على صفوف استلام صالحة داخل الملف')
    return rows


def build_receipt_cost_workbook(priced_rows, title='تقرير الاستلام بالأسعار', subtitle='', missing=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'الاستلام بالأسعار'
    ws.sheet_view.rightToLeft = True
    missing = missing or []

    dark = PatternFill('solid', start_color='2B1A13')
    gold = PatternFill('solid', start_color='D9A83D')
    soft = PatternFill('solid', start_color='FFF8EA')
    white = PatternFill('solid', start_color='FFFFFF')
    miss_fill = PatternFill('solid', start_color='FDEAEA')
    total_fill = PatternFill('solid', start_color='EAF7EF')
    thin = Side(style='thin', color='E7D2B8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center', wrap_text=True)

    ws.merge_cells('A1:H1')
    ws['A1'] = title
    ws['A1'].fill = dark
    ws['A1'].font = Font(name='Tahoma', bold=True, size=16, color='FFFFFF')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 34

    ws.merge_cells('A2:H2')
    ws['A2'] = subtitle or datetime.now().strftime('%Y-%m-%d')
    ws['A2'].fill = soft
    ws['A2'].font = Font(name='Tahoma', bold=True, size=11, color='8A5A13')
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 26

    headers = ['الصنف', 'التصنيف', 'الوحدة', 'الكمية المطلوبة', 'الكمية المستلمة', 'سعر وحدة الطلب', 'الإجمالي', 'حالة السعر']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(3, col, header)
        cell.fill = gold
        cell.font = Font(name='Tahoma', bold=True, size=11, color='1A1A1A')
        cell.alignment = center
        cell.border = border

    start = 4
    for idx, row in enumerate(priced_rows, start):
        has_price = row.get('order_unit_price') is not None
        fill = soft if idx % 2 else white
        if not has_price:
            fill = miss_fill
        values = [
            row.get('item_name') or '',
            row.get('category') or '',
            row.get('unit') or '',
            row.get('required') if row.get('required') is not None else '',
            row.get('received') or 0,
            row.get('order_unit_price') if has_price else '',
            row.get('total_cost') if has_price else '',
            'محسوب' if has_price else 'بدون سعر',
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(idx, col, value)
            cell.fill = fill
            cell.border = border
            cell.font = Font(name='Tahoma', size=10, bold=col in (4, 5, 6, 7))
            cell.alignment = right if col == 1 else center
            if col in (4, 5, 6, 7):
                cell.number_format = '#,##0.00'

    total_row = start + len(priced_rows)
    ws.cell(total_row, 6, 'إجمالي التكلفة').fill = total_fill
    ws.cell(total_row, 6).font = Font(name='Tahoma', bold=True, size=12)
    ws.cell(total_row, 6).alignment = center
    ws.cell(total_row, 7, f'=SUM(G{start}:G{total_row - 1})').fill = total_fill
    ws.cell(total_row, 7).font = Font(name='Tahoma', bold=True, size=12)
    ws.cell(total_row, 7).alignment = center
    ws.cell(total_row, 7).number_format = '#,##0.00'
    for col in range(1, 9):
        ws.cell(total_row, col).border = border

    if missing:
        miss_row = total_row + 2
        ws.merge_cells(start_row=miss_row, start_column=1, end_row=miss_row, end_column=8)
        ws.cell(miss_row, 1, 'أصناف بدون سعر في مركز الأسعار: ' + '، '.join(missing))
        ws.cell(miss_row, 1).fill = miss_fill
        ws.cell(miss_row, 1).font = Font(name='Tahoma', bold=True, color='B31210')
        ws.cell(miss_row, 1).alignment = right

    widths = [44, 18, 12, 16, 16, 16, 16, 16]
    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = 'A4'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
