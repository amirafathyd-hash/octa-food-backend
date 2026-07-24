"""
محطة التنقية من التقارير إلى اتخاذ القرار
==========================================

الفكرة: بترفع "فاتورة الكمية للمشتركين" الخام (نفس شيت Export اللي فيه
عمود لكل صنف/باقة/كمية) وبيرجع لك ملف بنفس شكل وترتيب ملف اليوم الجاهز
(زي "Octa Food Tue ... الثلاثاء.xlsx"): شيت Export فيه بياناتك المرفوعة
زي ما هي، شيت "Don't Use just refresh" فيه تفصيل كل صنف لـ Protein/Side
وربطه بالباقة النهائية، وشيت Update فيه الجدول المُجمّع (زي الـ Pivot
Table بالظبط): كل صنف (بروتين أو طبق جانبي) في صف، وكل باقة نهائية في
عمود بعدد ووزن، + إجمالي لكل صف وصف إجمالي كلي في الآخر.

مهم جدًا: الحساب هنا مبني على قاعدة عمل مؤكدة مع صاحب المشروع بعد فحص
ملف حقيقي (يوم الثلاثاء 21-7-2026) صف بصف:
  - كل صنف إنجليزي بيتقسم لـ Protein (والـ Side لو موجود) - قاموس ثابت.
  - كل باقة أصلية (زي "لايت دايت") بتتحول لتصنيف نهائي (زي "تكميم لايت")
    عبر شيت Packages.
  - 3 أصناف بس (ساندوتش البيض المسلوق / كلوب ساندوتش بالدجاج / ساندوتش
    الدجاج المشوي) بتتضاعف ×2 حصريًا لما تكون الباقة النهائية "تضخيم".
  - أي صنف جديد مش موجود في القاموس بيتعامل معاه النظام تلقائيًا باسم الصنف
    نفسه، ويتسجل في تقرير العملية عشان يتراجع ويتضاف للقاموس لاحقًا لو محتاج
    ترجمة أدق.

القاموس بيتوسع بمرور الوقت من ملفات مرجعية حقيقية، لكن التشغيل اليومي ما
بيقفش بسبب صنف جديد.
"""
import json
import os
import re
import tempfile
import zipfile
from collections import OrderedDict
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as BorderSide
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

LOOKUP_PATH = os.path.join(os.path.dirname(__file__), 'data', 'decision_station_lookup.json')

DAY_NAMES_BY_WEEKDAY = {
    5: 'السبت',    # Saturday
    6: 'الأحد',    # Sunday
    0: 'الاثنين',  # Monday
    1: 'الثلاثاء', # Tuesday
    2: 'الأربعاء', # Wednesday
    3: 'الخميس',   # Thursday
    4: 'الجمعة',   # Friday (نادرًا ما يستخدم فعليًا)
}

PREFERRED_PACKAGE_ORDER = ['جيم', 'تكميم لايت', 'تضخيم', 'غذاء العمل', 'سمارت دايت']

# ترتيب الصفوف المثبت من ملف المرجع:
# "Octa Food Tue 21 July 2026 الثلاثاء.xlsx".
# أي صنف جديد غير موجود هنا يضاف بعد الأصناف المعروفة بترتيب منطقي.
REFERENCE_PROTEIN_ORDER = [
    'ساندوتش الديك الرومي بالتوت',
    'ساندوتش البيض المسلوق',
    'كرواسون جبنة الفيتا والخضار',
    'شكشوكة تركية',
    'كلوب ساندوتش بالدجاج',
    'ساندوتش الدجاج المشوي',
    'ستيك دجاج',
    'دجاج بالفطر',
    'ستروجانوف باستا',
    'بيف أمانسي',
    'دجاج تكا',
    'فاصولياء بيضاء باللحم',
    'سمك بالسبانخ والليمون',
    'أوكتا بوكي بول الدجاج',
    'مقلوبة دجاج',
    'ساندوتش الدجاج بالباربيكيو بخبز الشيباتا',
    'سلطة الكينوا',
    'سلطة التانغو',
    'سلطة الفواكه',
    'تفاح أخضر',
    'تشيز كيك اللوتس',
    'كيكة الجزر',
    'كيك بالشكولاته واللوز',
    'مكسرات مشكلة',
]

REFERENCE_SIDE_ORDER_BY_PROTEIN = {
    'ستيك دجاج': ['خضار سوتيه', 'البطاطس المهروسة'],
    'بيف أمانسي': ['خضار سوتيه', 'الأرز بالزعفران'],
}

RAW_REQUIRED_HEADERS = {
    '#', 'الاسم الإنجليزي', 'الاسم العربي', 'الباقة', 'التصنيف',
    'مجموع الكارب', 'مجموع البروتين', 'الكمية',
}

EXPORT_COLUMNS = [
    '#', 'الاسم الإنجليزي', 'الاسم العربي', 'الباقة', 'المرجع',
    'التصنيف', 'الحجم', 'مجموع الكارب', 'مجموع البروتين', 'الكمية',
]

DONT_USE_COLUMNS = [
    'الاسم الإنجليزي', 'Protein', 'Side', 'الباقة', 'التصنيف',
    'مجموع الكارب', 'مجموع البروتين', 'الكمية', 'Total_GM',
    'Final_Package', 'Multiplier', 'Final_Count', 'Final_GM', 'Grams',
]


def load_lookup():
    with open(LOOKUP_PATH, encoding='utf-8') as fh:
        return json.load(fh)


def _category_rank(category):
    cat = category or ''
    if 'فطور' in cat:
        return 0
    if 'الوجبات الرئيسية' in cat or 'لو كارب' in cat:
        return 1
    if 'سلطات' in cat or 'فواكه' in cat or 'الإضافات' in cat:
        return 2
    if 'حلى' in cat or 'حلويات' in cat:
        return 3
    return 4


def _number(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _norm_text(value):
    return re.sub(r'\s+', ' ', str(value or '').strip()).casefold()


def _lookup_item_info(english, row, lookup):
    """يرجع بيانات الصنف من القاموس، ولو جديد يبني fallback آمن بدل إيقاف الملف."""
    items_map = lookup['items']
    info = items_map.get(english)
    if info is not None:
        return info, False

    english_norm = _norm_text(english)
    for known_name, known_info in items_map.items():
        if _norm_text(known_name) == english_norm:
            return known_info, False

    return {
        'protein': english,
        'side': None,
        'category': row.get('التصنيف') or '',
    }, True


def _rank_from_lookup_order(lookup, key, value):
    order = lookup.get(key) or []
    try:
        return order.index(value)
    except ValueError:
        return len(order) + 9999


def _dont_use_order_key(row):
    def stable_value(value):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    return json.dumps([
        row.get('الاسم الإنجليزي'),
        row.get('الباقة'),
        row.get('التصنيف'),
        stable_value(row.get('مجموع الكارب')),
        stable_value(row.get('مجموع البروتين')),
        stable_value(row.get('الكمية')),
    ], ensure_ascii=False, separators=(',', ':'))


def _text_set(values):
    return {
        _norm_text(value)
        for value in values or []
        if str(value or '').strip()
    }


def _special_grams(protein, final_pkg, final_count, default_gm, lookup):
    rules = lookup.get('special_grams_by_protein_package') or {}
    protein_rules = rules.get(protein) or rules.get(str(protein or '').strip())
    if not protein_rules:
        return default_gm
    grams_per_count = _number(protein_rules.get(final_pkg))
    if grams_per_count <= 0:
        return default_gm
    return final_count * grams_per_count


def _find_export_sheet(wb):
    """بيدور على أي شيت فيه أعمدة فاتورة المشتركين الخام (زي شيت Sheet1
    أو Export)، بغض النظر عن اسم الشيت."""
    for ws in wb.worksheets:
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = {
            str(val or '').strip(): idx
            for idx, val in enumerate(first_row, start=1)
        }
        if RAW_REQUIRED_HEADERS.issubset(headers):
            return ws, headers
    return None, None


def read_subscribers_invoice(file_storage):
    """بيقرا فاتورة المشتركين الخام ويرجع (rows, day_label).
    rows: list[dict] بنفس أعمدة EXPORT_COLUMNS.
    day_label: اسم اليوم بالعربي، مستنتج من تاريخ في اسم الملف لو موجود."""
    file_storage.seek(0)
    # read_only=False لأن القراءة العشوائية بالخلية (ws.cell) في وضع read_only
    # بطيئة/غير مضمونة على ملفات الشيتات الكبيرة زي دي.
    wb = load_workbook(file_storage, data_only=True, read_only=False)
    ws, headers = _find_export_sheet(wb)
    if ws is None:
        raise ValueError(
            "الملف المرفوع لازم يحتوي على أعمدة فاتورة المشتركين "
            "(الاسم الإنجليزي / الباقة / التصنيف / الكمية ... إلخ)"
        )

    col_index = {col: headers[col] for col in EXPORT_COLUMNS if col in headers}
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        english_idx = col_index.get('الاسم الإنجليزي')
        if english_idx is None or english_idx - 1 >= len(values):
            continue
        english = values[english_idx - 1]
        if not english or not str(english).strip():
            continue
        row = {}
        for col, idx in col_index.items():
            row[col] = values[idx - 1] if idx - 1 < len(values) else None
        for col in EXPORT_COLUMNS:
            row.setdefault(col, None)
        rows.append(row)
    wb.close()
    file_storage.seek(0)

    upload_name = (
        getattr(file_storage, 'filename', '') or
        os.path.basename(getattr(file_storage, 'name', '') or '')
    )
    day_label = _day_label_from_filename(upload_name)
    return rows, day_label


def _day_label_from_filename(filename):
    match = re.search(r'(20\d{2})[-_/](\d{1,2})[-_/](\d{1,2})', filename or '')
    if not match:
        return None
    date_value = datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    return DAY_NAMES_BY_WEEKDAY.get(date_value.weekday())


def _date_token_from_filename(filename):
    """يرجع التاريخ بصيغة ثابتة لاستخدامه في اسم ملف النتيجة داخل الـ ZIP."""
    match = re.search(r'(20\d{2})[-_/](\d{1,2})[-_/](\d{1,2})', filename or '')
    if not match:
        return None
    return f'{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}'


def compute_decision_tables(rows, lookup):
    """بياخد صفوف فاتورة المشتركين الخام ويرجع:
    dont_use_rows, pivot_rows, package_order, report

    شكل جدول Update الحقيقي مش عبارة عن صف واحد لكل اسم صنف بس - هو تجميع
    على مستويين زي أي Pivot Table فيها Row field مركّب:
      1) صف "Protein" بيجمع كل الكميات لنفس البروتين (زي "دجاج بالفطر")
         مهما اختلفت الأطباق الجانبية أو الباقات.
      2) تحته مباشرة، صف منفصل لكل طبق جانبي حقيقي (Side != '-') ظهر مع
         البروتين ده بالذات - حتى لو نفس اسم الطبق الجانبي ("خضار سوتيه"
         مثلاً) ظهر تحت بروتينات تانية، هيبقى له صف منفصل لكل بروتين
         (بالظبط زي ما بيحصل في ملف الإكسل الحقيقي - راجعنا ده صف بصف).
      3) لو الصنف مالوش طبق جانبي حقيقي، بيتضاف تحته صف "-" بنفس أرقامه،
         كما يظهر في ملف المرجع بالضبط.

    pivot_rows: list[(display_name, {final_package: [count, grams]})]
    بنفس ترتيب الظهور الحقيقي في ملف اليوم الجاهز."""
    package_map = lookup['package_map']
    double_items = set(lookup.get('double_in_bulking', []))
    double_proteins = _text_set(lookup.get('double_proteins_in_bulking'))
    bulking_pkg = lookup.get('bulking_final_package', 'تضخيم')
    dont_use_package_order = lookup.get('dont_use_package_order') or list(package_map.keys())
    original_package_rank = {
        name: idx
        for idx, name in enumerate(dont_use_package_order)
    }
    dont_use_row_rank = {
        key: idx
        for idx, key in enumerate(lookup.get('dont_use_row_order') or [])
    }

    inferred_items = set()
    inferred_packages = set()
    dont_use_rows = []
    package_order = []

    protein_order = []
    protein_category = {}  # protein -> أول تصنيف اتشاف بيه (لترتيب البلوكات زي الملف الأصلي)
    protein_data = {}  # protein -> {'totals': OrderedDict, 'side_order': [], 'sides': {side: OrderedDict}}

    for source_index, row in enumerate(rows):
        english = str(row.get('الاسم الإنجليزي') or '').strip()
        qty = _number(row.get('الكمية'))
        if not english or qty <= 0:
            continue
        item_info, inferred_item = _lookup_item_info(english, row, lookup)
        orig_pkg = str(row.get('الباقة') or '').strip()
        final_pkg = package_map.get(orig_pkg)

        if not final_pkg:
            final_pkg = orig_pkg or '(فارغ)'
            inferred_packages.add(final_pkg)
        if inferred_item:
            inferred_items.add(english)

        protein = item_info.get('protein')
        side = item_info.get('side') or '-'
        category = row.get('التصنيف') or item_info.get('category') or ''
        carb = _number(row.get('مجموع الكارب'))
        protein_g = _number(row.get('مجموع البروتين'))
        base_gm = carb or protein_g
        total_gm = base_gm if base_gm else qty

        multiplier = 2 if (
            final_pkg == bulking_pkg and
            (english in double_items or _norm_text(protein) in double_proteins)
        ) else 1
        final_count = qty * multiplier
        final_gm = total_gm * multiplier
        final_gm = _special_grams(protein, final_pkg, final_count, final_gm, lookup)

        dont_use_rows.append({
            'الاسم الإنجليزي': english,
            'Protein': protein,
            'Side': side,
            'الباقة': orig_pkg,
            'التصنيف': category,
            'مجموع الكارب': carb,
            'مجموع البروتين': protein_g,
            'الكمية': qty,
            'Total_GM': total_gm,
            'Final_Package': final_pkg,
            'Multiplier': multiplier,
            'Final_Count': final_count,
            'Final_GM': final_gm,
            'Grams': final_gm,
            '_source_index': source_index,
            '_original_package_rank': original_package_rank.get(orig_pkg, len(original_package_rank) + source_index),
        })
        dont_use_rows[-1]['_row_rank'] = dont_use_row_rank.get(
            _dont_use_order_key(dont_use_rows[-1]),
            len(dont_use_row_rank) + source_index,
        )

        if final_pkg not in package_order:
            package_order.append(final_pkg)

        if protein:
            if protein not in protein_data:
                protein_data[protein] = {'totals': OrderedDict(), 'side_order': [], 'sides': {}}
                protein_order.append(protein)
                protein_category[protein] = category
            pd = protein_data[protein]
            bucket = pd['totals'].setdefault(final_pkg, [0.0, 0.0])
            bucket[0] += final_count
            bucket[1] += final_gm

            if side:
                if side not in pd['sides']:
                    pd['sides'][side] = OrderedDict()
                    pd['side_order'].append(side)
                sbucket = pd['sides'][side].setdefault(final_pkg, [0.0, 0.0])
                sbucket[0] += final_count
                sbucket[1] += final_gm

    # pivot_rows: (display_name, totals_by_package, is_protein_level)
    # is_protein_level=True بس للصف الأب (البروتين) - ده اللي بيدخل في
    # حساب Grand Total، أما صفوف الطبق الجانبي فهي تفصيل تحت صف البروتين
    # (نفس قيمه جزئيًا) ومينفعش تتجمع تاني في الإجمالي الكلي وإلا
    # هيتضاعف الرقم.
    # نفس ترتيب "البلوكات" في ملف اليوم الجاهز: فطور، بعدين الوجبات
    # الرئيسية (وبنودها اللي بلو كارب)، بعدين سلطات/فواكه/إضافات، بعدين
    # حلى، وأي تصنيف تاني مش معروف يتحط في الآخر. الترتيب جوه كل بلوك
    # زي ترتيب أول ظهور في الملف المرفوع نفسه.
    reference_protein_rank = {
        name: idx
        for idx, name in enumerate(REFERENCE_PROTEIN_ORDER)
    }
    ordered_proteins = sorted(
        protein_order,
        key=lambda p: (
            reference_protein_rank.get(p, len(reference_protein_rank) + 9999),
            _rank_from_lookup_order(lookup, 'row_label_order', p),
            _category_rank(protein_category.get(p, '')),
            protein_order.index(p),
        ),
    )

    pivot_rows = []
    for protein in ordered_proteins:
        pd = protein_data[protein]
        pivot_rows.append((protein, pd['totals'], True))
        side_lookup_order = (
            REFERENCE_SIDE_ORDER_BY_PROTEIN.get(protein)
            or lookup.get('side_order_by_protein', {}).get(protein)
            or []
        )
        ordered_sides = sorted(
            pd['side_order'],
            key=lambda s: (
                side_lookup_order.index(s) if s in side_lookup_order else len(side_lookup_order) + 9999,
                pd['side_order'].index(s),
            ),
        )
        for side in ordered_sides:
            pivot_rows.append((side, pd['sides'][side], False))

    ordered_packages = list(PREFERRED_PACKAGE_ORDER)
    ordered_packages += [p for p in package_order if p not in ordered_packages]

    dont_use_rows.sort(key=lambda row: (
        row.get('_row_rank', 999999),
        row.get('_original_package_rank', 999999),
        row.get('_source_index', 999999),
    ))
    for row in dont_use_rows:
        row.pop('_row_rank', None)
        row.pop('_source_index', None)
        row.pop('_original_package_rank', None)

    report = {
        'source_rows': len(rows),
        'computed_rows': len(dont_use_rows),
        'row_labels': len(pivot_rows),
        'package_columns': ordered_packages,
        'inferred_items': sorted(inferred_items),
        'inferred_packages': sorted(inferred_packages),
        'inference_mode': bool(inferred_items or inferred_packages),
    }
    return dont_use_rows, pivot_rows, ordered_packages, report


# ---------------------------------------------------------------------------
# بناء ملف الإخراج بنفس شكل ملف اليوم الجاهز (Export / Don't Use / Update / Packages)
# ---------------------------------------------------------------------------

THIN = BorderSide(style='thin', color='DDDDDD')
BORDER_THIN = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def _plain_header_row(ws, columns, bold=False):
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx, value=col)
        cell.font = Font(bold=bold, size=11)


def _category_fill(category, lookup):
    palette = lookup.get('category_colors') or {}
    cat = category or ''
    key = None
    if 'فطور' in cat:
        key = 'فطور'
    elif 'حلى' in cat or 'حلويات' in cat:
        key = 'حلى'
    elif 'سلطات' in cat or 'فواكه' in cat or 'الإضافات' in cat:
        key = 'سلطات'
    elif 'الوجبات الرئيسية' in cat or 'لو كارب' in cat:
        key = 'الوجبات الرئيسية'
    spec = palette.get(key) or palette.get('__default__') or {'theme': 4, 'tint': 0.8}
    if 'rgb' in spec:
        from openpyxl.styles.colors import Color
        return PatternFill(patternType='solid', fgColor=Color(rgb=spec['rgb']))
    from openpyxl.styles.colors import Color
    return PatternFill(patternType='solid', fgColor=Color(theme=spec.get('theme', 4), tint=spec.get('tint', 0.8)))


def build_output_workbook(day_label, export_rows, dont_use_rows, pivot_rows,
                           package_order, lookup, out_path=None):
    """بتبني الملف بنفس ترتيب التابات ونفس الخطوط ونفس عرض الأعمدة ونفس
    الزوم ونفس تجميد الصفوف الموجودين في ملف اليوم الجاهز الأصلي بالظبط
    (اتقاسوا خلية خلية من ملف الثلاثاء المرجعي) - الحاجة الوحيدة اللي
    بتتغيّر هي القيم نفسها."""
    wb = Workbook()
    wb.remove(wb.active)  # هنعمل الشيتات بالترتيب الصح بدل الافتراضي

    ARIAL = 'Arial'
    TIMES = 'Times New Roman'

    # ---------------- Update (أول تاب في الملف الأصلي - نفس الترتيب) ----------------
    ws_up = wb.create_sheet('Update')
    ws_up.sheet_view.showGridLines = False
    ws_up.sheet_view.zoomScale = 40
    ws_up.freeze_panes = 'B1'
    ws_up.sheet_format.defaultRowHeight = 23
    ws_up.row_dimensions[6].height = 25.25
    ws_up.row_dimensions[7].height = 14.5
    ws_up.row_dimensions[8].height = 36.75

    header_font = Font(name=TIMES, bold=True, size=18)
    plain_font = Font(name=TIMES, bold=False, size=18)
    center = Alignment(horizontal='center', vertical='center')

    a6 = ws_up.cell(row=6, column=1, value=day_label or '')
    a6.font = header_font
    a6.fill = PatternFill(patternType='solid', fgColor='FFFFFF00')
    a6.alignment = center

    n_pkg = len(package_order)
    total_count_col = 2 + n_pkg * 2
    total_gm_col = total_count_col + 1

    for i, pkg in enumerate(package_order):
        col = 2 + i * 2
        for header_col in (col, col + 1):
            c1 = ws_up.cell(row=7, column=header_col, value=pkg)
            c1.font = header_font
            c1.alignment = center
        ws_up.cell(row=9, column=col, value='Count').font = plain_font
        ws_up.cell(row=9, column=col, value='Count').alignment = center
        ws_up.cell(row=9, column=col + 1, value='Grams').font = plain_font
        ws_up.cell(row=9, column=col + 1).alignment = center

    for total_col in (total_count_col, total_gm_col):
        cell = ws_up.cell(row=7, column=total_col, value='Grand Total')
        cell.font = header_font
        cell.alignment = center
    ws_up.cell(row=9, column=1, value='Row Labels').font = header_font
    ws_up.cell(row=9, column=1).alignment = center
    total_count_header = ws_up.cell(
        row=9, column=total_count_col, value='Grand Total (Count)'
    )
    total_gm_header = ws_up.cell(
        row=9, column=total_gm_col, value='Grand Total (Grams)'
    )
    for cell in (total_count_header, total_gm_header):
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=True
        )

    r = 10
    grand = [0.0] * (n_pkg * 2)
    parent_category = ''
    for name, bucket, is_protein_level in pivot_rows:
        if is_protein_level:
            parent_category = lookup.get('name_category', {}).get(name, '')
        category = (
            parent_category if not is_protein_level and name == '-'
            else lookup.get('name_category', {}).get(name, parent_category)
        )
        fill = _category_fill(category, lookup)
        name_cell = ws_up.cell(row=r, column=1, value=name)
        name_cell.font = header_font
        name_cell.alignment = center
        name_cell.fill = fill
        row_total_count = 0.0
        row_total_gm = 0.0
        for i, pkg in enumerate(package_order):
            count, gm = bucket.get(pkg, [0.0, 0.0])
            col = 2 + i * 2
            cc = ws_up.cell(row=r, column=col, value=count)
            gc = ws_up.cell(row=r, column=col + 1, value=gm)
            cc.font = header_font
            gc.font = header_font
            cc.alignment = center
            gc.alignment = center
            cc.fill = fill
            gc.fill = fill
            gc.border = Border(right=BorderSide(style='thick', color='FF000000'))
            if is_protein_level:
                grand[i * 2] += count
                grand[i * 2 + 1] += gm
            row_total_count += count
            row_total_gm += gm
        tc = ws_up.cell(row=r, column=total_count_col, value=row_total_count)
        tg = ws_up.cell(row=r, column=total_gm_col, value=row_total_gm)
        tc.font = header_font
        tg.font = header_font
        tc.alignment = center
        tg.alignment = center
        tc.fill = fill
        tg.fill = fill
        r += 1

    total_row = r
    gt_cell = ws_up.cell(row=total_row, column=1, value='Grand Total')
    gt_cell.font = header_font
    gt_cell.alignment = center
    gt_total_count, gt_total_gm = 0.0, 0.0
    for i in range(n_pkg):
        col = 2 + i * 2
        cc = ws_up.cell(row=total_row, column=col, value=grand[i * 2])
        gc = ws_up.cell(row=total_row, column=col + 1, value=grand[i * 2 + 1])
        cc.font = header_font
        gc.font = header_font
        cc.alignment = center
        gc.alignment = center
        gc.border = Border(right=BorderSide(style='thick', color='FF000000'))
        gt_total_count += grand[i * 2]
        gt_total_gm += grand[i * 2 + 1]
    tc = ws_up.cell(row=total_row, column=total_count_col, value=gt_total_count)
    tg = ws_up.cell(row=total_row, column=total_gm_col, value=gt_total_gm)
    tc.font = header_font
    tg.font = header_font
    tc.alignment = center
    tg.alignment = center

    # نفس عرض أعمدة ملف الثلاثاء المرجعي بالظبط (5 تصنيفات = نفس العدد
    # المعتاد كل يوم)؛ أي عمود زيادة (تصنيف بوفيه نادر مثلاً) ياخد عرض
    # قريب منطقي بدل ما يفضل بعرض افتراضي ضيق.
    up_widths = [44, 31.58, 9.91, 11.33, 9.91, 9.25, 10.08, 13.41, 9.91, 11.75, 9.91, 16.75, 17.58]
    for i, w in enumerate(up_widths, start=1):
        ws_up.column_dimensions[get_column_letter(i)].width = w
    for i in range(len(up_widths) + 1, total_gm_col + 1):
        ws_up.column_dimensions[get_column_letter(i)].width = 11

    # ---------------- Export (تاني تاب في الملف الأصلي) ----------------
    ws_export = wb.create_sheet('Export')
    ws_export.sheet_view.zoomScale = 90
    ws_export.sheet_format.defaultRowHeight = 14
    header_style_export = Font(name=ARIAL, size=11)
    data_style_export = Font(name=ARIAL, size=12)
    for i, col in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws_export.cell(row=1, column=i, value=col)
        cell.font = header_style_export
    for r, row in enumerate(export_rows, start=2):
        ws_export.row_dimensions[r].height = 15.5
        for c, col in enumerate(EXPORT_COLUMNS, start=1):
            cell = ws_export.cell(row=r, column=c, value=row.get(col))
            cell.font = data_style_export
    export_widths = {2: 13.5, 3: 30.0, 4: 27.83, 5: 8.33, 6: 9.0, 8: 13.16, 9: 13.33}
    for i, w in export_widths.items():
        ws_export.column_dimensions[get_column_letter(i)].width = w
    export_table_last_row = max(412, 1 + len(export_rows))
    export_table = Table(
        displayName='Table1',
        ref=f'A1:J{export_table_last_row}',
    )
    export_table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws_export.add_table(export_table)

    # ---------------- Don't Use Just Refresh (تالت تاب) ----------------
    ws_du = wb.create_sheet("Don't Use just refresh")
    ws_du.sheet_format.defaultRowHeight = 14
    header_style_du = Font(name=ARIAL, size=11)
    data_style_du = Font(name=ARIAL, size=11)
    for i, col in enumerate(DONT_USE_COLUMNS, start=1):
        cell = ws_du.cell(row=1, column=i, value=col)
        cell.font = header_style_du
    for r, row in enumerate(dont_use_rows, start=2):
        for c, col in enumerate(DONT_USE_COLUMNS, start=1):
            cell = ws_du.cell(row=r, column=c, value=row.get(col))
            cell.font = data_style_du
    table_last_row = max(412, 1 + len(dont_use_rows))
    for r in range(2 + len(dont_use_rows), table_last_row + 1):
        ws_du.cell(row=r, column=3, value='-').font = data_style_du
        ws_du.cell(row=r, column=11, value=1).font = data_style_du
    du_widths = {1: 40.75, 2: 24.5, 3: 16.33, 4: 17.58, 5: 23.58, 6: 9.5, 7: 10.58,
                 8: 4.25, 9: 11.0, 10: 15.25, 11: 10.33, 12: 13.0, 13: 10.66, 14: 8.5}
    for i, w in du_widths.items():
        ws_du.column_dimensions[get_column_letter(i)].width = w
    dont_use_table = Table(
        displayName='Table1_1',
        ref=f'A1:N{table_last_row}',
    )
    dont_use_table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium7',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws_du.add_table(dont_use_table)

    # ---------------- Packages (رابع تاب) ----------------
    ws_pkg = wb.create_sheet('Packages')
    ws_pkg.sheet_view.zoomScale = 160
    ws_pkg.sheet_format.defaultRowHeight = 14
    pkg_header_font = Font(name=ARIAL, bold=True, size=11)
    pkg_header_fill = PatternFill(patternType='solid', fgColor=Color(theme=4, tint=0))
    for i, col in enumerate(['Original_Package', 'Final_Package'], start=1):
        cell = ws_pkg.cell(row=1, column=i, value=col)
        cell.font = pkg_header_font
        cell.fill = pkg_header_fill
    for r, (orig, final) in enumerate(lookup['package_map'].items(), start=2):
        ws_pkg.cell(row=r, column=1, value=orig).font = Font(name=ARIAL, size=11)
        ws_pkg.cell(row=r, column=2, value=final).font = Font(name=ARIAL, size=11)
    ws_pkg.column_dimensions['A'].width = 17.66
    ws_pkg.column_dimensions['B'].width = 15.16
    ws_pkg.column_dimensions['M'].width = 25.5
    package_table = Table(
        displayName='tblPackageMap',
        ref=f"A1:B{1 + len(lookup['package_map'])}",
    )
    package_table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws_pkg.add_table(package_table)

    wb.active = 0  # التاب اللي بيفتح بيه الملف = Update، زي الأصلي بالظبط

    if out_path is None:
        out_path = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
    wb.save(out_path)
    return out_path


def process_subscribers_invoice(file_storage, day_label_override=None, out_path=None):
    """الدالة الرئيسية اللي بيستدعيها الـ endpoint: من ملف مرفوع لملف جاهز."""
    lookup = load_lookup()
    rows, detected_day = read_subscribers_invoice(file_storage)
    if not rows:
        raise ValueError('الملف المرفوع فاضي أو مالوش صفوف بكمية أكبر من صفر')

    day_label = day_label_override or detected_day
    if not day_label:
        raise ValueError(
            'مش قادر أحدد اسم اليوم. سمّي الملف وفيه تاريخ بصيغة YYYY-MM-DD '
            '(زي 2026-07-21) أو ابعت اسم اليوم صراحة.'
        )

    dont_use_rows, pivot_rows, package_order, report = compute_decision_tables(rows, lookup)
    out_path = build_output_workbook(
        day_label, rows, dont_use_rows, pivot_rows, package_order, lookup, out_path=out_path
    )
    report['day_label'] = day_label
    return out_path, report


def process_subscribers_invoices(file_storages, out_path=None):
    """يعالج ملفات أيام متعددة مع عزل حساب كل يوم في Workbook مستقل.

    كان دمج صفوف أكثر من يوم قبل الحساب يحوّلها كلها إلى Pivot واحد، وبالتالي
    يغيّر أرقام اليوم الصحيح. هنا كل ملف يمر على نفس مسار اليوم الواحد بدون
    أي مشاركة للصفوف أو المجاميع، ثم نجمع ملفات النتائج فقط داخل ZIP.
    """
    files = list(file_storages or [])
    if len(files) < 2:
        raise ValueError('معالجة الأيام المتعددة تحتاج ملفين على الأقل')

    if out_path is None:
        out_path = tempfile.NamedTemporaryFile(suffix='.zip', delete=False).name

    reports = []
    archive_names = set()
    temporary_outputs = []
    try:
        with zipfile.ZipFile(out_path, 'w', compression=zipfile.ZIP_DEFLATED) as archive:
            for index, file_storage in enumerate(files, start=1):
                source_name = (
                    getattr(file_storage, 'filename', '') or
                    os.path.basename(getattr(file_storage, 'name', '') or '') or
                    f'day-{index}.xlsx'
                )
                result_path, report = process_subscribers_invoice(file_storage)
                temporary_outputs.append(result_path)

                date_token = _date_token_from_filename(source_name)
                name_token = date_token or f'day-{index}'
                base_name = f"Octa_Food_Decision_{name_token}_{report['day_label']}"
                archive_name = f'{base_name}.xlsx'
                duplicate_no = 2
                while archive_name in archive_names:
                    archive_name = f'{base_name}_{duplicate_no}.xlsx'
                    duplicate_no += 1
                archive_names.add(archive_name)
                archive.write(result_path, archive_name)

                reports.append({
                    **report,
                    'source_file': source_name,
                    'output_file': archive_name,
                })
    except Exception:
        if os.path.exists(out_path):
            os.unlink(out_path)
        raise
    finally:
        for result_path in temporary_outputs:
            try:
                os.unlink(result_path)
            except OSError:
                pass

    return out_path, {
        'mode': 'multiple',
        'file_count': len(reports),
        'computed_rows': sum(item['computed_rows'] for item in reports),
        'days': reports,
    }
