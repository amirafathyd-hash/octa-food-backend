import io
import re
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def normalize_price_item_name(value):
    text = str(value or '').strip()
    text = re.sub(r'\s+', ' ', text)
    return text


def price_item_key(value):
    text = normalize_price_item_name(value)
    text = re.sub(r'\s*[-–—]\s*', ' - ', text)
    return text.casefold()


def _to_decimal(value):
    if value is None or value == '':
        return None
    if isinstance(value, Decimal):
        return value
    try:
        text = str(value).replace(',', '').strip()
        if not text:
            return None
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _clean_header(value):
    return re.sub(r'\s+', '', str(value or '').strip().casefold())


def _find_columns(ws):
    name_col = None
    price_col = None
    header_row = 1
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True), 1):
        for col_idx, value in enumerate(row, 1):
            h = _clean_header(value)
            if h in {'items', 'item', 'itemname', 'name', 'الصنف', 'اسمالصنف', 'المنتج', 'اسمالمنتج'}:
                name_col = col_idx
            if (
                'سعروحدةالطلب' in h
                or 'priceoforderingunit' in h
                or h in {'price', 'unitprice', 'orderunitprice', 'السعر'}
            ):
                price_col = col_idx
        if name_col and price_col:
            header_row = row_idx
            break
    if not name_col:
        name_col = 1
    if not price_col:
        raise ValueError('لم يتم العثور على عمود سعر وحدة الطلب داخل الملف')
    return header_row, name_col, price_col


def extract_price_items_from_workbook(file_obj):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    header_row, name_col, price_col = _find_columns(ws)

    rows = []
    seen = set()
    skipped = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        raw_name = row[name_col - 1] if len(row) >= name_col else None
        raw_price = row[price_col - 1] if len(row) >= price_col else None
        item_name = normalize_price_item_name(raw_name)
        price = _to_decimal(raw_price)
        if not item_name:
            continue
        if price is None:
            skipped += 1
            continue
        key = price_item_key(item_name)
        if key in seen:
            continue
        seen.add(key)
        rows.append({
            'item_name': item_name,
            'item_key': key,
            'order_unit_price': float(price),
        })
    if not rows:
        raise ValueError('لم يتم العثور على أصناف وأسعار صالحة داخل الملف')
    return rows, skipped


def build_price_items_workbook(items):
    wb = Workbook()
    ws = wb.active
    ws.title = 'مركز الأسعار'
    ws.sheet_view.rightToLeft = True

    title_fill = PatternFill('solid', start_color='1C1512')
    header_fill = PatternFill('solid', start_color='D8A83D')
    soft_fill = PatternFill('solid', start_color='FFF8EA')
    white_fill = PatternFill('solid', start_color='FFFFFF')
    thin = Side(style='thin', color='E7D2B8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center', wrap_text=True)

    ws.merge_cells('A1:B1')
    title = ws['A1']
    title.value = 'مركز أسعار الخضروات'
    title.fill = title_fill
    title.font = Font(name='Tahoma', bold=True, size=15, color='FFFFFF')
    title.alignment = center
    ws.row_dimensions[1].height = 32

    headers = ['اسم الصنف', 'سعر وحدة الطلب']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(name='Tahoma', bold=True, size=12, color='1A1A1A')
        cell.alignment = center
        cell.border = border
    ws.column_dimensions['A'].width = 46
    ws.column_dimensions['B'].width = 18

    for idx, item in enumerate(items, 3):
        fill = soft_fill if idx % 2 else white_fill
        name_cell = ws.cell(row=idx, column=1, value=item.get('item_name') or '')
        price_cell = ws.cell(row=idx, column=2, value=float(item.get('order_unit_price') or 0))
        for cell in (name_cell, price_cell):
            cell.fill = fill
            cell.border = border
            cell.font = Font(name='Tahoma', bold=True if cell.column == 2 else False, size=11)
        name_cell.alignment = right
        price_cell.alignment = center
        price_cell.number_format = '#,##0.0000'

    ws.freeze_panes = 'A3'
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
