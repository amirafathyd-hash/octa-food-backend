"""
محرك "Tokyo Ordering" — واجهة إدخال سهلة فوق ملف الإكسل الأصلي بتاعك.

مهم: الكود ده ميعملش أي حساب بديل عن الإكسل. هو بس بيقرا ويكتب في خليتين فاضيتين
لكل وجبة لكل يوم (عدد الوجبات والوزن بالجرام) في شيت All_Ingredients، ويرجّع نسخة
محدّثة من نفس الملف بالماكرو زي ما هو، عشان تفتحها في إكسل وتدوس على زرار
"Update" بتاعك زي العادة — الحساب الحقيقي يفضل في الماكرو نفسه، إحنا غيّرنا بس
طريقة إدخال الأرقام.
"""
import shutil
import tempfile
import os
import re
from datetime import datetime
from difflib import SequenceMatcher
from openpyxl import load_workbook

DAY_NAMES = {1: 'السبت', 2: 'الأحد', 3: 'الاثنين', 4: 'الثلاثاء', 5: 'الأربعاء', 6: 'الخميس'}

DAY_NO_COL = 36   # AJ
SHEET_NAME_COL = 37  # AK
COUNT_COL = 44    # AR
GRAMS_COL = 45    # AS
SAFETY_COL = 39   # AM

# Recipe tabs whose operational upload name is deliberately different from
# the English tab name. Values are searched in both Arabic and English names
# read from the daily workbook.
SHEET_INPUT_ALIASES = {
    'Almond Chicken': ['Almond chicken in the oven with potato wedges', 'دجاج باللوز في الفرن'],
    'Pasta with Vegetable(Pasta)': ['Chicken pasta with vegetables', 'دجاج مكروني بالخضار'],
    'Daoud Basha': ['Daoud Basha with saffron rice', 'داوود باشا'],
    'Fish with cream': ['Fish with cream and mashed potatoes', 'سمك بالكريمة'],
    'Moussaka Vegi': ['High protein Meat moussaka with white rice', 'مسقعة اللحم عالية البروتين'],
    'Moussaka Meat': ['High protein Meat moussaka with white rice', 'مسقعة اللحم عالية البروتين'],
    'Laham Oriental': ['Oriental beef with nuts rice', 'لحم أوريانتل'],
    'Beef Oriental rice': ['Oriental beef with nuts rice', 'الأرز بالمكسرات'],
    'Chicken Caesar Salad': ['Caesar salad', 'سلطة السيزر'],
    'Chicken Fajita': ['Chicken Fajita sandwich served with oat bread', 'ساندوتش فاهيتا الدجاج بخبز الشوفان'],
    'Mached Potato(3)': ['Mashed potato', 'Mashed potatoes', 'بطاطس مهروسة'],
    'Potato Wedges': ['Potato wedges', 'بطاطس ويدجز'],
    'Oven Vegetables (3)': ['Sauteed vegetables', 'خضار سوتيه'],
}

# The website's "Repeat Update" export contains one row per package/size.
# Each menu meal must be split into the exact Tokyo recipe tabs it drives.
# ``protein`` reads column I, ``side`` reads column H and ``count`` uses the
# subscriber quantity itself (sandwich recipes are count-driven in Tokyo).
RAW_TOKYO_COMPONENT_MAP = {
    'chicken mushroom with white rice': [
        ('Chicken Mushroom', 'protein'),
    ],
    'chicken with mushrooms with sauteed vegetables': [
        ('Chicken Mushroom', 'protein'),
        ('Sautee Vegetables (4)', 'side'),
    ],
    'beef m&c with saffron rice': [
        ('Beef Amansi', 'protein'),
    ],
    'beef amansi with sauteed vegetables': [
        ('Beef Amansi', 'protein'),
        ('Sautee Vegetables (4)', 'side'),
    ],
    'chicken steak with mashed potatoes': [
        ('Chicken Steak Sauce', 'protein'),
        ('Chicken Steak Topping', 'protein'),
        ('Mached Potato(4)', 'side'),
    ],
    'chicken steak with sauteed vegetables': [
        ('Chicken Steak Sauce', 'protein'),
        ('Chicken Steak Topping', 'protein'),
        ('Sautee Vegetables (4)', 'side'),
    ],
    'navy bean with meat and white rice': [
        ('Lahm Fasooliya', 'protein'),
    ],
    'stroganoff pasta': [
        ('Stroganoff Beef', 'protein'),
        ('Stroganoff pasta', 'side'),
    ],
    'lemon shrimp with orzo and spinach': [
        ('Lemon Fish', 'protein'),
        ('Orzo Pasta', 'side'),
    ],
    'lemon fish with orzo and spinach': [
        ('Lemon Fish', 'protein'),
        ('Orzo Pasta', 'side'),
    ],
    'tikka chicken with buryani rice': [
        ('Chicken Tikka', 'protein'),
    ],
    'octa chicken poke bowl (spicy)': [
        ('Octa Poki Bowl', 'protein'),
    ],
    'chicken maqluba': [
        ('Chicken Makloba', 'protein'),
        ('Makloba Veggi', 'side'),
    ],
    'bbq chicken sandwich in ciabatta bread': [
        ('BBQ Chicken', 'count'),
    ],
}

RAW_REQUIRED_HEADERS = {
    'الاسم الإنجليزي', 'الاسم العربي', 'التصنيف',
    'مجموع الكارب', 'مجموع البروتين', 'الكمية',
}


def read_current_inputs(template_path):
    wb = load_workbook(template_path, data_only=False, keep_vba=True, read_only=False)
    ws = wb['All_Ingredients']
    days = {}
    for r in range(2, ws.max_row + 1):
        day_no = ws.cell(row=r, column=DAY_NO_COL).value
        sheet_name = ws.cell(row=r, column=SHEET_NAME_COL).value
        if not day_no or not sheet_name or sheet_name == 'Butchery':
            continue
        count = ws.cell(row=r, column=COUNT_COL).value
        grams = ws.cell(row=r, column=GRAMS_COL).value
        day_key = str(int(day_no))
        days.setdefault(day_key, {'dayNo': int(day_no), 'dayName': DAY_NAMES.get(int(day_no), f'يوم {day_no}'), 'meals': []})
        days[day_key]['meals'].append({
            'row': r,
            'sheetName': sheet_name,
            'count': count if isinstance(count, (int, float)) else 0,
            'grams': grams if isinstance(grams, (int, float)) else 0,
        })
    wb.close()
    return [days[k] for k in sorted(days.keys(), key=int)]


DAY_NAME_TO_NO = {v: k for k, v in DAY_NAMES.items()}


def _number(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _day_no_from_filename(filename):
    match = re.search(r'(20\d{2})[-_/](\d{1,2})[-_/](\d{1,2})', filename or '')
    if not match:
        raise ValueError('اسم ملف ابديت تكرار لازم يحتوي على التاريخ بصيغة YYYY-MM-DD لتحديد يوم توكيو')
    date_value = datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    # Python: Monday=0. Tokyo: Saturday=1 ... Thursday=6.
    day_no = {5: 1, 6: 2, 0: 3, 1: 4, 2: 5, 3: 6}.get(date_value.weekday())
    if not day_no:
        raise ValueError('ملف يوم الجمعة غير مدعوم في خطة توكيو الحالية')
    return day_no


def _find_repeat_update_sheet(wb):
    for ws in wb.worksheets:
        headers = {
            str(ws.cell(row=1, column=column).value or '').strip(): column
            for column in range(1, min(ws.max_column, 40) + 1)
        }
        if RAW_REQUIRED_HEADERS.issubset(headers):
            return ws, headers
    return None, None


def _read_repeat_update(wb, file_storage):
    ws, headers = _find_repeat_update_sheet(wb)
    if ws is None:
        return None
    upload_name = (
        getattr(file_storage, 'filename', '') or
        os.path.basename(getattr(file_storage, 'name', '') or '')
    )
    day_no = _day_no_from_filename(upload_name)
    totals = {}
    source_names = set()
    ignored_names = set()
    unmapped_production = set()
    source_rows = 0

    for row in range(2, ws.max_row + 1):
        english = str(ws.cell(row, headers['الاسم الإنجليزي']).value or '').strip()
        if not english:
            continue
        category = str(ws.cell(row, headers['التصنيف']).value or '').strip()
        quantity = _number(ws.cell(row, headers['الكمية']).value)
        if quantity <= 0:
            continue
        source_rows += 1
        normalized = re.sub(r'\s+', ' ', english).strip().lower()
        components = RAW_TOKYO_COMPONENT_MAP.get(normalized)
        is_production_meal = 'الوجبات الرئيسية' in category or 'لو كارب' in category
        if not components:
            if is_production_meal:
                unmapped_production.add(english)
            else:
                ignored_names.add(english)
            continue

        source_names.add(english)
        carb_grams = _number(ws.cell(row, headers['مجموع الكارب']).value)
        protein_grams = _number(ws.cell(row, headers['مجموع البروتين']).value)
        for target, weight_kind in components:
            count, grams = totals.get(target, (0.0, 0.0))
            if weight_kind == 'protein':
                row_grams = protein_grams
            elif weight_kind == 'side':
                row_grams = carb_grams
            else:
                row_grams = quantity
            totals[target] = (count + quantity, grams + row_grams)

    if unmapped_production:
        raise ValueError(
            'توجد وجبات إنتاج جديدة غير مربوطة بتوكيو: ' +
            ', '.join(sorted(unmapped_production))
        )
    if not totals:
        raise ValueError('لم يتم العثور على وجبات إنتاج قابلة لتحديث توكيو في ملف ابديت تكرار')

    report = {
        'kind': 'repeat_update',
        'sheet_name': ws.title,
        'source_rows': source_rows,
        'source_meals': len(source_names),
        'target_recipes': len(totals),
        'ignored_meals': len(ignored_names),
        'ignored_names': sorted(ignored_names),
    }
    return day_no, totals, report


def read_day_file_payload(file_storage):
    """بتقرا شيت 'Update' من ملف يوم واحد (زي Octa_Food_Sat_...xlsx) وترجع:
    (day_no, {اسم الصنف: (Total Count, Total Grams)}, input_report)
    - اسم اليوم بيتقرا من الخلية A6 (زي 'السبت') وبيتحول لرقم اليوم.
    - أو يحول ملف "ابديت تكرار" الخام إلى وصفات توكيو أولًا."""
    file_storage.seek(0)
    wb = load_workbook(file_storage, data_only=True, read_only=True)
    if 'Update' not in wb.sheetnames:
        raw_result = _read_repeat_update(wb, file_storage)
        wb.close()
        file_storage.seek(0)
        if raw_result:
            return raw_result
        raise ValueError("الملف لا يحتوي على شيت Update ولا أعمدة ابديت تكرار المطلوبة")
    ws = wb['Update']

    day_label = str(ws['A6'].value or '').strip()
    day_no = DAY_NAME_TO_NO.get(day_label)
    if not day_no:
        raise ValueError(f"مش عارف أحدد اليوم من الخلية A6 (لقيت: '{day_label}')")

    arabic_totals = {}
    for r in range(10, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name:
            continue
        name = str(name).strip()
        count = ws.cell(row=r, column=12).value   # L = Total Count
        grams = ws.cell(row=r, column=13).value   # M = Total Grams
        count = count if isinstance(count, (int, float)) else None
        grams = grams if isinstance(grams, (int, float)) else None
        if count is None and grams is None:
            continue
        if name.strip().lower() == 'grand total':
            continue
        old_count, old_grams = arabic_totals.get(name, (0, 0))
        arabic_totals[name] = (
            old_count + (count or 0),
            old_grams + (grams or 0),
        )

    # The operational Update pivot is Arabic-only. This helper sheet carries
    # the matching English product names, so recipe tab names can be matched
    # safely even when Arabic labels or the weekly menu order change.
    english_to_arabic = {}
    if "Don't Use just refresh" in wb.sheetnames:
        raw = wb["Don't Use just refresh"]
        for r in range(3, raw.max_row + 1):
            english = raw.cell(row=r, column=1).value
            arabic = raw.cell(row=r, column=2).value
            if english and arabic:
                english_to_arabic[str(english).strip()] = str(arabic).strip()

    meals = dict(arabic_totals)
    for english, arabic in english_to_arabic.items():
        if arabic in arabic_totals:
            meals[english] = arabic_totals[arabic]
    wb.close()
    file_storage.seek(0)
    return day_no, meals, {
        'kind': 'prepared_update',
        'sheet_name': 'Update',
        'source_meals': len(arabic_totals),
        'target_recipes': len(meals),
        'ignored_meals': 0,
    }


def read_day_file_meals(file_storage):
    day_no, meals, _ = read_day_file_payload(file_storage)
    return day_no, meals


def validate_raw_targets_for_day(template_path, day_no, meals_by_name):
    """Stop a raw export if its filename points to the wrong Tokyo day."""
    wb = load_workbook(template_path, data_only=False, keep_vba=True, read_only=True)
    ws = wb['All_Ingredients']
    day_sheets = set()
    for row in range(2, ws.max_row + 1):
        try:
            same_day = int(ws.cell(row, DAY_NO_COL).value) == int(day_no)
        except (TypeError, ValueError):
            same_day = False
        if same_day:
            name = str(ws.cell(row, SHEET_NAME_COL).value or '').strip()
            if name and name != 'Butchery':
                day_sheets.add(name)
    wb.close()
    wrong_day = sorted(set(meals_by_name).difference(day_sheets))
    if wrong_day:
        raise ValueError(
            f'تاريخ الملف يشير إلى {DAY_NAMES.get(day_no, day_no)} لكن الوصفات لا تطابق هذا اليوم: ' +
            ', '.join(wrong_day)
        )
    return len(set(meals_by_name).intersection(day_sheets))


def read_day_safety_fields(template_path, day_no):
    """Return the editable Safety inputs for exactly one production day.

    The row number is the stable identifier used by the UI and is validated
    again while merging, so a value can never be written to another day.
    """
    wb = load_workbook(template_path, data_only=False, keep_vba=True, read_only=True)
    ws = wb['All_Ingredients']
    fields = []
    for r in range(2, ws.max_row + 1):
        row_day = ws.cell(row=r, column=DAY_NO_COL).value
        try:
            same_day = int(row_day) == int(day_no)
        except (TypeError, ValueError):
            same_day = False
        if not same_day:
            continue
        sheet_name = str(ws.cell(row=r, column=SHEET_NAME_COL).value or '').strip()
        if not sheet_name or sheet_name == 'Butchery':
            continue
        safety = ws.cell(row=r, column=SAFETY_COL).value
        fields.append({
            'id': str(r),
            'row': r,
            'name': sheet_name,
            'value': float(safety) if isinstance(safety, (int, float)) else 0,
        })
    wb.close()
    return fields


def _normalize_meal_name(s):
    """بتشيل الإيموجي والمسافات الزيادة عشان المطابقة تنجح حتى لو فيه رمز
    حار 🌶️ أو مسافات مختلفة بين النسختين."""
    import re
    s = str(s or '')
    s = re.sub(r'[\U0001F300-\U0001FAFF\u2600-\u27BF]', '', s)  # إيموجي
    return re.sub(r'\s+', ' ', s).strip()


def merge_day_into_template(template_path, day_no, meals_by_name, out_path=None,
                            safety_overrides=None):
    """بتاخد قاموس {اسم الصنف: (count, grams)} من ملف يوم واحد، وتحدّث بيه
    صفوف نفس اليوم (AJ=day_no) بس في شيت All_Ingredients، بالمطابقة على
    عمود AQ (Meal name). بترجع (out_path, report) - الـreport بيوضح كل صنف
    اتطابق واتحدّث، وكل صنف في اليوم ده متلقاش ليه مطابقة (يفضل زي ما هو،
    من غير ما يتصفّر غلط)."""
    wb = load_workbook(template_path, data_only=False, keep_vba=True, read_only=False)
    ws = wb['All_Ingredients']

    norm_lookup = {_normalize_meal_name(k): v for k, v in meals_by_name.items()}
    clean_safety = {}
    for row, value in (safety_overrides or {}).items():
        try:
            row_no = int(row)
            number = float(value)
        except (TypeError, ValueError):
            raise ValueError(f'قيمة Safety غير صحيحة للصف {row}')
        if number < 0:
            raise ValueError(f'قيمة Safety لا يمكن أن تكون سالبة للصف {row_no}')
        clean_safety[row_no] = number

    matched, unmatched = [], []
    applied_safety = []
    for r in range(2, ws.max_row + 1):
        row_day = ws.cell(row=r, column=DAY_NO_COL).value
        if not row_day or int(row_day) != day_no:
            continue
        meal_name = ws.cell(row=r, column=43).value  # AQ
        sheet_name = str(ws.cell(row=r, column=SHEET_NAME_COL).value or '').strip()
        if not meal_name and not sheet_name:
            continue
        if r in clean_safety:
            ws.cell(row=r, column=SAFETY_COL, value=clean_safety[r])
            applied_safety.append({
                'row': r,
                'name': sheet_name or str(meal_name or '').strip(),
                'value': clean_safety[r],
            })
        # Safety controls the recipe formulas independently of count/grams.
        # Apply it even when a daily-file meal name did not match, otherwise
        # the generated recipe could silently keep an old Safety value.
        if sheet_name in wb.sheetnames:
            safety_value = ws.cell(row=r, column=SAFETY_COL).value
            if isinstance(safety_value, (int, float)):
                wb[sheet_name]['AB1'] = float(safety_value)
        key = str(meal_name or '').strip()
        norm_key = _normalize_meal_name(key)

        found = None
        matched_input = None
        candidates = [sheet_name, *SHEET_INPUT_ALIASES.get(sheet_name, [])]
        for candidate in candidates:
            normalized = _normalize_meal_name(candidate)
            if candidate in meals_by_name:
                found, matched_input = meals_by_name[candidate], candidate
                break
            if normalized in norm_lookup:
                found, matched_input = norm_lookup[normalized], candidate
                break

        if found is None and sheet_name:
            english_keys = [k for k in meals_by_name if any('a' <= ch.lower() <= 'z' for ch in str(k))]
            fuzzy = max(
                ((SequenceMatcher(None, sheet_name.lower(), str(candidate).lower()).ratio(), candidate)
                 for candidate in english_keys),
                default=(0, None),
            )
            if fuzzy[0] >= 0.62:
                matched_input = fuzzy[1]
                found = meals_by_name[matched_input]

        # Last-resort compatibility for older templates whose AQ mapping is
        # known to be correct. It intentionally comes after recipe-name match.
        if found is None:
            if key in meals_by_name:
                found, matched_input = meals_by_name[key], key
            elif norm_key in norm_lookup:
                found, matched_input = norm_lookup[norm_key], key
        if found is None:
            unmatched.append(sheet_name or key)
            continue
        count, grams = found
        if count is not None:
            ws.cell(row=r, column=COUNT_COL, value=float(count))
        if grams is not None:
            ws.cell(row=r, column=GRAMS_COL, value=float(grams))
        # نفس وظيفة ماكرو UpdateRecipeData: نقل مدخلات التشغيل إلى خلايا
        # التحكم داخل شيت الوصفة. تشغيل VBA غير متاح على خادم Linux، لذلك
        # ننقل القيم مباشرة من غير تغيير أي معادلة أو تنسيق في الوصفة.
        if sheet_name in wb.sheetnames:
            recipe = wb[sheet_name]
            if grams is not None:
                recipe['Z1'] = float(grams)
            if count is not None:
                recipe['AD1'] = float(count)
        matched.append({
            'row': r, 'name': sheet_name or key, 'input_name': matched_input,
            'count': count, 'grams': grams,
        })

    if out_path is None:
        out_path = tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False).name
    wb.save(out_path)
    wb.close()

    report = {
        'day_no': day_no,
        'day_name': DAY_NAMES.get(day_no, str(day_no)),
        'matched_count': len(matched),
        'unmatched_count': len(unmatched),
        'matched': matched,
        'unmatched': unmatched,  # أصناف في اليوم ده مالقتش ليها رقم في الملف المرفوع - اتسابت زي ما هي
        'safety_count': len(applied_safety),
        'safety': applied_safety,
    }
    return out_path, report


def write_updated_workbook(template_path, days_payload, out_path=None):
    """days_payload: نفس شكل read_current_inputs الناتج، بعد تعديل count/grams من المستخدم."""
    wb = load_workbook(template_path, data_only=False, keep_vba=True, read_only=False)
    ws = wb['All_Ingredients']

    for day in days_payload or []:
        for meal in day.get('meals') or []:
            row = meal.get('row')
            if not row:
                continue
            count = meal.get('count')
            grams = meal.get('grams')
            if count is not None:
                try:
                    ws.cell(row=row, column=COUNT_COL, value=float(count))
                except (TypeError, ValueError):
                    pass
            if grams is not None:
                try:
                    ws.cell(row=row, column=GRAMS_COL, value=float(grams))
                except (TypeError, ValueError):
                    pass

    if out_path is None:
        out_path = tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False).name
    wb.save(out_path)
    wb.close()
    return out_path
