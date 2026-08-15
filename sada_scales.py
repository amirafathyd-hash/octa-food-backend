import re
from copy import copy
from datetime import datetime
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import PatternFill


DAY_OPTIONS = {
    'السبت': (1, 'السبت'),
    'الأحد': (2, 'الأحد'),
    'الاحد': (2, 'الأحد'),
    'الإثنين': (3, 'الاثنين'),
    'الاثنين': (3, 'الاثنين'),
    'الثلاثاء': (4, 'الثلاثاء'),
    'الأربعاء': (5, 'الاربعاء'),
    'الاربعاء': (5, 'الاربعاء'),
    'الخميس': (6, 'خميس'),
    'خميس': (6, 'خميس'),
}

ARABIC_RE = re.compile(r'[\u0600-\u06FF]+(?:[\s\u0600-\u06FF]+)*')
PART_RE = re.compile(r'(\d+)\s*/\s*(\d+)')
STANDALONE_TOTAL_RE = re.compile(r'^total\b', re.I)
IGNORE_DETAIL_TERMS = {'قبل الطبخ', 'بعد الطبخ'}
STOP_WORDS = {
    'قبل', 'بعد', 'الطبخ', 'دفعه', 'دفعة', 'مع', 'بدون', 'الصوص', 'صوص',
    'الاضافي', 'اضافي', 'اضافى', 'بال', 'و',
}
GROUP_FILLS = [
    'EAF4FF',
    'FFF2CC',
    'E2F0D9',
    'FCE4D6',
    'EDE7F6',
    'DDEBF7',
    'F4CCCC',
    'E2F0CB',
]


def _cell(ws, row, col):
    value = ws.cell(row, col).value
    if isinstance(value, float):
        return round(value, 6)
    return value


def _number(value):
    try:
        n = float(value or 0)
    except (TypeError, ValueError):
        return None
    if abs(n) < 0.000001:
        return None
    return round(n, 3)


def _norm_text(value):
    text = str(value or '').strip()
    text = text.replace('إ', 'ا').replace('أ', 'ا').replace('آ', 'ا')
    text = text.replace('ة', 'ه').replace('ى', 'ي')
    text = text.replace('أمانى', 'امانسي').replace('اماني', 'امانسي')
    text = text.replace('بالجار', 'بالبخار').replace('باجار', 'بخار')
    text = text.replace('مهروسة', 'مهروسه')
    text = re.sub(r'[ـ،,()\[\]{}]+', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.lower().strip()


def _batch_from_any_text(value, rtl_template=False):
    match = PART_RE.search(str(value or ''))
    if not match:
        return ''
    left, right = match.group(1), match.group(2)
    # In Arabic RTL labels, "2/1" is visually "1 of 2".
    return f'{right}/{left}' if rtl_template else f'{left}/{right}'


def _strip_batch(value):
    text = str(value or '')
    text = re.sub(r'دفعة\s*\d+\s*/\s*\d+', ' ', text)
    text = re.sub(r'\d+\s*/\s*\d+', ' ', text)
    return re.sub(r'\s+', ' ', text).strip()


def _extract_parenthetical_detail(value):
    text = str(value or '')
    for match in re.findall(r'[\(\[]([^)\]]+)[\)\]]', text):
        detail = _norm_text(match)
        if detail and detail not in IGNORE_DETAIL_TERMS:
            return detail
    return ''


def _clean_base_text(value):
    text = str(value or '')
    text = re.sub(r'[\(\[](?:قبل|بعد)\s+الطبخ[\)\]]', ' ', text)
    text = re.sub(r'[\(\[][^\)\]]+[\)\]]', ' ', text)
    return _strip_batch(text)


def _clean_output_name(value):
    """Remove parenthetical scale qualifiers without changing the meal name."""
    text = str(value or '')
    text = re.sub(r'[\(\[][^\)\]]+[\)\]]', ' ', text)
    text = _strip_batch(text)
    text = re.sub(r'\s+([،,])', r'\1', text)
    return re.sub(r'\s+', ' ', text).strip(' -—')


def _display_name_with_batch(item_name, batch):
    name = str(item_name or '').strip()
    batch = str(batch or '').strip()
    if batch and not PART_RE.search(name):
        return f'{name} — دفعة {batch}'
    return name


def _arabic_part(value):
    text = str(value or '').strip()
    matches = ARABIC_RE.findall(text)
    if not matches:
        return ''
    return _norm_text(matches[-1])


def _name_tokens(value):
    return {
        token for token in re.split(r'\s+', _norm_text(value))
        if len(token) > 1 and token not in STOP_WORDS
    }


def _names_match(left, right):
    left_tokens = _name_tokens(left)
    right_tokens = _name_tokens(right)
    if not left_tokens or not right_tokens:
        return _norm_text(left) == _norm_text(right)
    overlap = left_tokens & right_tokens
    needed = min(len(left_tokens), len(right_tokens))
    return len(overlap) >= max(1, needed - 1)


def _arabic_name(ws):
    preferred_cells = ((2, 37), (62, 2), (63, 2), (2, 2), (25, 18))
    for row, col in preferred_cells:
        value = _cell(ws, row, col)
        if isinstance(value, str):
            arabic = _arabic_part(value)
            if arabic:
                return arabic
    return None


def _find_day_meals(wb, day_no):
    if 'All_Ingredients' not in wb.sheetnames:
        return []
    ws = wb['All_Ingredients']
    day_col = name_col = header_row = None
    for row in range(1, min(ws.max_row, 200) + 1):
        for col in range(1, min(ws.max_column, 80) + 1):
            value = str(ws.cell(row, col).value or '').strip()
            if value == 'Day No.':
                day_col = col
                header_row = row
            elif value == 'Sheet Name':
                name_col = col
        if day_col and name_col:
            break
    if not day_col or not name_col or not header_row:
        return []

    names = []
    for row in range(header_row + 1, min(ws.max_row, 500) + 1):
        try:
            same_day = int(ws.cell(row, day_col).value or 0) == int(day_no)
        except (TypeError, ValueError):
            same_day = False
        name = str(ws.cell(row, name_col).value or '').strip()
        if same_day and name and name in wb.sheetnames and name not in names:
            names.append(name)
    return names


def _find_nearest_banner(ws, before_row):
    for row in range(before_row - 1, max(1, before_row - 30), -1):
        value = ws.cell(row, 2).value
        if isinstance(value, str) and PART_RE.search(value):
            arabic = ws.cell(row - 1, 2).value
            return {
                'arabic': str(arabic or '').strip(),
                'english': value.strip(),
                'batch': _batch_from_any_text(value),
            }
    return None


def _total_rows(ws):
    rows = []
    part = 0
    for row in range(1, ws.max_row + 1):
        b = ws.cell(row, 2).value
        c = ws.cell(row, 3).value
        c_text = str(c or '').strip().lower()
        if c_text == 'total sauce':
            part += 1
            value_row = row + 1
            rows.append({
                'part': part,
                'simple': False,
                'banner': _find_nearest_banner(ws, row),
                'label': b or 'Protein',
                'protein': _number(ws.cell(value_row, 2).value),
                'total_sauce': _number(ws.cell(value_row, 3).value),
                'mix': _number(ws.cell(value_row, 5).value),
                'topping': _number(ws.cell(value_row, 7).value),
                'protein_mix': _number(ws.cell(value_row, 8).value),
            })
        elif (
            isinstance(b, str)
            and STANDALONE_TOTAL_RE.search(b.strip())
            and b.strip().lower() != 'total'
            and (c is None or str(c).strip() == '')
        ):
            part += 1
            value_row = row + 1
            rows.append({
                'part': part,
                'simple': True,
                'banner': _find_nearest_banner(ws, row),
                'label': b.strip(),
                'value': _number(ws.cell(value_row, 2).value),
            })
    return [r for r in rows if any(_number(v) for v in r.values() if isinstance(v, (int, float)))]


def _simple_batch_rows(ws):
    found = []
    for row in range(1, ws.max_row + 1):
        label = str(ws.cell(row, 2).value or '').strip().lower()
        match = re.match(r'^batch\s+(\d+)$', label)
        if not match:
            continue
        value = _number(ws.cell(row, 4).value)
        if value is not None:
            found.append({'part': int(match.group(1)), 'value': value})
    total = len(found)
    return [
        {'batch': f"{item['part']}/{total}", 'value': item['value']}
        for item in found
        if item.get('value') is not None
    ]


def _simple_ingredient_rows(ws):
    rows = []
    for row in range(1, ws.max_row + 1):
        ingredient = ws.cell(row, 2).value
        value = _number(ws.cell(row, 8).value)
        if not ingredient or value is None:
            continue
        detail = _arabic_part(ingredient)
        if not detail:
            continue
        rows.append({'detail': detail, 'value': value})
    return rows


def _component_for_template_name(name):
    compact = _norm_text(_strip_batch(name))
    if 'قبل الطبخ' in compact:
        return 'protein'
    if _extract_parenthetical_detail(name):
        return 'ingredient'
    if 'صوص الاضافي' in compact or 'صوص اضافي' in compact:
        return 'topping'
    if 'بدون صوص' in compact:
        return 'protein'
    if 'مع الصوص' in compact:
        return 'protein_mix'
    if 'صوص' in compact:
        return 'total_sauce'
    return 'simple'


def _base_for_template_name(name):
    text = _clean_base_text(name)
    for phrase in ('صوص الاضافي', 'صوص اضافي', 'بدون صوص', 'مع الصوص', 'صوص', 'قبل الطبخ', 'بعد الطبخ'):
        text = text.replace(phrase, ' ')
    return _norm_text(text)


def _candidate_key(base, batch, component, detail=''):
    return f'{_norm_text(base)}|{batch or ""}|{component}|{_norm_text(detail)}'


def _build_tokyo_value_index(wb, day_no):
    index = {}
    diagnostics = []
    for sheet_name in _find_day_meals(wb, day_no):
        ws = wb[sheet_name]
        sheet_ar = _arabic_name(ws) or sheet_name
        simple_rows = _simple_batch_rows(ws)
        for item in simple_rows:
            value = item.get('value')
            batch = item.get('batch') or ''
            if value is not None:
                index[_candidate_key(sheet_ar, batch, 'simple')] = value
                diagnostics.append({'name': sheet_ar, 'batch': batch, 'component': 'simple', 'value': value})
        if simple_rows:
            total_value = round(sum(item['value'] for item in simple_rows), 3)
            index[_candidate_key(sheet_ar, '', 'simple')] = total_value
            diagnostics.append({'name': sheet_ar, 'batch': '', 'component': 'simple', 'value': total_value})
        for item in _simple_ingredient_rows(ws):
            value = item.get('value')
            detail = item.get('detail') or ''
            if value is not None and detail:
                index[_candidate_key(sheet_ar, '', 'ingredient', detail)] = value
                diagnostics.append({'name': sheet_ar, 'batch': '', 'component': 'ingredient', 'detail': detail, 'value': value})
        for item in _total_rows(ws):
            banner = item.get('banner') or {}
            base_ar = _strip_batch(banner.get('arabic') or sheet_ar)
            batch = banner.get('batch') or (f"{item['part']}/1" if item.get('part') else '')
            if item.get('simple'):
                value = item.get('value')
                if value is not None:
                    index[_candidate_key(base_ar, batch, 'simple')] = value
                    index[_candidate_key(base_ar, '', 'simple')] = value
                    diagnostics.append({'name': base_ar, 'batch': batch, 'component': 'simple', 'value': value})
                continue
            for component in ('protein', 'total_sauce', 'topping', 'protein_mix'):
                value = item.get(component)
                if value is not None:
                    index[_candidate_key(base_ar, batch, component)] = value
                    diagnostics.append({'name': base_ar, 'batch': batch, 'component': component, 'value': value})
    return index, diagnostics


def _resolve_planned_value(item, value_index, diagnostics):
    # A bare meal name in the scales log represents the ready meal.  Tokyo
    # stores that cooking target in the protein column for composed recipes.
    # Prefer it, while retaining the simple value as a fallback for genuinely
    # simple recipes that do not have a protein total.
    if item.get('component') == 'simple':
        protein_item = dict(item)
        protein_item['component'] = 'protein'
        protein_item['key'] = _candidate_key(
            item.get('base') or '',
            item.get('batch') or '',
            'protein',
            '',
        )
        protein_value = _resolve_planned_value(protein_item, value_index, diagnostics)
        if protein_value is not None:
            return protein_value

    planned = value_index.get(item['key'])
    if planned is not None:
        return planned

    wanted_batch = str(item.get('batch') or '').strip()
    wanted_component = item.get('component') or ''
    wanted_detail = item.get('detail') or ''

    related = [
        candidate for candidate in diagnostics
        if candidate.get('component') == wanted_component
        and (not wanted_detail or _names_match(wanted_detail, candidate.get('detail') or ''))
        and _names_match(item.get('base') or '', candidate.get('name') or '')
    ]
    has_batch_specific = any(str(candidate.get('batch') or '').strip() for candidate in related)

    # A row that came from the scales with a batch number must never borrow a
    # no-batch total while Tokyo has batch-specific rows for the same component.
    if not (wanted_batch and has_batch_specific):
        planned = value_index.get(_candidate_key(item['base'], '', wanted_component, wanted_detail))
        if planned is not None:
            return planned

    exact_batch_matches = []
    same_part_matches = []
    loose_matches = []
    wanted_part = _batch_part(wanted_batch)
    for candidate in related:
        if candidate.get('component') != wanted_component:
            continue
        candidate_batch = str(candidate.get('batch') or '').strip()
        if wanted_batch:
            if candidate_batch == wanted_batch:
                exact_batch_matches.append(candidate)
            elif wanted_part is not None and _batch_part(candidate_batch) == wanted_part:
                same_part_matches.append(candidate)
            elif not candidate_batch and not has_batch_specific:
                loose_matches.append(candidate)
        else:
            loose_matches.append(candidate)
    if len(exact_batch_matches) == 1:
        return exact_batch_matches[0].get('value')
    if len(same_part_matches) == 1:
        return same_part_matches[0].get('value')
    if len(loose_matches) == 1:
        return loose_matches[0].get('value')
    return None


def _build_weight_index(entries):
    index = {}
    rows = {}
    for row in entries or []:
        if row.get('deleted'):
            continue
        item_name = row.get('item_name') or ''
        batch = str(row.get('batch_no') or '').strip()
        if not batch:
            batch = _batch_from_any_text(item_name, rtl_template=True)
        # Parenthetical labels such as (قبل الطبخ), (جزر) and (ذرة) are scale
        # qualifiers, not separate Sada output meals.  Remove them before
        # grouping so all readings return to the base meal and batch.
        clean_name = _clean_output_name(item_name)
        base = _base_for_template_name(clean_name)
        component = _component_for_template_name(clean_name)
        detail = ''
        try:
            weight = float(row.get('weight') or 0)
        except (TypeError, ValueError):
            continue
        if weight <= 0:
            continue
        key = _candidate_key(base, batch, component, detail)
        index[key] = round(index.get(key, 0) + weight, 3)
        rows[key] = {
            'key': key,
            'item_name': _display_name_with_batch(clean_name, batch),
            'base': base,
            'batch': batch,
            'component': component,
            'detail': detail,
            'weight': index[key],
        }
    return index, rows


def _batch_sort_value(batch):
    match = PART_RE.search(str(batch or ''))
    if not match:
        return (999, 999)
    return (int(match.group(2)), int(match.group(1)))


def _batch_part(batch):
    match = PART_RE.search(str(batch or ''))
    return int(match.group(1)) if match else None


def _component_sort_value(component):
    return {
        'protein': 1,
        'total_sauce': 2,
        'protein_mix': 3,
        'topping': 4,
        'simple': 5,
    }.get(component, 9)


def _copy_row_style(ws, source_row, target_row):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)


def _clear_row_values(ws, row):
    for col in range(1, ws.max_column + 1):
        ws.cell(row, col).value = None


def _apply_group_fill(ws, row, group_index):
    fill = PatternFill('solid', fgColor=GROUP_FILLS[group_index % len(GROUP_FILLS)])
    for col in range(1, 4):
        ws.cell(row, col).fill = fill


def _last_template_item_row(ws):
    for row in range(ws.max_row, 3, -1):
        if ws.cell(row, 1).value:
            return row
    return max(4, ws.max_row)


def _copy_sheet_to_single_workbook(template_path, template_sheet):
    wb = load_workbook(template_path)
    if template_sheet not in wb.sheetnames:
        raise ValueError(f'قالب صدى لا يحتوي على شيت اليوم: {template_sheet}')
    keep = wb[template_sheet]
    for ws in list(wb.worksheets):
        if ws.title != keep.title:
            wb.remove(ws)
    return wb, keep


def _set_cell_value(cell, value):
    cell.value = value
    if value is not None:
        cell.number_format = '#,##0.00'


def build_sada_scales_workbook(tokyo_path, template_path, day_name, output_date, weight_entries):
    day_key = str(day_name or '').strip()
    if day_key not in DAY_OPTIONS:
        raise ValueError('اختار يوم صحيح لموازين صدى')
    day_no, template_sheet = DAY_OPTIONS[day_key]

    tokyo_wb = load_workbook(tokyo_path, data_only=True, keep_vba=False)
    value_index, diagnostics = _build_tokyo_value_index(tokyo_wb, day_no)
    tokyo_wb.close()

    wb, ws = _copy_sheet_to_single_workbook(template_path, template_sheet)
    ws.cell(3, 2).value = 'الوزن للطبخ طوكيو'
    ws.cell(3, 3).value = 'الوزن الفعلي بعد الطبخ'
    if output_date:
        try:
            dt = datetime.strptime(output_date, '%Y-%m-%d')
            ws.title = output_date
            ws['A2'] = dt
            ws['A2'].number_format = 'yyyy-mm-dd'
        except ValueError:
            ws.title = str(output_date)[:31]
            ws['A2'] = output_date

    _weight_index, weight_rows = _build_weight_index(weight_entries)
    output_rows = list(weight_rows.values())
    output_rows.sort(key=lambda r: (_norm_text(r['base']), _batch_sort_value(r['batch']), _component_sort_value(r['component']), _norm_text(r['item_name'])))

    template_last_row = _last_template_item_row(ws)
    existing_slots = max(1, template_last_row - 3)
    needed_slots = max(1, len(output_rows))
    if needed_slots > existing_slots:
        ws.insert_rows(template_last_row + 1, amount=needed_slots - existing_slots)
    elif existing_slots > needed_slots:
        ws.delete_rows(4 + needed_slots, existing_slots - needed_slots)

    for row in range(4, 4 + needed_slots):
        if row != 4:
            _copy_row_style(ws, 4, row)
        _clear_row_values(ws, row)

    matched_tokyo = matched_actual = 0
    missing_tokyo = []
    group_ids = {}
    for offset, item in enumerate(output_rows):
        target_row = 4 + offset
        group_key = item.get('base') or item.get('item_name') or ''
        if group_key not in group_ids:
            group_ids[group_key] = len(group_ids)
        _apply_group_fill(ws, target_row, group_ids[group_key])
        ws.cell(target_row, 1).value = item['item_name']

        planned = _resolve_planned_value(item, value_index, diagnostics)
        if planned is not None:
            _set_cell_value(ws.cell(target_row, 2), planned)
            matched_tokyo += 1
        else:
            ws.cell(target_row, 2).value = None
            missing_tokyo.append(item['item_name'])

        _set_cell_value(ws.cell(target_row, 3), item['weight'])
        matched_actual += 1

    out = BytesIO()
    wb.save(out)
    wb.close()
    out.seek(0)
    return out, {
        'matched_tokyo': matched_tokyo,
        'matched_actual': matched_actual,
        'missing_tokyo_count': len(missing_tokyo),
        'missing_actual_count': 0,
        'added_actual_rows': len(output_rows),
    }
