"""
Octa Food - Excel Writer
=========================
بيبني ملف الإكسيل النهائي ("جمّع الملفات وطلّع الشيت النهائي") من صفوف جدول
daily_items في Supabase مباشرة — تاب منفصل لكل شهر، بالتنسيق المتفق عليه:

التاريخ | الصنف | الوحدة | كمية الأوردر | كمية الفاتورة | كمية المستلم |
الفرق | سعر الفاتورة | المخزون الحالي

كل صف في daily_items بيتبني تدريجيًا من 3 مصادر بيتحدثوا فوق بعض بالـ upsert
(item_date, item_key هما مفتاح المطابقة):
  - /api/upload-order    -> name_en, name_ar, unit, qty_box, qty_needed, current_inventory
  - /api/upload-invoice  -> invoice_qty, invoice_price, invoice_unit_label
  - /api/upload-received -> qty_received, rec_unit
"""
import tempfile
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_BLUE = "1F4E78"
HEADER_GREEN = "2E7D32"
FILL_GREEN_LIGHT = "E2F0D9"
FILL_ROW_ALT = "DCE6F1"
FILL_ROW_WHITE = "FFFFFF"
FILL_TOTAL = "FFE699"

HEADERS = ["التاريخ", "الصنف", "الوحدة", "كمية الأوردر", "كمية الفاتورة",
           "كمية المستلم", "الفرق", "سعر الفاتورة", "المخزون الحالي"]


def _month_key(item_date):
    try:
        dt = datetime.strptime(item_date, "%Y-%m-%d")
        return (dt.year, dt.month)
    except (TypeError, ValueError):
        return (0, 0)  # صفوف بدون تاريخ صالح تتجمع في تاب واحد آخر السجل


def build_workbook(rows, out_path=None):
    """rows: قايمة dicts من daily_items (Supabase) - كل صف ممكن يحتوي على أي
    مجموعة من الحقول حسب إيه اللي اترفع لحد دلوقتي (أوردر/فاتورة/مستلم).
    بيرجع مسار ملف الإكسيل المحفوظ (مش كائن Workbook) عشان يتبعت مباشرة بـ send_file."""
    wb = Workbook()
    wb.remove(wb.active)

    thick = Side(style='medium', color='1B5E20')

    by_month = defaultdict(list)
    for r in rows:
        by_month[_month_key(r.get('item_date'))].append(r)

    for (year, month), month_rows in sorted(by_month.items()):
        sheet_name = datetime(year, month, 1).strftime('%B %Y') if year else 'بدون تاريخ'
        ws = wb.create_sheet(sheet_name[:31])
        ws.sheet_view.rightToLeft = True

        for col, htext in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=htext)
            cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if col == 6:  # كمية المستلم
                cell.fill = PatternFill('solid', start_color=HEADER_GREEN)
                cell.border = Border(top=thick, bottom=thick, left=thick, right=thick)
            else:
                cell.fill = PatternFill('solid', start_color=HEADER_BLUE)
        ws.row_dimensions[1].height = 30

        month_rows.sort(key=lambda r: (r.get('item_date') or '', r.get('name_en') or r.get('item_key') or ''))

        r_idx = 2
        for row in month_rows:
            item_label = f"{row.get('name_en') or row.get('item_key') or ''} — {row.get('name_ar') or ''}".strip(' —')
            qty_received = row.get('qty_received')
            values = [
                row.get('item_date'),
                item_label,
                row.get('unit') or row.get('rec_unit') or '',
                row.get('qty_needed'),
                row.get('invoice_qty'),
                qty_received,
                None,  # الفرق - معادلة
                row.get('invoice_price'),
                row.get('current_inventory'),
            ]
            band = FILL_ROW_WHITE if (r_idx % 2 == 0) else FILL_ROW_ALT
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=r_idx, column=col, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if col == 6:
                    cell.fill = PatternFill('solid', start_color=FILL_GREEN_LIGHT if qty_received is not None else band)
                    cell.border = Border(top=thick, bottom=thick, left=thick, right=thick)
                else:
                    cell.fill = PatternFill('solid', start_color=band)
            g = ws.cell(row=r_idx, column=7, value=f"=E{r_idx}-F{r_idx}")
            g.font = Font(name="Arial", size=10)
            g.alignment = Alignment(horizontal='center')
            g.fill = PatternFill('solid', start_color=band)
            r_idx += 1

        last_data_row = r_idx - 1
        if last_data_row >= 2:
            total_row = r_idx
            ws.cell(row=total_row, column=2, value="الإجمالي").font = Font(bold=True, name="Arial")
            for col in (4, 5, 6, 7):
                letter = get_column_letter(col)
                c = ws.cell(row=total_row, column=col, value=f"=SUM({letter}2:{letter}{last_data_row})")
                c.font = Font(bold=True, name="Arial")
                c.fill = PatternFill('solid', start_color=FILL_TOTAL)
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=total_row, column=col).border = Border(top=Side(style='double'))

        widths = [12, 34, 9, 13, 13, 13, 10, 13, 14]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"

    if not wb.sheetnames:
        wb.create_sheet("لا توجد بيانات")

    if out_path is None:
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        out_path = tmp.name
        tmp.close()
    wb.save(out_path)
    return out_path
