import io
import html
import json
import os
import re
import tempfile
import uuid
import zipfile
from collections import OrderedDict, defaultdict
from copy import copy
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from decision_station import (
    build_output_workbook,
    compute_decision_tables,
    load_lookup,
    read_subscribers_invoice,
)
from tokyo_ordering import merge_day_into_template, read_day_file_payload
from tokyo_production_reports import build_tokyo_day_package


TOKYO_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'tokyo_ordering_template.xlsm')
ARCHIVE_DIR = os.path.join(os.path.dirname(__file__), 'data', 'day_operations_archive')
DAY_OPS_TEMPLATE_DIR = os.path.join(os.path.dirname(__file__), 'data', 'day_ops_templates')
DAY_OPS_TEMPLATES = {
    'breakfast': os.path.join(DAY_OPS_TEMPLATE_DIR, 'Tokyo_Breakfast.xlsm'),
    'dessert': os.path.join(DAY_OPS_TEMPLATE_DIR, 'Tokyo_Dessert_Ordering.xlsm'),
    'rice': os.path.join(DAY_OPS_TEMPLATE_DIR, 'Rice_Ordering_Sheet.xlsm'),
    'sauce': os.path.join(DAY_OPS_TEMPLATE_DIR, 'Tokyo_Sauce.xlsm'),
    'tokyo': os.path.join(DAY_OPS_TEMPLATE_DIR, 'tokyo_ordering_template.xlsm'),
}
PUBLIC_SITE_BASE = (
    os.environ.get('PUBLIC_SITE_BASE')
    or os.environ.get('APP_PUBLIC_URL')
    or 'https://octafood-system.com'
).rstrip('/')

TOKYO_STATION_SHEETS = OrderedDict([
    ('Hot Section', 'All_Ingredients'),
    ('Marination', 'Marination_Ordering'),
])
PRODUCE_CATEGORY_LABELS = {'خضروات', 'خضراوات', 'فاكهة', 'فاكهه', 'فواكه'}


PACKAGE_ORDER = ['تضخيم', 'تكميم لايت', 'جيم', 'سمارت دايت', 'غذاء العمل']
DAY_LABEL_TO_NO = {
    'السبت': 1,
    'الأحد': 2,
    'الاحد': 2,
    'الاثنين': 3,
    'الثلاثاء': 4,
    'الأربعاء': 5,
    'الاربعاء': 5,
    'الخميس': 6,
    'الجمعة': 7,
}

DAY_OPS_NAME_ALIASES = {
    'areeka': 'Areekah',
    'club sandwich with chicken': 'Chicken Club Sandwich',
    'club sandwich': 'Club Sandwich',
    'saffron cake': 'Saffron Cake',
    'orange cake': 'Orange Cake',
    'أومليت خضار': 'Omelette Vegetable and Mushroom',
    'اومليت خضار': 'Omelette Vegetable and Mushroom',
    'egg and oatmeal pie': 'Oatmeal Egg Pie',
    'oatmeal egg pie': 'Oatmeal Egg Pie',
    'turkey croissant': 'Croissant Turkey',
    'banoffee pie': 'Baboffee Pie',
    'banoffee': 'Baboffee Pie',
    'baboffee pie': 'Baboffee Pie',
    'san sebastian cheese cake': 'Sabestien',
    'san sebastian cheesecake': 'Sabestien',
    'san sebastian': 'Sabestien',
    'rosemary brownie bites': 'Brownie',
    'brownie bites': 'Brownie',
    'chocolate cake': 'Chocolate Profiterole',
    'chocolate profiterole': 'Chocolate Profiterole',
    'greek yogurt with oats and blueberries': 'Greek Yogurt with Oats',
    'greek yogurt with blueberries': 'Greek Yogurt with Oats',
    'tateema': "Ta'teema",
    'ta teema': "Ta'teema",
    'tuna sandwich': 'Tuna Sandwich',
    'egg and cheese sandwich': 'Egg and Cheese Sandwich',
    'croissant turkey': 'Croissant Turkey',
    'رز ابيض': 'رز أبيض',
    'white rice': 'رز أبيض',
    'plain rice': 'رز أبيض',
    'رز مندي': 'رز  المندى',
    'رز المندي': 'رز  المندى',
    'mandi rice': 'رز  المندى',
    'rice mandi': 'رز  المندى',
    'رز كابلي': 'رز الكابلى',
    'رز كابلى': 'رز الكابلى',
    'kabli rice': 'رز الكابلى',
    'kabli': 'رز الكابلى',
    'رز كبسة': 'رز كبسه',
    'kabsa rice': 'رز كبسه',
    'biryani rice': 'رز بريانى-رز زعفران',
    'saffron rice': 'رز بريانى-رز زعفران',
    'رز زعفران': 'رز بريانى-رز زعفران',
    'رز برياني': 'رز بريانى-رز زعفران',
    'رز بريانى': 'رز بريانى-رز زعفران',
    'sayadiya rice': 'رز صيادية',
    'رز صيادية': 'رز صيادية',
    'jollof rice': 'رز جولوف',
    'رز جولوف': 'رز جولوف',
    'zurbian rice': 'رز الزربيان',
    'رز زربيان': 'رز الزربيان',
    'saleeq rice': 'رز السليق',
    'رز سليق': 'رز السليق',
    'bukhari rice': 'رز  البخارى',
    'رز بخاري': 'رز  البخارى',
    'mexican rice': 'رز مكسيكي',
    'رز مكسيكي': 'رز مكسيكي',
    'rice with nuts': 'رز بالمكسرات',
    'رز بالمكسرات': 'رز بالمكسرات',
}

WORKER_LINKS = [
    ('رابط الخضار الساخن', 'veg-hot.html', 'استلام خضار القسم الساخن'),
    ('رابط السلطة', 'veg-salad.html', 'استلام خضار قسم السلطة'),
    ('رابط الصوص', 'sauce-receipt.html', 'استلام الصوص'),
]


def _absolute_worker_url(url):
    raw = str(url or '').strip()
    if not raw:
        return ''
    if raw.startswith(('http://', 'https://', 'mailto:', 'tel:')):
        return raw
    if raw.startswith('/'):
        return f'{PUBLIC_SITE_BASE}{raw}'
    return f'{PUBLIC_SITE_BASE}/{raw.lstrip("./")}'


def _static_worker_links():
    return [
        {
            'title': title,
            'url': _absolute_worker_url(href),
            'description': desc,
            'worker_name': 'تشغيل النظام',
            'username': '',
            'source': 'system',
        }
        for title, href, desc in WORKER_LINKS
    ]


def _public_worker_links(worker_links):
    public_links = []
    for link in worker_links:
        public_links.append({
            'title': link.get('title') or '',
            'url': _absolute_worker_url(link.get('url')),
            'description': link.get('description') or '',
            'worker_name': link.get('worker_name') or 'تشغيل النظام',
            'username': link.get('username') or '',
            'source': link.get('source') or 'system',
        })
    return public_links


def _load_worker_links():
    links = _static_worker_links()
    try:
        from db import execute_with_retry, get_client
        sb = get_client()
        res = execute_with_retry(
            sb.table('worker_link_assignments')
            .select('worker_name, username, task_title, target_url, active, created_at')
            .eq('active', True)
            .order('created_at', desc=True)
        )
        for row in res.data or []:
            links.append({
                'title': row.get('task_title') or 'مهمة عامل',
                'url': _absolute_worker_url(row.get('target_url')),
                'description': 'رابط عامل نشط',
                'worker_name': row.get('worker_name') or '',
                'username': row.get('username') or '',
                'source': 'worker',
            })
    except Exception:
        pass
    return links


def _num(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _norm_name(value):
    return re.sub(r'\s+', ' ', str(value or '').replace('\u00a0', ' ')).strip().lower()


def _clean_sheet_title(value, fallback='Sheet'):
    safe = ''.join(ch for ch in str(value or fallback) if ch not in r'[]:*?/\\').strip()
    return (safe or fallback)[:31]


def _station_key(category):
    text = category or ''
    if 'فطور' in text:
        return 'breakfast'
    if 'حلى' in text or 'حلويات' in text:
        return 'dessert'
    if 'سلطات' in text:
        return 'salads'
    if 'فواكه' in text or 'الإضافات' in text:
        return 'addons'
    if 'الوجبات الرئيسية' in text or 'لو كارب' in text:
        return 'main_kitchen'
    return 'unclassified'


def _station_title(key):
    return {
        'breakfast': 'محطة الفطار',
        'main_kitchen': 'المطبخ الرئيسي',
        'salads': 'محطة السلطات',
        'dessert': 'محطة الحلى',
        'addons': 'الفواكه والإضافات',
        'unclassified': 'غير مصنف',
    }.get(key, key)


def _style_sheet(ws, title=None):
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'
    thin = Side(style='thin', color='E5D6C8')
    border = Border(top=thin, right=thin, bottom=thin, left=thin)
    header_fill = PatternFill('solid', fgColor='7A2118')
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=12)
    data_font = Font(name='Arial', size=11, color='211713')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            cell.font = data_font
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 12
        for cell in ws[letter]:
            max_len = max(max_len, len(str(cell.value or '')) + 2)
        ws.column_dimensions[letter].width = min(max_len, 38)
    if title:
        ws.title = _clean_sheet_title(title)


def _copy_cell_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


def _copy_station_daily_sheet(src_ws, out_ws):
    out_ws.sheet_view.rightToLeft = False
    out_row = 1
    for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row, min_col=1, max_col=4):
        daily_weight = row[3].value if len(row) > 3 else None
        if isinstance(daily_weight, (int, float)) and not isinstance(daily_weight, bool) and daily_weight == 0:
            continue
        category = row[1].value if len(row) > 1 else None
        if category and str(category).strip() in PRODUCE_CATEGORY_LABELS:
            continue
        for cell in row:
            new_cell = out_ws.cell(row=out_row, column=cell.column, value=cell.value)
            _copy_cell_style(cell, new_cell)
        out_row += 1
    for col_letter in ['A', 'B', 'C', 'D']:
        if col_letter in src_ws.column_dimensions:
            out_ws.column_dimensions[col_letter].width = src_ws.column_dimensions[col_letter].width
    out_ws.freeze_panes = 'A2'


def _append_daily_summary(wb):
    rows_by_name = OrderedDict()
    for ws in wb.worksheets:
        if ws.title == 'Summary':
            continue
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4, values_only=True):
            name, category, unit, daily_weight = (list(row) + [None] * 4)[:4]
            if not name or str(name).strip().lower() == 'items':
                continue
            if not isinstance(daily_weight, (int, float)) or daily_weight == 0:
                continue
            key = str(name).strip()
            bucket = rows_by_name.setdefault(key, {
                'name': key,
                'category': category or '',
                'unit': unit or '',
                'daily_weight': 0.0,
            })
            bucket['daily_weight'] += daily_weight
            if not bucket['category'] and category:
                bucket['category'] = category
            if not bucket['unit'] and unit:
                bucket['unit'] = unit

    ws = wb.create_sheet('Summary')
    _append_rows(ws, ['ITEMS', 'Category', 'Unit', 'Daily Weight'], [
        [row['name'], row['category'], row['unit'], row['daily_weight']]
        for row in rows_by_name.values()
    ])
    _style_sheet(ws)
    ws.sheet_view.rightToLeft = False
    return len(rows_by_name)


def _build_daily_ordering_from_tokyo(tokyo_workbook_path):
    from openpyxl import load_workbook

    src_wb = load_workbook(tokyo_workbook_path, data_only=True, keep_vba=True)
    wb = Workbook()
    wb.remove(wb.active)
    try:
        for title, sheet_name in TOKYO_STATION_SHEETS.items():
            if sheet_name not in src_wb.sheetnames:
                continue
            out_ws = wb.create_sheet(title)
            _copy_station_daily_sheet(src_wb[sheet_name], out_ws)
        if not wb.sheetnames:
            return None, {'error': 'ملف توكيو لا يحتوي تابات Daily Ordering المطلوبة'}
        summary_count = _append_daily_summary(wb)
        return wb, {'stations': list(TOKYO_STATION_SHEETS.keys()), 'summary_items': summary_count}
    finally:
        src_wb.close()


def _build_weekly_order_from_tokyo(tokyo_workbook_path):
    from openpyxl import load_workbook

    src_wb = load_workbook(tokyo_workbook_path, data_only=True, keep_vba=True)
    rows = OrderedDict()
    try:
        for station_label, sheet_name in TOKYO_STATION_SHEETS.items():
            if sheet_name not in src_wb.sheetnames:
                continue
            ws = src_wb[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5, values_only=True):
                name, category, unit, _daily_weight, weekly = (list(row) + [None] * 5)[:5]
                if not name or str(name).strip().lower() == 'items':
                    continue
                if not isinstance(weekly, (int, float)) or weekly == 0:
                    continue
                key = str(name).strip()
                bucket = rows.setdefault(key, {
                    'name': key,
                    'category': category or '',
                    'unit': unit or '',
                    'stations': defaultdict(float),
                    'total': 0.0,
                })
                bucket['stations'][station_label] += weekly
                bucket['total'] += weekly
                if not bucket['category'] and category:
                    bucket['category'] = category
                if not bucket['unit'] and unit:
                    bucket['unit'] = unit
    finally:
        src_wb.close()

    if not rows:
        return None, {'error': 'ملف توكيو لا يحتوي Weekly Order قابل للتجميع'}

    wb = Workbook()
    ws = wb.active
    ws.title = 'Weekly Order'
    headers = ['ITEMS', 'Category', 'Unit', *TOKYO_STATION_SHEETS.keys(), 'Weekly Total']
    station_labels = list(TOKYO_STATION_SHEETS.keys())
    _append_rows(ws, headers, [
        [
            row['name'],
            row['category'],
            row['unit'],
            *[row['stations'].get(label, 0) for label in station_labels],
            row['total'],
        ]
        for row in rows.values()
    ])
    _style_sheet(ws)
    ws.sheet_view.rightToLeft = False
    return wb, {'stations': station_labels, 'items': len(rows)}


def _write_workbook_with_images(zf, wb, folder, workbook_name, image_prefix, day_label, full_report, report_key):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    workbook_arc = f'{folder}/{workbook_name}'
    zf.writestr(workbook_arc, buf.getvalue())
    full_report['files'].append(workbook_arc)
    try:
        from xlsx_to_images import add_workbook_images_to_zip
        image_date = datetime.now().strftime('%Y-%m-%d')
        saved = add_workbook_images_to_zip(
            zf,
            wb,
            image_date,
            folder=f'{folder}/صور PNG',
            prefix=image_prefix,
            day_num_override=_day_no(day_label),
        )
        full_report[report_key] = {
            **(full_report.get(report_key) or {}),
            'images_count': len(saved),
        }
        full_report['files'].append(f'{folder}/صور PNG')
    except Exception as exc:
        full_report[report_key] = {
            **(full_report.get(report_key) or {}),
            'images_error': str(exc),
        }
        zf.writestr(f'{folder}/ملاحظة - تعذر توليد الصور.txt', f'تعذر توليد الصور: {exc}')
        full_report['files'].append(f'{folder}/ملاحظة - تعذر توليد الصور.txt')


def _summaries(dont_use_rows):
    package_totals = OrderedDict()
    protein_totals = OrderedDict()
    station_totals = OrderedDict()
    for row in dont_use_rows:
        pkg = row.get('Final_Package') or 'غير محدد'
        protein = row.get('Protein') or row.get('الاسم الإنجليزي') or 'غير محدد'
        category = row.get('التصنيف') or ''
        station = _station_key(category)
        count = _num(row.get('Final_Count'))
        grams = _num(row.get('Grams'))

        package_totals.setdefault(pkg, {'count': 0.0, 'grams': 0.0, 'rows': 0})
        package_totals[pkg]['count'] += count
        package_totals[pkg]['grams'] += grams
        package_totals[pkg]['rows'] += 1

        protein_totals.setdefault(protein, {'category': category, 'count': 0.0, 'grams': 0.0, 'packages': defaultdict(float)})
        protein_totals[protein]['count'] += count
        protein_totals[protein]['grams'] += grams
        protein_totals[protein]['packages'][pkg] += count

        station_totals.setdefault(station, {'count': 0.0, 'grams': 0.0, 'items': set(), 'rows': 0})
        station_totals[station]['count'] += count
        station_totals[station]['grams'] += grams
        station_totals[station]['items'].add(protein)
        station_totals[station]['rows'] += 1
    return package_totals, protein_totals, station_totals


def _day_no(day_label):
    return DAY_LABEL_TO_NO.get(str(day_label or '').strip()) or 1


def _clean_side_name(value):
    text = str(value or '').strip()
    if not text or text in {'-', 'بدون', 'بدون صوص', 'None', 'nan'}:
        return ''
    return text


def _looks_like_rice(value):
    text = _norm_name(value)
    return bool(text) and ('رز' in text or 'rice' in text)


def _looks_like_sauce(value):
    text = _norm_name(value)
    sauce_terms = (
        'صوص', 'sauce', 'طحينة', 'tahina', 'tahini', 'دقوس',
        'زبادي', 'yogurt', 'كوكتيل', 'cocktail', 'coctail',
        'herbal', 'red sauce', 'cucumber',
    )
    return bool(text) and any(term in text for term in sauce_terms)


def _counts_by_station(dont_use_rows):
    counts = {
        'breakfast': defaultdict(float),
        'dessert': defaultdict(float),
        'rice': defaultdict(float),
        'sauce': defaultdict(float),
    }
    for row in dont_use_rows:
        category = str(row.get('التصنيف') or '')
        protein = row.get('Protein') or row.get('الاسم الإنجليزي') or ''
        if not protein:
            continue
        count = _num(row.get('Final_Count'))
        if count <= 0:
            continue
        if 'فطور' in category:
            counts['breakfast'][str(protein).strip()] += count
        elif 'حلى' in category or 'حلويات' in category:
            counts['dessert'][str(protein).strip()] += count
        side = _clean_side_name(row.get('Side') or row.get('الصوص / الجانب') or row.get('Final_Side'))
        if side and _looks_like_rice(side):
            counts['rice'][side] += count
        if side and _looks_like_sauce(side):
            counts['sauce'][side] += count
    return counts


def _sheet_lookup_candidates(name):
    key = _norm_name(name)
    compact = re.sub(r'[^a-z0-9\u0600-\u06ff]+', '', key)
    return key, compact


def _match_sheet_name(sheet_by_norm, wanted):
    key, compact = _sheet_lookup_candidates(wanted)
    if key in sheet_by_norm:
        return sheet_by_norm[key]
    if not compact:
        return None
    for sheet_key, sheet in sheet_by_norm.items():
        sheet_compact = re.sub(r'[^a-z0-9\u0600-\u06ff]+', '', sheet_key)
        if compact == sheet_compact or compact in sheet_compact or sheet_compact in compact:
            return sheet
    return None


def _template_sheet_lookup(template_path, include_ordering=False):
    from openpyxl import load_workbook

    if not os.path.exists(template_path):
        return {}
    wb = load_workbook(template_path, read_only=True, data_only=True, keep_vba=True)
    try:
        return {
            _norm_name(name): name
            for name in wb.sheetnames
            if include_ordering or name != 'Ordering'
        }
    finally:
        wb.close()


def _dessert_ordering_lookup(template_path):
    from openpyxl import load_workbook

    if not os.path.exists(template_path):
        return {}
    wb = load_workbook(template_path, read_only=True, data_only=True, keep_vba=True)
    try:
        ws = wb['Ordering']
        rows_by_norm = {}
        for row in range(3, ws.max_row + 1):
            sheet_name = ws[f'AA{row}'].value
            meal_name = ws[f'AF{row}'].value
            for value in (sheet_name, meal_name):
                key = _norm_name(value)
                if key:
                    rows_by_norm[key] = row
        return rows_by_norm
    finally:
        wb.close()


def _count_value(pair, prefer='count'):
    if isinstance(pair, (list, tuple)) and len(pair) >= 2:
        count, grams = pair[0], pair[1]
    else:
        count, grams = pair, 0
    count = _num(count)
    grams = _num(grams)
    if prefer == 'grams':
        return grams or count
    return count or grams


def _merge_tokyo_component_counts(counts, tokyo_totals):
    """Use the one uploaded day sheet as a second source for station templates.

    The subscriber invoice has breakfast/dessert categories, but rice/sauce
    often arrive only through the Tokyo recipe mapping. This merge is additive
    only where a station has no direct count yet, so existing station behavior
    stays stable.
    """
    added = defaultdict(list)
    if not tokyo_totals:
        return added

    station_lookups = {
        'breakfast': _template_sheet_lookup(DAY_OPS_TEMPLATES['breakfast']),
        'rice': _template_sheet_lookup(DAY_OPS_TEMPLATES['rice']),
        'sauce': _template_sheet_lookup(DAY_OPS_TEMPLATES['sauce']),
    }
    dessert_lookup = _dessert_ordering_lookup(DAY_OPS_TEMPLATES['dessert'])

    direct_station_has_data = {
        'breakfast': bool(counts.get('breakfast')),
        'dessert': bool(counts.get('dessert')),
    }

    for target, pair in (tokyo_totals or {}).items():
        name = str(target or '').strip()
        if not name:
            continue
        wanted = DAY_OPS_NAME_ALIASES.get(_norm_name(name), name)

        rice_sheet = _match_sheet_name(station_lookups['rice'], wanted)
        if rice_sheet:
            counts['rice'][name] += _count_value(pair)
            added['rice'].append(name)
            continue

        sauce_sheet = _match_sheet_name(station_lookups['sauce'], wanted)
        if sauce_sheet:
            counts['sauce'][name] += _count_value(pair)
            added['sauce'].append(name)
            continue

        if not direct_station_has_data['breakfast']:
            breakfast_sheet = _match_sheet_name(station_lookups['breakfast'], wanted)
            if breakfast_sheet:
                counts['breakfast'][name] += _count_value(pair)
                added['breakfast'].append(name)
                continue

        if not direct_station_has_data['dessert']:
            dessert_row = _match_sheet_name(dessert_lookup, wanted)
            if dessert_row:
                counts['dessert'][name] += _count_value(pair)
                added['dessert'].append(name)

    return added


def _match_template_sheets(template_path, counts, target_cell='V1'):
    if not os.path.exists(template_path) or not counts:
        return [], sorted(counts)
    sheet_by_norm = _template_sheet_lookup(template_path)
    edits = []
    unmatched = []
    for name, count in counts.items():
        wanted = DAY_OPS_NAME_ALIASES.get(_norm_name(name), name)
        sheet = _match_sheet_name(sheet_by_norm, wanted)
        if not sheet:
            unmatched.append(name)
            continue
        edits.append({'sheet': sheet, 'address': target_cell, 'value': count})
    return edits, unmatched


def _match_dessert_ordering_edits(template_path, counts):
    if not os.path.exists(template_path) or not counts:
        return [], sorted(counts)
    rows_by_norm = _dessert_ordering_lookup(template_path)
    edits = []
    unmatched = []
    for name, count in counts.items():
        wanted = DAY_OPS_NAME_ALIASES.get(_norm_name(name), name)
        row = _match_sheet_name(rows_by_norm, wanted)
        if not row:
            unmatched.append(name)
            continue
        edits.append({'sheet': 'Ordering', 'address': f'AG{row}', 'value': count})
    return edits, unmatched


def _export_visible_sheet_pdf(workbook_path, sheet_name='Ordering'):
    import shutil
    import subprocess
    from openpyxl import load_workbook

    wb = load_workbook(workbook_path)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f'الشيت {sheet_name} غير موجود في ملف المحطة')
        for name in wb.sheetnames:
            ws = wb[name]
            ws.sheet_state = 'visible' if name == sheet_name else 'hidden'
        ws = wb[sheet_name]
        ws.print_area = f'A1:D{ws.max_row}'
        ws.page_setup.orientation = 'portrait'
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_margins.left = 0.2
        ws.page_margins.right = 0.2
        ws.page_margins.top = 0.35
        ws.page_margins.bottom = 0.35
        out_xlsx = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
        wb.save(out_xlsx)
    finally:
        wb.close()

    out_dir = tempfile.mkdtemp(prefix='day_ops_pdf_')
    profile_dir = tempfile.mkdtemp(prefix='day_ops_lo_profile_')
    soffice = os.environ.get('SOFFICE_BIN') or shutil.which('soffice') or 'soffice'
    proc = subprocess.run([
        soffice,
        f'-env:UserInstallation=file://{profile_dir}',
        '--headless',
        '--convert-to',
        'pdf',
        '--outdir',
        out_dir,
        out_xlsx,
    ], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=120)
    if proc.returncode != 0:
        raise RuntimeError((proc.stderr or proc.stdout or 'LibreOffice PDF export failed').strip())
    return os.path.join(out_dir, os.path.splitext(os.path.basename(out_xlsx))[0] + '.pdf')


def _updated_template_workbook(template_path, edits, day_no=None):
    from openpyxl import load_workbook
    from breakfast_ordering import recalc_workbook_to_xlsx

    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    try:
        if day_no and 'Ordering' in wb.sheetnames:
            wb['Ordering']['R1'] = day_no
        for edit in edits:
            sheet_name = edit.get('sheet')
            address = edit.get('address') or 'V1'
            if sheet_name in wb.sheetnames:
                wb[sheet_name][address] = edit.get('value') or 0
        out_path = tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False).name
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = 'auto'
        wb.save(out_path)
    finally:
        wb.close()
    return recalc_workbook_to_xlsx(out_path)


def _generate_station_pdfs(day_label, dont_use_rows, file_storage=None):
    outputs = []
    image_sources = []
    reports = []
    counts = _counts_by_station(dont_use_rows)
    day_no = _day_no(day_label)
    tokyo_component_report = None

    if file_storage is not None:
        try:
            file_storage.seek(0)
            _tokyo_day_no, tokyo_totals, tokyo_input_report = read_day_file_payload(file_storage)
            tokyo_added = _merge_tokyo_component_counts(counts, tokyo_totals)
            tokyo_component_report = {
                'station': 'tokyo_component_source',
                'internal': True,
                'source_report': tokyo_input_report,
                'added_counts': {key: len(value) for key, value in tokyo_added.items()},
            }
        except Exception as exc:
            tokyo_component_report = {
                'station': 'tokyo_component_source',
                'internal': True,
                'error': str(exc),
            }
        finally:
            try:
                file_storage.seek(0)
            except Exception:
                pass

    breakfast_template = DAY_OPS_TEMPLATES['breakfast']
    breakfast_edits, breakfast_unmatched = _match_template_sheets(breakfast_template, counts['breakfast'])
    if breakfast_edits:
        try:
            from breakfast_ordering import export_breakfast_excel_with_edits, export_breakfast_pdf_with_edits
            xlsx_path, report = export_breakfast_excel_with_edits(
                breakfast_edits,
                template_path=breakfast_template,
            )
            pdf_path, pdf_report = export_breakfast_pdf_with_edits(
                breakfast_edits,
                day_no=day_no,
                template_path=breakfast_template,
            )
            outputs.append((f'Day{day_no}_Breakfast.pdf', pdf_path))
            image_sources.append(('Breakfast', xlsx_path))
            reports.append({
                'station': 'breakfast',
                'matched_count': len(breakfast_edits),
                'unmatched': breakfast_unmatched,
                **(report or {}),
                'pdf': pdf_report,
            })
        except Exception as exc:
            reports.append({
                'station': 'breakfast',
                'matched_count': len(breakfast_edits),
                'unmatched': breakfast_unmatched,
                'error': str(exc),
            })
    elif counts['breakfast']:
        reports.append({'station': 'breakfast', 'matched_count': 0, 'unmatched': breakfast_unmatched})

    dessert_template = DAY_OPS_TEMPLATES['dessert']
    dessert_edits, dessert_unmatched = _match_dessert_ordering_edits(dessert_template, counts['dessert'])
    if dessert_edits:
        try:
            from dessert_ordering import export_dessert_excel_with_edits, export_dessert_pdf_with_edits
            xlsx_path, report = export_dessert_excel_with_edits(
                dessert_edits,
                template_path=dessert_template,
            )
            pdf_path, pdf_report = export_dessert_pdf_with_edits(
                dessert_edits,
                template_path=dessert_template,
            )
            outputs.append((f'Day{day_no}_Dessert.pdf', pdf_path))
            image_sources.append(('Dessert', xlsx_path))
            reports.append({
                'station': 'dessert',
                'matched_count': len(dessert_edits),
                'unmatched': dessert_unmatched,
                **(report or {}),
                'pdf': pdf_report,
            })
        except Exception as exc:
            reports.append({
                'station': 'dessert',
                'matched_count': len(dessert_edits),
                'unmatched': dessert_unmatched,
                'error': str(exc),
            })
    elif counts['dessert']:
        reports.append({'station': 'dessert', 'matched_count': 0, 'unmatched': dessert_unmatched})

    rice_template = DAY_OPS_TEMPLATES['rice']
    rice_edits, rice_unmatched = _match_template_sheets(rice_template, counts['rice'], target_cell='Z1')
    if rice_edits:
        try:
            xlsx_path = _updated_template_workbook(rice_template, rice_edits, day_no=day_no)
            pdf_path = _export_visible_sheet_pdf(xlsx_path, 'Ordering')
            outputs.append((f'Day{day_no}_Rice.pdf', pdf_path))
            image_sources.append(('Rice', xlsx_path))
            reports.append({
                'station': 'rice',
                'matched_count': len(rice_edits),
                'unmatched': rice_unmatched,
            })
        except Exception as exc:
            reports.append({
                'station': 'rice',
                'matched_count': len(rice_edits),
                'unmatched': rice_unmatched,
                'error': str(exc),
            })
    elif counts['rice']:
        reports.append({'station': 'rice', 'matched_count': 0, 'unmatched': rice_unmatched})

    sauce_template = DAY_OPS_TEMPLATES['sauce']
    sauce_edits, sauce_unmatched = _match_template_sheets(sauce_template, counts['sauce'], target_cell='Q19')
    if sauce_edits:
        try:
            from sauce_ordering import export_sauce_excel_with_edits, export_sauce_pdf_with_edits
            xlsx_path, report = export_sauce_excel_with_edits(
                sauce_edits,
                template_path=sauce_template,
            )
            pdf_path, pdf_report = export_sauce_pdf_with_edits(
                sauce_edits,
                day_no=day_no,
                template_path=sauce_template,
            )
            outputs.append((f'Day{day_no}_Sauce.pdf', pdf_path))
            image_sources.append(('Sauce', xlsx_path))
            reports.append({
                'station': 'sauce',
                'matched_count': len(sauce_edits),
                'unmatched': sauce_unmatched,
                **(report or {}),
                'pdf': pdf_report,
            })
        except Exception as exc:
            reports.append({
                'station': 'sauce',
                'matched_count': len(sauce_edits),
                'unmatched': sauce_unmatched,
                'error': str(exc),
            })
    elif counts['sauce']:
        reports.append({'station': 'sauce', 'matched_count': 0, 'unmatched': sauce_unmatched})

    if tokyo_component_report:
        reports.insert(0, tokyo_component_report)

    return outputs, image_sources, reports


def _append_rows(ws, headers, rows):
    ws.append(headers)
    for row in rows:
        ws.append(row)


def _activate_url_cells(ws, column=4, start_row=2):
    for row_idx in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=column)
        url = _absolute_worker_url(cell.value)
        if not url:
            continue
        cell.value = url
        cell.hyperlink = url
        cell.font = Font(name='Arial', color='0563C1', underline='single', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _build_operations_workbook(day_label, dont_use_rows, pivot_rows, package_order, report, worker_links):
    package_totals, protein_totals, station_totals = _summaries(dont_use_rows)
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet('Dashboard')
    _append_rows(ws, ['المؤشر', 'القيمة'], [
        ['اليوم', day_label],
        ['صفوف فاتورة المشتركين', report.get('source_rows')],
        ['صفوف التشغيل المحسوبة', report.get('computed_rows')],
        ['صفوف Pivot', report.get('row_labels')],
        ['الباقات النهائية', '، '.join(package_order)],
        ['أصناف جديدة تم التعامل معها تلقائيا', len(report.get('inferred_items') or [])],
        ['باقات جديدة تم التعامل معها تلقائيا', len(report.get('inferred_packages') or [])],
    ])
    _style_sheet(ws)

    ws = wb.create_sheet('Packages')
    pkg_rows = []
    for pkg in package_order:
        total = package_totals.get(pkg, {'count': 0, 'grams': 0, 'rows': 0})
        pkg_rows.append([pkg, total['rows'], total['count'], total['grams']])
    for pkg, total in package_totals.items():
        if pkg not in package_order:
            pkg_rows.append([pkg, total['rows'], total['count'], total['grams']])
    _append_rows(ws, ['الباقة النهائية', 'عدد الصفوف', 'إجمالي العدد', 'إجمالي الجرام'], pkg_rows)
    _style_sheet(ws)

    ws = wb.create_sheet('Stations')
    station_rows = []
    preferred_stations = ['breakfast', 'main_kitchen', 'salads', 'dessert', 'addons', 'unclassified']
    for key in preferred_stations:
        total = station_totals.get(key)
        if not total:
            continue
        station_rows.append([
            _station_title(key), total['rows'], len(total['items']),
            total['count'], total['grams'],
        ])
    _append_rows(ws, ['المحطة', 'صفوف التشغيل', 'عدد الأصناف', 'إجمالي العدد', 'إجمالي الجرام'], station_rows)
    _style_sheet(ws)

    ws = wb.create_sheet('By Protein')
    protein_rows = []
    for protein, total in protein_totals.items():
        package_breakdown = ' | '.join(
            f'{pkg}: {round(qty, 2)}'
            for pkg, qty in total['packages'].items()
            if qty
        )
        protein_rows.append([
            protein, total['category'], _station_title(_station_key(total['category'])),
            total['count'], total['grams'], package_breakdown,
        ])
    _append_rows(ws, ['الصنف / البروتين', 'التصنيف', 'المحطة', 'إجمالي العدد', 'إجمالي الجرام', 'تفصيل الباقات'], protein_rows)
    _style_sheet(ws)

    ws = wb.create_sheet('Worker Links')
    _append_rows(ws, ['العامل / القسم', 'الحساب', 'المهمة / الوجهة', 'الرابط', 'الوصف', 'المصدر'], [
        [
            link.get('worker_name'),
            link.get('username'),
            link.get('title'),
            link.get('url'),
            link.get('description'),
            'رابط عامل' if link.get('source') == 'worker' else 'رابط نظام',
        ]
        for link in worker_links
    ])
    _style_sheet(ws)
    _activate_url_cells(ws)

    inferred = (report.get('inferred_items') or [])
    if inferred:
        ws = wb.create_sheet('New Items')
        _append_rows(ws, ['صنف جديد تم تمريره تلقائيا'], [[item] for item in inferred])
        _style_sheet(ws)

    return wb, {
        'packages': [
            {'name': row[0], 'rows': row[1], 'count': row[2], 'grams': row[3]}
            for row in pkg_rows
        ],
        'stations': [
            {'name': row[0], 'rows': row[1], 'items': row[2], 'count': row[3], 'grams': row[4]}
            for row in station_rows
        ],
        'top_items': [
            {'name': row[0], 'category': row[1], 'station': row[2], 'count': row[3], 'grams': row[4]}
            for row in sorted(protein_rows, key=lambda r: r[3] or 0, reverse=True)[:12]
        ],
    }


def _build_worker_links_workbook(day_label, worker_links):
    wb = Workbook()
    ws = wb.active
    ws.title = 'روابط العاملين'
    _append_rows(ws, ['العامل / القسم', 'الحساب', 'المهمة / الوجهة', 'الرابط', 'الوصف', 'المصدر'], [
        [
            link.get('worker_name'),
            link.get('username'),
            link.get('title'),
            link.get('url'),
            link.get('description'),
            'رابط عامل' if link.get('source') == 'worker' else 'رابط نظام',
        ]
        for link in worker_links
    ])
    _style_sheet(ws)
    _activate_url_cells(ws)

    info = wb.create_sheet('معلومات')
    _append_rows(info, ['البند', 'القيمة'], [
        ['اليوم', day_label],
        ['عدد الروابط', len(worker_links)],
        ['ملاحظة', 'يتم إضافة روابط العاملين النشطة من قاعدة البيانات إذا كان الاتصال متاحا، مع روابط التشغيل الأساسية.'],
    ])
    _style_sheet(info)
    return wb


def _build_station_outputs_workbook(day_label, dont_use_rows):
    grouped = OrderedDict()
    station_order = ['breakfast', 'main_kitchen', 'salads', 'dessert', 'addons', 'unclassified']
    for key in station_order:
        grouped[key] = OrderedDict()

    for row in dont_use_rows:
        station = _station_key(row.get('التصنيف') or '')
        protein = row.get('Protein') or row.get('الاسم الإنجليزي') or 'غير محدد'
        side = row.get('Side') or '-'
        pkg = row.get('Final_Package') or 'غير محدد'
        key = (protein, side, pkg)
        grouped.setdefault(station, OrderedDict())
        bucket = grouped[station].setdefault(key, {
            'protein': protein,
            'side': side,
            'package': pkg,
            'category': row.get('التصنيف') or '',
            'count': 0.0,
            'grams': 0.0,
            'source_rows': 0,
        })
        bucket['count'] += _num(row.get('Final_Count'))
        bucket['grams'] += _num(row.get('Grams'))
        bucket['source_rows'] += 1

    wb = Workbook()
    wb.remove(wb.active)
    for station in station_order:
        rows = list(grouped.get(station, {}).values())
        ws = wb.create_sheet(_station_title(station))
        _append_rows(ws, ['الصنف / البروتين', 'الصوص / الجانب', 'الباقة النهائية', 'التصنيف', 'العدد', 'الجرام', 'عدد صفوف المصدر'], [
            [
                item['protein'],
                item['side'],
                item['package'],
                item['category'],
                item['count'],
                item['grams'],
                item['source_rows'],
            ]
            for item in rows
        ])
        _style_sheet(ws)

    ws = wb.create_sheet('الصوصات')
    main_rows = [
        item
        for item in grouped.get('main_kitchen', {}).values()
        if item['count'] or item['grams']
    ]
    _append_rows(ws, ['وجبة مرتبطة بالصوص', 'الجانب', 'الباقة النهائية', 'عدد الوجبات', 'ملاحظة'], [
        [item['protein'], item['side'], item['package'], item['count'], 'يحتاج ربط وصفة الصوص من صفحة الصوصات لو مطلوب جرامات صوص دقيقة']
        for item in main_rows
    ])
    _style_sheet(ws)

    ws = wb.create_sheet('ملاحظات')
    _append_rows(ws, ['البند', 'التفاصيل'], [
        ['اليوم', day_label],
        ['مصدر الملف', 'فاتورة المشتركين فقط'],
        ['Weekly Purchasing', 'يحتاج ملفات/مصادر المشتريات الأصلية، لذلك لم يتم توليد ملف وهمي من فاتورة المشتركين'],
        ['استخراج البروتين والصوصات', 'تم تجهيز قائمة تشغيل من الفاتورة، أما الجرامات الدقيقة للصوص تعتمد على وصفات/قوالب الصوص الموجودة في صفحة المحطة'],
        ['Tokyo Production', 'لوحة توكيو تعتمد على ملف توكيو الرئيسي وفحص الماكرو، لذلك تظل كرابط تشغيل منفصل'],
    ])
    _style_sheet(ws)
    return wb


def _worker_links_html(day_label, worker_links):
    rows_parts = []
    for link in worker_links:
        url = html.escape(_absolute_worker_url(link.get('url')) or '#', quote=True)
        title = html.escape(str(link.get('title') or ''), quote=True)
        worker = html.escape(str(link.get('worker_name') or 'تشغيل النظام'), quote=True)
        desc = html.escape(str(link.get('description') or ''), quote=True)
        rows_parts.append(
            f'<a class="link" href="{url}" target="_blank" rel="noopener">'
            f'<b>{title}</b><span>{worker} · {desc}</span></a>'
        )
    rows = '\n'.join(rows_parts)
    return f"""<!doctype html>
<html lang="ar" dir="rtl">
<meta charset="utf-8">
<title>روابط تشغيل {day_label}</title>
<style>
body{{margin:0;background:#17100b;color:#fff3df;font-family:Tahoma,Arial,sans-serif;padding:28px}}
.wrap{{max-width:900px;margin:auto}}h1{{margin:0 0 8px;font-size:28px}}p{{color:#cdbca9}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:12px;margin-top:20px}}
.link{{display:block;text-decoration:none;color:#fff;background:#2a1b13;border:1px solid #6f3a22;border-radius:14px;padding:16px}}
.link b{{display:block;font-size:16px;margin-bottom:7px}}.link span{{color:#d8c2a9;font-size:13px;line-height:1.6}}
</style>
<div class="wrap"><h1>روابط تشغيل اليوم - {day_label}</h1><p>ملف سريع للمدير أو المشرف يفتح منه صفحات التشغيل المطلوبة.</p><div class="grid">{rows}</div></div>
"""


def process_day_operations(file_storage, day_label_override=None):
    lookup = load_lookup()
    worker_links = _load_worker_links()
    rows, detected_day = read_subscribers_invoice(file_storage)
    if not rows:
        raise ValueError('الملف المرفوع فاضي أو مالوش صفوف تشغيل')

    day_label = day_label_override or detected_day
    if not day_label:
        raise ValueError('مش قادر أحدد اليوم. اكتب اسم اليوم أو سمّي الملف بتاريخ YYYY-MM-DD.')

    dont_use_rows, pivot_rows, package_order, report = compute_decision_tables(rows, lookup)
    report['day_label'] = day_label

    decision_path = build_output_workbook(
        day_label, rows, dont_use_rows, pivot_rows, package_order, lookup,
        out_path=tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name,
    )
    ops_wb, ops_summary = _build_operations_workbook(day_label, dont_use_rows, pivot_rows, package_order, report, worker_links)
    station_outputs_wb = _build_station_outputs_workbook(day_label, dont_use_rows)
    worker_links_wb = _build_worker_links_workbook(day_label, worker_links)

    full_report = {
        **report,
        'generated_at': datetime.now().isoformat(timespec='seconds'),
        'files': [
            f'01_محطة_القرار/ملف اتخاذ القرار - {day_label}.xlsx',
            f'00_لوحة_تشغيل_اليوم/ملخص تشغيل اليوم - {day_label}.xlsx',
            f'06_ملخص_المحطات_من_الشيت/تجميع المحطات من الفاتورة - {day_label}.xlsx',
            f'07_روابط_العاملين/روابط العاملين - {day_label}.xlsx',
            f'07_روابط_العاملين/روابط تشغيل اليوم - {day_label}.html',
        ],
        'operations': ops_summary,
        'worker_links_count': len(worker_links),
        'worker_links': _public_worker_links(worker_links),
        'station_outputs_note': (
            'مخرجات المحطات يتم توليدها من قوالب Breakfast/Dessert/Rice/Sauce عند وجود أسماء '
            'قابلة للمطابقة من شيت اليوم، وأي صنف غير مطابق يظهر في manifest للمراجعة.'
        ),
    }

    station_pdf_outputs, station_image_sources, station_pdf_reports = _generate_station_pdfs(
        day_label,
        dont_use_rows,
        file_storage=file_storage,
    )
    full_report['station_pdfs'] = [
        item for item in station_pdf_reports
        if item.get('station') != 'tokyo_component_source' and not item.get('internal')
    ]
    internal_station_reports = [
        item for item in station_pdf_reports
        if item.get('station') == 'tokyo_component_source' or item.get('internal')
    ]
    if internal_station_reports:
        full_report['station_pdf_source'] = internal_station_reports[0]
    for filename, _path in station_pdf_outputs:
        full_report['files'].append(f'02_PDF_المحطات_الجاهزة/{filename}')

    tokyo_path = None
    tokyo_report = None
    tokyo_error = None
    tokyo_zip_path = None
    tokyo_template_path = DAY_OPS_TEMPLATES['tokyo'] if os.path.exists(DAY_OPS_TEMPLATES['tokyo']) else TOKYO_TEMPLATE_PATH
    if os.path.exists(tokyo_template_path):
        try:
            file_storage.seek(0)
            tokyo_zip_path, tokyo_path, tokyo_report = build_tokyo_day_package(
                tokyo_template_path,
                file_storage,
            )
        except Exception as exc:
            tokyo_error = str(exc)
        finally:
            file_storage.seek(0)
    else:
        tokyo_error = 'ملف قالب توكيو غير موجود على السيرفر'

    full_report['tokyo'] = tokyo_report or {'error': tokyo_error}
    if tokyo_report and tokyo_report.get('files'):
        for name in reversed(tokyo_report['files']):
            full_report['files'].insert(1, f'03_توكيو_الإنتاج/{name}')
    elif tokyo_path:
        full_report['files'].insert(1, f'03_توكيو_الإنتاج/شيت توكيو المحدث - {day_label}.xlsm')

    daily_wb = None
    weekly_wb = None
    if tokyo_path:
        try:
            daily_wb, daily_report = _build_daily_ordering_from_tokyo(tokyo_path)
            full_report['daily_ordering'] = daily_report
        except Exception as exc:
            full_report['daily_ordering'] = {'error': str(exc)}
        try:
            weekly_wb, weekly_report = _build_weekly_order_from_tokyo(tokyo_path)
            full_report['weekly_order'] = weekly_report
        except Exception as exc:
            full_report['weekly_order'] = {'error': str(exc)}
    else:
        full_report['daily_ordering'] = {'error': 'Daily Ordering يحتاج شيت توكيو محدث؛ لم يتم توليد توكيو من الملف المرفوع'}
        full_report['weekly_order'] = {'error': 'Weekly Order يحتاج شيت توكيو محدث؛ لم يتم توليد توكيو من الملف المرفوع'}

    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        ops_buf = io.BytesIO()
        ops_wb.save(ops_buf)
        ops_buf.seek(0)
        zf.writestr(f'00_لوحة_تشغيل_اليوم/ملخص تشغيل اليوم - {day_label}.xlsx', ops_buf.getvalue())

        with open(decision_path, 'rb') as fh:
            zf.writestr(f'01_محطة_القرار/ملف اتخاذ القرار - {day_label}.xlsx', fh.read())

        station_buf = io.BytesIO()
        station_outputs_wb.save(station_buf)
        station_buf.seek(0)
        zf.writestr(f'06_ملخص_المحطات_من_الشيت/تجميع المحطات من الفاتورة - {day_label}.xlsx', station_buf.getvalue())

        links_buf = io.BytesIO()
        worker_links_wb.save(links_buf)
        links_buf.seek(0)
        zf.writestr(f'07_روابط_العاملين/روابط العاملين - {day_label}.xlsx', links_buf.getvalue())
        zf.writestr(
            f'07_روابط_العاملين/روابط تشغيل اليوم - {day_label}.html',
            _worker_links_html(day_label, worker_links).encode('utf-8'),
        )

        for filename, path in station_pdf_outputs:
            with open(path, 'rb') as fh:
                zf.writestr(f'02_PDF_المحطات_الجاهزة/{filename}', fh.read())
        if tokyo_zip_path and os.path.exists(tokyo_zip_path):
            with zipfile.ZipFile(tokyo_zip_path, 'r') as tokyo_zip:
                for member in tokyo_zip.infolist():
                    if member.is_dir():
                        continue
                    zf.writestr(f'03_توكيو_الإنتاج/{member.filename}', tokyo_zip.read(member.filename))
        elif tokyo_path:
            with open(tokyo_path, 'rb') as fh:
                zf.writestr(f'03_توكيو_الإنتاج/شيت توكيو المحدث - {day_label}.xlsm', fh.read())
        if daily_wb:
            _write_workbook_with_images(
                zf,
                daily_wb,
                f'04_Daily_Ordering_طلبات_اليوم',
                f'Daily_Ordering_{day_label}.xlsx',
                'DailyOrdering_',
                day_label,
                full_report,
                'daily_ordering',
            )
        if weekly_wb:
            _write_workbook_with_images(
                zf,
                weekly_wb,
                f'05_Weekly_Order_طلب_الأسبوع',
                f'Weekly_Order_{day_label}.xlsx',
                'WeeklyOrder_',
                day_label,
                full_report,
                'weekly_order',
            )
        full_report['files'].append('manifest.json')
        zf.writestr('manifest.json', json.dumps(full_report, ensure_ascii=False, indent=2).encode('utf-8'))
    zip_buf.seek(0)
    return zip_buf, full_report


def _archive_id(day_label):
    stamp = datetime.now().strftime('%Y%m%d-%H%M%S')
    safe_day = re.sub(r'[^\w\u0600-\u06FF-]+', '-', str(day_label or 'day')).strip('-')[:40]
    return f'{stamp}-{safe_day}-{uuid.uuid4().hex[:6]}'


def _ensure_archive_dir():
    os.makedirs(ARCHIVE_DIR, exist_ok=True)
    return ARCHIVE_DIR


def save_day_operations_archive(file_storage, day_label_override=None):
    zip_buf, report = process_day_operations(file_storage, day_label_override=day_label_override)
    archive_id = _archive_id(report.get('day_label'))
    archive_dir = _ensure_archive_dir()
    zip_name = f'{archive_id}.zip'
    zip_path = os.path.join(archive_dir, zip_name)
    report = {
        **report,
        'archive_id': archive_id,
        'saved_at': datetime.now().isoformat(timespec='seconds'),
        'zip_name': zip_name,
    }
    with open(zip_path, 'wb') as fh:
        fh.write(zip_buf.getvalue())
    with open(os.path.join(archive_dir, f'{archive_id}.json'), 'w', encoding='utf-8') as fh:
        json.dump(report, fh, ensure_ascii=False, indent=2)
    return report


def list_day_operations_archives():
    archive_dir = _ensure_archive_dir()
    items = []
    for name in os.listdir(archive_dir):
        if not name.endswith('.json'):
            continue
        try:
            with open(os.path.join(archive_dir, name), encoding='utf-8') as fh:
                report = json.load(fh)
        except Exception:
            continue
        archive_id = report.get('archive_id') or name[:-5]
        zip_path = os.path.join(archive_dir, report.get('zip_name') or f'{archive_id}.zip')
        items.append({
            'archive_id': archive_id,
            'day_label': report.get('day_label'),
            'saved_at': report.get('saved_at') or report.get('generated_at'),
            'source_rows': report.get('source_rows'),
            'computed_rows': report.get('computed_rows'),
            'row_labels': report.get('row_labels'),
            'files': report.get('files') or [],
            'worker_links_count': report.get('worker_links_count') or 0,
            'tokyo': report.get('tokyo') or {},
            'size': os.path.getsize(zip_path) if os.path.exists(zip_path) else 0,
        })
    items.sort(key=lambda item: item.get('saved_at') or '', reverse=True)
    return items


def get_day_operations_archive_path(archive_id):
    safe = re.sub(r'[^A-Za-z0-9_\-\u0600-\u06FF]+', '', str(archive_id or ''))
    path = os.path.join(_ensure_archive_dir(), f'{safe}.zip')
    if not safe or not os.path.exists(path):
        raise FileNotFoundError('الأرشيف غير موجود')
    return path
