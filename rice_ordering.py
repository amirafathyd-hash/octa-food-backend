"""Rice production engine driven by the workbook's own day/reference table.

The daily file supplies grams.  ``Ordering!AA:AD`` decides which rice sheet
belongs to each day and ``AC`` decides which source gram cells feed it.  No
day-specific meal names or worksheet positions are hardcoded in Python.
"""
import io
import os
import re
import shutil
import subprocess
import tempfile
import zipfile
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from rice_storage import RICE_TEMPLATE_PATH
from tokyo_ordering import read_day_file_payload


DAY_NAMES = {
    1: 'السبت', 2: 'الأحد', 3: 'الاثنين', 4: 'الثلاثاء',
    5: 'الأربعاء', 6: 'الخميس', 7: 'الجمعة',
}
SOURCE_COLUMN_MAP = {'AS': 'AQ', 'AX': 'AV'}
ARABIC_FONT = 'IBM Plex Sans Arabic'
LATIN_FONT = 'Noto Sans'
APP_DIR = os.path.dirname(os.path.abspath(__file__))
BUNDLED_FONTS_DIRS = (
    os.path.join(APP_DIR, 'fonts'),
    os.path.join(APP_DIR, 'data', 'fonts'),
)


def _number(value):
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value or '').replace(',', '').strip())
    except (TypeError, ValueError):
        return 0.0


def _name_key(value):
    text = str(value or '').strip().casefold()
    text = re.sub(r'[\U0001F300-\U0001FAFF\u2600-\u27BF]', '', text)
    text = re.sub(r'[\u064b-\u065f\u0670]', '', text)
    text = text.replace('ـ', '').replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا')
    text = text.replace('ى', 'ي').replace('ة', 'ه')
    return re.sub(r'[^a-z0-9\u0600-\u06ff]+', ' ', text).strip()


def _mapping_rows(wb, day_no=None):
    if 'Ordering' not in wb.sheetnames:
        raise ValueError('شيت الأرز الأساسي لا يحتوي على Ordering')
    ws = wb['Ordering']
    mappings = []
    for row in range(3, ws.max_row + 1):
        # Worksheet names may intentionally start/end with spaces.  Keep the
        # exact value from the approved mapping table so these recipes are not
        # silently dropped (for example the Day 3 saffron rice sheet).
        sheet_name = str(ws[f'AA{row}'].value or '')
        mapped_day = int(_number(ws[f'AB{row}'].value))
        if not sheet_name or not mapped_day or sheet_name not in wb.sheetnames:
            continue
        if day_no is not None and mapped_day != int(day_no):
            continue
        mappings.append({
            'row': row,
            'sheet': sheet_name,
            'day_no': mapped_day,
            'formula': str(ws[f'AC{row}'].value or ''),
            # Safety is never taken from the master workbook.  The selected
            # day's value is supplied explicitly by the dashboard per recipe.
            'safety_grams': 0.0,
        })
    return mappings


def _meal_grams_lookup(meals):
    lookup = {}
    for name, pair in (meals or {}).items():
        if isinstance(pair, (list, tuple)) and len(pair) >= 2:
            grams = _number(pair[1])
        else:
            grams = _number(pair)
        key = _name_key(name)
        if key and key not in lookup:
            lookup[key] = grams
    return lookup


def _compute_day_grams(wb, day_no, meals):
    ordering = wb['Ordering']
    lookup = _meal_grams_lookup(meals)
    results = []
    missing = []
    for mapping in _mapping_rows(wb, day_no):
        total = 0.0
        sources = []
        counted_source_names = set()
        references = re.findall(r'\b(AS|AX)(\d+)\b', mapping['formula'], flags=re.I)
        for value_column, row_text in references:
            value_column = value_column.upper()
            source_row = int(row_text)
            label_column = SOURCE_COLUMN_MAP[value_column]
            source_name = str(ordering[f'{label_column}{source_row}'].value or '').strip()
            source_key = _name_key(source_name)
            grams = lookup.get(source_key)
            found = grams is not None
            grams = _number(grams)
            # The daily Update parser already aggregates duplicate meal names.
            # Some approved formulas reference two rows carrying the same meal
            # name, so counting by reference would double the daily grams.
            if source_key not in counted_source_names:
                total += grams
                counted_source_names.add(source_key)
            sources.append({'name': source_name, 'grams': grams, 'found': found})
            if source_name and source_name != '-' and not found:
                missing.append(source_name)
            ordering[f'{value_column}{source_row}'] = grams
        results.append({**mapping, 'input_grams': round(total, 3), 'sources': sources})
    return results, sorted(set(missing), key=_name_key)


def _soffice_bin():
    return os.environ.get('SOFFICE_BIN') or shutil.which('soffice') or 'soffice'


def _soffice_env():
    env = os.environ.copy()
    font_dirs = [path for path in BUNDLED_FONTS_DIRS if os.path.isdir(path)]
    if font_dirs:
        existing = str(env.get('SAL_FONTPATH') or '').strip()
        env['SAL_FONTPATH'] = os.pathsep.join(font_dirs + ([existing] if existing else []))
    return env


def _recalculate_to_xlsx(source_path):
    out_dir = tempfile.mkdtemp(prefix='rice_recalc_')
    profile_dir = tempfile.mkdtemp(prefix='rice_lo_profile_')
    proc = subprocess.run([
        _soffice_bin(), f'-env:UserInstallation=file://{profile_dir}',
        '--headless', '--convert-to', 'xlsx', '--outdir', out_dir, source_path,
    ], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=150, env=_soffice_env())
    if proc.returncode != 0:
        raise RuntimeError((proc.stderr or proc.stdout or 'LibreOffice failed').strip())
    result = os.path.join(out_dir, os.path.splitext(os.path.basename(source_path))[0] + '.xlsx')
    if not os.path.exists(result):
        raise RuntimeError('تعذر إعادة حساب ملف الأرز')
    return result


def _export_to_pdf(workbook_path):
    out_dir = tempfile.mkdtemp(prefix='rice_pdf_')
    profile_dir = tempfile.mkdtemp(prefix='rice_lo_pdf_profile_')
    proc = subprocess.run([
        _soffice_bin(), f'-env:UserInstallation=file://{profile_dir}',
        '--headless', '--convert-to', 'pdf', '--outdir', out_dir, workbook_path,
    ], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=150, env=_soffice_env())
    if proc.returncode != 0:
        raise RuntimeError((proc.stderr or proc.stdout or 'LibreOffice PDF export failed').strip())
    result = os.path.join(out_dir, os.path.splitext(os.path.basename(workbook_path))[0] + '.pdf')
    if not os.path.exists(result):
        raise RuntimeError('تعذر إنشاء PDF الأرز')
    return result


def _write_inputs(day_no, inputs, template_path=RICE_TEMPLATE_PATH):
    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    try:
        if 'Ordering' not in wb.sheetnames:
            raise ValueError('شيت الأرز الأساسي لا يحتوي على Ordering')
        wb['Ordering']['R1'] = int(day_no)
        values = {str(item.get('sheet') or ''): max(0.0, _number(item.get('input_grams'))) for item in inputs}
        safety_values = {str(item.get('sheet') or ''): max(0.0, _number(item.get('safety_grams'))) for item in inputs}
        for mapping in _mapping_rows(wb, day_no):
            wb[mapping['sheet']]['Z1'] = values.get(mapping['sheet'], 0.0)
            wb['Ordering'][f'AD{mapping["row"]}'] = safety_values.get(mapping['sheet'], 0.0)
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = 'auto'
        output = tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False).name
        wb.save(output)
    finally:
        wb.close()
    return output


def _batch_rows(ws):
    rows = []
    for row in range(10, min(ws.max_row, 30) + 1):
        label = str(ws[f'B{row}'].value or '').strip()
        if not label.casefold().startswith('batch'):
            continue
        factor = _number(ws[f'C{row}'].value)
        final_grams = _number(ws[f'D{row}'].value)
        if factor > 0 and final_grams > 0:
            rows.append({'label': label, 'factor': factor, 'final_grams': final_grams})
    return rows


def _ingredient_rows(ws):
    rows = []
    started = False
    for row in range(5, min(ws.max_row, 45) + 1):
        name = ws[f'B{row}'].value
        unit = ws[f'C{row}'].value
        if not name or not unit:
            if started:
                break
            continue
        started = True
        rows.append({
            'name': str(name).strip(),
            'unit': str(unit).strip(),
            'base': _number(ws[f'D{row}'].value),
            'scale': _number(ws[f'F{row}'].value) or 1.0,
        })
    return rows


def _clean_arabic_title(sheet_name):
    return re.sub(r'\s*\(\d+\)\s*$', '', str(sheet_name or '')).strip()


def _english_title(value, fallback):
    text = str(value or '').strip()
    parts = re.split(r'\s[-–—]\s', text, maxsplit=1)
    return (parts[0] if parts else text).strip() or fallback


def _build_pdf_source(calculated_path, day_no):
    source = load_workbook(calculated_path, data_only=True, read_only=False)
    report = Workbook()
    report.remove(report.active)
    brown = PatternFill('solid', fgColor='93440B')
    peach = PatternFill('solid', fgColor='F5C9AA')
    green = PatternFill('solid', fgColor='00B050')
    white = PatternFill('solid', fgColor='FFFFFF')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    pages = []
    try:
        mappings = _mapping_rows(source, day_no)
        for mapping in mappings:
            recipe = source[mapping['sheet']]
            batches = _batch_rows(recipe)
            ingredients = _ingredient_rows(recipe)
            for batch_index, batch in enumerate(batches, 1):
                pages.append((mapping, recipe, ingredients, batch, batch_index, len(batches)))

        total_pages = len(pages)
        for page_index, (mapping, recipe, ingredients, batch, batch_index, batch_count) in enumerate(pages, 1):
            safe_name = f'{mapping["day_no"]}-{mapping["row"]}-{batch_index}'[:31]
            ws = report.create_sheet(safe_name)
            ws.sheet_view.showGridLines = False
            arabic_title = _clean_arabic_title(mapping['sheet'])
            english_title = _english_title(recipe['B2'].value, mapping['sheet'])

            # Balance the title between an empty left column and the Day cell
            # so it is visually centred across the whole printed page.
            ws.merge_cells('B1:F1')
            ws['B1'] = arabic_title
            ws['B1'].font = Font(name=ARABIC_FONT, bold=True, size=19)
            ws['B1'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
            ws['G1'] = f'Day {int(day_no)}'
            ws['G1'].font = Font(name=LATIN_FONT, bold=True, size=15)
            ws['G1'].alignment = Alignment(horizontal='right', vertical='center')
            ws.row_dimensions[1].height = 38

            ws.merge_cells('A3:G3')
            ws['A3'] = f'{arabic_title}   {batch_index} / {batch_count}'
            ws['A3'].fill = brown
            ws['A3'].font = Font(name=ARABIC_FONT, color='FFFFFF', bold=True, size=14)
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
            ws.row_dimensions[3].height = 27

            ws.merge_cells('A4:G4')
            ws['A4'] = f'{english_title}   {batch_index} / {batch_count}'
            ws['A4'].fill = brown
            ws['A4'].font = Font(name=LATIN_FONT, color='FFFFFF', bold=True, size=12)
            ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[4].height = 24

            headers = [
                'Ingredient - الأصناف', 'Unit - الوحدة',
                'Base Recipe (30kg Yield)\n— الوصفة الأساسية (إنتاج 30 كجم)',
                'Corrected Conversion Factor\n— معامل التحويل المصحح',
                'Final Cooked Amount\n— الكمية النهائية المطبوخة',
                'Scaling Factor —\nمعامل القياس',
                'Linear Scaled Amount - الكمية المعدلة\nحسابياً',
            ]
            for column, header in enumerate(headers, 1):
                cell = ws.cell(5, column, header)
                cell.fill = green if column in (4, 5, 6) else peach
                cell.font = Font(name=ARABIC_FONT, bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            ws.row_dimensions[5].height = 52

            for row_offset, item in enumerate(ingredients, 6):
                linear_amount = item['base'] * batch['factor']
                values = [
                    item['name'], item['unit'], item['base'], batch['factor'],
                    batch['final_grams'], item['scale'], linear_amount,
                ]
                for column, value in enumerate(values, 1):
                    cell = ws.cell(row_offset, column, value)
                    cell.fill = white
                    cell.border = border
                    cell.font = Font(name=ARABIC_FONT, size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    if column == 4:
                        cell.number_format = '0.000'
                    elif column == 5:
                        cell.number_format = '#,##0'
                    elif column == 6:
                        cell.number_format = '0%'
                    elif isinstance(value, (int, float)):
                        cell.number_format = '#,##0.0#'
                ws.row_dimensions[row_offset].height = 22

            widths = [31, 16, 29, 23, 25, 23, 38]
            for column, width in enumerate(widths, 1):
                ws.column_dimensions[chr(64 + column)].width = width
            last_row = 5 + len(ingredients)
            ws.print_area = f'A1:G{max(last_row + 17, 34)}'
            ws.page_setup.orientation = 'landscape'
            ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            ws.page_margins.left = 0.25
            ws.page_margins.right = 0.25
            ws.page_margins.top = 0.35
            ws.page_margins.bottom = 0.35
            ws.oddFooter.center.text = f'Page {page_index} of {total_pages}'
            ws.oddFooter.center.size = 10
            ws.oddFooter.center.font = f'{LATIN_FONT},Bold'

        if not report.sheetnames:
            raise ValueError('لا توجد كميات أرز موجبة لهذا اليوم')
        report.active = 0
        output = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
        report.save(output)
    finally:
        source.close()
        report.close()
    return output, {'day_no': int(day_no), 'page_count': len(pages)}


def _state_from_workbook(path, day_no=None):
    wb = load_workbook(path, data_only=True, read_only=False)
    try:
        active_day = int(day_no or _number(wb['Ordering']['R1'].value) or 1)
        items = []
        for mapping in _mapping_rows(wb, active_day):
            ws = wb[mapping['sheet']]
            items.append({
                'sheet': mapping['sheet'],
                'day_no': active_day,
                'input_grams': round(_number(ws['Z1'].value), 3),
                'final_grams': round(_number(ws['E1'].value), 3),
                'safety_grams': mapping['safety_grams'],
                'batches': _batch_rows(ws),
            })
        return {'day_no': active_day, 'day_name': DAY_NAMES.get(active_day, ''), 'items': items}
    finally:
        wb.close()


def get_rice_template_state(template_path=RICE_TEMPLATE_PATH, day_no=None):
    if not os.path.exists(template_path):
        raise FileNotFoundError('ملف الأرز الأساسي غير موجود')
    state = _state_from_workbook(template_path, day_no)
    state['template_updated_at'] = datetime.fromtimestamp(os.path.getmtime(template_path)).isoformat(timespec='seconds')
    return state


def _files_from_inputs(day_no, inputs, template_path=RICE_TEMPLATE_PATH):
    macro_workbook = _write_inputs(day_no, inputs, template_path)
    calculated = _recalculate_to_xlsx(macro_workbook)
    pdf_source, report = _build_pdf_source(calculated, day_no)
    pdf_path = _export_to_pdf(pdf_source)
    report['state'] = _state_from_workbook(calculated, day_no)
    report['matched_count'] = sum(1 for item in inputs if _number(item.get('input_grams')) > 0)
    return calculated, pdf_path, report


def build_rice_day_files(file_storage, template_path=RICE_TEMPLATE_PATH, safety_items=None, expected_day_no=None):
    day_no, meals, input_report = read_day_file_payload(file_storage)
    if expected_day_no is not None and int(_number(expected_day_no)) != int(day_no):
        raise ValueError(f'ملف اليوم يخص يوم {day_no} بينما اليوم المختار في لوحة الأرز هو {int(_number(expected_day_no))}')
    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    try:
        inputs, missing = _compute_day_grams(wb, day_no, meals)
    finally:
        wb.close()
    safety_lookup = {
        str(item.get('sheet') or ''): max(0.0, _number(item.get('safety_grams')))
        for item in (safety_items or [])
    }
    for item in inputs:
        item['safety_grams'] = safety_lookup.get(item['sheet'], 0.0)
    excel_path, pdf_path, report = _files_from_inputs(day_no, inputs, template_path)
    report.update({'input_report': input_report, 'missing_sources': missing, 'inputs': inputs})
    return excel_path, pdf_path, report


def build_rice_manual_files(day_no, items, template_path=RICE_TEMPLATE_PATH):
    day_no = int(_number(day_no))
    if day_no not in DAY_NAMES:
        raise ValueError('رقم يوم الأرز غير صحيح')
    allowed_wb = load_workbook(template_path, data_only=False, read_only=True, keep_vba=True)
    try:
        allowed = {mapping['sheet'] for mapping in _mapping_rows(allowed_wb, day_no)}
    finally:
        allowed_wb.close()
    inputs = []
    for item in items or []:
        sheet = str(item.get('sheet') or '')
        if sheet in allowed:
            inputs.append({
                'sheet': sheet,
                'input_grams': max(0.0, _number(item.get('input_grams'))),
                'safety_grams': max(0.0, _number(item.get('safety_grams'))),
            })
    return _files_from_inputs(day_no, inputs, template_path)


def package_rice_files(excel_path, pdf_path, day_no):
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', compression=zipfile.ZIP_DEFLATED) as archive:
        archive.write(excel_path, f'Day{int(day_no)}_Rice_Updated.xlsx')
        archive.write(pdf_path, f'Day{int(day_no)}_Rice_Batches.pdf')
    buffer.seek(0)
    return buffer


def replace_rice_template(file_storage, template_path=RICE_TEMPLATE_PATH):
    suffix = os.path.splitext(file_storage.filename or '')[1].lower()
    if suffix != '.xlsm':
        raise ValueError('الشيت الأساسي للأرز لازم يكون XLSM')
    upload_path = tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False).name
    file_storage.seek(0)
    file_storage.save(upload_path)
    wb = load_workbook(upload_path, data_only=False, keep_vba=True)
    try:
        mappings = _mapping_rows(wb)
        if not mappings:
            raise ValueError('الشيت الجديد لا يحتوي على خريطة الأرز في Ordering!AA:AD')
        for mapping in mappings:
            wb[mapping['sheet']]['Z1']
    finally:
        wb.close()
    shutil.copy2(upload_path, template_path)
    return get_rice_template_state(template_path), {'template_file': os.path.basename(template_path)}
