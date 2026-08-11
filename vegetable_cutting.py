"""Extract and combine vegetable cutting instructions from Tokyo workbooks."""

from collections import OrderedDict
import io
import math
import os
import re

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont
import arabic_reshaper
try:
    from bidi.algorithm import get_display
except ImportError:  # python-bidi 0.6.11+ exposes the Rust implementation here
    from bidi import get_display
from flask import Blueprint, jsonify, request, send_file
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import A1, A2, landscape
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas


vegetable_cutting_bp = Blueprint("vegetable_cutting", __name__)

DAY_NAMES_AR = {
    1: "السبت",
    2: "الأحد",
    3: "الإثنين",
    4: "الثلاثاء",
    5: "الأربعاء",
    6: "الخميس",
    7: "الجمعة",
}

DAY_NAMES_EN = {
    1: "Saturday",
    2: "Sunday",
    3: "Monday",
    4: "Tuesday",
    5: "Wednesday",
    6: "Thursday",
    7: "Friday",
}

_METHOD_WORDS = (
    "cut", "slice", "julienne", "wedge", "dice", "cube", "chop",
    "mince", "mash", "puree", "ring", "crescent", "shred", "clean",
)
_NON_METHODS = {
    "category", "cutting method", "protein", "dairy", "pate",
    "vegetables", "spices & seasonings", "bread & bakery", "filling",
}

DAY_NO_COL = 36       # AJ
SHEET_NAME_COL = 37   # AK

_FONT_DIR = os.path.join(os.path.dirname(__file__), "fonts")
_PDF_FONT_REGULAR = "OctaArabic"
_PDF_FONT_BOLD = "OctaArabicBold"


def _find_pdf_font(*names):
    search_dirs = (
        _FONT_DIR,
        "/usr/share/fonts/truetype/freefont",
        "/usr/share/fonts/truetype/dejavu",
        "/usr/local/share/fonts",
        "/Library/Fonts",
        "/System/Library/Fonts/Supplemental",
    )
    for directory in search_dirs:
        for name in names:
            path = os.path.join(directory, name)
            if os.path.isfile(path):
                return path
    return None


def _register_pdf_fonts():
    regular_path = _find_pdf_font(
        "IBMPlexSansArabic-Regular.ttf", "FreeSans.ttf", "DejaVuSans.ttf", "Arial Unicode.ttf", "Arial.ttf",
    )
    bold_path = _find_pdf_font(
        "IBMPlexSansArabic-Bold.ttf", "FreeSansBold.ttf", "DejaVuSans-Bold.ttf", "Arial Bold.ttf",
    ) or regular_path
    if not regular_path or not bold_path:
        raise CuttingWorkbookError(
            "خط PDF العربي غير متاح على الخادم. أعد نشر الـ backend باستخدام Dockerfile المرفق."
        )
    if _PDF_FONT_REGULAR not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(_PDF_FONT_REGULAR, regular_path))
    if _PDF_FONT_BOLD not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(TTFont(_PDF_FONT_BOLD, bold_path))


def _rtl(value):
    return get_display(arabic_reshaper.reshape(_clean_text(value)))


def _pdf_icon(pdf, kind, x, y, width=68, height=38):
    """Draw a compact vector cutting-method icon without external images."""
    pdf.saveState()
    pdf.translate(x, y)
    pdf.scale(width / 68.0, height / 38.0)
    x, y, width, height = 0, 0, 68, 38
    green = HexColor("#176047")
    brass = HexColor("#B38636")
    pdf.setStrokeColor(green)
    pdf.setFillColor(HexColor("#EEF5E8"))
    pdf.roundRect(x, y, width, height, 10, stroke=0, fill=1)
    pdf.setLineWidth(1.8)

    left, right = x + 12, x + width - 12
    bottom, top = y + 9, y + height - 9

    if kind in ("julienne", "slices", "grated"):
        count = 4 if kind != "grated" else 5
        for idx in range(count):
            yy = bottom + idx * ((top - bottom) / max(count - 1, 1))
            offset = 4 if idx % 2 else 0
            pdf.line(left + offset, yy, right - 5, yy)
        pdf.setStrokeColor(brass)
        pdf.line(right - 3, bottom - 1, right + 2, top + 1)
    elif kind == "dice":
        size = 9
        for row in range(2):
            for col in range(3):
                pdf.rect(left + col * 15, bottom + row * 13, size, size, stroke=1, fill=0)
    elif kind == "rectangles":
        pdf.rect(left, top - 8, 18, 8, stroke=1, fill=0)
        pdf.rect(left + 23, top - 8, 23, 8, stroke=1, fill=0)
        pdf.rect(left + 5, bottom, 25, 8, stroke=1, fill=0)
        pdf.rect(left + 35, bottom, 14, 8, stroke=1, fill=0)
    elif kind == "rings":
        for cx in (left + 11, left + 34):
            pdf.circle(cx, y + height / 2, 9, stroke=1, fill=0)
            pdf.circle(cx, y + height / 2, 4, stroke=1, fill=0)
    elif kind == "wedges":
        for start in (left, left + 25):
            path = pdf.beginPath()
            path.moveTo(start, bottom)
            path.lineTo(start + 11, top)
            path.lineTo(start + 21, bottom)
            path.close()
            pdf.drawPath(path, stroke=1, fill=0)
    elif kind == "crescent":
        pdf.arc(left, bottom - 1, left + 23, top + 1, 70, 220)
        pdf.arc(left + 8, bottom + 1, left + 27, top - 1, 75, 210)
        pdf.arc(left + 27, bottom, right + 2, top, 70, 220)
    elif kind in ("minced", "chunks"):
        points = ((16, 24), (28, 27), (40, 22), (21, 12), (35, 10), (49, 15))
        for px, py in points:
            pdf.circle(x + px, y + py, 2.2 if kind == "minced" else 3.2, stroke=1, fill=0)
    elif kind == "puree":
        path = pdf.beginPath()
        path.moveTo(left, bottom + 2)
        path.curveTo(left + 5, top + 4, left + 19, top + 2, left + 24, top - 3)
        path.curveTo(left + 34, top + 5, right - 2, top - 5, right, bottom + 2)
        path.close()
        pdf.drawPath(path, stroke=1, fill=0)
        pdf.setStrokeColor(brass)
        pdf.line(right - 2, bottom, right + 3, top + 3)
    elif kind == "clean":
        path = pdf.beginPath()
        path.moveTo(left, bottom)
        path.curveTo(left + 5, top + 5, right - 6, top + 4, right, top)
        path.curveTo(right - 6, bottom - 4, left + 14, bottom - 2, left, bottom)
        pdf.drawPath(path, stroke=1, fill=0)
        pdf.line(left + 5, bottom + 2, right - 4, top - 2)
    else:
        pdf.line(left, bottom, right - 5, top)
        pdf.setStrokeColor(brass)
        pdf.line(right - 7, top - 2, right + 2, top + 5)
    pdf.restoreState()


def _row_methods(row):
    methods = row.get("methods") or []
    if methods:
        return methods
    return [{
        "method": row.get("method") or "غير محدد",
        "weight_grams": float(row.get("weight_grams") or 0),
        "icon": row.get("icon") or "knife",
    }]


def _pdf_row_height(row):
    method_count = len(_row_methods(row))
    return max(34, 12 + method_count * 17)


def build_cutting_pdf(payload):
    _register_pdf_fonts()
    rows = payload.get("rows") or []
    if not rows:
        raise CuttingWorkbookError("لا توجد بيانات لإنشاء ملف PDF")
    if len(rows) > 500:
        raise CuttingWorkbookError("عدد الصفوف أكبر من الحد المسموح")

    margin = 40
    header_height = 126
    columns_height = 40
    footer_height = 28
    row_heights = [_pdf_row_height(row) for row in rows]
    required_height = (margin * 2) + header_height + columns_height + footer_height + sum(row_heights)
    a2_width, a2_height = landscape(A2)
    if required_height <= a2_height:
        page_width, page_height = a2_width, a2_height
    else:
        a1_width, a1_height = landscape(A1)
        if required_height <= a1_height:
            page_width, page_height = a1_width, a1_height
        else:
            # Keep the promise of one page even for unusually large uploads.
            page_width, page_height = a1_width, required_height + 20
    output = io.BytesIO()
    pdf = canvas.Canvas(output, pagesize=(page_width, page_height), pageCompression=1)
    pdf.setTitle(f"Vegetable Cutting - Day {payload.get('day_number', '')}")
    pdf.setAuthor("Octa Food")

    table_left = margin
    table_width = page_width - (margin * 2)
    cutting_width = min(510, table_width * 0.43)
    weight_width = min(170, table_width * 0.14)
    ingredient_width = table_width - cutting_width - weight_width

    # Header band
    header_y = page_height - margin - header_height
    pdf.setFillColor(HexColor("#164F3C"))
    pdf.roundRect(margin, header_y, table_width, header_height, 14, stroke=0, fill=1)
    pdf.setFillColor(HexColor("#B9D94A"))
    pdf.setFont(_PDF_FONT_BOLD, 13)
    pdf.drawRightString(page_width - margin - 28, header_y + 94, _rtl("كشف تجهيز المطبخ"))
    pdf.setFillColor(HexColor("#FFFFFF"))
    pdf.setFont(_PDF_FONT_BOLD, 32)
    day_ar = payload.get("day_name_ar") or f"يوم {payload.get('day_number', '')}"
    pdf.drawRightString(page_width - margin - 28, header_y + 53, _rtl(day_ar))
    pdf.setFont(_PDF_FONT_BOLD, 14)
    pdf.setFillColor(HexColor("#CFDED7"))
    pdf.drawRightString(page_width - margin - 28, header_y + 29, _clean_text(payload.get("day_name_en")))

    pdf.setFillColor(HexColor("#FFFFFF"))
    pdf.setFont(_PDF_FONT_BOLD, 24)
    pdf.drawString(margin + 28, header_y + 68, f"{len(rows):,}")
    pdf.setFont(_PDF_FONT_REGULAR, 11)
    pdf.setFillColor(HexColor("#CFDED7"))
    pdf.drawString(margin + 28, header_y + 48, _rtl("صنف مجمّع"))
    pdf.setFillColor(HexColor("#FFFFFF"))
    pdf.setFont(_PDF_FONT_BOLD, 24)
    total_weight = sum(float(row.get("weight_grams") or 0) for row in rows)
    pdf.drawString(margin + 155, header_y + 68, f"{total_weight:,.0f}")
    pdf.setFont(_PDF_FONT_REGULAR, 11)
    pdf.setFillColor(HexColor("#CFDED7"))
    pdf.drawString(margin + 155, header_y + 48, _rtl("إجمالي جرام"))

    columns_y = header_y - columns_height
    pdf.setFillColor(HexColor("#E4EDCC"))
    pdf.rect(table_left, columns_y, table_width, columns_height, stroke=0, fill=1)
    pdf.setFillColor(HexColor("#234B39"))
    pdf.setFont(_PDF_FONT_BOLD, 13)
    pdf.drawRightString(table_left + ingredient_width - 16, columns_y + 14, _rtl("الصنف"))
    pdf.drawCentredString(table_left + ingredient_width + weight_width / 2, columns_y + 14, _rtl("الإجمالي (جرام)"))
    pdf.drawRightString(table_left + table_width - 18, columns_y + 14, _rtl("طريقة وشكل التقطيع"))

    row_top = columns_y
    cutting_left = table_left + ingredient_width + weight_width
    cutting_right = table_left + table_width
    for row_index, (row, row_height) in enumerate(zip(rows, row_heights)):
        row_y = row_top - row_height
        pdf.setFillColor(HexColor("#FFFFFF") if row_index % 2 == 0 else HexColor("#FBFCFA"))
        pdf.rect(table_left, row_y, table_width, row_height, stroke=0, fill=1)
        pdf.setStrokeColor(HexColor("#E7ECE8"))
        pdf.setLineWidth(0.55)
        pdf.line(table_left, row_y, table_left + table_width, row_y)

        text_y = row_y + (row_height / 2) - 4
        pdf.setFillColor(HexColor("#17211D"))
        pdf.setFont(_PDF_FONT_BOLD, 12.5)
        pdf.drawRightString(table_left + ingredient_width - 16, text_y, _rtl(row.get("ingredient"))[:100])

        pdf.setFillColor(HexColor("#164F3C"))
        pdf.setFont(_PDF_FONT_BOLD, 16)
        weight = float(row.get("weight_grams") or 0)
        pdf.drawCentredString(table_left + ingredient_width + weight_width / 2, text_y, f"{weight:,.0f}")

        methods = _row_methods(row)
        line_gap = 17
        methods_height = len(methods) * line_gap
        method_y = row_y + (row_height + methods_height) / 2 - 13
        icon_w, icon_h = 42, 24
        icon_x = cutting_right - icon_w - 14
        method_right = icon_x - 12
        for method_row in methods:
            _pdf_icon(pdf, method_row.get("icon") or "knife", icon_x, method_y - 6, icon_w, icon_h)
            pdf.setFillColor(HexColor("#164F3C"))
            pdf.setFont(_PDF_FONT_BOLD, 11.5)
            pdf.drawRightString(method_right, method_y, _rtl(method_row.get("method"))[:52])
            pdf.setFillColor(HexColor("#718078"))
            pdf.setFont(_PDF_FONT_REGULAR, 9)
            method_weight = float(method_row.get("weight_grams") or 0)
            pdf.drawString(cutting_left + 14, method_y + 1, f"{method_weight:,.0f} g")
            method_y -= line_gap
        row_top = row_y

    pdf.setFillColor(HexColor("#87938C"))
    pdf.setFont(_PDF_FONT_REGULAR, 10)
    pdf.drawString(margin, 18, "OCTA FOOD - VEGETABLE PREP")
    pdf.drawRightString(page_width - margin, 18, _rtl("تقرير صفحة واحدة"))
    pdf.showPage()

    pdf.save()
    output.seek(0)
    return output


def build_cutting_xlsx(payload):
    rows = payload.get("rows") or []
    if not rows:
        raise CuttingWorkbookError("لا توجد بيانات لإنشاء ملف Excel")
    if len(rows) > 500:
        raise CuttingWorkbookError("عدد الصفوف أكبر من الحد المسموح")

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Vegetable Cutting"
    worksheet.sheet_view.rightToLeft = True
    worksheet.freeze_panes = "A5"

    dark_green = "123D2E"
    lime = "B8D83E"
    light_green = "EFF5F1"
    white = "FFFFFF"
    border_side = Side(style="thin", color="CCD8D1")

    worksheet.merge_cells("A1:E1")
    title_cell = worksheet["A1"]
    title_cell.value = "كشف تقطيع الخضار"
    title_cell.fill = PatternFill("solid", fgColor=dark_green)
    title_cell.font = Font(color=white, bold=True, size=18)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    worksheet.row_dimensions[1].height = 34

    worksheet.merge_cells("A2:E2")
    day_number = payload.get("day_number") or ""
    day_ar = _clean_text(payload.get("day_name_ar")) or f"اليوم {day_number}"
    day_en = _clean_text(payload.get("day_name_en"))
    worksheet["A2"] = f"{day_ar} - {day_en}" if day_en else day_ar
    worksheet["A2"].font = Font(color=dark_green, bold=True, size=12)
    worksheet["A2"].alignment = Alignment(horizontal="center", vertical="center")

    headers = (
        "الصنف", "إجمالي الوزن (جرام)", "طريقة التقطيع",
        "وزن الطريقة (جرام)", "المصدر",
    )
    for column, value in enumerate(headers, start=1):
        cell = worksheet.cell(row=4, column=column, value=value)
        cell.fill = PatternFill("solid", fgColor=lime)
        cell.font = Font(color=dark_green, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=border_side, right=border_side,
            top=border_side, bottom=border_side,
        )

    output_row = 5
    for item in rows:
        methods = _row_methods(item)
        sources = "، ".join(_clean_text(value) for value in (item.get("sources") or []) if value)
        for method in methods:
            values = (
                _clean_text(item.get("ingredient")),
                float(item.get("weight_grams") or 0),
                _clean_text(method.get("method")),
                float(method.get("weight_grams") or 0),
                sources,
            )
            for column, value in enumerate(values, start=1):
                cell = worksheet.cell(row=output_row, column=column, value=value)
                cell.fill = PatternFill("solid", fgColor=white if output_row % 2 else light_green)
                cell.alignment = Alignment(
                    horizontal="right" if column in (1, 3, 5) else "center",
                    vertical="center", wrap_text=True,
                )
                cell.border = Border(
                    left=border_side, right=border_side,
                    top=border_side, bottom=border_side,
                )
                if column in (2, 4):
                    cell.number_format = '#,##0.00'
            output_row += 1

    worksheet.auto_filter.ref = f"A4:E{max(output_row - 1, 4)}"
    for column, width in {"A": 34, "B": 21, "C": 30, "D": 21, "E": 34}.items():
        worksheet.column_dimensions[column].width = width
    worksheet.print_title_rows = "1:4"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def build_cutting_png(payload):
    rows = payload.get("rows") or []
    if not rows:
        raise CuttingWorkbookError("لا توجد بيانات لإنشاء الصورة")
    if len(rows) > 500:
        raise CuttingWorkbookError("عدد الصفوف أكبر من الحد المسموح")

    regular_path = _find_pdf_font(
        "IBMPlexSansArabic-Regular.ttf", "FreeSans.ttf", "DejaVuSans.ttf",
    )
    bold_path = _find_pdf_font(
        "IBMPlexSansArabic-Bold.ttf", "FreeSansBold.ttf", "DejaVuSans-Bold.ttf",
    ) or regular_path
    if not regular_path or not bold_path:
        raise CuttingWorkbookError("الخط العربي غير متاح على الخادم")

    regular = ImageFont.truetype(regular_path, 25)
    small = ImageFont.truetype(regular_path, 21)
    bold = ImageFont.truetype(bold_path, 27)
    title_font = ImageFont.truetype(bold_path, 48)
    subtitle_font = ImageFont.truetype(regular_path, 25)

    width = 1800
    margin = 52
    header_height = 170
    columns_height = 66
    footer_height = 54
    row_heights = [max(76, 30 + len(_row_methods(row)) * 34) for row in rows]
    height = margin + header_height + columns_height + sum(row_heights) + footer_height + margin
    if height > 50000:
        raise CuttingWorkbookError("التقرير أكبر من الحد المسموح للصورة")

    dark_green = "#123D2E"
    medium_green = "#176047"
    lime = "#B8D83E"
    pale = "#F1F6F3"
    grid = "#CCD8D1"
    muted = "#6D7B74"
    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)

    def rtl_text(xy, value, font, fill=dark_green, anchor="ra"):
        text = _clean_text(value)
        try:
            # Pillow on the deployed Docker image has libraqm.  Give it the
            # original Unicode text so Arabic is shaped exactly once.
            draw.text(
                xy, text, font=font, fill=fill, anchor=anchor,
                direction="rtl", language="ar",
            )
        except (KeyError, TypeError, ValueError):
            # Environments without libraqm still work using presentation
            # forms prepared by arabic-reshaper/python-bidi.
            draw.text(xy, _rtl(text), font=font, fill=fill, anchor=anchor)

    table_left = margin
    table_right = width - margin
    ingredient_left = 1110
    weight_left = 835

    draw.rounded_rectangle(
        (margin, margin, width - margin, margin + header_height - 12),
        radius=30, fill=dark_green,
    )
    rtl_text((width - margin - 40, margin + 58), "كشف تقطيع الخضار", title_font, "white")
    day_number = payload.get("day_number") or ""
    day_ar = payload.get("day_name_ar") or f"اليوم {day_number}"
    day_en = _clean_text(payload.get("day_name_en"))
    rtl_text((width - margin - 42, margin + 116), day_ar, subtitle_font, lime)
    draw.text((margin + 42, margin + 105), day_en, font=subtitle_font, fill="white", anchor="la")

    y = margin + header_height
    draw.rectangle((table_left, y, table_right, y + columns_height), fill=lime)
    rtl_text((table_right - 24, y + columns_height / 2), "الصنف", bold, anchor="rm")
    rtl_text(((weight_left + ingredient_left) / 2, y + columns_height / 2), "إجمالي الوزن", bold, anchor="mm")
    rtl_text((weight_left - 24, y + columns_height / 2), "طريقة التقطيع والوزن", bold, anchor="rm")
    y += columns_height

    for index, (row, row_height) in enumerate(zip(rows, row_heights)):
        row_bottom = y + row_height
        draw.rectangle(
            (table_left, y, table_right, row_bottom),
            fill="white" if index % 2 == 0 else pale,
            outline=grid,
        )
        draw.line((weight_left, y, weight_left, row_bottom), fill=grid, width=2)
        draw.line((ingredient_left, y, ingredient_left, row_bottom), fill=grid, width=2)

        rtl_text((table_right - 24, y + row_height / 2), row.get("ingredient"), bold, anchor="rm")
        draw.text(
            ((weight_left + ingredient_left) / 2, y + row_height / 2),
            f'{float(row.get("weight_grams") or 0):,.0f} g',
            font=bold, fill=medium_green, anchor="mm",
        )

        methods = _row_methods(row)
        line_y = y + (row_height - len(methods) * 34) / 2 + 17
        for method in methods:
            method_weight = f'{float(method.get("weight_grams") or 0):,.0f} g'
            draw.text((table_left + 24, line_y), method_weight, font=small, fill=muted, anchor="lm")
            rtl_text((weight_left - 24, line_y), method.get("method"), regular, anchor="rm")
            line_y += 34
        y = row_bottom

    draw.rectangle((table_left, y, table_right, y + footer_height), fill=dark_green)
    draw.text((table_left + 24, y + footer_height / 2), "OCTA FOOD - VEGETABLE PREP", font=small, fill="white", anchor="lm")
    rtl_text((table_right - 24, y + footer_height / 2), f"عدد الأصناف: {len(rows)}", small, "white", "rm")

    output = io.BytesIO()
    image.save(output, format="PNG", optimize=True)
    output.seek(0)
    return output


class CuttingWorkbookError(ValueError):
    pass


def _clean_text(value):
    return re.sub(r"\s+", " ", str(value or "").replace("ـ", " ")).strip()


def _as_day(value):
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    if not numeric.is_integer():
        return None
    day = int(numeric)
    return day if day in DAY_NAMES_AR else None


def _as_weight(value):
    if isinstance(value, bool):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number) or number <= 0:
        return None
    return number


def _is_cutting_method(value):
    method = _clean_text(value)
    if not method or method.casefold() in _NON_METHODS:
        return False
    if re.search(r"[\u0600-\u06ff]", method):
        return True
    lowered = method.casefold()
    return any(word in lowered for word in _METHOD_WORDS)


def _icon_kind(method):
    text = _clean_text(method).casefold()
    if any(word in text for word in ("مكعب", "dice", "cube")):
        return "dice"
    if any(word in text for word in ("جوليان", "julienne", "شعيرات", "shred")):
        return "julienne"
    if any(word in text for word in ("ويدجز", "أرباع", "wedge")):
        return "wedges"
    if any(word in text for word in ("نص قمر", "crescent")):
        return "crescent"
    if any(word in text for word in ("دائر", "دوائر", "ring")):
        return "rings"
    if any(word in text for word in ("مبشور", "grated", "grate")):
        return "grated"
    if any(word in text for word in ("مستطيل", "rectangle")):
        return "rectangles"
    if any(word in text for word in ("عشوائي", "random")):
        return "chunks"
    if any(word in text for word in ("مهروس", "بورية", "معجون", "mash", "puree")):
        return "puree"
    if any(word in text for word in ("مفروم", "مقطع ناعم", "mince", "chop")):
        return "minced"
    if any(word in text for word in ("تنضيف", "تنظيف", "clean")):
        return "clean"
    if any(word in text for word in ("شرائح", "slice")):
        return "slices"
    return "knife"


def _read_day_number(workbook):
    preferred = ("All_Ingredients", "Ordering", "Marination_Ordering")
    for sheet_name in preferred:
        if sheet_name in workbook.sheetnames:
            day = _as_day(workbook[sheet_name]["R1"].value)
            if day:
                return day, sheet_name
    for worksheet in workbook.worksheets:
        day = _as_day(worksheet["R1"].value)
        if day:
            return day, worksheet.title
    raise CuttingWorkbookError("تعذر قراءة رقم اليوم من الخلية R1")


def _summary_rows(workbook, day_number):
    if "Cutting_Shapes_Ordering" not in workbook.sheetnames:
        return None
    worksheet = workbook["Cutting_Shapes_Ordering"]
    rows = []
    # The prepared output block is F:I: day, ingredient, weight, cutting method.
    for day_value, ingredient, weight, method in worksheet.iter_rows(
        min_row=3, min_col=6, max_col=9, values_only=True
    ):
        if _as_day(day_value) != day_number:
            continue
        numeric_weight = _as_weight(weight)
        ingredient_text = _clean_text(ingredient)
        method_text = _clean_text(method)
        if ingredient_text and numeric_weight and _is_cutting_method(method_text):
            rows.append((ingredient_text, numeric_weight, method_text))
    return rows


def _day_recipe_sheet_names(workbook, day_number):
    """Return recipe tabs assigned to ``day_number`` in All_Ingredients.

    The Tokyo workbook keeps the day-to-tab mapping in AJ:AK.  The mapping
    currently extends well past row 40 (day 6 starts around row 77), so this
    must always follow the worksheet's actual last row instead of a fixed
    range.

    ``None`` means the workbook has no AJ:AK mapping (for example the
    breakfast workbook), while an empty list means a mapping exists but does
    not contain the requested day.
    """
    if "All_Ingredients" not in workbook.sheetnames:
        return None

    worksheet = workbook["All_Ingredients"]
    sheet_names = []
    seen = set()
    for row_number in range(1, worksheet.max_row + 1):
        if _as_day(worksheet.cell(row=row_number, column=DAY_NO_COL).value) != day_number:
            continue
        sheet_name = _clean_text(
            worksheet.cell(row=row_number, column=SHEET_NAME_COL).value
        )
        if not sheet_name or sheet_name in seen or sheet_name not in workbook.sheetnames:
            continue
        seen.add(sheet_name)
        sheet_names.append(sheet_name)
    return sheet_names


def _recipe_rows(workbook, day_number):
    rows = []
    ignored_sheets = {
        "Ordering", "All_Ingredients", "Marination_Ordering",
        "Cutting_Shapes_Ordering", "Butchery",
    }
    day_sheet_names = _day_recipe_sheet_names(workbook, day_number)
    if day_sheet_names == []:
        raise CuttingWorkbookError(
            f"لا توجد تابات مرتبطة باليوم {day_number} في العمودين AJ/AK"
        )

    worksheets = (
        [workbook[sheet_name] for sheet_name in day_sheet_names]
        if day_sheet_names is not None
        else [
            worksheet for worksheet in workbook.worksheets
            if worksheet.title not in ignored_sheets
        ]
    )
    for worksheet in worksheets:
        for row in worksheet.iter_rows(
            min_row=1, min_col=1, max_col=8, values_only=True
        ):
            method, ingredient, weight = row[0], row[1], row[7]
            numeric_weight = _as_weight(weight)
            ingredient_text = _clean_text(ingredient)
            method_text = _clean_text(method)
            if ingredient_text and numeric_weight and _is_cutting_method(method_text):
                rows.append((ingredient_text, numeric_weight, method_text))
    return rows


def extract_workbook(file_storage):
    try:
        workbook = openpyxl.load_workbook(
            file_storage.stream,
            read_only=True,
            data_only=True,
            keep_links=False,
        )
    except Exception as exc:
        raise CuttingWorkbookError(f"تعذر فتح الملف: {exc}") from exc

    try:
        day_number, day_sheet = _read_day_number(workbook)
        rows = _summary_rows(workbook, day_number)
        extraction_mode = "summary" if rows is not None else "recipes"
        if rows is None:
            rows = _recipe_rows(workbook, day_number)
        return {
            "filename": file_storage.filename or "workbook.xlsm",
            "day_number": day_number,
            "day_sheet": day_sheet,
            "mode": extraction_mode,
            "rows": rows,
        }
    finally:
        workbook.close()


def combine_workbooks(extracted):
    days = {item["day_number"] for item in extracted}
    if len(days) != 1:
        details = "، ".join(
            f'{item["filename"]}: يوم {item["day_number"]}' for item in extracted
        )
        raise CuttingWorkbookError(f"الملفان ليسا لنفس يوم التشغيل ({details})")

    combined = OrderedDict()
    for source in extracted:
        for ingredient, weight, method in source["rows"]:
            key = _clean_text(ingredient).casefold()
            if key not in combined:
                combined[key] = {
                    "ingredient": ingredient,
                    "weight_grams": 0.0,
                    "sources": [],
                    "methods": OrderedDict(),
                }
            combined[key]["weight_grams"] += weight
            method_key = _clean_text(method).casefold()
            if method_key not in combined[key]["methods"]:
                combined[key]["methods"][method_key] = {
                    "method": method,
                    "weight_grams": 0.0,
                    "icon": _icon_kind(method),
                }
            combined[key]["methods"][method_key]["weight_grams"] += weight
            if source["filename"] not in combined[key]["sources"]:
                combined[key]["sources"].append(source["filename"])

    rows = []
    for index, row in enumerate(combined.values(), start=1):
        row["id"] = index
        row["weight_grams"] = round(row["weight_grams"], 2)
        methods = list(row["methods"].values())
        for method_row in methods:
            method_row["weight_grams"] = round(method_row["weight_grams"], 2)
        row["methods"] = methods
        row["method"] = " / ".join(method_row["method"] for method_row in methods)
        row["icon"] = methods[0]["icon"] if len(methods) == 1 else "multiple"
        rows.append(row)
    return next(iter(days)), rows


@vegetable_cutting_bp.route("/api/vegetable-cutting/extract", methods=["POST"])
def vegetable_cutting_extract():
    files = [item for item in request.files.getlist("files") if item and item.filename]
    if len(files) != 2:
        return jsonify({"error": "ارفع ملف توكيو الرئيسي وملف الإفطار معًا"}), 400

    for item in files:
        if not item.filename.lower().endswith((".xlsx", ".xlsm")):
            return jsonify({"error": f"صيغة الملف غير مدعومة: {item.filename}"}), 400

    try:
        extracted = [extract_workbook(item) for item in files]
        day_number, rows = combine_workbooks(extracted)
    except CuttingWorkbookError as exc:
        return jsonify({"error": str(exc)}), 400

    if not rows:
        return jsonify({"error": "لم يتم العثور على أصناف لها طريقة تقطيع وكمية"}), 400

    return jsonify({
        "ok": True,
        "day_number": day_number,
        "day_name_ar": DAY_NAMES_AR[day_number],
        "day_name_en": DAY_NAMES_EN[day_number],
        "rows": rows,
        "total_weight_grams": round(sum(row["weight_grams"] for row in rows), 2),
        "sources": [
            {
                "filename": item["filename"],
                "day_number": item["day_number"],
                "day_cell": f'{item["day_sheet"]}!R1',
                "extraction_mode": item["mode"],
                "matched_rows": len(item["rows"]),
            }
            for item in extracted
        ],
    })


@vegetable_cutting_bp.route("/api/vegetable-cutting/export-pdf", methods=["POST"])
def vegetable_cutting_export_pdf():
    payload = request.get_json(silent=True) or {}
    try:
        output = build_cutting_pdf(payload)
    except (CuttingWorkbookError, TypeError, ValueError) as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({
            "error": f"تعذر إنشاء PDF على الخادم: {str(exc)[:180]}"
        }), 500
    day_number = _as_day(payload.get("day_number")) or 1
    return send_file(
        output,
        as_attachment=True,
        download_name=f"Vegetable_Cutting_Day_{day_number}.pdf",
        mimetype="application/pdf",
    )


@vegetable_cutting_bp.route("/api/vegetable-cutting/export-xlsx", methods=["POST"])
def vegetable_cutting_export_xlsx():
    payload = request.get_json(silent=True) or {}
    try:
        output = build_cutting_xlsx(payload)
    except (CuttingWorkbookError, TypeError, ValueError) as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({
            "error": f"تعذر إنشاء ملف Excel على الخادم: {str(exc)[:180]}"
        }), 500
    day_number = _as_day(payload.get("day_number")) or 1
    return send_file(
        output,
        as_attachment=True,
        download_name=f"Vegetable_Cutting_Day_{day_number}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@vegetable_cutting_bp.route("/api/vegetable-cutting/export-png", methods=["POST"])
def vegetable_cutting_export_png():
    payload = request.get_json(silent=True) or {}
    try:
        output = build_cutting_png(payload)
    except (CuttingWorkbookError, TypeError, ValueError) as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({
            "error": f"تعذر إنشاء الصورة على الخادم: {str(exc)[:180]}"
        }), 500
    day_number = _as_day(payload.get("day_number")) or 1
    return send_file(
        output,
        as_attachment=True,
        download_name=f"Vegetable_Cutting_Day_{day_number}.png",
        mimetype="image/png",
    )
