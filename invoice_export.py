"""
استخراج بيانات الفاتورة الكاملة من ملف PDF، مع تصحيح ترتيب الحروف العربي.

كتير من مولّدات الفواتير العربية (زي اللي بتستخدمها الموردين هنا) بتخزّن النص العربي
داخل ملف PDF بترتيب حروف معكوس (right-to-left storage order بدون إعادة ترتيب صحيحة) —
فلو قرأت النص "زي ما هو" بيطلع مقلوب. الدالة fix_line بتاخد كل سطر وتصحح ترتيب الحروف
العربية فيه فقط، وتسيب الأرقام والحروف اللاتينية كما هي (عشان الأرقام لا تتقلب).
"""
import re
import unicodedata
import tempfile
import pdfplumber

ARABIC_RUN = re.compile(r'[\u0600-\u06FF\uFB50-\uFDFF\uFE70-\uFEFF]+')
ARABIC_CHAR = re.compile(r'[\u0600-\u06FF\uFB50-\uFDFF\uFE70-\uFEFF]')


def _fix_token(token):
    # يقسّم الكلمة لقطع: عربي / غير عربي، ويقلب ترتيب القطع (لأن التخزين كان معكوسًا)،
    # وبيقلب ترتيب الحروف داخل القطعة العربية فقط (الأرقام/اللاتيني تفضل كما هي).
    # مهم: العكس ده بيحصل على شكل الجليفات الأصلي (presentation forms) قبل أي تطبيع NFKC،
    # عشان الحروف المدمجة (ligatures) اللي بتتفك لحرفين منفصلين تفضل بترتيبها الصحيح
    # الداخلي بعد العكس (لو طبّعنا الأول وبعدين عكسنا، بنعكس ترتيب حروف الـ ligature
    # الداخلي بالغلط، وده اللي كان بيسبب لخبطة زي "شركة" -> "رشكة").
    segments = re.findall(r'[\u0600-\u06FF\uFB50-\uFDFF\uFE70-\uFEFF]+|[^\u0600-\u06FF\uFB50-\uFDFF\uFE70-\uFEFF]+', token)
    fixed = []
    for seg in reversed(segments):
        if ARABIC_RUN.fullmatch(seg):
            fixed.append(seg[::-1])
        else:
            fixed.append(seg)
    return unicodedata.normalize('NFKC', ''.join(fixed))


def fix_line(line):
    line = re.sub(r'[\ue000-\uf8ff]', '', line)  # شيل رموز الأيقونات (private-use glyphs)
    tokens = line.split(' ')
    tokens = [_fix_token(t) for t in tokens if t]
    tokens.reverse()
    return ' '.join(tokens)


def extract_fixed_lines(pdf_path):
    raw_lines = []
    fixed_lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            for line in text.split('\n'):
                raw_lines.append(line)
                fixed_lines.append(fix_line(line))
    fixed_lines = resolve_known_cid_glyphs(fixed_lines)
    return raw_lines, fixed_lines


# بعض فونتات الـ PDF مالها تعريف يونيكود لحروف عربية متصلة معيّنة، فبتطلع كرمز
# "(cid:NUMBER)" بدل الحرف الحقيقي. عبارة "المركز الرئيسي" ثابتة وموجودة في كل
# فاتورة، فبنستخدمها كمرجع نكتشف بيه إن الرمز ده بيمثّل الحروف "مر" — ولو ظهر
# نفس الرمز في مكان تاني بالفاتورة (زي اسم صنف "تمر")، نستبدله بالحروف الصحيحة
# بدل ما نمسحه ويضيع الاسم.
_CID_BOILERPLATE_RE = re.compile(r'ال\(cid:(\d+)\)كز الرئيسي')


def resolve_known_cid_glyphs(lines):
    cid_map = {}
    for line in lines:
        m = _CID_BOILERPLATE_RE.search(line)
        if m:
            cid_map[m.group(1)] = 'مر'
    if not cid_map:
        return lines

    def repl(m):
        return cid_map.get(m.group(1), m.group(0))

    return [re.sub(r'\(cid:(\d+)\)', repl, line) for line in lines]


NUM = r'[\d.,]+'


def _to_float(s):
    if s is None:
        return 0.0
    if isinstance(s, (int, float)):
        return float(s)
    s = s.replace(',', '').strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


# بعض بنود الفاتورة بتتقسم بشكل غريب على PDF: جزء الضريبة (%النسبة القيمة) والقوس
# الفاتح بيطلعوا على سطر مستقل قبل وبعد سطر البند الرئيسي، بدل ما يكونوا كلهم على
# سطر واحد. ده بيخلي سطر البند الرئيسي يفقد الجزء ده تمامًا (مفيش % ولا قوس فيه)
# فما بيتطابقش مع row_re، والبند بيضيع كامل أو يندمج غلط مع اللي بعده. الدالة دي
# بتلمّ الأجزاء المنفصلة دي وترجّعها سطر واحد كامل قبل أي تحليل.
_ORPHAN_TAX_RE = re.compile(
    r'^(?:[\d.,]+%\s+)?%(?P<tax_pct>\d+)\s+(?P<tax_val>' + NUM + r')\)$'
)
_BROKEN_ROW_RE = re.compile(
    r'^(?P<idx>\d+)\s+(?P<mid>.+?)\s+(?P<before_tax>' + NUM + r')\s+(?P<total>' + NUM + r')$'
)
_ORPHAN_PAREN_RE = re.compile(r'^' + NUM + r'\s*\($')


def reassemble_split_rows(lines):
    out = list(lines)
    n = len(out)
    for i in range(n):
        line = out[i].strip()
        if not line:
            continue
        m = _ORPHAN_TAX_RE.match(line)
        if not m or i + 1 >= n:
            continue
        nxt = out[i + 1].strip()
        bm = _BROKEN_ROW_RE.match(nxt)
        if not bm:
            continue
        merged = (
            f"{bm.group('idx')} {bm.group('mid')} {bm.group('before_tax')} "
            f"%{m.group('tax_pct')} {m.group('tax_val')}) ( {bm.group('total')}"
        )
        out[i + 1] = merged
        out[i] = ''
        if i + 2 < n and _ORPHAN_PAREN_RE.match(out[i + 2].strip()):
            out[i + 2] = ''
    return out


def parse_items(fixed_lines):
    """
    كل بند بياخد عادة 2-3 سطور بعد التصحيح:
      - سطر اسم الصنف (ممكن يكون فاضي لو الاسم قصير وداخل سطر الأرقام)
      - سطر الأرقام: # سعر_الوحدة كمية+وحدة  اجمالي_قبل_الضريبة %نسبة_الضريبة قيمة_الضريبة) ( اجمالي
      - سطر تكملة الاسم (لو الاسم طويل) + قيمة الخصم (عادة 0.00)
    """
    # نمط السطر الرقمي: من الآخر للأول (الجزء الثابت أوضح من اليمين)
    row_re = re.compile(
        r'^(?P<idx>\d+)\s+(?P<mid>.*?)\s+'
        r'(?P<before_tax>' + NUM + r')\s+%(?P<tax_pct>\d+)\s+'
        r'(?P<tax_val>' + NUM + r')\)\s*\(\s*(?P<total>' + NUM + r')$'
    )
    qty_unit_re = re.compile(r'^(?P<qty>' + NUM + r')(?P<unit>[\u0600-\u06FF]+)$')

    fixed_lines = reassemble_split_rows(fixed_lines)

    items = []
    pending_name_parts = []
    header_seen = False
    header_keywords = re.compile(r'(فاتورة|بيانات|الرقم|المنتج|الكمية|الخصم|الضريبة|الاجمالي|اﻻﺠﻤﺎﻟﻲ|الرضيبة|االجمالي|المجموع|قيمة|العميل|موافق|الرئيس)')

    i = 0
    n = len(fixed_lines)
    while i < n:
        clean = fixed_lines[i].strip()
        if not clean:
            i += 1
            continue
        m = row_re.match(clean)
        if not m:
            if header_keywords.search(clean):
                header_seen = True
                pending_name_parts = []  # أي كلام قبل/فوق رأس الجدول مالوش دعوة بأي صنف
            elif header_seen and re.search(r'[\u0600-\u06FF]', clean):
                pending_name_parts.append(clean)
            i += 1
            continue

        mid_tokens = m.group('mid').split()
        qty_unit_token = None
        for t in reversed(mid_tokens):
            if qty_unit_re.match(t):
                qty_unit_token = t
                break
        qty, unit = '', ''
        if qty_unit_token:
            qm = qty_unit_re.match(qty_unit_token)
            qty, unit = qm.group('qty'), qm.group('unit')
            idx_in_mid = mid_tokens.index(qty_unit_token)
            before_qty = mid_tokens[:idx_in_mid]
        else:
            before_qty = mid_tokens

        unit_price = ''
        inline_name_tokens = []
        for t in reversed(before_qty):
            if re.match(r'^' + NUM + r'$', t) and not unit_price:
                unit_price = t
            else:
                inline_name_tokens.insert(0, t)

        inline_name = ' '.join(inline_name_tokens).strip()
        name = inline_name if inline_name else ' '.join(pending_name_parts).strip()
        name = re.sub(r'\d+\.?\d*%', '', name).strip()
        pending_name_parts = []
        consumed_next = False

        # نلحق سطر تكملة الاسم اللي بعد سطر الأرقام (لو موجود وغير رقمي خالص ومش سطر بند جديد).
        # بدون مسافة بين الاسم وتكملته، لأن التقسيم بيحصل وسط الكلمة نفسها غالبًا
        # (مثلاً "خس مدور امر" + "يكي" = "خس مدور امريكي"، مش "امر يكي").
        if i + 1 < n:
            nxt = fixed_lines[i + 1].strip()
            if nxt and re.search(r'[\u0600-\u06FF]', nxt) and not row_re.match(nxt) and not header_keywords.search(nxt):
                extra = re.sub(r'[\d.]+', '', nxt).strip()
                if extra:
                    name = f'{name}{extra}'.strip()
                    consumed_next = True  # السطر ده اتستخدم، السطر اللي بعده ميتعالجش تاني كأنه اسم صنف جديد

        # شيل رموز الحروف اللي مالها تعريف يونيكود في الخط المستخدم (عيب في ملف PDF نفسه،
        # مش في الترتيب)، ولو الاسم بقى قصير جدًا بعد الشيل، نعلّم عليه إنه يحتاج مراجعة يدوية
        needs_review = False
        if re.search(r'\(cid:\d+\)', name):
            name = re.sub(r'\(cid:\d+\)', '', name).strip()
            needs_review = True
        if len(re.sub(r'[^\u0600-\u06FF]', '', name)) <= 1:
            needs_review = True

        items.append({
            'name': name or '(بدون اسم)',
            'unit_price': _to_float(unit_price),
            'qty': _to_float(qty),
            'unit': unit,
            'total_before_tax': _to_float(m.group('before_tax')),
            'tax_pct': _to_float(m.group('tax_pct')),
            'tax_value': _to_float(m.group('tax_val')),
            'total': _to_float(m.group('total')),
            'needs_review': needs_review,
        })

        i += 2 if consumed_next else 1

    return items


def build_invoices_workbook(invoices):
    """يبني ملف إكسل منسّق بالكامل (ألوان، حدود، عرض أعمدة، خط عريض للعناوين
    والإجماليات) من بيانات الفواتير — بعد ما المستخدم يراجعها ويعدّلها في الواجهة."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    RED = 'C0392B'
    LIGHT = 'FDF2F0'
    WHITE = 'FFFFFF'
    INK = '241A17'

    header_fill = PatternFill('solid', fgColor=RED)
    header_font = Font(color=WHITE, bold=True, size=11)
    title_font = Font(bold=True, size=13, color=RED)
    total_font = Font(bold=True, size=11, color=INK)
    normal_font = Font(size=10.5, color=INK)
    light_fill = PatternFill('solid', fgColor=LIGHT)
    thin = Side(style='thin', color='D8B9B0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = '#,##0.00'

    def style_sheet_rtl(ws):
        ws.sheet_view.rightToLeft = True

    def header_row(ws, row_idx, headers, widths):
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=row_idx, column=i, value=h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = ws.cell(row=row_idx + 1, column=1)

    def write_row(ws, row_idx, values, money_cols=(), bold=False, fill=None):
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=i, value=v)
            c.font = total_font if bold else normal_font
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            if fill:
                c.fill = fill
            if i in money_cols and isinstance(v, (int, float)):
                c.number_format = money_fmt

    wb = Workbook()

    # ---- ملخص الفواتير ----
    ws = wb.active
    ws.title = 'ملخص الفواتير'
    style_sheet_rtl(ws)
    headers = ['الملف', 'التاريخ', 'رقم الفاتورة', 'المورد / العميل', 'قبل الضريبة', 'الضريبة', 'الإجمالي', 'عدد البنود', 'ملاحظات']
    widths = [30, 13, 16, 28, 13, 11, 13, 11, 30]
    header_row(ws, 1, headers, widths)
    for r, inv in enumerate(invoices, start=2):
        items = inv.get('items') or []
        write_row(ws, r, [
            inv.get('fileName', ''), inv.get('date', ''), inv.get('number', ''),
            inv.get('party', ''), _to_float(inv.get('subtotal')), _to_float(inv.get('vat')),
            None, len(items), inv.get('notes', ''),
        ], money_cols=(5, 6, 7), fill=light_fill if r % 2 == 0 else None)
        # الإجمالي = قبل الضريبة + الضريبة — معادلة شغالة
        c7 = ws.cell(row=r, column=7, value=f'=E{r}+F{r}')
        c7.font = normal_font
        c7.border = border
        c7.alignment = Alignment(horizontal='center', vertical='center')
        c7.number_format = money_fmt
        if r % 2 == 0:
            c7.fill = light_fill
    last_summary_row = len(invoices) + 1

    # ---- تجميع يومي ----
    ws2 = wb.create_sheet('تجميع يومي')
    style_sheet_rtl(ws2)
    header_row(ws2, 1, ['التاريخ', 'عدد الفواتير', 'إجمالي اليوم'], [16, 14, 16])
    dates_sorted = sorted({inv.get('date') or 'بدون تاريخ' for inv in invoices})
    r = 2
    for key in dates_sorted:
        c1 = ws2.cell(row=r, column=1, value=key)
        c1.font = normal_font
        c1.border = border
        c1.alignment = Alignment(horizontal='center', vertical='center')
        # عدد الفواتير وإجمالي اليوم بيتحسبوا أوتوماتيك من شيت "ملخص الفواتير"
        c2 = ws2.cell(row=r, column=2, value=f"=COUNTIF('ملخص الفواتير'!B2:B{last_summary_row},A{r})")
        c3 = ws2.cell(row=r, column=3, value=f"=SUMIF('ملخص الفواتير'!B2:B{last_summary_row},A{r},'ملخص الفواتير'!G2:G{last_summary_row})")
        for c in (c2, c3):
            c.font = normal_font
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
        c3.number_format = money_fmt
        if r % 2 == 0:
            for c in (c1, c2, c3):
                c.fill = light_fill
        r += 1
    if dates_sorted:
        last_daily_row = r - 1
        total_fill = PatternFill('solid', fgColor='F1D8D6')
        c1 = ws2.cell(row=r, column=1, value='الإجمالي')
        c2 = ws2.cell(row=r, column=2, value=f'=SUM(B2:B{last_daily_row})')
        c3 = ws2.cell(row=r, column=3, value=f'=SUM(C2:C{last_daily_row})')
        for c in (c1, c2, c3):
            c.font = total_font
            c.border = border
            c.alignment = Alignment(horizontal='center')
            c.fill = total_fill
        c3.number_format = money_fmt

    # ---- تاب لكل فاتورة ----
    used_names = set()

    def safe_sheet_name(name):
        base = (name or 'فاتورة').replace('/', '-').replace('\\', '-').replace('*', '-') \
            .replace('?', '-').replace('[', '-').replace(']', '-').replace(':', '-')
        base = base.strip()[:28] or 'فاتورة'
        candidate, i = base, 2
        while candidate in used_names:
            candidate = f'{base[:25]}-{i}'
            i += 1
        used_names.add(candidate)
        return candidate

    for idx, inv in enumerate(invoices):
        title = f"{inv.get('date') or f'فاتورة-{idx+1}'} {inv.get('number') or ''}".strip()
        wsi = wb.create_sheet(safe_sheet_name(title))
        style_sheet_rtl(wsi)
        wsi.column_dimensions['A'].width = 38
        wsi.column_dimensions['B'].width = 13
        wsi.column_dimensions['C'].width = 14
        wsi.column_dimensions['D'].width = 15

        labels = [
            ('اسم الملف', inv.get('fileName', '')),
            ('التاريخ', inv.get('date', '')),
            ('رقم الفاتورة', inv.get('number', '')),
            ('المورد / العميل', inv.get('party', '')),
        ]
        row = 1
        for label, value in labels:
            c1 = wsi.cell(row=row, column=1, value=label)
            c1.font = Font(bold=True, size=11, color=INK)
            c2 = wsi.cell(row=row, column=2, value=value)
            c2.font = normal_font
            wsi.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            row += 1

        row += 1
        header_row(wsi, row, ['الصنف / البيان', 'الكمية', 'سعر الوحدة', 'إجمالي البند'], [38, 13, 14, 15])
        table_header_row = row
        row += 1
        first_item_row = row
        for j, item in enumerate(inv.get('items') or []):
            qty = _to_float(item.get('qty'))
            price = _to_float(item.get('unitPrice'))
            write_row(wsi, row, [item.get('item', ''), qty, price, None], money_cols=(3,),
                      fill=light_fill if j % 2 == 0 else None)
            # إجمالي البند = الكمية × سعر الوحدة — معادلة شغالة، يتغير أوتوماتيك لو عدّلت أي رقم
            c4 = wsi.cell(row=row, column=4, value=f'=B{row}*C{row}')
            c4.font = normal_font
            c4.border = border
            c4.alignment = Alignment(horizontal='center', vertical='center')
            c4.number_format = money_fmt
            if j % 2 == 0:
                c4.fill = light_fill
            row += 1
        last_item_row = row - 1

        wsi.freeze_panes = wsi.cell(row=table_header_row + 1, column=1)

        row += 1
        items_total_row = row
        subtotal_row = row + 1
        vat_row = row + 2
        total_row = row + 3
        totals = [
            ('إجمالي البنود', f'=SUM(D{first_item_row}:D{last_item_row})' if last_item_row >= first_item_row else 0),
            ('قبل الضريبة', _to_float(inv.get('subtotal'))),
            ('الضريبة', _to_float(inv.get('vat'))),
            ('الإجمالي', f'=D{subtotal_row}+D{vat_row}'),
        ]
        for label, value in totals:
            c1 = wsi.cell(row=row, column=1, value=label)
            c1.font = total_font
            c1.border = border
            c4 = wsi.cell(row=row, column=4, value=value)
            c4.font = total_font
            c4.number_format = money_fmt
            c4.border = border
            c4.alignment = Alignment(horizontal='center')
            if label == 'الإجمالي':
                fill_final = PatternFill('solid', fgColor='F1D8D6')
                c1.fill = fill_final
                c4.fill = fill_final
            row += 1
        if inv.get('notes'):
            row += 1
            c1 = wsi.cell(row=row, column=1, value='ملاحظات')
            c1.font = Font(bold=True, size=10.5, color='B55A00')
            c2 = wsi.cell(row=row, column=2, value=inv.get('notes'))
            wsi.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)

    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    return tmp.name


def parse_invoice_full(pdf_path, file_name=''):
    raw_lines, fixed_lines = extract_fixed_lines(pdf_path)
    raw_text = '\n'.join(raw_lines)

    date_m = re.search(r'(20\d{2}-\d{2}-\d{2})', raw_text)
    number_m = re.search(r'\b([A-Za-z]{2,}\d{3,})\b', raw_text)
    tax_numbers = re.findall(r'\b(\d{15})\b', raw_text)

    non_empty_fixed = [l.strip() for l in fixed_lines if l.strip()]
    supplier = non_empty_fixed[0] if non_empty_fixed else ''
    customer = ''
    for l in non_empty_fixed[1:]:
        if re.search(r'(شركة|رشكة|مؤسسة|موسسة|مطعم|مصنع)', l) and l != supplier:
            customer = l
            break

    def find_amount(keyword_patterns, exclude=None):
        for l in fixed_lines:
            s = l.strip()
            if not s:
                continue
            if exclude and re.search(exclude, s):
                continue
            if re.search(keyword_patterns, s):
                m = re.search(r'(' + NUM + r')\s*$', s)
                if m:
                    return _to_float(m.group(1))
        return 0.0

    subtotal = find_amount(r'قبل')
    vat = find_amount(r'قيمة')
    total = find_amount(r'المجموع', exclude=r'المستحق')

    BOILERPLATE = ['موافق عليه', 'المركز الرئيسي', 'فاتورة مبيعات', 'بيانات العميل']

    def clean_party(s):
        for b in BOILERPLATE:
            s = s.replace(b, '')
        return re.sub(r'\s+', ' ', s).strip()

    items = parse_items(fixed_lines)

    review_items = [it['name'] for it in items if it.get('needs_review')]
    notes = ''
    if not items:
        notes = 'لم يتم استخراج بنود — قد تحتاج مراجعة يدوية'
    elif review_items:
        notes = f"تنبيه: اسم {len(review_items)} صنف غير واضح بسبب عيب في الخط المستخدم بملف PDF (السعر/الكمية صحيحين، الاسم يحتاج تأكيد يدوي): {', '.join(review_items)}"

    return {
        'fileName': file_name,
        'date': date_m.group(1) if date_m else '',
        'number': number_m.group(1) if number_m else '',
        'party': clean_party(customer) or clean_party(supplier),
        'supplier': clean_party(supplier),
        'customer': clean_party(customer),
        'taxNumbers': tax_numbers,
        'subtotal': subtotal,
        'vat': vat,
        'total': total,
        'items': [
            {
                'item': it['name'],
                'qty': it['qty'],
                'unit': it['unit'],
                'unitPrice': it['unit_price'],
                'total': it['total'],
                'needsReview': it.get('needs_review', False),
            } for it in items
        ],
        'notes': notes,
    }
