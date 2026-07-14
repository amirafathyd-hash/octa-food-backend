import os
import shutil
import subprocess
import tempfile
from collections import defaultdict
from copy import deepcopy

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SALADS_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "data", "Salads_Order_Costing.xlsm")
_TEMPLATE_STATE_CACHE = {}


def _as_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _rounded(value):
    if isinstance(value, float):
        return round(value, 3)
    return value


def _fill_rgb(cell):
    color = cell.fill.fgColor
    if color and color.type == "rgb" and color.rgb:
        return "#" + color.rgb[-6:]
    return None


def _cell_payload(ws_formula, ws_values, row, col):
    formula_cell = ws_formula.cell(row=row, column=col)
    value_cell = ws_values.cell(row=row, column=col)
    formula = formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else None
    return {
        "address": formula_cell.coordinate,
        "row": row,
        "col": col,
        "value": _rounded(value_cell.value if formula else formula_cell.value),
        "formula": formula,
        "editable": True,
        "fill": _fill_rgb(formula_cell),
        "bold": bool(formula_cell.font.bold),
        "align": formula_cell.alignment.horizontal,
        "number_format": formula_cell.number_format,
    }


def _soffice_bin():
    return os.environ.get("SOFFICE_BIN") or shutil.which("soffice") or "soffice"


def recalc_workbook_to_xlsx(xlsm_path):
    out_dir = tempfile.mkdtemp(prefix="salads_recalc_")
    profile_dir = tempfile.mkdtemp(prefix="salads_lo_profile_")
    cmd = [
        _soffice_bin(),
        f"-env:UserInstallation=file://{profile_dir}",
        "--headless",
        "--convert-to",
        "xlsx",
        "--outdir",
        out_dir,
        xlsm_path,
    ]
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=120)
    if proc.returncode != 0:
        raise RuntimeError((proc.stderr or proc.stdout or "LibreOffice failed").strip())
    return os.path.join(out_dir, os.path.splitext(os.path.basename(xlsm_path))[0] + ".xlsx")


def export_workbook_to_pdf(workbook_path):
    out_dir = tempfile.mkdtemp(prefix="salads_pdf_")
    profile_dir = tempfile.mkdtemp(prefix="salads_lo_pdf_profile_")
    cmd = [
        _soffice_bin(),
        f"-env:UserInstallation=file://{profile_dir}",
        "--headless",
        "--convert-to",
        "pdf",
        "--outdir",
        out_dir,
        workbook_path,
    ]
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=120)
    if proc.returncode != 0:
        raise RuntimeError((proc.stderr or proc.stdout or "LibreOffice PDF export failed").strip())
    return os.path.join(out_dir, os.path.splitext(os.path.basename(workbook_path))[0] + ".pdf")


def _apply_edits(wb, edits):
    for edit in edits or []:
        sheet = edit.get("sheet")
        address = edit.get("address")
        value = edit.get("value")
        if not sheet or not address or sheet not in wb.sheetnames:
            continue
        number = _as_number(value)
        wb[sheet][address] = number if number is not None and str(value).strip() != "" else value


def _norm_text(value):
    return " ".join(str(value or "").replace("\u00a0", " ").split()).strip().lower()


def _extract_uploaded_counts(upload_path, known_salads):
    known = {_norm_text(item["name"]): item for item in known_salads if item.get("name")}
    matched = {}
    if not known:
        return matched
    wb = load_workbook(upload_path, data_only=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                row_values = [cell.value for cell in row]
                for idx, value in enumerate(row_values):
                    key = _norm_text(value)
                    if key not in known or key in matched:
                        continue
                    candidates = row_values[idx + 1:idx + 5] + row_values[max(0, idx - 3):idx]
                    for candidate in candidates:
                        number = _as_number(candidate)
                        if number is not None:
                            matched[key] = number
                            break
    finally:
        wb.close()
    return matched


def extract_workbook_state(workbook_path):
    wb_formula = load_workbook(workbook_path, data_only=False, keep_vba=True)
    wb_values = load_workbook(workbook_path, data_only=True, keep_vba=True)
    sheets = []
    for sheet_name in wb_formula.sheetnames:
        ws_formula = wb_formula[sheet_name]
        ws_values = wb_values[sheet_name]
        rows = []
        for row in range(1, ws_formula.max_row + 1):
            rows.append([_cell_payload(ws_formula, ws_values, row, col) for col in range(1, ws_formula.max_column + 1)])
        sheets.append({
            "name": sheet_name,
            "max_row": ws_formula.max_row,
            "max_col": ws_formula.max_column,
            "columns": [get_column_letter(col) for col in range(1, ws_formula.max_column + 1)],
            "rows": rows,
        })
    wb_formula.close()
    wb_values.close()
    return {"sheets": sheets}


def extract_dashboard_state(workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    user = wb["User"]
    ordering = wb["Ordering"]
    usage = wb["Usage"]

    salads = []
    for row in range(4, user.max_row + 1):
        name = user[f"G{row}"].value
        if not name:
            continue
        salads.append({
            "row": row,
            "name": name,
            "count": _rounded(user[f"H{row}"].value),
        })

    ingredients = []
    for row in range(1, ordering.max_row + 1):
        item = ordering[f"A{row}"].value
        category = ordering[f"B{row}"].value
        if row != 1 and not item and not category:
            continue
        ingredients.append({
            "row": row,
            "item": item,
            "category": category,
            "unit": ordering[f"C{row}"].value,
            "daily_weight": _rounded(ordering[f"D{row}"].value),
            "weekly_weight": _rounded(ordering[f"E{row}"].value),
            "daily_order": _rounded(ordering[f"L{row}"].value),
            "fill": _fill_rgb(ordering[f"A{row}"]),
        })

    costs = defaultdict(float)
    for row in range(2, usage.max_row + 1):
        salad = usage[f"B{row}"].value
        cost = _as_number(usage[f"L{row}"].value)
        if salad and cost is not None:
            costs[salad] += cost
    for salad in salads:
        count = _as_number(salad["count"]) or 0
        total = costs.get(salad["name"], 0)
        salad["unit_cost"] = round(total / count, 3) if count else round(total, 3)
        salad["total_cost"] = round(total, 3)

    day = _rounded(user["K4"].value)
    wb.close()
    return {"salads": salads, "ingredients": ingredients, "day": day}


def get_salads_template_state(template_path=SALADS_TEMPLATE_PATH):
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Salads_Order_Costing.xlsm غير موجود في data")
    stat = os.stat(template_path)
    cache_key = (template_path, stat.st_mtime_ns, stat.st_size)
    if cache_key in _TEMPLATE_STATE_CACHE:
        return deepcopy(_TEMPLATE_STATE_CACHE[cache_key])
    state = extract_dashboard_state(template_path)
    state.update(extract_workbook_state(template_path))
    _TEMPLATE_STATE_CACHE.clear()
    _TEMPLATE_STATE_CACHE[cache_key] = deepcopy(state)
    return state


def recalculate_salads_with_edits(edits, template_path=SALADS_TEMPLATE_PATH):
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Salads_Order_Costing.xlsm غير موجود في data")
    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    _apply_edits(wb, edits)
    out_path = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
    wb.save(out_path)
    wb.close()
    recalculated = recalc_workbook_to_xlsx(out_path)
    state = extract_dashboard_state(recalculated)
    state.update(extract_workbook_state(recalculated))
    return state


def update_salads_counts_from_upload(file_storage, template_path=SALADS_TEMPLATE_PATH):
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Salads_Order_Costing.xlsm غير موجود في data")
    suffix = os.path.splitext(file_storage.filename or "")[1].lower()
    if suffix not in (".xlsx", ".xlsm", ".xls"):
        raise ValueError("ملف الأعداد لازم يكون Excel")
    upload_path = tempfile.NamedTemporaryFile(suffix=suffix or ".xlsx", delete=False).name
    file_storage.seek(0)
    file_storage.save(upload_path)

    current_xlsx = recalc_workbook_to_xlsx(template_path)
    current_state = extract_dashboard_state(current_xlsx)
    matched = _extract_uploaded_counts(upload_path, current_state["salads"])
    if not matched:
        raise ValueError("ملف الأعداد مفيهوش أسماء سلطات مطابقة للشيت الرئيسي")

    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    try:
        ws = wb["User"]
        changed = 0
        for salad in current_state["salads"]:
            key = _norm_text(salad["name"])
            if key not in matched:
                continue
            ws[f"H{salad['row']}"] = matched[key]
            changed += 1
        out_path = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
        wb.save(out_path)
    finally:
        wb.close()

    recalculated = recalc_workbook_to_xlsx(out_path)
    state = extract_dashboard_state(recalculated)
    state.update(extract_workbook_state(recalculated))
    return state, {"matched_count": changed}


def _updated_workbook(edits, template_path=SALADS_TEMPLATE_PATH):
    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    _apply_edits(wb, edits)
    out_path = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
    day = _as_number(wb["User"]["K4"].value) or 1
    wb.save(out_path)
    wb.close()
    return recalc_workbook_to_xlsx(out_path), int(day)


def export_salads_excel_with_edits(edits, template_path=SALADS_TEMPLATE_PATH):
    xlsx, day = _updated_workbook(edits, template_path)
    return xlsx, {"day_no": day}


def export_salads_pdf_with_edits(edits, template_path=SALADS_TEMPLATE_PATH):
    xlsx, day = _updated_workbook(edits, template_path)
    return export_workbook_to_pdf(xlsx), {"day_no": day}


def replace_salads_template(file_storage, template_path=SALADS_TEMPLATE_PATH):
    suffix = os.path.splitext(file_storage.filename or "")[1].lower()
    if suffix != ".xlsm":
        raise ValueError("الشيت الرئيسي لازم يكون .xlsm")
    upload_path = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
    file_storage.seek(0)
    file_storage.save(upload_path)
    wb = load_workbook(upload_path, data_only=False, keep_vba=True)
    try:
        for required in ["User", "Ordering", "Usage"]:
            if required not in wb.sheetnames:
                raise ValueError(f"الشيت الجديد لازم يحتوي على {required}")
    finally:
        wb.close()
    shutil.copyfile(upload_path, template_path)
    return get_salads_template_state(template_path), {"template_file": os.path.basename(template_path)}


def export_salads_cost_report_with_edits(edits, template_path=SALADS_TEMPLATE_PATH):
    xlsx, day = _updated_workbook(edits, template_path)
    source = load_workbook(xlsx, data_only=True)
    state = extract_dashboard_state(xlsx)

    report = Workbook()
    summary = report.active
    summary.title = "Summary"
    summary.sheet_view.showGridLines = False
    summary.append(["Metric", "Value"])
    summary.append(["Day", day])
    summary.append(["Salads Count", len(state["salads"])])
    summary.append(["Total Cost", round(sum(s["total_cost"] for s in state["salads"]), 3)])

    salad_ws = report.create_sheet("Salad Costs")
    salad_ws.append(["Salad", "Count", "Unit Cost", "Total Cost"])
    for item in state["salads"]:
        salad_ws.append([item["name"], item["count"], item["unit_cost"], item["total_cost"]])

    ing_ws = report.create_sheet("Ordering Map")
    ing_ws.append(["Item", "Category", "Unit", "Daily Weight", "Weekly Weight", "Daily Order"])
    for item in state["ingredients"][1:]:
        ing_ws.append([item["item"], item["category"], item["unit"], item["daily_weight"], item["weekly_weight"], item["daily_order"]])

    usage = source["Usage"]
    usage_ws = report.create_sheet("Usage Details")
    usage_ws.append(["Salad", "Ingredient", "Weight", "Daily Qty", "Weekly Qty", "Cost"])
    for row in range(2, usage.max_row + 1):
        if usage[f"B{row}"].value:
            usage_ws.append([usage[f"B{row}"].value, usage[f"C{row}"].value, usage[f"D{row}"].value, usage[f"G{row}"].value, usage[f"H{row}"].value, usage[f"L{row}"].value])

    for ws in report.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="70306F")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.000"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25

    if salad_ws.max_row > 1:
        chart = BarChart()
        chart.title = "Total Cost by Salad"
        chart.y_axis.title = "Cost"
        data = Reference(salad_ws, min_col=4, min_row=1, max_row=salad_ws.max_row)
        cats = Reference(salad_ws, min_col=1, min_row=2, max_row=salad_ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend = None
        chart.height = 8
        chart.width = 18
        summary.add_chart(chart, "D2")

    out_path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
    report.save(out_path)
    source.close()
    return out_path, {"day_no": day, "salads_count": len(state["salads"])}


def export_salads_cost_report_pdf_with_edits(edits, template_path=SALADS_TEMPLATE_PATH):
    report, meta = export_salads_cost_report_with_edits(edits, template_path)
    return export_workbook_to_pdf(report), meta
