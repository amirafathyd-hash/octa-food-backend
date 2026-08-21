"""Weekly packaging-order extraction and inventory endpoints."""

from collections import OrderedDict
from datetime import datetime, timezone
import base64
import csv
import hashlib
import io
import json
import math
import os
import re
import uuid

from flask import Blueprint, jsonify, request, send_file
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont
import arabic_reshaper
import requests
try:
    from bidi.algorithm import get_display
except ImportError:
    from bidi import get_display

from db import execute_with_retry, get_client


packaging_orders_bp = Blueprint("packaging_orders", __name__)

DAY_ORDER = ("Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday")
DAY_AR = {
    "Saturday": "السبت", "Sunday": "الأحد", "Monday": "الإثنين",
    "Tuesday": "الثلاثاء", "Wednesday": "الأربعاء", "Thursday": "الخميس",
}
PACKING_ALIASES = {
    "morning": ("Morning Packing", "Morn Packing", "Packing Morn"),
    "evening": ("Evening Packing", "Even Packing", "Packing Even"),
}
COUNT_ITEM_COLUMNS = ((3, 4), (5, 6), (7, 8), (9, 10), (11, 12))
FONT_DIR = os.path.join(os.path.dirname(__file__), "fonts")
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp"}
CUSTOMER_COUNT_VISION_PROMPT = """\
اقرأ صورة جدول التشغيل هذه واستخرج فقط يوم الأسبوع وإجمالي عدد العملاء.
ابحث عن صف المجموع النهائي في عمود «عدد التطبيق» أو «عدد العملاء». قد يظهر
المجموع أكثر من مرة في الصورة؛ إذا كانت القيم متطابقة استخدم القيمة مرة واحدة.
اليوم يجب أن يكون واحدًا من Saturday, Sunday, Monday, Tuesday, Wednesday, Thursday.
ارجع JSON فقط بدون شرح بهذا الشكل:
{"day":"Monday","customer_count":879}
"""


class PackagingWorkbookError(ValueError):
    pass


def _clean(value):
    return re.sub(r"\s+", " ", str(value or "").strip())


def _key(value):
    text = _clean(value).casefold()
    text = text.translate(str.maketrans("أإآىةؤئ", "ااايهوي"))
    text = re.sub(r"[\sـ_\-]+", "", text)
    return text


def _number(value):
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean(value).replace(",", "")
    try:
        return float(text)
    except (TypeError, ValueError):
        return 0.0


def _canonical_item(value):
    raw = _clean(value)
    key = _key(raw)
    digits = "".join(re.findall(r"\d+", key))
    if "مستطيل" in key:
        size = "16" if "16" in digits else ("28" if "28" in digits else "")
        return f"صحن مستطيل {size} أونز".strip()
    if "ساندوتش" in key or "ساندوتش" in key:
        return "صحن ساندوتش 32 أونز"
    if "مقسم" in key:
        return "صحن مقسم 32 أونز"
    if "دائري" in key or "دايري" in key:
        size = "20" if "20" in digits else ("32" if "32" in digits else "")
        white = " أبيض" if "ابيض" in key else ""
        return f"صحن دائري{white} {size} أونز".strip()
    if "مكرون" in key:
        return "علبة مكرونة"
    if "سلط" in key and "250" in digits:
        return "علبة سلطة 250 مل"
    if "فواكه" in key and "200" in digits:
        return "علبة فواكه 200 مل"
    if "كيس" in key and "5" in digits and "شفاف" in key:
        return "كيس شفاف 5×5"
    if "ذهبي" in key and "2" in digits:
        return "علبة ذهبية 2 أونز"
    if ("علب" in key or "علبه" in key) and "2" in digits and "اونز" in key:
        return "علبة 2 أونز"
    return raw


def _find_sheet(wb, aliases):
    exact = {_key(name): name for name in wb.sheetnames}
    for alias in aliases:
        if _key(alias) in exact:
            return exact[_key(alias)]
    # Structural fallback: only accept sheets whose first cell identifies a packing schedule.
    for name in wb.sheetnames:
        title = _key(wb[name].cell(1, 1).value)
        if "pack" in _key(name) and ("schedule" in title or "جدول" in title):
            wanted_morning = any("morn" in _key(alias) for alias in aliases)
            is_morning = "morn" in _key(name) or "صباح" in title
            if wanted_morning == is_morning:
                return name
    return None


def _read_day_name(ws):
    haystack = " ".join(_clean(ws.cell(row, col).value) for row in range(1, 4) for col in range(1, 4))
    lowered = haystack.casefold()
    for day in DAY_ORDER:
        if day.casefold() in lowered:
            return day
    for day, arabic in DAY_AR.items():
        if _key(arabic) in _key(haystack):
            return day
    return ""


def _sheet_items(ws):
    items = OrderedDict()
    max_row = min(ws.max_row or 0, 5000)
    for values in ws.iter_rows(min_row=4, max_row=max_row, min_col=1, max_col=12, values_only=True):
        first = _key(values[0] if values else None)
        if first in {"total", "totals", "الاجمالي", "الاجماليات"}:
            continue
        for qty_col, item_col in COUNT_ITEM_COLUMNS:
            qty = _number(values[qty_col - 1] if len(values) >= qty_col else None)
            raw_item = _clean(values[item_col - 1] if len(values) >= item_col else None)
            if qty <= 0 or not raw_item:
                continue
            item = _canonical_item(raw_item)
            items[item] = items.get(item, 0.0) + qty
    return items


def _customer_count_from_workbook(wb):
    """Read the daily customer total by structure, never by a fixed tab name."""
    count_headers = (
        "عدد التطبيق", "عدد العملاء", "اجمالي العملاء", "إجمالي العملاء",
        "customer count", "customers", "application count",
    )
    total_labels = ("المجموع", "الاجمالي", "الإجمالي", "grand total", "total")
    for ws in wb.worksheets:
        max_row = min(ws.max_row or 0, 500)
        max_col = min(ws.max_column or 0, 40)
        if not max_row or not max_col:
            continue
        has_count_header = False
        direct_candidates = []
        total_candidates = []
        # Sequential iteration is dramatically faster for read-only workbooks.
        # Random ws.cell() access re-reads worksheet XML repeatedly and made a
        # six-day upload appear to hang for several minutes.
        for values in ws.iter_rows(
            min_row=1, max_row=max_row, min_col=1, max_col=max_col,
            values_only=True,
        ):
            values = list(values)
            keys = [_key(value) for value in values]
            for col, key in enumerate(keys):
                if any(_key(label) in key for label in count_headers):
                    has_count_header = True
                    for offset in (1, -1, 2):
                        index = col + offset
                        if 0 <= index < len(values):
                            number = _number(values[index])
                            if number > 0:
                                direct_candidates.append(number)
            if any(any(_key(label) == key or _key(label) in key for label in total_labels) for key in keys):
                numbers = [_number(value) for value in values]
                total_candidates.extend(number for number in numbers if 0 < number < 1_000_000)
        if has_count_header:
            candidates = total_candidates or direct_candidates
            if candidates:
                return int(round(max(candidates)))
    return 0


def _day_from_workbook(wb):
    """Find the operating day from content across all sheets, not a tab name."""
    for ws in wb.worksheets:
        day = _read_day_name(ws)
        if day:
            return day
    return ""


def _day_from_filename(filename):
    text = _clean(filename)
    lowered = text.casefold()
    for day in DAY_ORDER:
        if day.casefold() in lowered:
            return day
    for day, arabic in DAY_AR.items():
        if _key(arabic) in _key(text):
            return day
    return ""


def _customer_count_from_image(raw, filename):
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise PackagingWorkbookError("قراءة الصور تحتاج ANTHROPIC_API_KEY في إعدادات Railway")
    extension = os.path.splitext(filename or "")[1].lower()
    media_type = {
        ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
        ".webp": "image/webp",
    }.get(extension, "image/jpeg")
    payload = {
        "model": os.environ.get("ANTHROPIC_VISION_MODEL") or os.environ.get("ANTHROPIC_MODEL") or "claude-sonnet-4-20250514",
        "max_tokens": 500,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type,
                 "data": base64.b64encode(raw).decode("ascii")}},
                {"type": "text", "text": CUSTOMER_COUNT_VISION_PROMPT},
            ],
        }],
    }
    response = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
        json=payload, timeout=90,
    )
    if response.status_code != 200:
        try:
            message = response.json().get("error", {}).get("message") or response.text
        except Exception:
            message = response.text
        raise PackagingWorkbookError(f"تعذر قراءة الصورة {filename}: {message[:300]}")
    content = response.json().get("content", [])
    text = "".join(block.get("text", "") for block in content if block.get("type") == "text")
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if not match:
        raise PackagingWorkbookError(f"تعذر استخراج عدد العملاء من الصورة {filename}")
    try:
        data = json.loads(match.group(0))
    except json.JSONDecodeError as exc:
        raise PackagingWorkbookError(f"نتيجة قراءة الصورة {filename} غير صالحة") from exc
    day_text = _clean(data.get("day"))
    day = next((item for item in DAY_ORDER if item.casefold() == day_text.casefold()), "")
    day = day or _day_from_filename(filename)
    count = max(0, int(round(_number(data.get("customer_count")))))
    if not day:
        raise PackagingWorkbookError(f"تعذر تحديد يوم التشغيل في الصورة {filename}")
    if not count:
        raise PackagingWorkbookError(f"تعذر تحديد إجمالي عدد العملاء في الصورة {filename}")
    return day, count


def extract_customer_count_files(files):
    counts = OrderedDict()
    sources = []
    for uploaded in files:
        filename = uploaded.filename or "file"
        extension = os.path.splitext(filename)[1].lower()
        raw = uploaded.read()
        uploaded.seek(0)
        if extension in IMAGE_EXTENSIONS:
            day, count = _customer_count_from_image(raw, filename)
            source_type = "image"
        elif extension in {".xlsx", ".xlsm"}:
            try:
                workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            except Exception as exc:
                raise PackagingWorkbookError(f"تعذر قراءة ملف Excel {filename}: {exc}") from exc
            try:
                day = _day_from_workbook(workbook) or _day_from_filename(filename)
                count = max(0, int(round(_customer_count_from_workbook(workbook))))
            finally:
                workbook.close()
            source_type = "excel"
            if not day:
                raise PackagingWorkbookError(f"تعذر تحديد يوم التشغيل داخل {filename}")
            if not count:
                raise PackagingWorkbookError(f"تعذر العثور على إجمالي عدد العملاء داخل {filename}")
        else:
            raise PackagingWorkbookError(f"صيغة الملف {filename} غير مدعومة؛ ارفع صورة أو Excel")
        if day in counts:
            raise PackagingWorkbookError(f"تم رفع أكثر من ملف ليوم {DAY_AR.get(day, day)}")
        counts[day] = count
        sources.append({"filename": filename, "type": source_type, "day": day,
                        "day_ar": DAY_AR.get(day, day), "customers": count})
    ordered = OrderedDict((day, counts[day]) for day in DAY_ORDER if day in counts)
    return ordered, sources


def _supply_plan(customer_counts, spoon_carton_size=200, mode="weekly"):
    spoon_carton_size = 300 if int(_number(spoon_carton_size)) == 300 else 200
    normalized = OrderedDict()
    for day in DAY_ORDER:
        if day in (customer_counts or {}):
            normalized[day] = max(0, int(round(_number(customer_counts.get(day)))))
    daily = []
    for day, customers in normalized.items():
        daily.append({
            "day": day,
            "day_ar": DAY_AR.get(day, day),
            "customers": customers,
            "bag_cartons": int(math.ceil(customers / 200.0)) if customers else 0,
            "spoon_cartons": int(math.ceil((customers * 1.2) / spoon_carton_size)) if customers else 0,
        })
    total_customers = sum(item["customers"] for item in daily)
    weekly = {
        "customers": total_customers,
        "bag_cartons": int(math.ceil(total_customers / 200.0)) if total_customers else 0,
        "spoon_cartons": int(math.ceil((total_customers * 1.2) / spoon_carton_size)) if total_customers else 0,
    }
    daily_totals = {
        "customers": total_customers,
        "bag_cartons": sum(item["bag_cartons"] for item in daily),
        "spoon_cartons": sum(item["spoon_cartons"] for item in daily),
    }
    return {
        "mode": "daily" if mode == "daily" else "weekly",
        "spoon_carton_size": spoon_carton_size,
        "bag_carton_size": 200,
        "spoon_safety_percent": 20,
        "daily": daily,
        "weekly": weekly,
        "daily_totals": daily_totals,
    }


def extract_packaging_files(files):
    days = OrderedDict()
    sources = []
    seen_files = set()
    for uploaded in files:
        raw = uploaded.read()
        uploaded.seek(0)
        signature = (len(raw), raw[:4096], raw[-4096:])
        if signature in seen_files:
            continue
        seen_files.add(signature)
        try:
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        except Exception as exc:
            raise PackagingWorkbookError(f"تعذر قراءة الملف {uploaded.filename}: {exc}") from exc
        matched = []
        customer_count = _customer_count_from_workbook(wb)
        day_name = ""
        file_items = OrderedDict()
        for shift, aliases in PACKING_ALIASES.items():
            sheet_name = _find_sheet(wb, aliases)
            if not sheet_name:
                continue
            ws = wb[sheet_name]
            day_name = day_name or _read_day_name(ws)
            shift_items = _sheet_items(ws)
            for item, qty in shift_items.items():
                file_items[item] = file_items.get(item, 0.0) + qty
            matched.append({"shift": shift, "sheet": sheet_name, "units": round(sum(shift_items.values()), 3)})
        wb.close()
        if len(matched) != 2:
            raise PackagingWorkbookError(
                f"الملف {uploaded.filename} لازم يحتوي شيت صباح وشيت مساء للتغليف"
            )
        if not day_name:
            raise PackagingWorkbookError(f"تعذر تحديد يوم التشغيل داخل {uploaded.filename}")
        if day_name in days:
            raise PackagingWorkbookError(f"تم رفع أكثر من ملف ليوم {DAY_AR[day_name]}")
        days[day_name] = file_items
        sources.append({
            "filename": uploaded.filename, "day": day_name, "day_ar": DAY_AR[day_name],
            "sheets": matched, "customers": customer_count,
        })
    if not days:
        raise PackagingWorkbookError("لم يتم العثور على ملفات تغليف صالحة")
    ordered_days = OrderedDict((day, days[day]) for day in DAY_ORDER if day in days)
    all_items = sorted({item for rows in ordered_days.values() for item in rows}, key=_key)
    rows = []
    for item in all_items:
        daily = {day: round(ordered_days[day].get(item, 0.0), 3) for day in ordered_days}
        rows.append({"item": item, "daily": daily, "total": round(sum(daily.values()), 3), "unit": "قطعة"})
    return ordered_days, rows, sources


def _read_inventory_state():
    try:
        rows = execute_with_retry(
            get_client().table("upload_log").select("message,created_at")
            .eq("file_type", "packaging_inventory").order("created_at", desc=True).limit(1)
        ).data or []
        if rows:
            data = json.loads(rows[0].get("message") or "{}")
            return data if isinstance(data, dict) else {}
    except Exception:
        pass
    return {}


def _save_inventory_state(items):
    state = {
        "items": {str(name): max(0, _number(value)) for name, value in items.items() if _clean(name)},
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    execute_with_retry(get_client().table("upload_log").insert({
        "file_type": "packaging_inventory", "file_name": "packaging_inventory",
        "item_date": None, "message": json.dumps(state, ensure_ascii=False), "level": "info",
    }))
    return state


def _inventory_rows_from_matrix(matrix):
    """Find an item/quantity pair without depending on exact sheet names or columns."""
    if not matrix:
        return {}
    item_words = ("الصنف", "اسم الصنف", "item", "items", "product", "name")
    qty_words = ("المخزون", "الكمية", "الرصيد", "stock", "inventory", "quantity", "qty")
    header_index = item_col = qty_col = None
    for row_index, row in enumerate(matrix[:25]):
        keys = [_key(value) for value in row]
        found_item = next((i for i, key in enumerate(keys) if any(_key(word) in key for word in item_words)), None)
        found_qty = next((i for i, key in enumerate(keys) if any(_key(word) in key for word in qty_words)), None)
        if found_item is not None and found_qty is not None and found_item != found_qty:
            header_index, item_col, qty_col = row_index, found_item, found_qty
            break
    if header_index is None:
        # Fallback: score every column as text/item or numeric/quantity.
        width = max((len(row) for row in matrix), default=0)
        best = None
        for text_col in range(width):
            for number_col in range(width):
                if text_col == number_col:
                    continue
                score = 0
                for row in matrix[:100]:
                    name = _clean(row[text_col] if text_col < len(row) else "")
                    raw_qty = row[number_col] if number_col < len(row) else None
                    qty = _number(raw_qty)
                    if name and not re.fullmatch(r"[\d\s.,]+", name) and qty > 0:
                        score += 1
                if best is None or score > best[0]:
                    best = (score, text_col, number_col)
        if not best or best[0] < 1:
            raise PackagingWorkbookError("لم أجد عمود اسم الصنف وعمود المخزون في الملف")
        header_index, item_col, qty_col = -1, best[1], best[2]
    items = OrderedDict()
    for row in matrix[header_index + 1:]:
        name = _clean(row[item_col] if item_col < len(row) else "")
        qty = _number(row[qty_col] if qty_col < len(row) else None)
        if not name or qty < 0:
            continue
        key = _key(name)
        if any(_key(word) == key for word in item_words) or any(word in key for word in ("total", "اجمالي")):
            continue
        canonical = _canonical_item(name)
        items[canonical] = items.get(canonical, 0.0) + qty
    if not items:
        raise PackagingWorkbookError("الملف لا يحتوي أصناف مخزون صالحة")
    return dict(items)


def _read_inventory_spreadsheet(upload):
    filename = _clean(upload.filename)
    ext = os.path.splitext(filename)[1].lower()
    raw = upload.read()
    if ext == ".csv":
        text = raw.decode("utf-8-sig", errors="replace")
        matrix = list(csv.reader(io.StringIO(text)))
    else:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        matrix = []
        for ws in wb.worksheets:
            candidate = [list(row) for row in ws.iter_rows(values_only=True)]
            try:
                return _inventory_rows_from_matrix(candidate)
            except PackagingWorkbookError:
                matrix.extend(candidate)
    return _inventory_rows_from_matrix(matrix)


def _read_inventory_image(upload):
    api_key = os.environ.get("ANTHROPIC_API_KEY", "")
    if not api_key:
        raise PackagingWorkbookError("قراءة الصور تحتاج ANTHROPIC_API_KEY على السيرفر")
    raw = upload.read()
    ext = os.path.splitext(_clean(upload.filename))[1].lower()
    media_type = {".png": "image/png", ".webp": "image/webp"}.get(ext, "image/jpeg")
    prompt = """اقرأ هذه الصورة كقائمة مخزون تغليف. استخرج اسم كل صنف والكمية المتاحة فقط.
أعد JSON فقط بهذا الشكل: {\"items\":[{\"name\":\"اسم الصنف\",\"quantity\":10}]}.
لا تخمن رقما غير ظاهر، وتجاهل العناوين والإجماليات والتواريخ."""
    response = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={"x-api-key": api_key, "anthropic-version": "2023-06-01", "content-type": "application/json"},
        json={
            "model": os.environ.get("ANTHROPIC_VISION_MODEL") or os.environ.get("ANTHROPIC_MODEL") or "claude-sonnet-4-20250514",
            "max_tokens": 3000,
            "messages": [{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": base64.b64encode(raw).decode("ascii")}},
                {"type": "text", "text": prompt},
            ]}],
        }, timeout=90,
    )
    if response.status_code != 200:
        raise PackagingWorkbookError(f"تعذر قراءة الصورة: HTTP {response.status_code}")
    text = "".join(block.get("text", "") for block in response.json().get("content", []) if block.get("type") == "text")
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if not match:
        raise PackagingWorkbookError("لم أستطع استخراج قائمة مخزون من الصورة")
    data = json.loads(match.group(0))
    items = {}
    for row in data.get("items", []):
        name = _canonical_item(row.get("name"))
        if _clean(name):
            items[name] = items.get(name, 0.0) + max(0, _number(row.get("quantity")))
    if not items:
        raise PackagingWorkbookError("الصورة لا تحتوي أصناف مخزون واضحة")
    return items


def _save_week_event(record):
    execute_with_retry(get_client().table("upload_log").insert({
        "file_type": "packaging_week_record", "file_name": record["id"],
        "item_date": record.get("date") or None,
        "message": json.dumps(record, ensure_ascii=False), "level": "info",
    }))


def _week_records():
    rows = (execute_with_retry(
        get_client().table("upload_log").select("file_name,message,created_at")
        .eq("file_type", "packaging_week_record").order("created_at", desc=True).limit(500)
    ).data or [])
    latest = OrderedDict()
    for row in rows:
        try:
            record = json.loads(row.get("message") or "{}")
        except Exception:
            continue
        record_id = _clean(record.get("id") or row.get("file_name"))
        if record_id and record_id not in latest:
            record["created_at"] = row.get("created_at")
            latest[record_id] = record
    return list(latest.values())


def _consume_inventory(current, rows):
    updated = {str(key): max(0, _number(value)) for key, value in (current or {}).items()}
    saved_rows = []
    for row in rows or []:
        item = _canonical_item(row.get("item"))
        total = max(0, _number(row.get("total")))
        before = max(0, _number(updated.get(item, 0)))
        used = min(before, total)
        remaining = max(0, total - used)
        updated[item] = round(before - used, 3)
        saved_rows.append({
            "item": item, "total": round(total, 3), "inventory_before": round(before, 3),
            "used_inventory": round(used, 3), "remaining": round(remaining, 3),
            "daily": {
                str(day): round(max(0, _number(value)), 3)
                for day, value in (row.get("daily") or {}).items()
            },
            "unit": _clean(row.get("unit")) or "قطعة",
        })
    return updated, saved_rows


def _rtl(value):
    return get_display(arabic_reshaper.reshape(_clean(value)))


def _fonts():
    regular = os.path.join(FONT_DIR, "IBMPlexSansArabic-Regular.ttf")
    bold = os.path.join(FONT_DIR, "IBMPlexSansArabic-Bold.ttf")
    return regular, bold


def build_supplies_png(payload):
    """Render the standalone bags-and-cutlery calculator as a branded PNG."""
    settings = payload.get("supplies") or {}
    supplies = _supply_plan(
        payload.get("customer_counts") or {},
        settings.get("spoon_carton_size", 200),
        settings.get("mode", "weekly"),
    )
    if not supplies["weekly"]["customers"]:
        raise PackagingWorkbookError("لا توجد أعداد عملاء لإنشاء التقرير")

    daily_mode = supplies["mode"] == "daily"
    daily_rows = supplies["daily"] if daily_mode else []
    width = 1600
    margin = 55
    header_h = 175
    cards_top = 215
    cards_h = 230
    table_top = 485
    table_head_h = 58
    row_h = 58
    footer_h = 58
    height = (table_top + table_head_h + row_h * len(daily_rows) + footer_h + 35) if daily_mode else 585
    image = Image.new("RGB", (width, height), "#F7F4EE")
    draw = ImageDraw.Draw(image)
    regular_path, bold_path = _fonts()
    f_title = ImageFont.truetype(bold_path, 42)
    f_sub = ImageFont.truetype(regular_path, 20)
    f_card = ImageFont.truetype(bold_path, 31)
    f_value = ImageFont.truetype(bold_path, 62)
    f_head = ImageFont.truetype(bold_path, 21)
    f_cell = ImageFont.truetype(regular_path, 20)
    f_num = ImageFont.truetype(bold_path, 22)

    def rtl_text(xy, value, font, fill="#183B42", anchor="ra"):
        text = _clean(value)
        try:
            draw.text(xy, text, font=font, fill=fill, anchor=anchor, direction="rtl", language="ar")
        except (KeyError, TypeError, ValueError):
            draw.text(xy, _rtl(text), font=font, fill=fill, anchor=anchor)

    draw.rounded_rectangle((margin, 32, width - margin, header_h), 28, fill="#163B47")
    rtl_text((width - margin - 36, 78), "تقرير احتياج الأكياس والملاعق", f_title, "#FFFFFF")
    mode_text = "تجميع يومي" if daily_mode else "تجميع أسبوعي"
    rtl_text((width - margin - 36, 125), f"{mode_text} · سعة كرتونة الملاعق {supplies['spoon_carton_size']}", f_sub, "#CDE3DF")
    draw.rounded_rectangle((margin + 30, 66, margin + 300, 145), 18, fill="#FFB84D")
    draw.text((margin + 165, 91), f"{supplies['weekly']['customers']:,}", font=f_card, fill="#163B47", anchor="mm")
    rtl_text((margin + 165, 127), "إجمالي العملاء", f_sub, "#163B47", anchor="mm")

    shown = supplies["daily_totals"] if daily_mode else supplies["weekly"]
    gap = 28
    card_w = (width - (2 * margin) - gap) / 2

    def card(x0, fill, title, value, note, kind):
        x1 = x0 + card_w
        draw.rounded_rectangle((x0, cards_top, x1, cards_top + cards_h), 28, fill=fill)
        icon_x, icon_y = x1 - 112, cards_top + cards_h / 2
        draw.rounded_rectangle((icon_x - 62, icon_y - 62, icon_x + 62, icon_y + 62), 26, fill="#FFFFFF")
        if kind == "bag":
            draw.rounded_rectangle((icon_x - 25, icon_y - 19, icon_x + 25, icon_y + 35), 6, outline="#523500", width=6)
            draw.arc((icon_x - 17, icon_y - 40, icon_x + 17, icon_y - 4), 180, 360, fill="#523500", width=6)
            ink = "#523500"
        else:
            draw.ellipse((icon_x - 12, icon_y - 43, icon_x + 12, icon_y - 10), outline="#0E5556", width=6)
            draw.line((icon_x, icon_y - 10, icon_x, icon_y + 43), fill="#0E5556", width=7)
            ink = "#0E5556"
        content_left = x0 + 34
        content_right = icon_x - 82
        content_x = (content_left + content_right) / 2
        rtl_text((content_x, cards_top + 48), title, f_card, ink, anchor="mm")
        draw.text((content_x, cards_top + 116), f"{int(value):,}", font=f_value, fill=ink, anchor="mm")
        rtl_text((content_x, cards_top + 160), "كرتونة مطلوبة", f_head, ink, anchor="mm")
        rtl_text((content_x, cards_top + 205), note, f_sub, ink, anchor="mm")

    card(margin, "#FFD078", "أكياس التغليف", shown["bag_cartons"], "كل كرتونة تحتوي 200 كيس", "bag")
    card(margin + card_w + gap, "#A7E7DC", "الملاعق", shown["spoon_cartons"], f"كل كرتونة تحتوي {supplies['spoon_carton_size']} ملعقة · احتياطي 20%", "spoon")

    if daily_mode:
        columns = [
            ("اليوم", 430), ("عدد العملاء", 330),
            ("كراتين الأكياس", 350), ("كراتين الملاعق", 350),
        ]
        x = margin
        for label, col_w in columns:
            draw.rectangle((x, table_top, x + col_w, table_top + table_head_h), fill="#163B47", outline="#315B63", width=2)
            rtl_text((x + col_w / 2, table_top + table_head_h / 2), label, f_head, "#FFFFFF", anchor="mm")
            x += col_w
        for index, item in enumerate(daily_rows):
            y0 = table_top + table_head_h + index * row_h
            fill = "#FFFFFF" if index % 2 == 0 else "#EAF4F2"
            values = (item["day_ar"], item["customers"], item["bag_cartons"], item["spoon_cartons"])
            x = margin
            for col_index, ((_, col_w), value) in enumerate(zip(columns, values)):
                draw.rectangle((x, y0, x + col_w, y0 + row_h), fill=fill, outline="#D3E0DE", width=1)
                if col_index == 0:
                    rtl_text((x + col_w / 2, y0 + row_h / 2), value, f_cell, "#183B42", anchor="mm")
                else:
                    draw.text((x + col_w / 2, y0 + row_h / 2), f"{int(value):,}", font=f_num, fill="#183B42", anchor="mm")
                x += col_w

    footer_y = height - footer_h - 18
    draw.rounded_rectangle((margin, footer_y, width - margin, footer_y + footer_h), 16, fill="#163B47")
    draw.text((margin + 24, footer_y + footer_h / 2), "OCTA FOOD · DAILY SUPPLIES", font=f_sub, fill="#FFFFFF", anchor="lm")
    rtl_text((width - margin - 24, footer_y + footer_h / 2), "طلبات التغليف", f_sub, "#FFFFFF", anchor="rm")
    output = io.BytesIO()
    image.save(output, format="PNG", optimize=True)
    output.seek(0)
    return output


def build_packaging_png(payload):
    rows = payload.get("rows") or []
    days = [day for day in DAY_ORDER if day in (payload.get("days") or [])]
    if not rows:
        raise PackagingWorkbookError("لا توجد بيانات لإنشاء الصورة")
    supply_settings = payload.get("supplies") or {}
    supplies = _supply_plan(
        payload.get("customer_counts") or {},
        supply_settings.get("spoon_carton_size", 200),
        supply_settings.get("mode", "weekly"),
    )
    show_supplies = supplies["weekly"]["customers"] > 0
    supply_h = (238 + (56 * len(supplies["daily"]) if supplies["mode"] == "daily" else 0)) if show_supplies else 0
    width = 1800
    header_h, table_head_h, row_h, footer_h = 166, 72, 62, 58
    height = header_h + supply_h + table_head_h + row_h * len(rows) + footer_h + 70
    image = Image.new("RGB", (width, height), "#F7F4EE")
    draw = ImageDraw.Draw(image)

    def rtl_text(xy, value, font, fill="#183B42", anchor="ra"):
        text = _clean(value)
        try:
            # On the deployed image Pillow uses libraqm, so pass the original
            # Unicode text and let it shape Arabic exactly once.
            draw.text(
                xy, text, font=font, fill=fill, anchor=anchor,
                direction="rtl", language="ar",
            )
        except (KeyError, TypeError, ValueError):
            # Fallback for Pillow builds without libraqm.
            draw.text(xy, _rtl(text), font=font, fill=fill, anchor=anchor)

    def item_icon(cx, cy, value):
        """Draw a compact line symbol matching the packaging item shape."""
        key = _key(value)
        ink, bg = "#176A69", "#E7F4F1"
        draw.rounded_rectangle((cx - 23, cy - 23, cx + 23, cy + 23), 10, fill=bg, outline="#BED9D3", width=2)
        if "مقسم" in key:
            draw.ellipse((cx - 14, cy - 14, cx + 14, cy + 14), outline=ink, width=3)
            draw.line((cx, cy - 14, cx, cy), fill=ink, width=3)
            draw.line((cx, cy, cx - 11, cy + 9), fill=ink, width=3)
            draw.line((cx, cy, cx + 11, cy + 9), fill=ink, width=3)
        elif "دائري" in key or "دايري" in key:
            draw.ellipse((cx - 15, cy - 15, cx + 15, cy + 15), outline=ink, width=3)
            draw.ellipse((cx - 9, cy - 9, cx + 9, cy + 9), outline=ink, width=2)
        elif "مستطيل" in key:
            draw.rounded_rectangle((cx - 18, cy - 12, cx + 18, cy + 12), 5, outline=ink, width=3)
            draw.rounded_rectangle((cx - 12, cy - 7, cx + 12, cy + 7), 3, outline=ink, width=2)
        elif "ساندوتش" in key:
            draw.arc((cx - 18, cy - 14, cx + 18, cy + 13), 180, 360, fill=ink, width=3)
            draw.rounded_rectangle((cx - 18, cy, cx + 18, cy + 12), 3, outline=ink, width=3)
        elif "سلط" in key or "فواكه" in key:
            draw.line((cx - 17, cy - 7, cx + 17, cy - 7), fill=ink, width=3)
            draw.polygon(((cx - 14, cy - 4), (cx + 14, cy - 4), (cx + 10, cy + 14), (cx - 10, cy + 14)), outline=ink)
        else:
            draw.rounded_rectangle((cx - 16, cy - 13, cx + 16, cy + 14), 5, outline=ink, width=3)
            draw.line((cx - 12, cy - 3, cx + 12, cy - 3), fill=ink, width=2)

    regular_path, bold_path = _fonts()
    f_title = ImageFont.truetype(bold_path, 45)
    f_sub = ImageFont.truetype(regular_path, 21)
    f_head = ImageFont.truetype(bold_path, 22)
    f_cell = ImageFont.truetype(regular_path, 20)
    f_num = ImageFont.truetype(bold_path, 21)
    margin = 45
    draw.rounded_rectangle((margin, 35, width - margin, header_h - 12), 28, fill="#163B47")
    rtl_text((width - margin - 34, 82), "طلبات التغليف الأسبوعية", f_title, "white")
    draw.text((margin + 34, 91), f"{len(rows)}", font=f_title, fill="#FFB84D", anchor="lm")
    rtl_text((margin + 115, 96), "صنف تغليف", f_sub, "white", anchor="lm")

    if show_supplies:
        supply_y = header_h + 12
        draw.rounded_rectangle((margin, supply_y, width - margin, supply_y + supply_h - 18), 24, fill="#FFFFFF", outline="#D7E0DE", width=2)
        rtl_text((width - margin - 28, supply_y + 28), "احتياج الأكياس والملاعق", f_head, "#183B42")
        card_top = supply_y + 68
        gap = 20
        card_w = (width - (2 * margin) - 60 - gap) / 2

        def supply_card(x0, color, title, cartons, detail, kind):
            x1 = x0 + card_w
            draw.rounded_rectangle((x0, card_top, x1, card_top + 128), 20, fill=color)
            icon_x, icon_y = x1 - 76, card_top + 64
            draw.ellipse((icon_x - 36, icon_y - 36, icon_x + 36, icon_y + 36), fill="#FFFFFF")
            if kind == "bag":
                draw.rounded_rectangle((icon_x - 18, icon_y - 14, icon_x + 18, icon_y + 22), 5, outline="#183B42", width=4)
                draw.arc((icon_x - 12, icon_y - 27, icon_x + 12, icon_y - 3), 180, 360, fill="#183B42", width=4)
            else:
                draw.ellipse((icon_x - 7, icon_y - 27, icon_x + 7, icon_y - 8), outline="#183B42", width=4)
                draw.line((icon_x, icon_y - 8, icon_x, icon_y + 25), fill="#183B42", width=5)
            rtl_text((x1 - 130, card_top + 29), title, f_head, "#183B42")
            rtl_text((x1 - 130, card_top + 70), f"{cartons:,} كرتونة", f_title, "#183B42")
            rtl_text((x1 - 130, card_top + 110), detail, f_sub, "#3F666B")

        shown_totals = supplies["daily_totals"] if supplies["mode"] == "daily" else supplies["weekly"]
        supply_card(margin + 20, "#FFF1D8", "أكياس التغليف", shown_totals["bag_cartons"], "200 كيس في الكرتونة", "bag")
        supply_card(margin + 40 + card_w, "#DDF4EF", "ملاعق", shown_totals["spoon_cartons"], f"{supplies['spoon_carton_size']} ملعقة في الكرتونة · احتياطي 20%", "spoon")
        if supplies["mode"] == "daily":
            row_y = card_top + 148
            for item in supplies["daily"]:
                draw.rounded_rectangle((margin + 22, row_y, width - margin - 22, row_y + 44), 11, fill="#F3F7F6")
                rtl_text((width - margin - 42, row_y + 22), item["day_ar"], f_cell, "#183B42", anchor="rm")
                draw.text((width / 2 + 180, row_y + 22), f"{item['customers']:,}", font=f_num, fill="#183B42", anchor="mm")
                rtl_text((width / 2 + 95, row_y + 22), "عميل", f_cell, "#60787B", anchor="rm")
                rtl_text((width / 2 - 40, row_y + 22), f"{item['bag_cartons']} أكياس", f_cell, "#9A5C00", anchor="rm")
                rtl_text((margin + 250, row_y + 22), f"{item['spoon_cartons']} ملاعق", f_cell, "#0F6765", anchor="rm")
                row_y += 56

    table_x0, table_x1 = margin, width - margin
    item_w, total_w, stock_w, needed_w = 420, 155, 155, 180
    day_w = ((table_x1 - table_x0 - item_w - total_w - stock_w - needed_w) / len(days)) if days else 0
    columns = [("item", item_w, "الصنف")]
    columns += [(day, day_w, DAY_AR[day]) for day in days]
    columns += [("total", total_w, "الإجمالي"), ("used_inventory", stock_w, "استخدام المخزون"), ("remaining", needed_w, "المطلوب")]
    y = header_h + supply_h
    x = table_x0
    for _, col_w, label in columns:
        draw.rectangle((x, y, x + col_w, y + table_head_h), fill="#FFB84D", outline="#E6C58D", width=2)
        rtl_text((x + col_w / 2, y + table_head_h / 2), label, f_head, "#15323B", anchor="mm")
        x += col_w
    for index, row in enumerate(rows):
        y0 = y + table_head_h + index * row_h
        fill = "#FFFFFF" if index % 2 == 0 else "#EDF5F3"
        x = table_x0
        for key, col_w, _ in columns:
            draw.rectangle((x, y0, x + col_w, y0 + row_h), fill=fill, outline="#D7E0DE", width=1)
            if key == "item":
                item_icon(x + 35, y0 + row_h / 2, row.get("item"))
                rtl_text((x + col_w - 18, y0 + row_h / 2), row.get("item"), f_cell, "#183B42", anchor="rm")
            else:
                if key in days:
                    value = (row.get("daily") or {}).get(key, 0)
                else:
                    value = row.get(key, 0)
                draw.text((x + col_w / 2, y0 + row_h / 2), f"{float(value or 0):,.0f}", font=f_num, fill="#183B42", anchor="mm")
            x += col_w
    footer_y = y + table_head_h + len(rows) * row_h
    draw.rectangle((margin, footer_y, width - margin, footer_y + footer_h), fill="#163B47")
    draw.text((margin + 20, footer_y + footer_h / 2), "OCTA FOOD · PACKAGING ORDERS", font=f_sub, fill="white", anchor="lm")
    output = io.BytesIO()
    image.save(output, format="PNG", optimize=True)
    output.seek(0)
    return output


def build_packaging_xlsx(payload):
    rows = payload.get("rows") or []
    days = [day for day in DAY_ORDER if day in (payload.get("days") or [])]
    if not rows:
        raise PackagingWorkbookError("لا توجد بيانات لإنشاء Excel")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packaging Orders"
    columns = ["الصنف"] + [DAY_AR[day] for day in days] + ["الإجمالي", "المستخدم من المخزون", "المطلوب بعد المخزون", "الوحدة"]
    ws.append(columns)
    for row in rows:
        ws.append([row.get("item")] + [(row.get("daily") or {}).get(day, 0) for day in days] + [
            row.get("total", 0), row.get("used_inventory", 0), row.get("remaining", 0), row.get("unit", "قطعة")
        ])
    dark, gold, pale = "163B47", "FFB84D", "EDF5F3"
    thin = Side(style="thin", color="CBD8D5")
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=gold)
        cell.font = Font(bold=True, color=dark, size=12)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=pale if cell.row % 2 == 0 else "FFFFFF")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 34
    for index in range(2, len(columns) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width = 16
    ws.freeze_panes = "B2"
    ws.sheet_view.rightToLeft = True
    supply_settings = payload.get("supplies") or {}
    supplies = _supply_plan(
        payload.get("customer_counts") or {},
        supply_settings.get("spoon_carton_size", 200),
        supply_settings.get("mode", "weekly"),
    )
    if supplies["weekly"]["customers"] > 0:
        supply_ws = wb.create_sheet("Bags and Spoons")
        supply_ws.sheet_view.rightToLeft = True
        supply_ws.append(["اليوم", "عدد العملاء", "كراتين الأكياس", "كراتين الملاعق", "سعة كرتونة الملاعق"])
        for item in supplies["daily"]:
            supply_ws.append([item["day_ar"], item["customers"], item["bag_cartons"], item["spoon_cartons"], supplies["spoon_carton_size"]])
        shown_totals = supplies["daily_totals"] if supplies["mode"] == "daily" else supplies["weekly"]
        total_label = "إجمالي الطلب اليومي" if supplies["mode"] == "daily" else "الإجمالي الأسبوعي"
        supply_ws.append([total_label, shown_totals["customers"], shown_totals["bag_cartons"], shown_totals["spoon_cartons"], supplies["spoon_carton_size"]])
        for cell in supply_ws[1]:
            cell.fill = PatternFill("solid", fgColor=gold)
            cell.font = Font(bold=True, color=dark, size=12)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in supply_ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for letter, width_value in zip(("A", "B", "C", "D", "E"), (20, 18, 20, 20, 24)):
            supply_ws.column_dimensions[letter].width = width_value
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@packaging_orders_bp.route("/api/packaging-orders/customer-counts", methods=["POST"])
def packaging_customer_counts():
    files = [file for file in request.files.getlist("files") if file and file.filename]
    if not 1 <= len(files) <= 6:
        return jsonify({"error": "ارفع من ملف واحد إلى 6 ملفات، صور أو Excel"}), 400
    try:
        counts, sources = extract_customer_count_files(files)
        return jsonify({
            "ok": True,
            "days": list(counts.keys()),
            "customer_counts": counts,
            "sources": sources,
            "supplies": _supply_plan(counts, 200, "weekly"),
        })
    except PackagingWorkbookError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"error": f"تعذر قراءة أعداد العملاء: {str(exc)[:180]}"}), 500


@packaging_orders_bp.route("/api/packaging-orders/extract", methods=["POST"])
def packaging_orders_extract():
    files = [file for file in request.files.getlist("files") if file and file.filename]
    if not 1 <= len(files) <= 6:
        return jsonify({"error": "ارفع من ملف واحد إلى 6 ملفات تشغيل"}), 400
    if any(not file.filename.lower().endswith((".xlsx", ".xlsm")) for file in files):
        return jsonify({"error": "مسموح فقط بملفات XLSX وXLSM"}), 400
    try:
        days, rows, sources = extract_packaging_files(files)
        inventory = (_read_inventory_state().get("items") or {})
        for row in rows:
            stock = max(0, _number(inventory.get(row["item"], 0)))
            row["inventory"] = round(stock, 3)
            row["used_inventory"] = round(min(stock, row["total"]), 3)
            row["remaining"] = round(max(0, row["total"] - stock), 3)
        fingerprint = json.dumps({"days": list(days.keys()), "rows": rows}, ensure_ascii=False, sort_keys=True)
        customer_counts = {source["day"]: source.get("customers", 0) for source in sources}
        return jsonify({
            "ok": True, "days": list(days.keys()), "rows": rows, "sources": sources,
            "total_units": round(sum(row["total"] for row in rows), 3),
            "customer_counts": customer_counts,
            "supplies": _supply_plan(customer_counts, 200, "weekly"),
            "run_id": hashlib.sha256(fingerprint.encode("utf-8")).hexdigest()[:24],
        })
    except PackagingWorkbookError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"error": f"تعذر تجهيز طلبات التغليف: {str(exc)[:180]}"}), 500


@packaging_orders_bp.route("/api/packaging-inventory", methods=["GET", "POST"])
def packaging_inventory():
    if request.method == "GET":
        return jsonify({"ok": True, **_read_inventory_state()})
    payload = request.get_json(silent=True) or {}
    items = payload.get("items") or {}
    if not isinstance(items, dict):
        return jsonify({"error": "بيانات المخزون غير صحيحة"}), 400
    try:
        return jsonify({"ok": True, **_save_inventory_state(items)})
    except Exception as exc:
        return jsonify({"error": f"تعذر حفظ المخزون: {str(exc)[:180]}"}), 500


@packaging_orders_bp.route("/api/packaging-inventory/import", methods=["POST"])
def packaging_inventory_import():
    upload = request.files.get("file")
    if not upload or not upload.filename:
        return jsonify({"error": "اختر صورة أو ملف مخزون"}), 400
    ext = os.path.splitext(upload.filename)[1].lower()
    try:
        if ext in (".xlsx", ".xlsm", ".csv"):
            items = _read_inventory_spreadsheet(upload)
        elif ext in (".png", ".jpg", ".jpeg", ".webp"):
            items = _read_inventory_image(upload)
        else:
            return jsonify({"error": "مسموح بصورة PNG/JPG/WEBP أو ملف XLSX/XLSM/CSV"}), 400
        state = _save_inventory_state(items)
        return jsonify({"ok": True, **state, "count": len(items), "filename": upload.filename})
    except PackagingWorkbookError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"error": f"تعذر استيراد المخزون: {str(exc)[:180]}"}), 500


@packaging_orders_bp.route("/api/packaging-orders/weeks", methods=["GET", "POST"])
def packaging_weeks():
    if request.method == "GET":
        try:
            records = [record for record in _week_records() if not record.get("deleted")]
            return jsonify({"ok": True, "records": records})
        except Exception as exc:
            return jsonify({"error": f"تعذر تحميل سجل الأسابيع: {str(exc)[:180]}"}), 500
    payload = request.get_json(silent=True) or {}
    rows = payload.get("rows") or []
    name = _clean(payload.get("name"))
    week_date = _clean(payload.get("date"))
    run_id = _clean(payload.get("run_id"))
    if not rows or not name or not week_date or not run_id:
        return jsonify({"error": "اكتب اسم الأسبوع وتاريخه أولًا"}), 400
    try:
        for existing in _week_records():
            if existing.get("run_id") == run_id and not existing.get("deleted"):
                return jsonify({"error": "تم اعتماد هذه الملفات وخصم مخزونها من قبل"}), 409
        current = (_read_inventory_state().get("items") or {})
        updated, saved_rows = _consume_inventory(current, rows)
        _save_inventory_state(updated)
        now = datetime.now(timezone.utc).isoformat()
        record = {
            "id": uuid.uuid4().hex, "run_id": run_id, "name": name, "date": week_date,
            "rows": saved_rows, "days": payload.get("days") or [], "created_at": now,
            "customer_counts": payload.get("customer_counts") or {},
            "supplies": _supply_plan(
                payload.get("customer_counts") or {},
                (payload.get("supplies") or {}).get("spoon_carton_size", 200),
                (payload.get("supplies") or {}).get("mode", "weekly"),
            ),
            "updated_at": now, "deleted": False,
        }
        _save_week_event(record)
        return jsonify({"ok": True, "record": record, "inventory": updated})
    except Exception as exc:
        return jsonify({"error": f"تعذر اعتماد الأسبوع: {str(exc)[:180]}"}), 500


@packaging_orders_bp.route("/api/packaging-orders/weeks/<record_id>", methods=["PUT", "DELETE"])
def packaging_week_record(record_id):
    try:
        record = next((row for row in _week_records() if row.get("id") == record_id), None)
        if not record or record.get("deleted"):
            return jsonify({"error": "السجل غير موجود"}), 404
        now = datetime.now(timezone.utc).isoformat()
        if request.method == "PUT":
            payload = request.get_json(silent=True) or {}
            record["name"] = _clean(payload.get("name")) or record.get("name")
            record["date"] = _clean(payload.get("date")) or record.get("date")
            record["updated_at"] = now
            _save_week_event(record)
            return jsonify({"ok": True, "record": record})
        # Deleting a weekly transaction reverses its exact inventory usage.
        current = (_read_inventory_state().get("items") or {})
        restored = {str(key): max(0, _number(value)) for key, value in current.items()}
        for row in record.get("rows") or []:
            item = _canonical_item(row.get("item"))
            restored[item] = round(restored.get(item, 0) + max(0, _number(row.get("used_inventory"))), 3)
        _save_inventory_state(restored)
        record["deleted"] = True
        record["updated_at"] = now
        _save_week_event(record)
        return jsonify({"ok": True, "inventory": restored})
    except Exception as exc:
        return jsonify({"error": f"تعذر تعديل السجل: {str(exc)[:180]}"}), 500


@packaging_orders_bp.route("/api/packaging-orders/export-png", methods=["POST"])
def packaging_orders_export_png():
    try:
        output = build_packaging_png(request.get_json(silent=True) or {})
        return send_file(output, as_attachment=True, download_name="Packaging_Orders.png", mimetype="image/png")
    except PackagingWorkbookError as exc:
        return jsonify({"error": str(exc)}), 400


@packaging_orders_bp.route("/api/packaging-orders/export-supplies-png", methods=["POST"])
def packaging_orders_export_supplies_png():
    try:
        output = build_supplies_png(request.get_json(silent=True) or {})
        return send_file(output, as_attachment=True, download_name="Bags_and_Spoons_Report.png", mimetype="image/png")
    except PackagingWorkbookError as exc:
        return jsonify({"error": str(exc)}), 400


@packaging_orders_bp.route("/api/packaging-orders/export-xlsx", methods=["POST"])
def packaging_orders_export_xlsx():
    try:
        output = build_packaging_xlsx(request.get_json(silent=True) or {})
        return send_file(output, as_attachment=True, download_name="Packaging_Orders.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except PackagingWorkbookError as exc:
        return jsonify({"error": str(exc)}), 400
