"""
محرك الطلب الذكي (Smart Ordering Engine)
==========================================
بياخد اسم الوجبة + عدد الأوردرات، ويحسب جدول التجهيز (Batches) بتاعها عن طريق
تشغيل صيغ الإكسل الحقيقية لنفس الوجبة (مش إعادة كتابة المعادلة بإيد) — عشان
نضمن إن كل وجبة من الـ88 بتتحسب بمعادلتها الخاصة بالظبط زي ما هي في الملف
الأصلي، حتى لو الصيغة مختلفة عن باقي الوجبات.

الفكرة: 
  1. Z1 ("Update") هي الخلية الوحيدة اللي بتتغيّر يدويًا في الإكسل الأصلي.
  2. بنحسبها هنا = عدد الأوردرات × حصة الوجبة بالجرام (مستخرجة من نفس الملف
     المرجعي، عمود Z1/Count الأصلي لكل وجبة).
  3. بنستخدم pycel (بيشغّل صيغ إكسل حقيقية، مش إعادة تنفيذ يدوي) عشان نحسب
     باقي الجدول من القيمة الجديدة دي.
"""
import os
import json
import hashlib
import re
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from pycel import ExcelCompiler

TOKYO_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'tokyo_ordering_template.xlsm')
PORTIONS_PATH = os.path.join(os.path.dirname(__file__), 'meal_portions_data.json')
PACKAGES_PATH = os.path.join(os.path.dirname(__file__), 'menu_packages.json')
EXPECTED_TEMPLATE_SHA256 = '0e110f2f45330cce3c1aea0f2f86542323614f3329e64ed935dc0be71ddb5d18'
EXPECTED_VBA_SHA256 = 'a6e4ee6fecaca26a334e05efba9422ca9f6d1ea539a638969fb4ab47d28c5758'
EXPECTED_SHEET_COUNT = 88
EXPECTED_FORMULA_COUNT = 31085

with open(PORTIONS_PATH, encoding='utf-8') as f:
    # { meal_name: [portion_grams_per_order, arabic_name] }
    _SAVED_MEAL_PORTIONS = json.load(f)

with open(PACKAGES_PATH, encoding='utf-8') as f:
    # { "Saturday": [["اسم المنيو", "اسم الشيت"], ...], ... }
    MENU_PACKAGES = json.load(f)

DAY_LABELS_AR = {
    'Saturday': 'السبت', 'Sunday': 'الأحد', 'Monday': 'الإثنين',
    'Tuesday': 'الثلاثاء', 'Wednesday': 'الأربعاء', 'Thursday': 'الخميس',
    'Friday': 'الجمعة',
}

_RENAMED_MEALS = {
    'Blankwet Fish': 'Blankwet Shrimp',
    'Fish with cream': 'Salmon with cream',
    'Lemon Fish': 'Lemon Shrimp',
    'Curry Fish': 'Curry Shrimp',
}
_RENAMED_ARABIC = {
    'Blankwet Fish': 'سمك بلانكويت',
    'Fish with cream': 'سمك بالكريمة والبازلاء',
    'Lemon Fish': 'سمك بالليمون',
    'Curry Fish': 'سمك بالكاري والكريمة',
}


def _load_live_meal_portions():
    """يقرأ وزن الحصة من نفس ملف توكيو المستخدم في الحساب.

    الوزن = قيمة Z1 الحالية / عدد الوجبات المرجعي في All_Ingredients!AN.
    بهذه الطريقة لا تظل الداشبورد مرتبطة بأرقام نسخة أقدم من ملف الإكسل.
    """
    if not os.path.exists(TOKYO_TEMPLATE_PATH):
        return dict(_SAVED_MEAL_PORTIONS)

    wb = load_workbook(TOKYO_TEMPLATE_PATH, data_only=True, read_only=False, keep_vba=True)
    try:
        ws = wb['All_Ingredients']
        portions = {}
        for row in range(2, ws.max_row + 1):
            sheet_name = ws.cell(row=row, column=37).value  # AK
            reference_count = ws.cell(row=row, column=40).value  # AN
            if not sheet_name or sheet_name == 'Butchery' or sheet_name not in wb.sheetnames:
                continue
            z1_value = wb[sheet_name]['Z1'].value
            if not isinstance(z1_value, (int, float)) or not isinstance(reference_count, (int, float)) or reference_count <= 0:
                continue
            saved_key = sheet_name if sheet_name in _SAVED_MEAL_PORTIONS else _RENAMED_MEALS.get(sheet_name)
            arabic_name = _RENAMED_ARABIC.get(sheet_name) or (_SAVED_MEAL_PORTIONS.get(saved_key) or [None, ''])[1]
            portions[str(sheet_name)] = [round(float(z1_value) / float(reference_count), 2), arabic_name]
        return portions or dict(_SAVED_MEAL_PORTIONS)
    finally:
        wb.close()


MEAL_PORTIONS = _load_live_meal_portions()


def get_template_integrity():
    """فحص سريع للقالب قبل السماح بأي حساب حساس."""
    if not os.path.exists(TOKYO_TEMPLATE_PATH):
        return {'ready': False, 'errors': ['ملف توكيو الرئيسي غير موجود على السيرفر']}

    errors = []
    with zipfile.ZipFile(TOKYO_TEMPLATE_PATH, 'r') as archive:
        names = set(archive.namelist())
        workbook_root = ET.fromstring(archive.read('xl/workbook.xml'))
        sheet_count = len(workbook_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet'))
        if 'xl/vbaProject.bin' not in names:
            errors.append('الماكرو غير موجود داخل الملف')
            vba_sha256 = None
        else:
            vba_sha256 = hashlib.sha256(archive.read('xl/vbaProject.bin')).hexdigest()
        formula_count = sum(
            len(re.findall(br'<f(?:\s|>)', archive.read(name)))
            for name in names if name.startswith('xl/worksheets/sheet') and name.endswith('.xml')
        )

    missing_package_sheets = sorted({
        sheet_name
        for items in MENU_PACKAGES.values()
        for _, sheet_name in items
        if sheet_name not in MEAL_PORTIONS
    })
    if missing_package_sheets:
        errors.append('وجبات غير متطابقة مع ملف الإكسل: ' + ', '.join(missing_package_sheets))
    if sheet_count != EXPECTED_SHEET_COUNT:
        errors.append(f'عدد الشيتات تغير: المتوقع {EXPECTED_SHEET_COUNT} والموجود {sheet_count}')
    if formula_count != EXPECTED_FORMULA_COUNT:
        errors.append(f'عدد المعادلات تغير: المتوقع {EXPECTED_FORMULA_COUNT} والموجود {formula_count}')

    with open(TOKYO_TEMPLATE_PATH, 'rb') as template_file:
        template_sha256 = hashlib.sha256(template_file.read()).hexdigest()
    if template_sha256 != EXPECTED_TEMPLATE_SHA256:
        errors.append('بصمة ملف توكيو لا تطابق النسخة المعتمدة')
    if vba_sha256 != EXPECTED_VBA_SHA256:
        errors.append('بصمة الماكرو لا تطابق النسخة المعتمدة')

    return {
        'ready': not errors,
        'errors': errors,
        'sheet_count': sheet_count,
        'meal_count': len(MEAL_PORTIONS),
        'formula_count': formula_count,
        'has_vba': bool(vba_sha256),
        'vba_sha256': vba_sha256,
        'template_sha256': template_sha256,
    }


def _sheet_xml_paths(archive):
    workbook_root = ET.fromstring(archive.read('xl/workbook.xml'))
    rels_root = ET.fromstring(archive.read('xl/_rels/workbook.xml.rels'))
    rel_targets = {
        rel.attrib['Id']: rel.attrib['Target'].lstrip('/')
        for rel in rels_root
    }
    paths = {}
    rel_attr = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id'
    for sheet in workbook_root.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheets/{http://schemas.openxmlformats.org/spreadsheetml/2006/main}sheet'):
        target = rel_targets.get(sheet.attrib.get(rel_attr), '')
        if target.startswith('worksheets/'):
            target = 'xl/' + target
        elif not target.startswith('xl/'):
            target = 'xl/' + target
        paths[sheet.attrib['name']] = target
    return paths


def build_macro_workbook(meal_orders):
    """ينشئ نسخة تشغيل من ملف xlsm بتعديل Z1 فقط داخل XML.

    لا نفتح الملف ثم نعيد حفظه بمكتبة جداول، وبالتالي يظل الماكرو وكل المعادلات
    والرسومات والإضافات الثنائية بنفس البايتات الموجودة في القالب الأصلي.
    """
    updates = {}
    for item in meal_orders or []:
        name = item.get('meal_name')
        count = item.get('order_count')
        if name not in MEAL_PORTIONS:
            raise ValueError(f'الوجبة "{name}" غير موجودة في القالب الحالي')
        try:
            count = int(count)
        except (TypeError, ValueError):
            raise ValueError(f'عدد طلبات "{name}" غير صحيح')
        if count <= 0:
            raise ValueError(f'عدد طلبات "{name}" لازم يكون أكبر من صفر')
        updates[name] = round(count * MEAL_PORTIONS[name][0], 2)
    if not updates:
        raise ValueError('اختار وجبة واحدة على الأقل')

    out_path = tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False).name
    with zipfile.ZipFile(TOKYO_TEMPLATE_PATH, 'r') as source:
        sheet_paths = _sheet_xml_paths(source)
        missing = sorted(name for name in updates if name not in sheet_paths)
        if missing:
            raise ValueError('شيتات غير موجودة في الملف: ' + ', '.join(missing))
        path_updates = {sheet_paths[name]: value for name, value in updates.items()}
        original_vba = hashlib.sha256(source.read('xl/vbaProject.bin')).hexdigest()
        with zipfile.ZipFile(out_path, 'w') as output:
            for info in source.infolist():
                data = source.read(info.filename)
                if info.filename in path_updates:
                    value = ('%.10f' % path_updates[info.filename]).rstrip('0').rstrip('.')
                    pattern = br'(<c\b[^>]*\br="Z1"[^>]*>.*?<v>)([^<]*)(</v>.*?</c>)'
                    data, changed = re.subn(pattern, lambda m: m.group(1) + value.encode('ascii') + m.group(3), data, count=1)
                    if changed != 1:
                        raise ValueError(f'تعذر تحديث خلية Z1 داخل {info.filename}')
                output.writestr(info, data)

    with zipfile.ZipFile(out_path, 'r') as result:
        result_vba = hashlib.sha256(result.read('xl/vbaProject.bin')).hexdigest()
        if result_vba != original_vba:
            os.unlink(out_path)
            raise ValueError('تم إيقاف التصدير لأن بصمة الماكرو تغيرت')
    return out_path

def _get_excel():
    """بنعمل نسخة جديدة من المحرك في كل مرة (مش بنعيد استخدام نسخة قديمة) —
    لأن pycel مابيعملش invalidation كامل لكل الخلايا المعتمدة على خلية اتغيّرت
    لو استخدمنا نفس الكائن أكتر من مرة، وده كان بيسبب أرقام قديمة/خاطئة في
    بعض الحالات. تحميل الملف بياخد وقت أطول شوية، بس ده ضمان الدقة المطلوب."""
    return ExcelCompiler(filename=TOKYO_TEMPLATE_PATH)


def list_available_meals():
    """بترجع قائمة بكل الوجبات المتاحة للطلب (اسم إنجليزي + عربي لو موجود)،
    مرتبة أبجديًا، عشان تتعرض في الـ dropdown بتاع الداش بورد."""
    out = []
    for name, (portion, arabic) in MEAL_PORTIONS.items():
        out.append({'name': name, 'arabic_name': arabic, 'portion_g': portion})
    out.sort(key=lambda m: m['name'].lower())
    return out


def list_menu_packages():
    """بترجع الوجبات منظّمة حسب باقات الأيام (زي المنيو اللي العملاء بيطلبوا
    منه فعليًا)، بالاسم اللي العميل شايفه + اسم الشيت الحقيقي اللي هيتحسب
    عليه. أي وجبة موجودة في الشيتات بس مش في أي باقة لسه، بتتحط في باقة
    "أخرى" عشان تفضل متاحة للإضافة من غير ما تضيع."""
    mapped_sheets = set()
    packages = []
    for day_key, items in MENU_PACKAGES.items():
        day_items = []
        for menu_name, sheet_name in items:
            info = MEAL_PORTIONS.get(sheet_name)
            if not info:
                continue  # لو الشيت اتشال من الملف الأصلي مستقبلًا
            mapped_sheets.add(sheet_name)
            day_items.append({
                'menu_name': menu_name,
                'sheet_name': sheet_name,
                'arabic_name': info[1],
                'portion_g': info[0],
            })
        packages.append({
            'day_key': day_key,
            'day_label_ar': DAY_LABELS_AR.get(day_key, day_key),
            'items': day_items,
        })

    extra_items = []
    for sheet_name, (portion, arabic) in MEAL_PORTIONS.items():
        if sheet_name not in mapped_sheets:
            extra_items.append({
                'menu_name': sheet_name,
                'sheet_name': sheet_name,
                'arabic_name': arabic,
                'portion_g': portion,
            })
    extra_items.sort(key=lambda m: m['sheet_name'].lower())
    if extra_items:
        packages.append({'day_key': 'Extra', 'day_label_ar': 'أصناف أخرى', 'items': extra_items})

    return packages


def _read_cell(excel, sheet, cell):
    try:
        return excel.evaluate(f"'{sheet}'!{cell}")
    except Exception:
        return None


def _read_block(excel, sheet):
    """بيرجّع جدول الباتشات الكامل (يمين وشمال، زي ما بنستخرجه في محطات
    التجهيز) بعد ما يكون Z1 اتغيّر للقيمة الجديدة."""
    rows = []
    labels_left = {51: 'For 10KG Sauce output', 52: 'Total', 53: 'Batch 1', 54: 'Batch 2', 55: 'Batch 3'}
    for r in range(51, 56):
        left_label = _read_cell(excel, sheet, f'B{r}') or labels_left.get(r, '')
        conv = _read_cell(excel, sheet, f'C{r}')
        final_kg = _read_cell(excel, sheet, f'D{r}')
        right_label = _read_cell(excel, sheet, f'G{r}')
        conv_r = _read_cell(excel, sheet, f'H{r}')
        uncooked = _read_cell(excel, sheet, f'I{r}')
        cooked = _read_cell(excel, sheet, f'J{r}')
        rows.append({
            'label': str(left_label) if left_label else '',
            'conversion_factor': conv,
            'final_kg': final_kg,
            'protein_label': str(right_label) if right_label else '',
            'protein_conversion_factor': conv_r,
            'uncooked_protein': uncooked,
            'cooked_protein': cooked,
        })
    return rows


def _has_batch_table(excel, sheet):
    """بعض الوجبات (زي السندوتشات) معندهاش جدول الباتشات (B51:J55) خالص —
    بتتحسب بطريقة أبسط (وزن كل مكوّن = عدد الأوردرات × كمية المكوّن في
    الوحدة الواحدة، من غير batches خالص). بنتأكد الأول قبل ما نحاول نقرا
    جدول مش موجود، عشان منرميش خطأ غلط."""
    val = _read_cell(excel, sheet, 'B51')
    return val is not None and str(val).strip() != ''


def _read_ingredient_list(excel, sheet, max_row=40):
    """بديل لجدول الباتشات للوجبات اللي مالهاش الجدول ده (زي السندوتشات) —
    بيرجّع كل صف مكوّن (B..H) بعد ما يكون Z1 اتغيّر، عشان نشوف وزن كل
    مكوّن للعدد الجديد من الأوردرات. بنوقف عند أول صف فاضي عشان منلقطش
    أقسام تانية تحت في نفس الشيت (زي مكونات التتبيلة المنفصلة)."""
    rows = []
    for r in range(5, max_row + 1):
        name = _read_cell(excel, sheet, f'B{r}')
        unit = _read_cell(excel, sheet, f'C{r}')
        if not name or not str(name).strip():
            if rows:
                break  # خلصنا أول قسم حقيقي، نوقف هنا
            continue
        if not unit:
            # صف عنوان قسم جديد (زي "Base Recipe (10kg Yield)") مش مكوّن حقيقي
            if rows:
                break
            continue
        amount = _read_cell(excel, sheet, f'H{r}') or _read_cell(excel, sheet, f'G{r}')
        if amount is None:
            continue
        rows.append({'label': str(name).strip(), 'unit': unit, 'amount': amount})
    return rows


def calculate_meal(meal_name, order_count):
    """بياخد اسم وجبة وعدد أوردرات، ويرجّع جدول الباتشات المحسوب بناءً على
    صيغ الإكسل الحقيقية لنفس الوجبة دي."""
    if meal_name not in MEAL_PORTIONS:
        raise ValueError(f'الوجبة "{meal_name}" مش موجودة في قائمة الوجبات المتاحة')
    if not isinstance(order_count, (int, float)) or order_count <= 0:
        raise ValueError('عدد الأوردرات لازم يكون رقم أكبر من صفر')

    portion_g, arabic_name = MEAL_PORTIONS[meal_name]
    z1_value = round(order_count * portion_g, 2)

    excel = _get_excel()
    z1_addr = f"'{meal_name}'!Z1"

    if _has_batch_table(excel, meal_name):
        # مهم جدًا: لازم نقرا كل الخلايا النهائية اللي محتاجينها الأول (قبل ما
        # نغيّر Z1)، عشان pycel يبني شجرة الاعتماديات الكاملة من Z1 لحدهم.
        target_cells = ['B', 'C', 'D', 'G', 'H', 'I', 'J']
        for r in range(51, 56):
            for col in target_cells:
                excel.evaluate(f"'{meal_name}'!{col}{r}")
        excel.set_value(z1_addr, z1_value)
        rows = _read_block(excel, meal_name)
        mode = 'batches'
    else:
        # مفيش جدول باتشات للوجبة دي (زي السندوتشات) — بنحسب قايمة المكونات
        # مباشرة بدل منه.
        for r in range(5, 40):
            for col in ['B', 'C', 'G', 'H']:
                excel.evaluate(f"'{meal_name}'!{col}{r}")
        excel.evaluate(z1_addr)
        excel.set_value(z1_addr, z1_value)
        ingredients = _read_ingredient_list(excel, meal_name)
        rows = [{
            'label': ing['label'],
            'conversion_factor': None,
            'final_kg': ing['amount'],
            'protein_label': '',
            'protein_conversion_factor': None,
            'uncooked_protein': None,
            'cooked_protein': None,
            'unit': ing['unit'],
        } for ing in ingredients]
        mode = 'ingredients'

    return {
        'meal_name': meal_name,
        'arabic_name': arabic_name,
        'order_count': order_count,
        'portion_g': portion_g,
        'z1_calculated': z1_value,
        'mode': mode,
        'rows': rows,
    }


def calculate_multiple(meal_orders):
    """meal_orders: [{'meal_name': ..., 'order_count': ...}, ...]
    بيحسب كل وجبة على حدة ويرجّع كل النتائج مع بعض."""
    results = []
    errors = []
    for item in meal_orders:
        name = item.get('meal_name')
        count = item.get('order_count')
        try:
            results.append(calculate_meal(name, count))
        except Exception as e:
            errors.append({'meal_name': name, 'error': str(e)})
    return {'results': results, 'errors': errors}
