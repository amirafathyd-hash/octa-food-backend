"""Create the daily Tokyo Hot Section PDF from the original workbook.

The ranges and ordering below mirror the workbook's own VBA PDF modules:
BatchPDF, SpecialTablesPDF, ActualsPDF and GarnishPDF.
LibreOffice is used only as the spreadsheet renderer/calculation engine; the
recipe cells, formulas, formatting and print tables remain those of the source
Tokyo workbook.
"""

from __future__ import annotations

import io
import os
import shutil
import subprocess
import tempfile
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas

from tokyo_ordering import (
    DAY_NAMES,
    merge_day_into_template,
    read_day_file_payload,
    read_day_file_shifts,
    validate_raw_targets_for_day,
)


SPECIAL_SHEETS = [
    'Asian Chicken Sandwich', 'Jollof Sauce', 'Almond Chicken',
    'Chicken Caesar Salad', 'Octa Poki Bowl', 'Makloba Veggi',
    'Lasagne Bachamel', 'Beef Lasagne Sauce',
    'Beef philly cheese steak', 'Sigapore Vegetables',
    'Mexican Vegetables', 'Beef Kebab Sandwich',
]

ACTUALS_TYPE_B = {
    'Herbal Potato Wedges', 'Sautee Vegetables (1)', 'Mached Potato(1)',
    'Grilled Vegetables(2)', 'Mached Potato(3)', 'Potato Wedges',
    'Oven Vegetables (3)', 'Mached Potato(4)', 'Sautee Vegetables (4)',
    'Grilled Vegetables(5)', 'Sigapore Vegetables', 'Mexican Vegetables',
    'Chicken Steak Topping', 'Beef Kebab Sandwich', 'Spaghetti pasta (3)',
    'Oven Vegetables (6)', 'Chicken Mandi (1)', 'Beef Zurbian',
    'Chicken Saleeq (3)', 'Chicken Mandi (3)', 'Chicken Makloba',
    'Beef Bokhary', 'Chicken Saleeq (6)', 'Beef Kabli', 'Jollof Sauce',
    'Spaghetti pasta', 'Spaghetti pasta (2)',
}


def _soffice_path() -> str:
    candidates = [
        os.environ.get('LIBREOFFICE_PATH'),
        shutil.which('libreoffice'),
        shutil.which('soffice'),
        '/usr/bin/libreoffice',
        '/opt/libreoffice/program/soffice',
        '/Users/mostafaabdo/.cache/codex-runtimes/codex-primary-runtime/dependencies/bin/override/soffice',
    ]
    for candidate in candidates:
        if candidate and os.path.exists(candidate):
            return candidate
    raise RuntimeError('LibreOffice غير موجود على السيرفر؛ لا يمكن إنشاء ملفات PDF')


def _run_soffice(source: Path, output_dir: Path, target: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    profile = tempfile.mkdtemp(prefix='tokyo-lo-profile-')
    try:
        cmd = [
            _soffice_path(), '--headless', '--nologo', '--nodefault',
            '--nolockcheck', '--nofirststartwizard',
            f'-env:UserInstallation=file://{profile}',
            '--convert-to', target, '--outdir', str(output_dir), str(source),
        ]
        completed = subprocess.run(cmd, capture_output=True, text=True, timeout=240)
        if completed.returncode != 0:
            raise RuntimeError((completed.stderr or completed.stdout or 'LibreOffice failed').strip())
        suffix = '.pdf' if target.startswith('pdf') else '.xlsx'
        result = output_dir / f'{source.stem}{suffix}'
        if not result.exists():
            raise RuntimeError(f'LibreOffice لم ينشئ الملف المتوقع: {result.name}')
        return result
    finally:
        shutil.rmtree(profile, ignore_errors=True)


def _day_sheets(workbook, master_name: str, day_no: int) -> list[str]:
    ws = workbook[master_name]
    names = []
    for row in range(2, min(ws.max_row, 500) + 1):
        value = ws.cell(row, 36).value
        name = ws.cell(row, 37).value
        try:
            same_day = int(value) == int(day_no)
        except (TypeError, ValueError):
            same_day = False
        if same_day and name:
            clean = str(name).strip()
            if clean in workbook.sheetnames and clean not in names:
                names.append(clean)
    return names


def _number(value) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _batch_ranges(ws_values) -> list[str]:
    ranges = []
    for header in range(62, min(ws_values.max_row, 500) + 1):
        if str(ws_values.cell(header, 2).value or '').strip().lower() != 'ingredient':
            continue
        if _number(ws_values.cell(header + 1, 5).value) <= 0:
            continue
        end = None
        next_header = min(ws_values.max_row, header + 80)
        for row in range(header + 1, next_header + 1):
            label = str(ws_values.cell(row, 2).value or '').strip()
            low = label.lower()
            if row > header + 1 and low == 'ingredient':
                break
            if low.startswith('total ') or low in {'protein', 'pasta', 'potato'}:
                end = row + 1
                break
        if end:
            ranges.append(f'A{max(1, header - 2)}:H{end}')
    return ranges


def _special_range(ws_values) -> str | None:
    ceiling = min(30, ws_values.max_row)
    for row in range(5, ceiling + 1):
        if 'garnish' in str(ws_values.cell(row, 2).value or '').lower():
            ceiling = row - 1
            break
    last = None
    for row in range(ceiling, 4, -1):
        label = str(ws_values.cell(row, 2).value or '').strip()
        if label and label.lower() not in {'ingredient', 'category', 'total'} and not label.lower().startswith('base recipe'):
            last = row
            break
    return f'B2:H{last}' if last else None


def _garnish_range(ws_values) -> str | None:
    if 'garnish' not in str(ws_values.cell(35, 2).value or '').lower():
        return None
    last = 36
    for row in range(37, min(ws_values.max_row, 60) + 1):
        if str(ws_values.cell(row, 2).value or '').strip():
            last = row
        elif last >= 37:
            break
    return f'B35:F{last}'


def _marination_range(ws_values) -> str | None:
    header_row = None
    header_col = None
    for row in range(1, min(ws_values.max_row, 40) + 1):
        for col in range(30, min(ws_values.max_column, 55) + 1):
            if str(ws_values.cell(row, col).value or '').strip().lower() == 'ingredient':
                header_row = row
                header_col = col
                break
        if header_row:
            break
    if not header_row or not header_col:
        return None

    last = header_row
    for row in range(header_row + 1, min(ws_values.max_row, header_row + 80) + 1):
        if str(ws_values.cell(row, header_col).value or '').strip():
            last = row
        elif last > header_row:
            break
    if last <= header_row:
        return None
    start_col = max(1, header_col - 1)
    end_col = min(ws_values.max_column, header_col + 9)
    start_letter = get_column_letter(start_col)
    end_letter = get_column_letter(end_col)
    return f'{start_letter}1:{end_letter}{last}'


def _page_setup(ws, day_no: int, section: str) -> None:
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4 if section in {'batch', 'special', 'marination', 'production'} else 0.5
    ws.page_margins.right = ws.page_margins.left
    ws.page_margins.top = 1.0
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    ws.print_options.horizontalCentered = True
    ws.oddHeader.center.text = f'&B&16{ws.title}'
    if section == 'marination':
        ws.oddHeader.right.text = f'&B&16Day {day_no}  |  Marination'
    else:
        ws.oddHeader.right.text = f'&B&12Day {day_no}'
    # Blocks are exported separately and merged afterwards. Their native
    # &N value is therefore only the number of pages in that block. Global
    # numbering is stamped after the final merge instead.
    ws.oddFooter.center.text = ''


def _make_block_workbook(source: Path, values_path: Path, destination: Path,
                         day_no: int, section: str, day_sheets: list[str]) -> tuple[Path, int]:
    wb = load_workbook(source, data_only=False, keep_links=True)
    values = load_workbook(values_path, data_only=True, keep_links=True)
    included = []
    page_total = 0

    for name in day_sheets:
        if name not in wb.sheetnames or name not in values.sheetnames:
            continue
        ws = wb[name]
        wsv = values[name]
        ranges = []
        if section == 'production':
            # Keep the recipe order from All_Ingredients. Each recipe prints
            # either its calculated batch table(s), or its special table when
            # that recipe does not use the standard batch layout.
            if name == 'Asian Chicken Sandwich':
                # The original tab contains the sandwich recipe followed by
                # Instant Marination. LibreOffice collapses multiple print
                # areas from one sheet into one page, so keep the original tab
                # for marination and add a dedicated printable copy of the
                # first table below.
                ranges = ['B9:H28']
            elif name == 'Herbal Chicken':
                # This recipe's production table lives in the lower printable
                # block and has no standard batch marker. Keep it as its own
                # production page so the morning report stays at 41 pages.
                ranges = ['A62:H65']
            else:
                ranges = _batch_ranges(wsv)
                if not ranges and name in SPECIAL_SHEETS:
                    found = _special_range(wsv)
                    ranges = [found] if found else []
        elif section == 'batch':
            ranges = _batch_ranges(wsv)
        elif section == 'special' and name in SPECIAL_SHEETS:
            found = _special_range(wsv)
            ranges = [found] if found else []
        elif section == 'actuals':
            end_col, end_row = ('T', 30) if name in ACTUALS_TYPE_B else ('V', 32) if name == 'Chicken Mushroom' else ('U', 32)
            ranges = [f'R25:{end_col}{end_row}']
        elif section == 'garnish':
            found = _garnish_range(wsv)
            ranges = [found] if found else []
        elif section == 'marination':
            found = _marination_range(wsv)
            ranges = [found] if found else []

        if not ranges:
            continue
        ws.print_area = ranges
        _page_setup(ws, day_no, section)
        included.append(name)
        # LibreOffice exports each included worksheet as one physical page in
        # these fitted report blocks, even when the worksheet has several
        # contiguous print ranges.
        page_total += 1

    if not included:
        wb.close()
        values.close()
        raise RuntimeError(f'لا توجد جداول قابلة للطباعة في قسم {section} لليوم المختار')

    # LibreOffice includes hidden sheets in whole-workbook PDF export. Freeze
    # the already-calculated values in the selected recipe sheets, then remove
    # non-selected tabs from this disposable print copy. The source XLSM and
    # its formulas are untouched.
    for name in included:
        ws = wb[name]
        value_ws = values[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    cell.value = value_ws[cell.coordinate].value

    if section == 'production' and 'Asian Chicken Sandwich' in included:
        # Make the first recipe table its own worksheet and place it directly
        # before the original marination worksheet. Copying after formulas are
        # frozen guarantees that both pages use the same calculated snapshot.
        marination_ws = wb['Asian Chicken Sandwich']
        recipe_ws = wb.copy_worksheet(marination_ws)
        recipe_ws.title = 'Asian Sandwich Recipe'
        recipe_ws.print_area = 'B2:H8'
        _page_setup(recipe_ws, day_no, section)
        recipe_ws.oddHeader.center.text = '&B&16Asian Chicken Sandwich'

        wb._sheets.remove(recipe_ws)
        marination_index = wb._sheets.index(marination_ws)
        wb._sheets.insert(marination_index, recipe_ws)
        included.insert(included.index('Asian Chicken Sandwich'), recipe_ws.title)
        page_total += 1

    included_set = set(included)
    for ws in list(wb.worksheets):
        if ws.title not in included_set:
            wb.remove(ws)
    wb.active = wb.sheetnames.index(included[0])
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = 'auto'
    wb.save(destination)
    wb.close()
    values.close()
    return destination, page_total


def _merge_pdfs(paths: list[Path], destination: Path) -> Path:
    writer = PdfWriter()
    for path in paths:
        writer.append(str(path))
    with destination.open('wb') as handle:
        writer.write(handle)
    writer.close()
    _stamp_global_page_numbers(destination)
    return destination


def _stamp_global_page_numbers(pdf_path: Path) -> None:
    """Stamp one continuous ``Page X of N`` footer on the merged PDF."""
    source = PdfReader(io.BytesIO(pdf_path.read_bytes()))
    writer = PdfWriter()
    total = len(source.pages)

    for page_no, page in enumerate(source.pages, start=1):
        width = float(page.mediabox.width)
        height = float(page.mediabox.height)
        overlay_stream = io.BytesIO()
        overlay_canvas = canvas.Canvas(overlay_stream, pagesize=(width, height))
        overlay_canvas.setFont('Helvetica-Bold', 8)
        overlay_canvas.drawCentredString(
            width / 2.0, 15.0, f'Page {page_no} of {total}'
        )
        overlay_canvas.save()
        overlay_stream.seek(0)
        overlay = PdfReader(overlay_stream).pages[0]
        page.merge_page(overlay)
        writer.add_page(page)

    temporary = pdf_path.with_suffix('.numbered.tmp.pdf')
    with temporary.open('wb') as handle:
        writer.write(handle)
    writer.close()
    temporary.replace(pdf_path)


def _make_libreoffice_copy(source: Path, destination: Path) -> Path:
    """Make a calculation-only XLSX copy.

    Microsoft 365 stores one Moussaka formula with the internal
    ``_xlfn._TRO_LEADING`` compatibility wrapper. LibreOffice interprets that
    wrapper as an unknown function. Removing the no-op wrapper in this
    disposable renderer copy lets the original formula calculate normally;
    the delivered macro workbook is never changed this way.
    """
    wb = load_workbook(source, data_only=False, keep_vba=False, keep_links=True)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith('=') and '_xlfn._TRO_LEADING(' in value:
                    cell.value = value.replace('_xlfn._TRO_LEADING(', '(')
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = 'auto'
    wb.save(destination)
    wb.close()
    return destination


def _stamp_selected_day(workbook_path: Path, day_no: int) -> None:
    wb = load_workbook(workbook_path, keep_vba=True, data_only=False)
    wb['All_Ingredients']['R1'] = day_no
    wb['Marination_Ordering']['R1'] = day_no
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = 'auto'
    wb.save(workbook_path)
    wb.close()


def _render_hot_section_pdf(updated_xlsm: Path, root: Path, day_no: int,
                            output_name: str, ordered=False):
    render_root = root / Path(output_name).stem.replace(' ', '_')
    render_root.mkdir(parents=True, exist_ok=True)
    calculation_source = _make_libreoffice_copy(
        updated_xlsm, render_root / 'calculation-source.xlsx'
    )
    recalculated = _run_soffice(
        calculation_source, render_root / 'recalculated', 'xlsx'
    )
    values = load_workbook(recalculated, data_only=False, read_only=False)
    hot_sheets = _day_sheets(values, 'All_Ingredients', day_no)
    values.close()

    blocks = []
    counts = {}
    sections = ('production', 'actuals', 'garnish') if ordered else (
        'batch', 'special', 'actuals', 'garnish'
    )
    for section in sections:
        block_book = render_root / f'{section}.xlsx'
        try:
            block_book, page_count = _make_block_workbook(
                recalculated, recalculated, block_book, day_no, section, hot_sheets
            )
            block_pdf = _run_soffice(block_book, render_root / 'pdf', 'pdf')
            blocks.append(block_pdf)
            counts[section] = page_count
        except RuntimeError as exc:
            if 'لا توجد جداول قابلة للطباعة' not in str(exc):
                raise
            counts[section] = 0

    if not blocks:
        raise RuntimeError('لا توجد أي جداول Hot Section قابلة للطباعة لليوم المختار')
    hot_pdf = _merge_pdfs(blocks, root / output_name)
    return hot_pdf, hot_sheets, counts


def _split_safety_values(safety_overrides, shift):
    if not isinstance(safety_overrides, dict):
        return safety_overrides
    nested = safety_overrides.get(shift)
    return nested if isinstance(nested, dict) else safety_overrides


def _total_safety_values(safety_overrides):
    """Combine independent shift Safety values for the persisted total master."""
    if not isinstance(safety_overrides, dict):
        return safety_overrides
    morning = safety_overrides.get('morning')
    evening = safety_overrides.get('evening')
    if not isinstance(morning, dict) and not isinstance(evening, dict):
        return safety_overrides

    combined = {}
    for values in (morning, evening):
        if not isinstance(values, dict):
            continue
        for row, value in values.items():
            combined[row] = combined.get(row, 0.0) + float(value or 0)
    return combined


def _merge_sheet1_snapshot(template_path, day_no, meals, safety_overrides):
    return merge_day_into_template(
        template_path,
        day_no,
        meals,
        safety_overrides=safety_overrides,
        zero_missing=True,
        # Sheet1 is authoritative. Never fall back to the old AQ labels;
        # those labels can belong to a previous operating-day layout.
        allow_legacy_aq_fallback=False,
        # Jollof has no standalone row in the supplied Sheet1. Preserve its
        # configured master input instead of stealing another meal's value.
        preserve_missing_sheets={'Jollof Sauce'},
    )


def build_tokyo_day_package(template_path: str, uploaded_file, output_dir: str | None = None,
                            safety_overrides=None):
    """Return ``(zip_path, updated_xlsm, report)`` for one uploaded day file."""
    root = Path(output_dir or tempfile.mkdtemp(prefix='tokyo-day-reports-'))
    root.mkdir(parents=True, exist_ok=True)

    split_result = read_day_file_shifts(uploaded_file)
    if split_result:
        day_no, shifts, input_report = split_result

        # Persist one total master after both shift reports have been built.
        updated_xlsm, match_report = _merge_sheet1_snapshot(
            template_path, day_no, shifts['total'], _total_safety_values(safety_overrides)
        )
        updated_xlsm = Path(updated_xlsm)
        _stamp_selected_day(updated_xlsm, day_no)

        shift_files = []
        shift_reports = {}
        for shift, label in (('morning', 'Morning'), ('evening', 'Evening')):
            shift_xlsm, shift_match = _merge_sheet1_snapshot(
                template_path,
                day_no,
                shifts[shift],
                _split_safety_values(safety_overrides, shift),
            )
            shift_xlsm = Path(shift_xlsm)
            _stamp_selected_day(shift_xlsm, day_no)
            pdf_name = f'Day{day_no}_Hot Section {label}.pdf'
            hot_pdf, hot_sheets, counts = _render_hot_section_pdf(
                shift_xlsm, root, day_no, pdf_name, ordered=True
            )
            shift_files.append(hot_pdf)
            shift_reports[shift] = {
                'matched_count': shift_match.get('matched_count', 0),
                'hot_sheets': len(hot_sheets),
                'pages': {**counts, 'hot_total': sum(counts.values())},
            }

        zip_path = root / f'Tokyo_Production_Day{day_no}_Morning_Evening.zip'
        with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as archive:
            for output in shift_files:
                archive.write(output, output.name)

        report = {
            **match_report,
            'input': input_report,
            'split': shift_reports,
            'files': [output.name for output in shift_files],
        }
        return str(zip_path), str(updated_xlsm), report

    day_no, meals, input_report = read_day_file_payload(uploaded_file)
    if input_report.get('kind') == 'repeat_update':
        validate_raw_targets_for_day(template_path, day_no, meals)
    updated_xlsm, match_report = merge_day_into_template(
        template_path,
        day_no,
        meals,
        safety_overrides=safety_overrides,
        # A daily upload is a complete snapshot for that Tokyo day. Clearing
        # missing rows prevents values from the previous run surviving a
        # refresh and appearing in the new production PDF.
        zero_missing=True,
    )
    updated_xlsm = Path(updated_xlsm)

    _stamp_selected_day(updated_xlsm, day_no)

    day_name = DAY_NAMES.get(day_no, f'Day {day_no}')
    hot_pdf, hot_sheets, counts = _render_hot_section_pdf(
        updated_xlsm, root, day_no, f'Tokyo_Hot_Section_Day{day_no}.pdf'
    )
    final_xlsm = root / f'Tokyo_Ordering_Updated_Day{day_no}.xlsm'
    shutil.copyfile(updated_xlsm, final_xlsm)

    zip_path = root / f'Tokyo_Production_{day_name}_Day{day_no}.zip'
    with zipfile.ZipFile(zip_path, 'w', compression=zipfile.ZIP_DEFLATED) as archive:
        archive.write(hot_pdf, hot_pdf.name)
        archive.write(final_xlsm, final_xlsm.name)

    report = {
        **match_report,
        'input': input_report,
        'hot_sheets': len(hot_sheets),
        'pages': {**counts, 'hot_total': sum(counts.values())},
        'files': [hot_pdf.name, final_xlsm.name],
    }
    return str(zip_path), str(final_xlsm), report
