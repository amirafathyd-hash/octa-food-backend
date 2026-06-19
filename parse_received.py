import openpyxl
from datetime import datetime


def _parse_date(val):
    if isinstance(val, datetime):
        return val.strftime('%-d-%b-%Y') if hasattr(val, 'strftime') else str(val)
    if isinstance(val, str):
        for fmt in ('%d-%b-%Y', '%Y-%m-%d', '%d/%m/%Y'):
            try:
                dt = datetime.strptime(val.strip(), fmt)
                return dt.strftime('%-d-%b-%Y')
            except ValueError:
                continue
    return None


def parse_received_xlsx(path):
    """
    Reads every sheet that has the standard header row and returns a flat list of:
    { date, name_ar, name_en, key, qty_box, qty_needed, unit, current_inventory,
      qty_received, rec_unit }
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row_idx = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), start=1):
            if row and row[0] == 'Date':
                header_row_idx = i
                break
        if header_row_idx is None:
            continue

        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or not row[0]:
                continue
            date_str = _parse_date(row[0])
            if not date_str:
                continue
            item_field = (row[1] or '').strip()
            if '—' in item_field:
                name_ar, name_en = [p.strip() for p in item_field.split('—', 1)]
            else:
                name_ar, name_en = '', item_field

            records.append({
                'date': date_str,
                'key': name_en.strip().upper(),
                'name_ar': name_ar,
                'name_en': name_en,
                'qty_box': row[2] if len(row) > 2 else None,
                'qty_needed': row[3] if len(row) > 3 else None,
                'unit': row[4] if len(row) > 4 else None,
                'current_inventory': row[5] if len(row) > 5 else None,
                'qty_received': row[6] if len(row) > 6 else None,
                'rec_unit': row[7] if len(row) > 7 else None,
            })
    return records


if __name__ == '__main__':
    import json
    import sys
    print(json.dumps(parse_received_xlsx(sys.argv[1])[:5], ensure_ascii=False, indent=2))
