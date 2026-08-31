"""Permanent count-driven production engine for the approved sauce workbook."""
import io
import json
import os
import re
import shutil
import subprocess
import tempfile
import zipfile
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from sauce_storage import SAUCE_MAPPING_PATH, SAUCE_TEMPLATE_PATH
from tokyo_ordering import read_day_file_payload


DAY_NAMES = {1: 'السبت', 2: 'الأحد', 3: 'الاثنين', 4: 'الثلاثاء', 5: 'الأربعاء', 6: 'الخميس'}
# The approved master is arranged in these six production-day blocks.
DAY_BLOCK_LENGTHS = (9, 7, 6, 7, 8, 4)


class SauceMappingRequiredError(ValueError):
    def __init__(self, missing_groups, available_meals):
        super().__init__('توجد أسماء وجبات جديدة تحتاج ربطها مرة واحدة بوصفات الصوص')
        self.missing_groups = missing_groups
        self.available_meals = available_meals


def _load_mapping_overrides():
    try:
        with open(SAUCE_MAPPING_PATH, 'r', encoding='utf-8') as stream:
            data = json.load(stream)
    except (OSError, ValueError, TypeError):
        data = {}
    aliases = data.get('aliases') if isinstance(data, dict) else {}
    return aliases if isinstance(aliases, dict) else {}


def save_sauce_mappings(mappings):
    aliases = _load_mapping_overrides()
    saved = 0
    for item in mappings or []:
        sheet = str(item.get('sheet') or '').strip()
        group_index = int(_number(item.get('group_index')))
        meal_name = str(item.get('meal_name') or '').strip()
        if not sheet or group_index < 0 or not meal_name:
            continue
        key = f'{sheet}::{group_index}'
        values = [str(value).strip() for value in aliases.get(key, []) if str(value).strip()]
        if meal_name not in values:
            values.append(meal_name)
        aliases[key] = values
        saved += 1
    os.makedirs(os.path.dirname(SAUCE_MAPPING_PATH), exist_ok=True)
    temp_path = tempfile.NamedTemporaryFile(
        mode='w', suffix='.json', delete=False, dir=os.path.dirname(SAUCE_MAPPING_PATH), encoding='utf-8'
    ).name
    with open(temp_path, 'w', encoding='utf-8') as stream:
        json.dump({'version': 1, 'aliases': aliases}, stream, ensure_ascii=False, indent=2)
    os.replace(temp_path, SAUCE_MAPPING_PATH)
    return {'saved_count': saved, 'mapping_file': os.path.basename(SAUCE_MAPPING_PATH)}


def _number(value):
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value or '').replace(',', '').strip())
    except (TypeError, ValueError):
        return 0.0


def _name_key(value):
    text = str(value or '').strip().casefold()
    text = re.sub(r'[\u064b-\u065f\u0670]', '', text).replace('ـ', '')
    text = text.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا')
    text = text.replace('ى', 'ي').replace('ة', 'ه')
    return re.sub(r'[^a-z0-9\u0600-\u06ff]+', ' ', text).strip()


def _recipe_sheet_names(wb):
    return [name for name in wb.sheetnames if name != 'Ordering']


def _day_sheets(wb, day_no):
    day_no = int(_number(day_no))
    if day_no not in DAY_NAMES:
        raise ValueError('رقم يوم الصوص غير صحيح')
    recipes = _recipe_sheet_names(wb)
    start = sum(DAY_BLOCK_LENGTHS[:day_no - 1])
    end = start + DAY_BLOCK_LENGTHS[day_no - 1]
    selected = recipes[start:end]
    if len(selected) != DAY_BLOCK_LENGTHS[day_no - 1]:
        raise ValueError('ترتيب شيتات الصوص لا يطابق الـMaster المعتمد')
    return selected


# Each tuple is one operational meal. Aliases in a tuple are the same row;
# different tuples are deliberately added together (for shared recipes).
SOURCE_GROUPS = {
    'Tahina Small': [('Foul', 'فول'), ('Fish', 'سمك')],
    'Red sauce': [('Foul', 'فول'), ('Fish', 'سمك')],
    'Cucumber yogurt sauce': [('Kabsa', 'كبسة')],
    'Herbal Sauce': [('Herbal Chicken', 'دجاج الأعشاب')],
    'Red sauce Big': [('Kabsa', 'كبسة'), ('Mandi', 'مندي')],
    'Tahina Big': [('Mandi', 'مندي')],
    'Coctail Sauce Weight': [
        ('Asian chicken sandwich with oat bread', 'ساندويتش الدجاج الآسيوي'),
        ('Chicken Awsal sandwich', 'ساندويتش دجاج اوصل'),
    ],
    'Lemon Herb Sauce Packed': [('Arabic Chicken Burger', 'برجر دجاج العربي')],
    'Lemon Herb Sauce unpacked': [('Arabic Chicken Burger', 'برجر دجاج العربي')],
    'Cucumber yogurt sauce (2)': [('Zurbian', 'زربيان')],
    'Red sauce Big (2)': [('Zurbian', 'زربيان')],
    'House Sauce Packed': [('Beef Burger BBQ', 'برجر لحم باربكيو')],
    'House Sauce unpacked': [('Beef Burger BBQ', 'برجر لحم باربكيو')],
    'Honey Mustard Sauce Packed': [('Chicken Burger Honey Mustard', 'برجر دجاج بالعسل والخردل')],
    'Honey Mustard Sauce unpacked': [('Chicken Burger Honey Mustard', 'برجر دجاج بالعسل والخردل')],
    'Tahina Big (2)': [('Mandi', 'مندي')],
    'Red sauce Big (3)': [('Mandi', 'مندي'), ('Chicken Saleeq', 'سليق دجاج', 'Saleeq')],
    'Coctail Sauce Dip In  (3)': [('Almond chicken in the oven with potato wedges', 'دجاج باللوز في الفرن', 'Almond Chicken', 'دجاج باللوز')],
    'Coctail Sauce Weight (2)': [
        ('Chicken Fajita Sandwich on Oat Bread', 'ساندوتش فاهيتا الدجاج بخبز الشوفان'),
        ('Chicken Bell Pepper Sandwich', 'ساندوتش الدجاج بالفلفل الرومي'),
    ],
    'Mixed Pickle': [
        ('Chicken Fajita Sandwich on Oat Bread', 'ساندوتش فاهيتا الدجاج بخبز الشوفان'),
        ('Chicken Bell Pepper Sandwich', 'ساندوتش الدجاج بالفلفل الرومي'),
    ],
    'Smoked Sauce Packed': [('Smoky Beef Burger', 'برجر سموكي لحم')],
    'Smoked Sauce unpacked': [('Smoky Beef Burger', 'برجر سموكي لحم')],
    'Cucumber yogurt sauce (3)': [('Chicken Maqluba', 'مقلوبة دجاج', 'Maklouba', 'مقلوبة')],
    'Shabat Sauce': [
        ('BBQ Chicken Sandwich in Ciabatta Bread', 'ساندوتش الدجاج بالباربيكيو بخبز الشيباتا'),
        ('Grilled chicken sandwich', 'ساندوتش الدجاج المشوي'),
    ],
    'Biryani Yoghrt sauce': [('Tikka chicken with Buryani rice', 'دجاج تكا', 'Chicken Tikka')],
    'Coctail Sauce Dip In  (4)': [('Octa Chicken Poke bowl (spicy)', 'أوكتا بوكي بول الدجاج', 'Octa PokiPowl')],
    'Coctail Sauce Weight (5)': [('Octa Chicken Poke bowl (spicy)', 'أوكتا بوكي بول الدجاج', 'Octa PokiPowl')],
    'Coleslaw Salad': [('Classic Chicken Burger', 'برجر الدجاج الكلاسيكي')],
    'Garlic Aioli (2)': [('Classic Chicken Burger', 'برجر الدجاج الكلاسيكي')],
    'Red sauce Big (4)': [('Boukhary Beef', 'بخارى لحم', 'Boukhary')],
    'Cucumber yogurt sauce (4)': [('Boukhary Beef', 'بخارى لحم', 'Boukhary')],
    'Coctail Sauce Weight (4)': [
        ('Philadelphia beef sandwich with oat bread', 'ساندويتش فلادلفيا لحم', 'ساندوتش فيلادلفيا لحم بخبز الشوفان'),
        ('Classic beef sandwich', 'Classic meat sandwich', 'ساندوتش اللحم الكلاسيكي'),
    ],
    'BBQ Sauce Packed': [('Chicken Burger BBQ', 'برجر دجاج باربكيو', 'برجر باربكيو الدجاج')],
    'BBQ Sauce unpacked': [('Chicken Burger BBQ', 'برجر دجاج باربكيو', 'برجر باربكيو الدجاج')],
    'Sumak Onion': [('Kebab Sandwich', 'ساندويتش كباب لحم', 'كباب اللحم')],
    'Tahina Small (2)': [('Fish with lemon', 'سمك بالليمون', 'سمك بالكاري والكريمة'), ('Kebab Sandwich', 'ساندويتش كباب لحم', 'كباب اللحم')],
    'Red sauce (2)': [('Fish with lemon', 'سمك بالليمون', 'سمك بالكاري والكريمة')],
    'Red sauce Big (5)': [('Chicken Saleeq', 'سليق دجاج', 'Saleeq')],
    'Cucumber yogurt sauce (5)': [('Kabli', 'كابلى', 'كابلي')],
    'Garlic Aioli': [('Beef Burger Arabic', 'برجر لحم عربي')],
    'Garlic Aioli unpacked': [('Beef Burger Arabic', 'برجر لحم عربي')],
}


def _meal_count_lookup(meals):
    result = {}
    for name, pair in (meals or {}).items():
        count = _number(pair[0] if isinstance(pair, (list, tuple)) else pair)
        key = _name_key(name)
        if key:
            result[key] = count
    return result


def _match_group(aliases, lookup):
    for alias in aliases:
        key = _name_key(alias)
        if key in lookup:
            return lookup[key], key
    candidates = []
    for alias in aliases:
        key = _name_key(alias)
        if len(key) < 5:
            continue
        for source_key, count in lookup.items():
            if key in source_key or source_key in key:
                candidates.append((source_key, count))
    unique = {key: value for key, value in candidates}
    if len(unique) == 1:
        key, value = next(iter(unique.items()))
        return value, key
    return 0.0, ''


def _compute_day_counts(wb, day_no, meals):
    lookup = _meal_count_lookup(meals)
    overrides = _load_mapping_overrides()
    results, missing = [], []
    for sheet_name in _day_sheets(wb, day_no):
        ws = wb[sheet_name]
        groups = SOURCE_GROUPS.get(sheet_name)
        if not groups:
            meal_title = str(ws['B3'].value or '').strip()
            parts = [part.strip(' ()') for part in re.split(r'\s[-–—]\s', meal_title) if part.strip(' ()')]
            groups = [tuple(parts or [meal_title])]
        total, sources = 0.0, []
        for group_index, base_aliases in enumerate(groups):
            override_aliases = overrides.get(f'{sheet_name}::{group_index}', [])
            aliases = tuple(base_aliases) + tuple(override_aliases)
            count, matched_key = _match_group(aliases, lookup)
            total += count
            sources.append({'aliases': list(aliases), 'count': count, 'matched': bool(matched_key)})
            if not matched_key:
                missing.append({
                    'sheet': sheet_name,
                    'sauce_name': str(ws['B2'].value or sheet_name).strip(),
                    'group_index': group_index,
                    'source_label': str(base_aliases[0]),
                })
        results.append({'sheet': sheet_name, 'input_count': round(total, 3), 'safety_count': 0.0, 'sources': sources})
    return results, missing


def _soffice_bin():
    return os.environ.get('SOFFICE_BIN') or shutil.which('soffice') or 'soffice'


def _convert(source_path, extension, prefix):
    out_dir = tempfile.mkdtemp(prefix=prefix)
    profile_dir = tempfile.mkdtemp(prefix=f'{prefix}profile_')
    proc = subprocess.run([
        _soffice_bin(), f'-env:UserInstallation=file://{profile_dir}', '--headless',
        '--convert-to', extension, '--outdir', out_dir, source_path,
    ], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=150)
    if proc.returncode != 0:
        raise RuntimeError((proc.stderr or proc.stdout or 'LibreOffice failed').strip())
    result = os.path.join(out_dir, os.path.splitext(os.path.basename(source_path))[0] + f'.{extension}')
    if not os.path.exists(result):
        raise RuntimeError(f'تعذر إنشاء ملف {extension.upper()} للصوص')
    return result


def _write_inputs(day_no, inputs, template_path=SAUCE_TEMPLATE_PATH):
    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    try:
        selected = set(_day_sheets(wb, day_no))
        values = {str(item.get('sheet') or ''): max(0.0, _number(item.get('input_count'))) for item in inputs}
        safety = {str(item.get('sheet') or ''): max(0.0, _number(item.get('safety_count'))) for item in inputs}
        wb['Ordering']['R1'] = int(day_no)
        for sheet_name in selected:
            wb[sheet_name]['V1'] = values.get(sheet_name, 0.0)
            wb[sheet_name]['Q19'] = safety.get(sheet_name, 0.0)
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = 'auto'
        output = tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False).name
        wb.save(output)
    finally:
        wb.close()
    return output


def _pdf_source(calculated_path, day_no):
    source = load_workbook(calculated_path, data_only=True, read_only=True)
    report = Workbook()
    report.remove(report.active)
    dark_fill = PatternFill('solid', fgColor='303D4D')
    green_fill = PatternFill('solid', fgColor='C6E0B4')
    yellow_fill = PatternFill('solid', fgColor='FFF200')
    white_fill = PatternFill('solid', fgColor='FFFFFF')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    mixed_font = 'IBM Plex Sans Arabic'
    try:
        selected = _day_sheets(source, day_no)
        total_pages = len(selected)
        for page_index, sheet_name in enumerate(selected, 1):
            values = source[sheet_name]
            ws = report.create_sheet(sheet_name[:31])
            ws.sheet_view.showGridLines = False

            ws.merge_cells('A1:G1')
            ws['A1'] = sheet_name
            ws['A1'].font = Font(name=mixed_font, bold=True, size=16)
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws['H1'] = f'Day {int(day_no)}'
            ws['H1'].font = Font(name=mixed_font, bold=True, size=14)
            ws['H1'].alignment = Alignment(horizontal='right', vertical='center')
            ws.row_dimensions[1].height = 31

            ws.merge_cells('A3:H3')
            ws['A3'] = str(values['B2'].value or sheet_name).strip()
            ws.merge_cells('A4:H4')
            ws['A4'] = str(values['B3'].value or '').strip()
            for row in (3, 4):
                cell = ws.cell(row, 1)
                cell.fill = dark_fill
                cell.font = Font(name=mixed_font, color='FFFFFF', bold=True, size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
                for column in range(1, 9):
                    ws.cell(row, column).fill = dark_fill
            ws.row_dimensions[3].height = 22
            ws.row_dimensions[4].height = 22

            headers = [
                'Category', 'Ingredient', 'Unit', 'Base Recipe\n(1 Portion)',
                'Corrected\nConversion Factor', 'Scaling Factor\n(1-10KG)',
                'Linear Scaled\nAmount', 'Scaled Amount Post Conversion Factor',
            ]
            for column, header in enumerate(headers, 1):
                cell = ws.cell(5, column, header)
                cell.fill = green_fill if column in (5, 6) else dark_fill
                cell.font = Font(
                    name=mixed_font,
                    color='000000' if column in (5, 6) else 'FFFFFF',
                    bold=True,
                    size=9,
                )
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
            ws.row_dimensions[5].height = 46

            output_row = 6
            for source_row in range(5, min(values.max_row, 50) + 1):
                ingredient = values.cell(source_row, 2).value
                if not ingredient:
                    continue
                is_total = str(ingredient).strip().casefold() == 'sauce quantity gm'
                for column in range(1, 9):
                    value = values.cell(source_row, column).value
                    cell = ws.cell(output_row, column, value)
                    cell.fill = yellow_fill if is_total and column in (7, 8) else white_fill
                    cell.font = Font(name=mixed_font, bold=is_total or column in (2, 8), size=9)
                    cell.alignment = Alignment(
                        horizontal='right' if column in (2, 8) else 'center',
                        vertical='center',
                        wrap_text=True,
                        readingOrder=2 if column == 2 else 0,
                    )
                    cell.border = border
                    if isinstance(value, (int, float)):
                        cell.number_format = '0%' if column == 6 else '#,##0.##'
                ws.row_dimensions[output_row].height = 21
                output_row += 1

            widths = [16, 36, 11, 15, 18, 16, 17, 34]
            for column, width in enumerate(widths, 1):
                ws.column_dimensions[chr(64 + column)].width = width
            last_row = output_row - 1
            ws.print_area = f'A1:H{max(last_row + 12, 28)}'
            ws.page_setup.orientation = 'landscape'
            ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            ws.page_margins.left = 0.25
            ws.page_margins.right = 0.25
            ws.page_margins.top = 0.35
            ws.page_margins.bottom = 0.35
            ws.oddFooter.center.text = f'Page {page_index} of {total_pages}'
            ws.oddFooter.center.size = 10
            ws.oddFooter.center.font = 'Arial,Bold'

        if not report.sheetnames:
            raise ValueError('لا توجد وصفات صوص مرتبطة بهذا اليوم')
        report.active = 0
        output = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False).name
        report.save(output)
        return output, len(report.sheetnames)
    finally:
        source.close()
        report.close()


def _state_from_workbook(path, day_no):
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        items = []
        for sheet_name in _day_sheets(wb, day_no):
            ws = wb[sheet_name]
            items.append({
                'sheet': sheet_name,
                'name': str(ws['B2'].value or sheet_name).strip(),
                'meal': str(ws['B3'].value or '').strip(),
                'input_count': round(_number(ws['V1'].value), 3),
                'safety_count': 0,
                'final_count': round(_number(ws['Q20'].value), 3),
            })
        return {'day_no': int(day_no), 'day_name': DAY_NAMES[int(day_no)], 'items': items}
    finally:
        wb.close()


def get_sauce_production_state(template_path=SAUCE_TEMPLATE_PATH, day_no=1):
    if not os.path.exists(template_path):
        raise FileNotFoundError('ملف الصوص الأساسي غير موجود')
    state = _state_from_workbook(template_path, int(_number(day_no) or 1))
    state['template_updated_at'] = datetime.fromtimestamp(os.path.getmtime(template_path)).isoformat(timespec='seconds')
    return state


def _files_from_inputs(day_no, inputs, template_path=SAUCE_TEMPLATE_PATH):
    macro_path = _write_inputs(day_no, inputs, template_path)
    calculated = _convert(macro_path, 'xlsx', 'sauce_recalc_')
    pdf_source, page_count = _pdf_source(calculated, day_no)
    pdf_path = _convert(pdf_source, 'pdf', 'sauce_pdf_')
    report = {
        'day_no': int(day_no), 'page_count': page_count,
        'matched_count': sum(1 for item in inputs if _number(item.get('input_count')) > 0),
        'state': _state_from_workbook(calculated, day_no),
    }
    return calculated, pdf_path, report


def build_sauce_day_files(file_storage, template_path=SAUCE_TEMPLATE_PATH, safety_items=None, expected_day_no=None):
    # Some operational exports leave Update!A6/B6 empty. The dashboard day
    # selection is an explicit, reliable fallback in that case.
    day_no, meals, input_report = read_day_file_payload(
        file_storage, fallback_day_no=expected_day_no
    )
    if expected_day_no is not None and int(_number(expected_day_no)) != int(day_no):
        raise ValueError(f'ملف اليوم يخص يوم {day_no} بينما اليوم المختار في لوحة الصوص هو {int(_number(expected_day_no))}')
    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    try:
        inputs, missing = _compute_day_counts(wb, day_no, meals)
    finally:
        wb.close()
    safety_lookup = {
        str(item.get('sheet') or ''): max(0.0, _number(item.get('safety_count')))
        for item in (safety_items or [])
    }
    for item in inputs:
        item['safety_count'] = safety_lookup.get(item['sheet'], 0.0)
    if missing:
        raise SauceMappingRequiredError(missing, list((meals or {}).keys()))
    excel_path, pdf_path, report = _files_from_inputs(day_no, inputs, template_path)
    report.update({'input_report': input_report, 'missing_sources': missing, 'inputs': inputs})
    return excel_path, pdf_path, report


def build_sauce_manual_files(day_no, items, template_path=SAUCE_TEMPLATE_PATH):
    day_no = int(_number(day_no))
    wb = load_workbook(template_path, data_only=False, read_only=True, keep_vba=True)
    try:
        allowed = set(_day_sheets(wb, day_no))
    finally:
        wb.close()
    inputs = []
    for item in items or []:
        sheet = str(item.get('sheet') or '')
        if sheet in allowed:
            inputs.append({
                'sheet': sheet,
                'input_count': max(0.0, _number(item.get('input_count'))),
                'safety_count': max(0.0, _number(item.get('safety_count'))),
            })
    return _files_from_inputs(day_no, inputs, template_path)


def package_sauce_files(excel_path, pdf_path, day_no):
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, 'w', compression=zipfile.ZIP_DEFLATED) as archive:
        archive.write(excel_path, f'Day{int(day_no)}_Sauce_Updated.xlsx')
        archive.write(pdf_path, f'Day{int(day_no)}_Sauce.pdf')
    buffer.seek(0)
    return buffer


def replace_sauce_production_template(file_storage, template_path=SAUCE_TEMPLATE_PATH):
    if os.path.splitext(file_storage.filename or '')[1].lower() != '.xlsm':
        raise ValueError('الشيت الأساسي للصوص لازم يكون XLSM')
    upload = tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False).name
    file_storage.seek(0)
    file_storage.save(upload)
    wb = load_workbook(upload, data_only=False, read_only=True, keep_vba=True)
    try:
        if 'Ordering' not in wb.sheetnames or len(_recipe_sheet_names(wb)) != sum(DAY_BLOCK_LENGTHS):
            raise ValueError('الشيت الجديد لا يطابق ترتيب Master الصوص المعتمد')
        for day_no in DAY_NAMES:
            for sheet_name in _day_sheets(wb, day_no):
                wb[sheet_name]['V1']
                wb[sheet_name]['Q19']
                wb[sheet_name]['Q20']
    finally:
        wb.close()
    shutil.copy2(upload, template_path)
    return get_sauce_production_state(template_path, 1), {'template_file': os.path.basename(template_path)}
