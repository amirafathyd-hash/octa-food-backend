import io
import os
import re
import tempfile
import unicodedata
from collections import defaultdict
from datetime import date, datetime

from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

from db import execute_with_retry, get_client
from invoice_export import parse_invoice_full
from parse_order import parse_order_pdf


veg_comparison_bp = Blueprint('veg_comparison', __name__)
_auth_checker = None


def configure_veg_comparison(auth_checker):
    global _auth_checker
    _auth_checker = auth_checker


def _require_auth():
    if _auth_checker is None:
        return jsonify({'error': 'إعدادات التحقق غير مكتملة'}), 500
    _, err = _auth_checker()
    return err


def _text(value):
    return str(value or '').strip()


def _number(value):
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r'[^0-9.\-]', '', str(value).replace(',', ''))
    try:
        return float(cleaned)
    except (TypeError, ValueError):
        return 0.0


def _date_text(value):
    if isinstance(value, (datetime, date)):
        return value.strftime('%Y-%m-%d')
    raw = _text(value)
    if not raw:
        return ''
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d-%b-%Y', '%d %b %Y'):
        try:
            return datetime.strptime(raw, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    match = re.search(r'(20\d{2})[-/.](\d{1,2})[-/.](\d{1,2})', raw)
    if match:
        return f'{match.group(1)}-{int(match.group(2)):02d}-{int(match.group(3)):02d}'
    return raw


def _header_key(value):
    value = unicodedata.normalize('NFKC', _text(value)).lower()
    value = re.sub(r'[\u064b-\u065f\u0670]', '', value)
    value = value.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا').replace('ى', 'ي').replace('ة', 'ه')
    return re.sub(r'[^a-z0-9\u0600-\u06ff]+', '', value)


ALIASES = {
    'date': {_header_key(x) for x in ('التاريخ', 'تاريخ', 'date', 'item date', 'invoice date')},
    'name': {_header_key(x) for x in ('الصنف', 'اسم الصنف', 'المنتج', 'اسم المنتج', 'الصنف / البيان', 'الصنف / المنتج', 'item', 'item name', 'product', 'product name', 'product / item')},
    'name_en': {_header_key(x) for x in ('الاسم الإنجليزي', 'الاسم الانجليزي', 'name en', 'english name', 'item english')},
    'name_ar': {_header_key(x) for x in ('الاسم العربي', 'name ar', 'arabic name', 'item arabic')},
    'unit': {_header_key(x) for x in ('الوحدة', 'وحدة', 'unit', 'uom')},
    'order_qty': {_header_key(x) for x in ('الكمية', 'كمية الطلب', 'كمية الأوردر', 'كميه الاوردر', 'طلب اليوم', 'order qty', 'order quantity', 'qty needed', 'daily order')},
    'invoice_qty': {_header_key(x) for x in ('الكمية', 'كمية الفاتورة', 'كميه الفاتوره', 'invoice qty', 'invoice quantity', 'qty', 'quantity')},
    'unit_price': {_header_key(x) for x in ('سعر الوحدة', 'سعر الوحده', 'سعر الفاتورة', 'unit price', 'price')},
    'total': {_header_key(x) for x in ('الإجمالي', 'الاجمالي', 'المجموع', 'إجمالي البند', 'اجمالي البند', 'invoice total', 'line total', 'total', 'amount')},
}


def _find_header(sheet, kind):
    qty_key = 'order_qty' if kind == 'order' else 'invoice_qty'
    for row_idx in range(1, min(sheet.max_row, 30) + 1):
        mapping = {}
        for col_idx in range(1, sheet.max_column + 1):
            key = _header_key(sheet.cell(row=row_idx, column=col_idx).value)
            if not key:
                continue
            for field, aliases in ALIASES.items():
                if key in aliases and field not in mapping:
                    mapping[field] = col_idx
        has_name = any(k in mapping for k in ('name', 'name_en', 'name_ar'))
        if has_name and qty_key in mapping:
            return row_idx, mapping
    return None, None


def _split_combined_name(value):
    value = _text(value)
    if not value:
        return '', ''
    latin = ' '.join(re.findall(r'[A-Za-z][A-Za-z0-9\s()/.&\-]*', value)).strip(' -—')
    arabic = ' '.join(re.findall(r'[\u0600-\u06FF][\u0600-\u06FF\s()/.&\-]*', value)).strip(' -—')
    return latin, arabic


def _unit(value):
    raw = unicodedata.normalize('NFKC', _text(value)).upper().replace(' ', '')
    if '-' in raw:
        raw = raw.split('-')[-1]
    aliases = {
        'KG': {'KG', 'KGS', 'KILO', 'KILOGRAM', 'كج', 'كجم', 'كيلو', 'كيلوجرام'},
        'GM': {'G', 'GM', 'GRAM', 'GRAMS', 'جرام', 'جم'},
        'PACK': {'PACK', 'PKT', 'PACKAGE', 'حزمه', 'حزمة', 'باكيت', 'ربطه', 'ربطة'},
        'BOX': {'BOX', 'CTN', 'CARTON', 'كرتون', 'صندوق'},
        'PC': {'PC', 'PCS', 'PIECE', 'حبه', 'حبة', 'قطعه', 'قطعة'},
        'TRAY': {'TRAY', 'صينيه', 'صينية'},
        'BTL': {'BTL', 'BOTTLE', 'زجاجه', 'زجاجة', 'عبوه', 'عبوة'},
        'LTR': {'L', 'LTR', 'LITER', 'LITRE', 'لتر'},
        'ML': {'ML', 'مل'},
    }
    normalized_ar = _header_key(raw)
    for canonical, values in aliases.items():
        if raw in values or normalized_ar in {_header_key(x) for x in values}:
            return canonical
    return raw or 'UNKNOWN'


def _canonical_qty(qty, unit):
    qty = _number(qty)
    unit = _unit(unit)
    if unit == 'GM':
        return qty / 1000.0, 'KG'
    if unit == 'ML':
        return qty / 1000.0, 'LTR'
    return qty, unit


def group_veg_daily_rows(rows, log_date=None):
    """يجمع صفوف يوم الخضار بنفس قاعدة الاسم والوحدة المستخدمة في المقارنة."""
    grouped = {}
    for row in rows or []:
        qty, unit = _canonical_qty(row.get('qty'), row.get('unit'))
        match_key = _norm(row.get('name_en') or row.get('name_ar'))
        key = (match_key, unit)
        if not match_key:
            continue
        if key not in grouped:
            grouped[key] = {
                'log_date': row.get('log_date') or log_date,
                'name_en': _text(row.get('name_en')),
                'name_ar': _text(row.get('name_ar')),
                'match_key': match_key,
                'qty': 0.0,
                'unit': unit,
            }
        item = grouped[key]
        item['qty'] += qty
        if not item['name_en'] and row.get('name_en'):
            item['name_en'] = _text(row.get('name_en'))
        if not item['name_ar'] and row.get('name_ar'):
            item['name_ar'] = _text(row.get('name_ar'))
    result = list(grouped.values())
    for item in result:
        item['qty'] = round(item['qty'], 3)
    return sorted(result, key=lambda item: (item.get('name_en') or item.get('name_ar') or '').lower())


def _parse_excel(path, kind, file_name):
    workbook = load_workbook(path, read_only=True, data_only=True)
    items = []
    errors = []
    qty_field = 'order_qty' if kind == 'order' else 'invoice_qty'
    for sheet in workbook.worksheets:
        header_row, mapping = _find_header(sheet, kind)
        if not mapping:
            continue
        sheet_default_date = _date_text(sheet.title)
        for row_values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            def cell(field):
                col = mapping.get(field)
                return row_values[col - 1] if col and col <= len(row_values) else None

            name_en = _text(cell('name_en'))
            name_ar = _text(cell('name_ar'))
            combined = _text(cell('name'))
            if combined and not (name_en or name_ar):
                name_en, name_ar = _split_combined_name(combined)
                if not (name_en or name_ar):
                    name_ar = combined
            if not (name_en or name_ar):
                continue
            if _header_key(combined or name_en or name_ar) in {
                _header_key('الإجمالي'), _header_key('المجموع'), _header_key('total')
            }:
                continue
            qty = _number(cell(qty_field))
            entry = {
                'name_en': name_en,
                'name_ar': name_ar,
                'name': combined or name_ar or name_en,
                'qty': qty,
                'unit': _unit(cell('unit')),
                'date': _date_text(cell('date')) or sheet_default_date,
                'source': file_name,
            }
            if kind == 'invoice':
                entry['unit_price'] = _number(cell('unit_price'))
                entry['total'] = _number(cell('total'))
                if not entry['total'] and entry['unit_price'] and qty:
                    entry['total'] = entry['unit_price'] * qty
            items.append(entry)
    workbook.close()
    if not items:
        errors.append(f'{file_name}: لم يتم العثور على جدول معروف. تأكد من وجود أعمدة الاسم والكمية والوحدة.')
    return items, errors


def _parse_order_file(path, suffix, file_name):
    if suffix == '.pdf':
        order = parse_order_pdf(path)
        log_date = _date_text(order.get('date'))
        items = []
        for section in ('salads', 'dressing'):
            for row in order.get(section) or []:
                items.append({
                    'name_en': _text(row.get('name_en')),
                    'name_ar': _text(row.get('name_ar')),
                    'name': _text(row.get('name_ar') or row.get('name_en')),
                    'qty': _number(row.get('qty_needed')),
                    'unit': _unit(row.get('unit')),
                    'date': log_date,
                    'source': file_name,
                })
        return items, ([] if items else [f'{file_name}: لم يتم استخراج أصناف من ملف الأوردر.'])
    return _parse_excel(path, 'order', file_name)


def _parse_invoice_file(path, suffix, file_name):
    if suffix == '.pdf':
        invoice = parse_invoice_full(path, file_name)
        items = []
        for row in invoice.get('items') or []:
            name = _text(row.get('item'))
            name_en, name_ar = _split_combined_name(name)
            if not (name_en or name_ar):
                name_ar = name
            items.append({
                'name_en': name_en,
                'name_ar': name_ar,
                'name': name,
                'qty': _number(row.get('qty')),
                'unit': _unit(row.get('unit')),
                'unit_price': _number(row.get('unitPrice')),
                'total': _number(row.get('total')),
                'date': _date_text(invoice.get('date')),
                'source': file_name,
                'needs_review': bool(row.get('needsReview')),
            })
        errors = []
        if not items:
            errors.append(f'{file_name}: لم يتم استخراج بنود من الفاتورة.')
        elif any(row.get('needs_review') for row in items):
            errors.append(f'{file_name}: توجد أسماء غير واضحة في ملف PDF وتم تعليمها للمراجعة.')
        return items, errors
    return _parse_excel(path, 'invoice', file_name)


def _stored_orders(date_from='', date_to=''):
    sb = get_client()
    page_size = 1000
    offset = 0
    rows = []
    while True:
        query = sb.table('veg_daily_log').select('log_date,name_en,name_ar,qty,unit,source_note').order('log_date')
        if date_from:
            query = query.gte('log_date', date_from)
        if date_to:
            query = query.lte('log_date', date_to)
        result = execute_with_retry(query.range(offset, offset + page_size - 1))
        batch = result.data or []
        rows.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
    return [{
        'name_en': _text(row.get('name_en')),
        'name_ar': _text(row.get('name_ar')),
        'name': _text(row.get('name_ar') or row.get('name_en')),
        'qty': _number(row.get('qty')),
        'unit': _unit(row.get('unit')),
        'date': _date_text(row.get('log_date')),
        'source': 'الأوردرات المحفوظة',
    } for row in rows]


def _norm(value):
    value = unicodedata.normalize('NFKC', _text(value)).lower()
    value = re.sub(r'[\u064b-\u065f\u0670]', '', value)
    value = value.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا').replace('ى', 'ي').replace('ة', 'ه')
    value = re.sub(r'\b(fresh|vegetable|vegetables|خضار|خضروات)\b', ' ', value)
    return re.sub(r'[^a-z0-9\u0600-\u06ff]+', ' ', value).strip()


def _variants(item):
    return {v for v in (_norm(item.get('name_en')), _norm(item.get('name_ar')), _norm(item.get('name'))) if v}


def _aggregate(items, kind):
    groups = {}
    for item in items:
        qty, unit = _canonical_qty(item.get('qty'), item.get('unit'))
        variants = _variants(item)
        if not variants:
            continue
        preferred = _norm(item.get('name_en')) or _norm(item.get('name_ar')) or sorted(variants)[0]
        existing_key = next((group_key for group_key, group in groups.items()
                             if group_key[1] == unit and group['variants'].intersection(variants)), None)
        key = existing_key or (preferred, unit)
        if key not in groups:
            groups[key] = {
                'name_en': _text(item.get('name_en')),
                'name_ar': _text(item.get('name_ar')),
                'name': _text(item.get('name') or item.get('name_ar') or item.get('name_en')),
                'qty': 0.0,
                'unit': unit,
                'total': 0.0,
                'unit_price': 0.0,
                'variants': set(),
                'dates': set(),
                'sources': set(),
                'needs_review': False,
            }
        group = groups[key]
        group['qty'] += qty
        group['total'] += _number(item.get('total'))
        group['variants'].update(variants)
        if item.get('date'):
            group['dates'].add(_date_text(item.get('date')))
        if item.get('source'):
            group['sources'].add(_text(item.get('source')))
        group['needs_review'] = group['needs_review'] or bool(item.get('needs_review'))
        if not group['name_en'] and item.get('name_en'):
            group['name_en'] = _text(item.get('name_en'))
        if not group['name_ar'] and item.get('name_ar'):
            group['name_ar'] = _text(item.get('name_ar'))
    for group in groups.values():
        if kind == 'invoice':
            group['unit_price'] = group['total'] / group['qty'] if group['qty'] else 0.0
        group['qty'] = round(group['qty'], 3)
        group['total'] = round(group['total'], 3)
        group['unit_price'] = round(group['unit_price'], 3)
    return list(groups.values())


def _match_score(order, invoice):
    scores = []
    for left in order['variants']:
        for right in invoice['variants']:
            if left == right:
                scores.append(100.0)
                continue
            scores.append(max(float(fuzz.ratio(left, right)), float(fuzz.token_set_ratio(left, right))))
    return max(scores or [0.0])


def _display_name(item):
    en = _text(item.get('name_en'))
    ar = _text(item.get('name_ar'))
    return ' — '.join(v for v in (en, ar) if v) or _text(item.get('name')) or 'بدون اسم'


def _join(values):
    return '، '.join(sorted(v for v in values if v))


def _compare(order_items, invoice_items):
    orders = _aggregate(order_items, 'order')
    invoices = _aggregate(invoice_items, 'invoice')
    candidates = []
    for order_idx, order in enumerate(orders):
        ranked = sorted(
            ((idx, _match_score(order, invoice)) for idx, invoice in enumerate(invoices)),
            key=lambda pair: pair[1], reverse=True,
        )
        best_idx, best_score = ranked[0] if ranked else (None, 0.0)
        second_score = ranked[1][1] if len(ranked) > 1 else 0.0
        candidates.append((best_score, best_score - second_score, order_idx, best_idx))

    used_invoices = set()
    rows = []
    for best_score, score_gap, order_idx, best_idx in sorted(candidates, reverse=True):
        order = orders[order_idx]
        invoice = invoices[best_idx] if best_idx is not None else None
        accepted = invoice is not None and best_idx not in used_invoices and best_score >= 90 and score_gap >= 7
        review = invoice is not None and best_idx not in used_invoices and best_score >= 82
        if accepted or review:
            used_invoices.add(best_idx)
            unit_match = order['unit'] == invoice['unit'] or 'UNKNOWN' in (order['unit'], invoice['unit'])
            difference = round(invoice['qty'] - order['qty'], 3)
            if review and not accepted:
                status = 'مراجعة المطابقة'
            elif invoice.get('needs_review'):
                status = 'مراجعة اسم الفاتورة'
            elif not unit_match:
                status = 'اختلاف وحدة'
            elif abs(difference) <= 0.001:
                status = 'مطابق'
            elif difference > 0:
                status = 'زيادة في الفاتورة'
            else:
                status = 'عجز في الفاتورة'
            rows.append({
                'order_date': _join(order['dates']), 'invoice_date': _join(invoice['dates']),
                'order_item': _display_name(order), 'invoice_item': _display_name(invoice),
                'order_unit': order['unit'], 'invoice_unit': invoice['unit'],
                'order_qty': order['qty'], 'invoice_qty': invoice['qty'], 'difference': difference,
                'unit_price': invoice['unit_price'], 'invoice_total': invoice['total'],
                'score': round(best_score, 1), 'status': status,
                'order_source': _join(order['sources']), 'invoice_source': _join(invoice['sources']),
            })
        else:
            rows.append({
                'order_date': _join(order['dates']), 'invoice_date': '',
                'order_item': _display_name(order), 'invoice_item': '',
                'order_unit': order['unit'], 'invoice_unit': '',
                'order_qty': order['qty'], 'invoice_qty': 0, 'difference': round(-order['qty'], 3),
                'unit_price': 0, 'invoice_total': 0, 'score': round(best_score, 1),
                'status': 'غير موجود في الفاتورة',
                'order_source': _join(order['sources']), 'invoice_source': '',
            })

    for idx, invoice in enumerate(invoices):
        if idx in used_invoices:
            continue
        rows.append({
            'order_date': '', 'invoice_date': _join(invoice['dates']),
            'order_item': '', 'invoice_item': _display_name(invoice),
            'order_unit': '', 'invoice_unit': invoice['unit'],
            'order_qty': 0, 'invoice_qty': invoice['qty'], 'difference': invoice['qty'],
            'unit_price': invoice['unit_price'], 'invoice_total': invoice['total'], 'score': 0,
            'status': 'غير موجود في الأوردر',
            'order_source': '', 'invoice_source': _join(invoice['sources']),
        })

    priority = {'مراجعة المطابقة': 0, 'مراجعة اسم الفاتورة': 0, 'اختلاف وحدة': 0,
                'غير موجود في الفاتورة': 1, 'غير موجود في الأوردر': 1,
                'عجز في الفاتورة': 2, 'زيادة في الفاتورة': 2, 'مطابق': 3}
    rows.sort(key=lambda row: (priority.get(row['status'], 9), row['order_item'] or row['invoice_item']))
    stats = {
        'rows': len(rows),
        'matched': sum(1 for row in rows if row['status'] == 'مطابق'),
        'differences': sum(1 for row in rows if row['status'] in ('عجز في الفاتورة', 'زيادة في الفاتورة')),
        'missing': sum(1 for row in rows if row['status'] == 'غير موجود في الفاتورة'),
        'extra': sum(1 for row in rows if row['status'] == 'غير موجود في الأوردر'),
        'review': sum(1 for row in rows if row['status'] in ('مراجعة المطابقة', 'مراجعة اسم الفاتورة', 'اختلاف وحدة')),
    }
    return rows, stats


def _compare_by_order_dates(order_items, invoice_items):
    """يقارن كل يوم أوردر بفواتير نفس اليوم فقط.

    ملف الفواتير الفعلي قد يحتوي عشرات التبويبات لشهر كامل؛ لذلك لا يجوز
    تجميع الشهر كله على أوردر يوم واحد. البنود خارج تواريخ الأوردر تُستبعد
    من الحساب ويُعاد عددها للواجهة كتوضيح للمستخدم.
    """
    order_dates = sorted({_date_text(item.get('date')) for item in order_items if _date_text(item.get('date'))})
    invoice_dates = sorted({_date_text(item.get('date')) for item in invoice_items if _date_text(item.get('date'))})
    if not order_dates:
        rows, stats = _compare(order_items, invoice_items)
        return rows, stats, {'ignored_invoice_items': 0, 'included_invoice_items': len(invoice_items),
                             'order_dates': [], 'invoice_dates': invoice_dates}

    rows = []
    included_invoice_items = 0
    for log_date in order_dates:
        day_orders = [item for item in order_items if _date_text(item.get('date')) == log_date]
        day_invoices = [item for item in invoice_items if _date_text(item.get('date')) == log_date]
        included_invoice_items += len(day_invoices)
        day_rows, _ = _compare(day_orders, day_invoices)
        rows.extend(day_rows)

    undated_orders = [item for item in order_items if not _date_text(item.get('date'))]
    if undated_orders:
        undated_invoices = [item for item in invoice_items if not _date_text(item.get('date'))]
        included_invoice_items += len(undated_invoices)
        undated_rows, _ = _compare(undated_orders, undated_invoices)
        rows.extend(undated_rows)

    stats = {
        'rows': len(rows),
        'matched': sum(1 for row in rows if row['status'] == 'مطابق'),
        'differences': sum(1 for row in rows if row['status'] in ('عجز في الفاتورة', 'زيادة في الفاتورة')),
        'missing': sum(1 for row in rows if row['status'] == 'غير موجود في الفاتورة'),
        'extra': sum(1 for row in rows if row['status'] == 'غير موجود في الأوردر'),
        'review': sum(1 for row in rows if row['status'] in ('مراجعة المطابقة', 'مراجعة اسم الفاتورة', 'اختلاف وحدة')),
    }
    return rows, stats, {
        'ignored_invoice_items': max(0, len(invoice_items) - included_invoice_items),
        'included_invoice_items': included_invoice_items,
        'order_dates': order_dates,
        'invoice_dates': invoice_dates,
    }


def _save_upload(upload):
    suffix = os.path.splitext(upload.filename or '')[1].lower()
    if suffix not in ('.pdf', '.xlsx', '.xlsm'):
        raise ValueError(f'{upload.filename}: النوع غير مدعوم. استخدم PDF أو XLSX أو XLSM.')
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as temp:
        upload.save(temp.name)
        return temp.name, suffix


@veg_comparison_bp.route('/api/veg-order-invoice-compare/preview', methods=['POST'])
def preview_comparison():
    err = _require_auth()
    if err:
        return err
    use_saved = request.form.get('use_saved_orders') == '1'
    order_files = request.files.getlist('order_files')
    invoice_files = request.files.getlist('invoice_files')
    if not use_saved and not order_files:
        return jsonify({'error': 'اختر الأوردرات المحفوظة أو ارفع ملف أوردر'}), 400
    if not invoice_files:
        return jsonify({'error': 'ارفع فاتورة واحدة على الأقل'}), 400

    orders = _stored_orders(request.form.get('from', ''), request.form.get('to', '')) if use_saved else []
    invoices = []
    errors = []
    for upload, kind in [(f, 'order') for f in order_files] + [(f, 'invoice') for f in invoice_files]:
        path = None
        try:
            path, suffix = _save_upload(upload)
            parsed, parse_errors = (
                _parse_order_file(path, suffix, upload.filename)
                if kind == 'order' else _parse_invoice_file(path, suffix, upload.filename)
            )
            (orders if kind == 'order' else invoices).extend(parsed)
            errors.extend(parse_errors)
        except Exception as exc:
            errors.append(f'{upload.filename}: {exc}')
        finally:
            if path:
                try:
                    os.unlink(path)
                except OSError:
                    pass

    if not orders:
        return jsonify({'error': 'لم يتم العثور على أصناف أوردر صالحة', 'details': errors}), 400
    if not invoices:
        return jsonify({'error': 'لم يتم العثور على بنود فاتورة صالحة', 'details': errors}), 400
    rows, stats, scope = _compare_by_order_dates(orders, invoices)
    if scope.get('order_dates') and not scope.get('included_invoice_items'):
        order_days = '، '.join(scope.get('order_dates') or [])
        invoice_days = '، '.join(scope.get('invoice_dates') or []) or 'غير محدد داخل الملف'
        return jsonify({
            'error': 'تاريخ الأوردر لا يطابق أي تاريخ داخل ملف الفاتورة.',
            'details': [f'تاريخ الأوردر: {order_days}', f'تواريخ الفاتورة: {invoice_days}'],
        }), 400
    if scope.get('ignored_invoice_items'):
        errors.append(
            f"تم تجاهل {scope['ignored_invoice_items']} بند فاتورة خارج تاريخ أو فترة الأوردر المحددة."
        )
    return jsonify({'rows': rows, 'stats': stats, 'warnings': errors,
                    'source_counts': {'orders': len(orders), 'invoices': len(invoices),
                                      'invoices_in_scope': scope.get('included_invoice_items', len(invoices))},
                    'scope': scope})


@veg_comparison_bp.route('/api/veg-order-invoice-compare/export', methods=['POST'])
def export_comparison():
    err = _require_auth()
    if err:
        return err
    payload = request.get_json(silent=True) or {}
    rows = payload.get('rows') or []
    stats = payload.get('stats') or {}
    if not rows or len(rows) > 5000:
        return jsonify({'error': 'لا توجد نتائج صالحة للتصدير'}), 400

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'المقارنة'
    sheet.sheet_view.rightToLeft = True
    sheet.freeze_panes = 'A7'
    sheet.sheet_view.showGridLines = False

    red = 'EC1510'
    dark = '3B221B'
    white = 'FFFFFF'
    line = Side(style='thin', color='E7D8D2')
    border = Border(bottom=line)
    status_colors = {
        'مطابق': 'DDF3E5', 'زيادة في الفاتورة': 'FFF0BF', 'عجز في الفاتورة': 'FFD9D5',
        'غير موجود في الفاتورة': 'F7C7C3', 'غير موجود في الأوردر': 'FCE3B5',
        'مراجعة المطابقة': 'E6D9FF', 'مراجعة اسم الفاتورة': 'E6D9FF', 'اختلاف وحدة': 'E6D9FF',
    }
    headers = ['#', 'تاريخ الأوردر', 'تاريخ الفاتورة', 'صنف الأوردر', 'صنف الفاتورة المطابق',
               'وحدة الأوردر', 'وحدة الفاتورة', 'كمية الأوردر', 'كمية الفاتورة', 'الفرق',
               'سعر الوحدة', 'إجمالي الفاتورة', 'دقة المطابقة %', 'الحالة', 'مصدر الأوردر', 'مصدر الفاتورة']

    sheet.merge_cells('A1:P1')
    sheet['A1'] = 'مقارنة أوردرات الخضار بالفواتير'
    sheet['A1'].fill = PatternFill('solid', fgColor=red)
    sheet['A1'].font = Font(color=white, bold=True, size=16)
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[1].height = 30
    summary_values = [('إجمالي الصفوف', stats.get('rows', len(rows))), ('مطابق', stats.get('matched', 0)),
                      ('فروق', stats.get('differences', 0)), ('ناقص', stats.get('missing', 0)),
                      ('زيادة', stats.get('extra', 0)), ('مراجعة', stats.get('review', 0))]
    for idx, (label, value) in enumerate(summary_values):
        col = idx * 2 + 1
        sheet.cell(3, col, label).font = Font(bold=True, color=dark)
        sheet.cell(3, col + 1, value).font = Font(bold=True, color=red)
        sheet.cell(3, col).alignment = sheet.cell(3, col + 1).alignment = Alignment(horizontal='center')

    for col, header in enumerate(headers, 1):
        cell = sheet.cell(6, col, header)
        cell.fill = PatternFill('solid', fgColor=dark)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    sheet.auto_filter.ref = f'A6:P{len(rows) + 6}'

    keys = ['order_date', 'invoice_date', 'order_item', 'invoice_item', 'order_unit', 'invoice_unit',
            'order_qty', 'invoice_qty', 'difference', 'unit_price', 'invoice_total', 'score', 'status',
            'order_source', 'invoice_source']
    for idx, row in enumerate(rows, 1):
        excel_row = idx + 6
        values = [idx] + [row.get(key, '') for key in keys]
        for col, value in enumerate(values, 1):
            cell = sheet.cell(excel_row, col, value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=col in (4, 5, 15, 16))
            if col in (8, 9, 10, 11, 12):
                cell.number_format = '#,##0.000'
        status = _text(row.get('status'))
        sheet.cell(excel_row, 14).fill = PatternFill('solid', fgColor=status_colors.get(status, white))
        sheet.cell(excel_row, 14).font = Font(bold=True, color=dark)

    widths = [7, 15, 15, 30, 30, 13, 13, 14, 14, 12, 14, 16, 14, 22, 25, 25]
    for col, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(col)].width = width

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    stamp = datetime.now().strftime('%Y-%m-%d')
    return send_file(output, as_attachment=True,
                     download_name=f'Veg_Order_Invoice_Comparison_{stamp}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
