import io
import json
import os
import re
import secrets
import tempfile
import uuid
from datetime import date, datetime, timezone

from flask import Blueprint, after_this_request, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from db import execute_with_retry, get_client
from invoice_export import build_invoices_workbook, parse_invoice_full
from veg_comparison import _compare


invoice_receipts_bp = Blueprint('invoice_receipts', __name__)

BUCKET_NAME = os.environ.get('INVOICE_RECEIPT_BUCKET', 'invoice-receipts')
WORKER_TOKEN = os.environ.get(
    'INVOICE_RECEIPT_TOKEN',
    'a4d4d597345943379e775f950d10f07d',
)
MAX_FILE_BYTES = int(os.environ.get('INVOICE_RECEIPT_MAX_MB', '20')) * 1024 * 1024

_auth_checker = None

DEPARTMENTS = {
    'hot': 'قسم الخضار الساخن',
    'salad': 'قسم خضار السلطة',
}


def configure_invoice_receipts(auth_checker):
    """Inject app.py's existing session/role checker without a circular import."""
    global _auth_checker
    _auth_checker = auth_checker


def _require_admin():
    if _auth_checker is None:
        return None, (jsonify({'error': 'إعداد التحقق من الجلسة غير مكتمل'}), 500)
    return _auth_checker()


def _require_worker_token():
    supplied = (
        request.headers.get('X-Invoice-Receipt-Token')
        or request.form.get('token')
        or request.args.get('token')
        or ''
    )
    if not WORKER_TOKEN or not secrets.compare_digest(str(supplied), str(WORKER_TOKEN)):
        return jsonify({'error': 'رابط استلام الفواتير غير صالح'}), 403
    return None


def _safe_file_name(value):
    raw = os.path.basename(str(value or 'invoice.pdf')).strip()
    stem, ext = os.path.splitext(raw)
    stem = re.sub(r'[^\w\u0600-\u06FF.-]+', '-', stem, flags=re.UNICODE).strip('.-')
    return f'{stem or "invoice"}{ext.lower() or ".pdf"}'


def _valid_iso_date(value):
    try:
        return date.fromisoformat(str(value)).isoformat()
    except (TypeError, ValueError):
        return None


def _clean_known_pdf_text(value):
    """Clean embedded-font placeholders in records saved before parser fixes."""
    if not isinstance(value, str):
        return value
    value = value.replace('(cid:53)', 'مر')
    value = value.replace('موسسة', 'مؤسسة').replace('ابراهيم', 'إبراهيم')
    return re.sub(r'\s+', ' ', value).strip()


def _normalize_saved_record(record):
    record = dict(record or {})
    for key in ('supplier_name', 'customer_name'):
        record[key] = _clean_known_pdf_text(record.get(key))
    parsed = dict(record.get('parsed_data') or {})
    for key in ('party', 'supplier', 'customer', 'notes'):
        parsed[key] = _clean_known_pdf_text(parsed.get(key))
    items = []
    for item in parsed.get('items') or []:
        normalized_item = dict(item)
        normalized_item['item'] = _clean_known_pdf_text(normalized_item.get('item'))
        items.append(normalized_item)
    parsed['items'] = items
    department = str(parsed.get('receipt_department') or record.get('department') or '').strip().lower()
    record['department'] = department if department in DEPARTMENTS else 'general'
    record['department_label'] = DEPARTMENTS.get(record['department'], 'غير محدد')
    record['parsed_data'] = parsed
    return record


def _valid_department(value, required=False):
    department = str(value or '').strip().lower()
    if department in DEPARTMENTS:
        return department
    return None if required else ''


def _next_month(month):
    try:
        year, month_no = [int(part) for part in str(month).split('-', 1)]
        if month_no < 1 or month_no > 12:
            return None
        if month_no == 12:
            return f'{year + 1}-01-01'
        return f'{year}-{month_no + 1:02d}-01'
    except (TypeError, ValueError):
        return None


def _filtered_query(payload=None):
    payload = payload or request.args
    sb = get_client()
    query = sb.table('invoice_receipts').select('*')

    ids = payload.get('ids') if hasattr(payload, 'get') else None
    if isinstance(ids, list) and ids:
        query = query.in_('id', [str(item) for item in ids[:1000]])
    else:
        exact_date = _valid_iso_date(payload.get('date'))
        date_from = _valid_iso_date(payload.get('date_from'))
        date_to = _valid_iso_date(payload.get('date_to'))
        month = str(payload.get('month') or '').strip()
        if exact_date:
            query = query.eq('receipt_date', exact_date)
        elif re.fullmatch(r'\d{4}-\d{2}', month) and _next_month(month):
            query = query.gte('receipt_date', f'{month}-01').lt('receipt_date', _next_month(month))
        else:
            if date_from:
                query = query.gte('receipt_date', date_from)
            if date_to:
                query = query.lte('receipt_date', date_to)

    return query.order('receipt_date', desc=True).order('created_at', desc=True).limit(5000)


def _list_records(payload=None):
    rows = execute_with_retry(_filtered_query(payload)).data or []
    records = [_normalize_saved_record(row) for row in rows]
    department = _valid_department((payload or request.args).get('department'))
    if department:
        records = [row for row in records if row.get('department') == department]
    return records


@invoice_receipts_bp.route('/api/invoice-receipts/link-status', methods=['GET'])
def invoice_receipts_link_status():
    err = _require_worker_token()
    if err:
        return err
    return jsonify({'ok': True, 'max_mb': MAX_FILE_BYTES // (1024 * 1024)})


@invoice_receipts_bp.route('/api/invoice-receipts/worker-link', methods=['GET'])
def invoice_receipts_worker_link():
    _, err = _require_admin()
    if err:
        return err
    return jsonify({'token': WORKER_TOKEN, 'path': 'invoice-receipt-center'})


@invoice_receipts_bp.route('/api/invoice-receipts/upload', methods=['POST'])
def invoice_receipts_upload():
    err = _require_worker_token()
    if err:
        return err

    receipt_date = _valid_iso_date(request.form.get('receipt_date'))
    if not receipt_date:
        return jsonify({'error': 'اختار تاريخ الاستلام الصحيح'}), 400

    files = request.files.getlist('files')
    if not files:
        return jsonify({'error': 'اختار فاتورة PDF واحدة على الأقل'}), 400
    if len(files) > 30:
        return jsonify({'error': 'الحد الأقصى 30 فاتورة في المرة الواحدة'}), 400

    department = _valid_department(request.form.get('department'), required=True)
    if not department:
        return jsonify({'error': 'حدد قسم الفواتير: الخضار الساخن أو خضار السلطة'}), 400

    uploader_name = (request.form.get('uploader_name') or '').strip()[:120]
    note = (request.form.get('note') or '').strip()[:500]
    sb = get_client()
    results = []
    errors = []

    for uploaded in files:
        original_name = _safe_file_name(uploaded.filename)
        if not original_name.lower().endswith('.pdf'):
            errors.append({'file_name': original_name, 'error': 'يُسمح بملفات PDF فقط'})
            continue

        file_bytes = uploaded.read()
        if not file_bytes:
            errors.append({'file_name': original_name, 'error': 'الملف فارغ'})
            continue
        if len(file_bytes) > MAX_FILE_BYTES:
            errors.append({
                'file_name': original_name,
                'error': f'حجم الملف أكبر من {MAX_FILE_BYTES // (1024 * 1024)}MB',
            })
            continue
        if not file_bytes.startswith(b'%PDF'):
            errors.append({'file_name': original_name, 'error': 'الملف ليس PDF صالحًا'})
            continue

        temp_path = None
        storage_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
                tmp.write(file_bytes)
                temp_path = tmp.name

            parsed = parse_invoice_full(temp_path, original_name)
            parsed['receipt_department'] = department
            parsed['receipt_department_label'] = DEPARTMENTS[department]
            # Supabase Storage rejects some Arabic/custom-font characters in
            # object keys. Keep the original display name in the database, and
            # use a stable ASCII-only object key internally.
            storage_path = (
                f'{department}/{receipt_date[0:4]}/{receipt_date[5:7]}/{receipt_date[8:10]}/'
                f'{uuid.uuid4().hex}.pdf'
            )
            sb.storage.from_(BUCKET_NAME).upload(
                storage_path,
                file_bytes,
                file_options={'content-type': 'application/pdf', 'upsert': 'false'},
            )

            inserted = execute_with_retry(sb.table('invoice_receipts').insert({
                'receipt_date': receipt_date,
                'invoice_date': _valid_iso_date(parsed.get('date')),
                'invoice_no': (parsed.get('number') or '')[:160],
                'supplier_name': (parsed.get('supplier') or parsed.get('party') or '')[:300],
                'customer_name': (parsed.get('customer') or '')[:300],
                'file_name': original_name,
                'storage_path': storage_path,
                'file_size': len(file_bytes),
                'parsed_data': parsed,
                'uploader_name': uploader_name,
                'note': note,
                'created_at': datetime.now(timezone.utc).isoformat(),
            }))
            record = (inserted.data or [{}])[0]
            results.append({
                'id': record.get('id'),
                'file_name': original_name,
                'receipt_date': receipt_date,
                'invoice_date': record.get('invoice_date'),
                'invoice_no': record.get('invoice_no'),
                'supplier_name': record.get('supplier_name'),
                'department': department,
                'department_label': DEPARTMENTS[department],
                'total': parsed.get('total') or 0,
                'items_count': len(parsed.get('items') or []),
            })
        except Exception as exc:
            if storage_path:
                try:
                    sb.storage.from_(BUCKET_NAME).remove([storage_path])
                except Exception:
                    pass
            errors.append({'file_name': original_name, 'error': str(exc)})
        finally:
            if temp_path and os.path.exists(temp_path):
                os.unlink(temp_path)

    status = 200 if results else 400
    return jsonify({'ok': bool(results), 'saved': results, 'errors': errors}), status


@invoice_receipts_bp.route('/api/invoice-receipts/list', methods=['GET'])
def invoice_receipts_list():
    _, err = _require_admin()
    if err:
        return err
    try:
        records = _list_records()
        total_value = sum(float((row.get('parsed_data') or {}).get('total') or 0) for row in records)
        days = sorted({row.get('receipt_date') for row in records if row.get('receipt_date')}, reverse=True)
        return jsonify({
            'records': records,
            'summary': {'files': len(records), 'days': len(days), 'total': total_value},
        })
    except Exception as exc:
        return jsonify({'error': f'تعذر تحميل أرشيف الفواتير: {exc}'}), 500


@invoice_receipts_bp.route('/api/invoice-receipts/<receipt_id>/pdf', methods=['GET'])
def invoice_receipts_pdf(receipt_id):
    _, err = _require_admin()
    if err:
        return err
    sb = get_client()
    result = execute_with_retry(
        sb.table('invoice_receipts').select('file_name, storage_path').eq('id', receipt_id).limit(1)
    )
    rows = result.data or []
    if not rows:
        return jsonify({'error': 'الفاتورة غير موجودة'}), 404
    row = rows[0]
    try:
        content = sb.storage.from_(BUCKET_NAME).download(row['storage_path'])
        return send_file(
            io.BytesIO(content),
            mimetype='application/pdf',
            as_attachment=request.args.get('inline') != '1',
            download_name=row.get('file_name') or 'invoice.pdf',
        )
    except Exception as exc:
        return jsonify({'error': f'تعذر تحميل ملف PDF: {exc}'}), 500


def _records_to_invoices(records):
    invoices = []
    for record in sorted(records, key=lambda item: (item.get('receipt_date') or '', item.get('created_at') or '')):
        parsed = dict(record.get('parsed_data') or {})
        parsed['fileName'] = record.get('file_name') or parsed.get('fileName') or 'invoice.pdf'
        invoice_date = record.get('invoice_date') or parsed.get('date') or ''
        parsed['date'] = record.get('receipt_date') or invoice_date
        parsed['receiptDate'] = record.get('receipt_date')
        notes = [str(parsed.get('notes') or '').strip(), str(record.get('note') or '').strip()]
        if invoice_date:
            notes.insert(0, f'تاريخ الفاتورة: {invoice_date}')
        parsed['notes'] = ' | '.join(part for part in notes if part)
        invoices.append(parsed)
    return invoices


def _latest_vegetable_receipt(receipt_date, department):
    rows = execute_with_retry(
        get_client().table('upload_log')
        .select('id,message,created_at')
        .eq('file_type', 'vegetables_receipt')
        .order('created_at', desc=True)
        .limit(1000)
    ).data or []
    for row in rows:
        try:
            payload = json.loads(row.get('message') or '{}')
        except (TypeError, ValueError):
            continue
        payload_department = str(payload.get('department') or 'general').strip().lower()
        payload_date = _valid_iso_date(payload.get('selected_date'))
        if payload_department == department and payload_date == receipt_date and payload.get('rows'):
            payload['log_id'] = row.get('id')
            return payload
    return None


def _receipt_comparison_items(payload, receipt_date):
    items = []
    for row in payload.get('rows') or []:
        name = str(row.get('items') or row.get('item') or row.get('name') or '').strip()
        if not name:
            continue
        quantity = row.get('received')
        if quantity in (None, ''):
            continue
        items.append({
            'name': name,
            'qty': quantity,
            'unit': row.get('order_unit') or row.get('unit') or '',
            'date': receipt_date,
            'source': DEPARTMENTS.get(payload.get('department'), 'استلام الخضروات'),
        })
    return items


def _invoice_comparison_items(records, receipt_date):
    items = []
    for record in records:
        parsed = record.get('parsed_data') or {}
        for row in parsed.get('items') or []:
            qty = row.get('qty') or 0
            total = row.get('total')
            if total in (None, ''):
                try:
                    total = float(row.get('unitPrice') or 0) * float(qty or 0)
                except (TypeError, ValueError):
                    total = 0
            items.append({
                'name': row.get('item') or '',
                'qty': qty,
                'unit': row.get('unit') or '',
                'total': total,
                'date': record.get('invoice_date') or receipt_date,
                'source': record.get('file_name') or 'invoice.pdf',
            })
    return items


def _build_department_comparison_workbook(rows, stats, department_label, receipt_date):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'المقارنة'
    sheet.sheet_view.rightToLeft = True
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = 'A7'

    red, dark, white = 'EC1510', '3B221B', 'FFFFFF'
    line = Side(style='thin', color='E7D8D2')
    border = Border(bottom=line)
    status_colors = {
        'مطابق': 'DDF3E5', 'زيادة في الفاتورة': 'FFF0BF', 'عجز في الفاتورة': 'FFD9D5',
        'غير موجود في الفاتورة': 'F7C7C3', 'غير موجود في الأوردر': 'FCE3B5',
        'مراجعة المطابقة': 'E6D9FF', 'مراجعة اسم الفاتورة': 'E6D9FF', 'اختلاف وحدة': 'E6D9FF',
    }
    headers = ['#', 'تاريخ الاستلام', 'تاريخ الفاتورة', 'صنف الاستلام', 'صنف الفاتورة المطابق',
               'وحدة الاستلام', 'وحدة الفاتورة', 'كمية الاستلام', 'كمية الفاتورة', 'الفرق',
               'سعر الوحدة', 'إجمالي الفاتورة', 'دقة المطابقة %', 'الحالة', 'مصدر الاستلام', 'مصدر الفاتورة']
    keys = ['order_date', 'invoice_date', 'order_item', 'invoice_item', 'order_unit', 'invoice_unit',
            'order_qty', 'invoice_qty', 'difference', 'unit_price', 'invoice_total', 'score', 'status',
            'order_source', 'invoice_source']

    sheet.merge_cells('A1:P1')
    sheet['A1'] = f'مقارنة {department_label} بالفواتير — {receipt_date}'
    sheet['A1'].fill = PatternFill('solid', fgColor=red)
    sheet['A1'].font = Font(color=white, bold=True, size=16)
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[1].height = 30
    summary_values = [('إجمالي الصفوف', stats.get('rows', len(rows))), ('مطابق', stats.get('matched', 0)),
                      ('فروق', stats.get('differences', 0)), ('ناقص', stats.get('missing', 0)),
                      ('زيادة', stats.get('extra', 0)), ('مراجعة', stats.get('review', 0))]
    for idx, (label, value) in enumerate(summary_values):
        col = idx * 2 + 1
        sheet.cell(3, col, label).font = Font(bold=True, color=dark)
        sheet.cell(3, col + 1, value).font = Font(bold=True, color=red)
        sheet.cell(3, col).alignment = sheet.cell(3, col + 1).alignment = Alignment(horizontal='center')
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(6, col, header)
        cell.fill = PatternFill('solid', fgColor=dark)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for idx, row in enumerate(rows, 1):
        excel_row = idx + 6
        for col, value in enumerate([idx] + [row.get(key, '') for key in keys], 1):
            cell = sheet.cell(excel_row, col, value)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=col in (4, 5, 15, 16))
            if col in (8, 9, 10, 11, 12):
                cell.number_format = '#,##0.000'
        status = str(row.get('status') or '')
        sheet.cell(excel_row, 14).fill = PatternFill('solid', fgColor=status_colors.get(status, white))
        sheet.cell(excel_row, 14).font = Font(bold=True, color=dark)

    total_row = len(rows) + 8
    sheet.cell(total_row, 10, 'إجمالي قيمة الفواتير').font = Font(bold=True, color=dark)
    sheet.cell(total_row, 10).fill = PatternFill('solid', fgColor='FFF1F0')
    sheet.cell(total_row, 12, f'=SUM(L7:L{max(7, len(rows) + 6)})')
    sheet.cell(total_row, 12).font = Font(bold=True, color=red)
    sheet.cell(total_row, 12).number_format = '#,##0.000'
    sheet.auto_filter.ref = f'A6:P{max(6, len(rows) + 6)}'
    for col, width in enumerate([7, 15, 15, 30, 30, 13, 13, 14, 14, 12, 14, 16, 14, 22, 25, 25], 1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@invoice_receipts_bp.route('/api/invoice-receipts/automatic-export', methods=['POST'])
def invoice_receipts_automatic_export():
    err = _require_worker_token()
    if err:
        return err
    payload = request.get_json(silent=True) or {}
    ids = payload.get('ids') or []
    department = _valid_department(payload.get('department'), required=True)
    receipt_date = _valid_iso_date(payload.get('receipt_date'))
    export_kind = str(payload.get('kind') or '').strip().lower()
    if not ids or len(ids) > 30 or not department or not receipt_date:
        return jsonify({'error': 'بيانات التجميع التلقائي غير مكتملة'}), 400
    records = _list_records({'ids': ids, 'department': department})
    if not records:
        return jsonify({'error': 'لم يتم العثور على فواتير القسم المرفوعة'}), 404

    safe_label = f'{department}-{receipt_date}'
    if export_kind == 'invoices':
        workbook_path = build_invoices_workbook(_records_to_invoices(records))

        @after_this_request
        def cleanup_auto_invoices(response):
            try:
                os.unlink(workbook_path)
            except OSError:
                pass
            return response

        return send_file(
            workbook_path,
            as_attachment=True,
            download_name=f'Invoices_{safe_label}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    if export_kind != 'comparison':
        return jsonify({'error': 'نوع الملف التلقائي غير صحيح'}), 400
    receipt = _latest_vegetable_receipt(receipt_date, department)
    if not receipt:
        return jsonify({'error': f'لا يوجد استلام مسجل لـ {DEPARTMENTS[department]} بتاريخ {receipt_date}'}), 404
    order_items = _receipt_comparison_items(receipt, receipt_date)
    invoice_items = _invoice_comparison_items(records, receipt_date)
    if not order_items:
        return jsonify({'error': 'سجل الاستلام لا يحتوي على كميات مستلمة صالحة للمقارنة'}), 400
    rows, stats = _compare(order_items, invoice_items)
    output = _build_department_comparison_workbook(rows, stats, DEPARTMENTS[department], receipt_date)
    return send_file(
        output,
        as_attachment=True,
        download_name=f'Comparison_{safe_label}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@invoice_receipts_bp.route('/api/invoice-receipts/export', methods=['POST'])
def invoice_receipts_export():
    _, err = _require_admin()
    if err:
        return err
    payload = request.get_json(silent=True) or {}
    try:
        records = _list_records(payload)
        if not records:
            return jsonify({'error': 'لا توجد فواتير في الاختيار الحالي'}), 404
        workbook_path = build_invoices_workbook(_records_to_invoices(records))

        @after_this_request
        def cleanup(response):
            try:
                os.unlink(workbook_path)
            except OSError:
                pass
            return response

        period_label = payload.get('date') or payload.get('month') or 'selected'
        label = '-'.join(part for part in (payload.get('department'), period_label) if part)
        label = re.sub(r'[^0-9A-Za-z_-]+', '-', str(label))
        return send_file(
            workbook_path,
            as_attachment=True,
            download_name=f'invoice-receipts-{label}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as exc:
        return jsonify({'error': f'تعذر إنشاء ملف Excel: {exc}'}), 500


@invoice_receipts_bp.route('/api/invoice-receipts/<receipt_id>', methods=['DELETE'])
def invoice_receipts_delete(receipt_id):
    _, err = _require_admin()
    if err:
        return err
    sb = get_client()
    found = execute_with_retry(
        sb.table('invoice_receipts').select('storage_path').eq('id', receipt_id).limit(1)
    ).data or []
    if not found:
        return jsonify({'error': 'الفاتورة غير موجودة'}), 404
    try:
        sb.storage.from_(BUCKET_NAME).remove([found[0]['storage_path']])
    except Exception:
        pass
    execute_with_retry(sb.table('invoice_receipts').delete().eq('id', receipt_id))
    return jsonify({'ok': True})
