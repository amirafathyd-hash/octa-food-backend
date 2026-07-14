import os
import shutil
import subprocess
import tempfile
from collections import defaultdict
from copy import copy

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SAUCE_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "data", "Tokyo_Sauce.xlsm")


def _as_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _rounded(value):
    if isinstance(value, float):
        return round(value, 3)
    return value


def _fill_rgb(cell):
    color = cell.fill.fgColor
    if color and color.type == "rgb" and color.rgb:
        return "#" + color.rgb[-6:]
    return None


def _cell_payload(ws_formula, ws_values, row, col):
    formula_cell = ws_formula.cell(row=row, column=col)
    value_cell = ws_values.cell(row=row, column=col)
    formula = formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else None
    return {
        "address": formula_cell.coordinate,
        "row": row,
        "col": col,
        "value": _rounded(value_cell.value if formula else formula_cell.value),
        "formula": formula,
        "editable": True,
        "fill": _fill_rgb(formula_cell),
        "bold": bool(formula_cell.font.bold),
        "align": formula_cell.alignment.horizontal,
        "number_format": formula_cell.number_format,
    }


def _soffice_bin():
    return os.environ.get("SOFFICE_BIN") or shutil.which("soffice") or "soffice"


def recalc_workbook_to_xlsx(xlsm_path):
    out_dir = tempfile.mkdtemp(prefix="sauce_recalc_")
    profile_dir = tempfile.mkdtemp(prefix="sauce_lo_profile_")
    cmd = [
        _soffice_bin(),
        f"-env:UserInstallation=file://{profile_dir}",
        "--headless",
        "--convert-to",
        "xlsx",
        "--outdir",
        out_dir,
        xlsm_path,
    ]
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=120)
    if proc.returncode != 0:
        raise RuntimeError((proc.stderr or proc.stdout or "LibreOffice failed").strip())
    return os.path.join(out_dir, os.path.splitext(os.path.basename(xlsm_path))[0] + ".xlsx")


def export_workbook_to_pdf(workbook_path):
    out_dir = tempfile.mkdtemp(prefix="sauce_pdf_")
    profile_dir = tempfile.mkdtemp(prefix="sauce_lo_pdf_profile_")
    cmd = [
        _soffice_bin(),
        f"-env:UserInstallation=file://{profile_dir}",
        "--headless",
        "--convert-to",
        "pdf",
        "--outdir",
        out_dir,
        workbook_path,
    ]
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=120)
    if proc.returncode != 0:
        raise RuntimeError((proc.stderr or proc.stdout or "LibreOffice PDF export failed").strip())
    return os.path.join(out_dir, os.path.splitext(os.path.basename(workbook_path))[0] + ".pdf")


def _apply_edits(wb, edits):
    for edit in edits or []:
        sheet = edit.get("sheet")
        address = edit.get("address")
        value = edit.get("value")
        if not sheet or not address or sheet not in wb.sheetnames:
            continue
        number = _as_number(value)
        wb[sheet][address] = number if number is not None and str(value).strip() != "" else value


def _recipe_sheet_names(wb):
    return [name for name in wb.sheetnames if name not in ("List of Meals", "Ordering")]


def _selected_day_recipe_sheets(workbook_path, day_no, per_day=7):
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        recipes = _recipe_sheet_names(wb)
    finally:
        wb.close()
    day = max(1, int(_as_number(day_no) or 1))
    start = (day - 1) * per_day
    selected = recipes[start:start + per_day]
    return selected or recipes[:per_day]


def _english_first_title(value):
    text = str(value or "").strip()
    if " - " not in text:
        return text
    left, right = [part.strip() for part in text.split(" - ", 1)]
    if any("A" <= ch <= "Z" or "a" <= ch <= "z" for ch in right):
        return f"{right} - {left}"
    return text


def _norm_text(value):
    return " ".join(str(value or "").replace("\u00a0", " ").split()).strip().lower()


def _extract_uploaded_counts(upload_path, known_sauces):
    known = {}
    for item in known_sauces:
        for value in (item.get("name"), item.get("sheet")):
            key = _norm_text(value)
            if key:
                known[key] = item
    matched = {}
    if not known:
        return matched
    wb = load_workbook(upload_path, data_only=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                row_values = [cell.value for cell in row]
                for idx, value in enumerate(row_values):
                    key = _norm_text(value)
                    if key not in known or key in matched:
                        continue
                    candidates = row_values[idx + 1:idx + 5] + row_values[max(0, idx - 3):idx]
                    for candidate in candidates:
                        number = _as_number(candidate)
                        if number is not None:
                            matched[key] = number
                            break
    finally:
        wb.close()
    return matched


def extract_workbook_state(workbook_path):
    wb_formula = load_workbook(workbook_path, data_only=False, keep_vba=True)
    wb_values = load_workbook(workbook_path, data_only=True, keep_vba=True)
    sheets = []
    for sheet_name in wb_formula.sheetnames:
        ws_formula = wb_formula[sheet_name]
        ws_values = wb_values[sheet_name]
        rows = []
        for row in range(1, ws_formula.max_row + 1):
            rows.append([_cell_payload(ws_formula, ws_values, row, col) for col in range(1, ws_formula.max_column + 1)])
        sheets.append({
            "name": sheet_name,
            "max_row": ws_formula.max_row,
            "max_col": ws_formula.max_column,
            "columns": [get_column_letter(col) for col in range(1, ws_formula.max_column + 1)],
            "rows": rows,
        })
    wb_formula.close()
    wb_values.close()
    return {"sheets": sheets}


def extract_dashboard_state(workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    ordering = wb["Ordering"]

    sauce = []
    recipe_sheet_names = _recipe_sheet_names(wb)
    for sheet_name in recipe_sheet_names:
        ws = wb[sheet_name]
        name = ws["B2"].value or sheet_name
        count = _rounded(ws["Q20"].value)
        extra_count = _rounded(ws["Q19"].value)
        total_cost = _as_number(ws["Q21"].value) or 0
        numeric_count = _as_number(count) or 0
        if not name and numeric_count == 0:
            continue
        sauce.append({
            "sheet": sheet_name,
            "row": 19,
            "name": name,
            "count": count,
            "extra_count": extra_count,
            "unit_cost": round(total_cost / numeric_count, 3) if numeric_count else round(total_cost, 3),
            "total_cost": round(total_cost, 3),
            "fill": _fill_rgb(ws["B2"]),
        })

    ingredients = []
    for row in range(1, ordering.max_row + 1):
        item = ordering[f"A{row}"].value
        category = ordering[f"B{row}"].value
        if row != 1 and not item and not category:
            continue
        ingredients.append({
            "row": row,
            "item": item,
            "category": category,
            "unit": ordering[f"C{row}"].value,
            "daily_weight": _rounded(ordering[f"D{row}"].value),
            "weekly_weight": _rounded(ordering[f"E{row}"].value),
            "daily_order": _rounded(ordering[f"L{row}"].value),
            "fill": _fill_rgb(ordering[f"A{row}"]),
        })

    wb.close()
    return {"sauce": sauce, "ingredients": ingredients, "day": 1}


def get_sauce_template_state(template_path=SAUCE_TEMPLATE_PATH):
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Tokyo_Sauce.xlsm غير موجود في data")
    recalculated = recalc_workbook_to_xlsx(template_path)
    state = extract_dashboard_state(recalculated)
    state.update(extract_workbook_state(recalculated))
    return state


def recalculate_sauce_with_edits(edits, template_path=SAUCE_TEMPLATE_PATH):
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Tokyo_Sauce.xlsm غير موجود في data")
    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    _apply_edits(wb, edits)
    out_path = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
    wb.save(out_path)
    wb.close()
    recalculated = recalc_workbook_to_xlsx(out_path)
    state = extract_dashboard_state(recalculated)
    state.update(extract_workbook_state(recalculated))
    return state


def update_sauce_counts_from_upload(file_storage, template_path=SAUCE_TEMPLATE_PATH):
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Tokyo_Sauce.xlsm غير موجود في data")
    suffix = os.path.splitext(file_storage.filename or "")[1].lower()
    if suffix not in (".xlsx", ".xlsm", ".xls"):
        raise ValueError("ملف الأعداد لازم يكون Excel")
    upload_path = tempfile.NamedTemporaryFile(suffix=suffix or ".xlsx", delete=False).name
    file_storage.seek(0)
    file_storage.save(upload_path)

    current_xlsx = recalc_workbook_to_xlsx(template_path)
    current_state = extract_dashboard_state(current_xlsx)
    matched = _extract_uploaded_counts(upload_path, current_state["sauce"])
    if not matched:
        raise ValueError("ملف الأعداد مفيهوش أسماء صوص مطابقة للشيت الرئيسي")

    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    try:
        changed = 0
        for item in current_state["sauce"]:
            keys = [_norm_text(item["name"]), _norm_text(item["sheet"])]
            value = next((matched[key] for key in keys if key in matched), None)
            if value is None or item["sheet"] not in wb.sheetnames:
                continue
            wb[item["sheet"]]["Q19"] = value
            changed += 1
        out_path = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
        wb.save(out_path)
    finally:
        wb.close()

    recalculated = recalc_workbook_to_xlsx(out_path)
    state = extract_dashboard_state(recalculated)
    state.update(extract_workbook_state(recalculated))
    return state, {"matched_count": changed}


def _updated_workbook(edits, template_path=SAUCE_TEMPLATE_PATH):
    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    _apply_edits(wb, edits)
    out_path = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
    day = 1
    wb.save(out_path)
    wb.close()
    return recalc_workbook_to_xlsx(out_path), int(day)


def export_sauce_excel_with_edits(edits, template_path=SAUCE_TEMPLATE_PATH):
    xlsx, day = _updated_workbook(edits, template_path)
    return xlsx, {"day_no": day}


def export_sauce_pdf_with_edits(edits, day_no=1, template_path=SAUCE_TEMPLATE_PATH):
    xlsx, day = _updated_workbook(edits, template_path)
    selected = _selected_day_recipe_sheets(xlsx, day_no)
    wb_values = load_workbook(xlsx, data_only=True)
    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    dark_fill = PatternFill("solid", fgColor="303D4D")
    green_fill = PatternFill("solid", fgColor="C6E0B4")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(color="000000", bold=True, size=16)
    body_font = Font(color="000000", size=10)
    body_bold = Font(color="000000", bold=True, size=10)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    try:
        total_pages = len(selected)
        for page_index, sheet_name in enumerate(selected, 1):
            vals = wb_values[sheet_name]
            ws = out_wb.create_sheet(sheet_name[:31])
            ws.sheet_view.showGridLines = False

            ws.merge_cells("A1:H1")
            ws["A1"] = sheet_name
            ws["A1"].font = title_font
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["I1"] = f"Day {int(_as_number(day_no) or day)}"
            ws["I1"].font = title_font
            ws["I1"].alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[1].height = 28

            sauce_title = _english_first_title(vals["B2"].value or sheet_name)
            meal_title = vals["B3"].value or ""
            ws.merge_cells("B4:H4")
            ws["B4"] = sauce_title
            ws["B4"].fill = dark_fill
            ws["B4"].font = white_font
            ws["B4"].alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells("B5:H5")
            ws["B5"] = meal_title
            ws["B5"].fill = dark_fill
            ws["B5"].font = white_font
            ws["B5"].alignment = Alignment(horizontal="center", vertical="center")
            for address in ("A4", "A5"):
                ws[address].fill = dark_fill
                ws[address].border = border

            headers = [
                "Category",
                "Ingredient",
                "Unit",
                "Base Recipe\n(1 Portion)",
                "Corrected\nConversion Factor",
                "Scaling Factor\n(1-10KG)",
                "Linear Scaled\nAmount",
                "Scaled Amount Post Conversion Factor",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=6, column=col)
                cell.value = header
                cell.fill = green_fill if col in (5, 6) else dark_fill
                cell.font = Font(color="000000" if col in (5, 6) else "FFFFFF", bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            out_row = 7
            for src_row in range(5, vals.max_row + 1):
                ingredient = vals.cell(src_row, 2).value
                if not ingredient:
                    continue
                values = [vals.cell(src_row, col).value for col in range(1, 9)]
                for col, value in enumerate(values, 1):
                    cell = ws.cell(out_row, col)
                    cell.value = value
                    cell.font = body_bold if col in (2, 8) else body_font
                    cell.alignment = Alignment(
                        horizontal="right" if col in (2, 8) else "center",
                        vertical="center",
                        wrap_text=False,
                    )
                    cell.border = border
                    if isinstance(value, (int, float)):
                        cell.number_format = "#,##0.##"
                out_row += 1

            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 37
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 17
            ws.column_dimensions["F"].width = 15
            ws.column_dimensions["G"].width = 16
            ws.column_dimensions["H"].width = 34
            ws.column_dimensions["I"].width = 14
            ws.row_dimensions[4].height = 20
            ws.row_dimensions[5].height = 20
            ws.row_dimensions[6].height = 52
            for row in range(7, out_row):
                ws.row_dimensions[row].height = 18

            ws.page_setup.orientation = "landscape"
            ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            ws.print_area = f"A1:I{max(out_row + 18, 42)}"
            ws.page_margins.left = 0.25
            ws.page_margins.right = 0.25
            ws.page_margins.top = 0.45
            ws.page_margins.bottom = 0.35
            ws.oddFooter.center.text = "Page &P of &N"
            ws.oddFooter.center.size = 10
            ws.oddFooter.center.font = "Arial,Bold"
            footer_row = max(out_row + 18, 42)
            ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=9)
            footer = ws.cell(footer_row, 1)
            footer.value = f"Page {page_index} of {total_pages}"
            footer.font = Font(color="000000", bold=True, size=10)
            footer.alignment = Alignment(horizontal="center", vertical="center")
        if out_wb.sheetnames:
            out_wb.active = 0
        out_path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
        out_wb.save(out_path)
    finally:
        wb_values.close()
        out_wb.close()
    return export_workbook_to_pdf(out_path), {"day_no": int(_as_number(day_no) or day), "sheets": selected}


def replace_sauce_template(file_storage, template_path=SAUCE_TEMPLATE_PATH):
    suffix = os.path.splitext(file_storage.filename or "")[1].lower()
    if suffix != ".xlsm":
        raise ValueError("الشيت الرئيسي لازم يكون .xlsm")
    upload_path = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
    file_storage.seek(0)
    file_storage.save(upload_path)
    wb = load_workbook(upload_path, data_only=False, keep_vba=True)
    try:
        for required in ["List of Meals", "Ordering"]:
            if required not in wb.sheetnames:
                raise ValueError(f"الشيت الجديد لازم يحتوي على {required}")
    finally:
        wb.close()
    shutil.copyfile(upload_path, template_path)
    return get_sauce_template_state(template_path), {"template_file": os.path.basename(template_path)}


def export_sauce_cost_report_with_edits(edits, template_path=SAUCE_TEMPLATE_PATH):
    xlsx, day = _updated_workbook(edits, template_path)
    source = load_workbook(xlsx, data_only=True)
    state = extract_dashboard_state(xlsx)

    report = Workbook()
    summary = report.active
    summary.title = "Summary"
    summary.sheet_view.showGridLines = False
    summary.append(["Metric", "Value"])
    summary.append(["Day", day])
    summary.append(["Sauce Count", len(state["sauce"])])
    summary.append(["Total Cost", round(sum(s["total_cost"] for s in state["sauce"]), 3)])

    salad_ws = report.create_sheet("Sauce Costs")
    salad_ws.append(["Sauce", "Count", "Extra Count", "Unit Cost", "Total Cost"])
    for item in state["sauce"]:
        salad_ws.append([item["name"], item["count"], item.get("extra_count"), item["unit_cost"], item["total_cost"]])

    ing_ws = report.create_sheet("Ordering Map")
    ing_ws.append(["Item", "Category", "Unit", "Daily Weight", "Weekly Weight", "Daily Order"])
    for item in state["ingredients"][1:]:
        ing_ws.append([item["item"], item["category"], item["unit"], item["daily_weight"], item["weekly_weight"], item["daily_order"]])

    usage_ws = report.create_sheet("Recipe Details")
    usage_ws.append(["Sauce", "Ingredient", "Unit", "Base Recipe", "Scaled Amount", "Ordering Qty", "Cost"])
    for sauce in state["sauce"]:
        ws = source[sauce["sheet"]]
        for row in range(5, ws.max_row + 1):
            ingredient = ws[f"B{row}"].value
            if not ingredient:
                continue
            usage_ws.append([
                sauce["name"],
                ingredient,
                ws[f"C{row}"].value,
                ws[f"D{row}"].value,
                ws[f"H{row}"].value,
                ws[f"K{row}"].value,
                ws[f"L{row}"].value,
            ])

    for ws in report.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="70306F")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.000"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25

    if salad_ws.max_row > 1:
        chart = BarChart()
        chart.title = "Total Cost by Sauce"
        chart.y_axis.title = "Cost"
        data = Reference(salad_ws, min_col=5, min_row=1, max_row=salad_ws.max_row)
        cats = Reference(salad_ws, min_col=1, min_row=2, max_row=salad_ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend = None
        chart.height = 8
        chart.width = 18
        summary.add_chart(chart, "D2")

    out_path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
    report.save(out_path)
    source.close()
    return out_path, {"day_no": day, "sauce_count": len(state["sauce"])}


def export_sauce_cost_report_pdf_with_edits(edits, template_path=SAUCE_TEMPLATE_PATH):
    report, meta = export_sauce_cost_report_with_edits(edits, template_path)
    return export_workbook_to_pdf(report), meta
