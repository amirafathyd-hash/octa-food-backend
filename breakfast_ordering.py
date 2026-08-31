import os
import re
import shutil
import subprocess
import tempfile
import unicodedata
from datetime import datetime, timezone
from difflib import SequenceMatcher
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BREAKFAST_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "data", "Tokyo_Breakfast.xlsm")


def _as_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _rounded(value):
    if isinstance(value, float):
        return round(value, 3)
    return value


def _fill_rgb(cell):
    color = cell.fill.fgColor
    if color and color.type == "rgb" and color.rgb:
        return "#" + color.rgb[-6:]
    return None


def _cell_payload(ws_formula, ws_values, row, col):
    formula_cell = ws_formula.cell(row=row, column=col)
    value_cell = ws_values.cell(row=row, column=col)
    formula = formula_cell.value if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=") else None
    return {
        "address": formula_cell.coordinate,
        "row": row,
        "col": col,
        "value": _rounded(value_cell.value if formula else formula_cell.value),
        "formula": formula,
        "editable": True,
        "fill": _fill_rgb(formula_cell),
        "bold": bool(formula_cell.font.bold),
        "align": formula_cell.alignment.horizontal,
        "number_format": formula_cell.number_format,
    }


def _soffice_bin():
    return os.environ.get("SOFFICE_BIN") or shutil.which("soffice") or "soffice"


def recalc_workbook_to_xlsx(xlsm_path):
    out_dir = tempfile.mkdtemp(prefix="breakfast_recalc_")
    profile_dir = tempfile.mkdtemp(prefix="breakfast_lo_profile_")
    cmd = [
        _soffice_bin(),
        f"-env:UserInstallation=file://{profile_dir}",
        "--headless",
        "--convert-to",
        "xlsx",
        "--outdir",
        out_dir,
        xlsm_path,
    ]
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=120)
    if proc.returncode != 0:
        raise RuntimeError((proc.stderr or proc.stdout or "LibreOffice failed").strip())
    return os.path.join(out_dir, os.path.splitext(os.path.basename(xlsm_path))[0] + ".xlsx")


def export_workbook_to_pdf(workbook_path):
    out_dir = tempfile.mkdtemp(prefix="breakfast_pdf_")
    profile_dir = tempfile.mkdtemp(prefix="breakfast_lo_pdf_profile_")
    cmd = [
        _soffice_bin(),
        f"-env:UserInstallation=file://{profile_dir}",
        "--headless",
        "--convert-to",
        "pdf",
        "--outdir",
        out_dir,
        workbook_path,
    ]
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=120)
    if proc.returncode != 0:
        raise RuntimeError((proc.stderr or proc.stdout or "LibreOffice PDF export failed").strip())
    return os.path.join(out_dir, os.path.splitext(os.path.basename(workbook_path))[0] + ".pdf")


def _apply_edits(wb, edits):
    for edit in edits or []:
        sheet = edit.get("sheet")
        address = edit.get("address")
        value = edit.get("value")
        if not sheet or not address or sheet not in wb.sheetnames:
            continue
        number = _as_number(value)
        wb[sheet][address] = number if number is not None and str(value).strip() != "" else value


def _recipe_sheet_names(wb):
    return [name for name in wb.sheetnames if name not in ("List of Meals", "Ordering", "NameNormalizationLog")]


def _resolve_sheet_name(wb, value):
    wanted = _norm_text(value)
    if not wanted:
        return None
    for name in wb.sheetnames:
        if _norm_text(name) == wanted:
            return name
    return None


def _day_recipe_entries(wb, day_no):
    """Read the authoritative breakfast day map from Ordering!AA:AC.

    Breakfast days do not contain a fixed number of recipes, so slicing the
    workbook tabs in groups of seven selects the wrong sheets from day 2
    onwards.  The workbook itself stores the exact recipe sheet, day and base
    count in columns AA, AB and AC (AC may be formula-backed from AG).
    """
    if "Ordering" not in wb.sheetnames:
        raise ValueError("الشيت الرئيسي لازم يحتوي على Ordering")
    ordering = wb["Ordering"]
    day = max(1, int(_as_number(day_no) or 1))
    entries = []
    seen = set()
    for row in range(3, ordering.max_row + 1):
        mapped_day = int(_as_number(ordering.cell(row, 28).value) or 0)  # AB
        if mapped_day != day:
            continue
        sheet_name = _resolve_sheet_name(wb, ordering.cell(row, 27).value)  # AA
        if not sheet_name or sheet_name in seen:
            continue
        count = _as_number(ordering.cell(row, 29).value)  # AC
        if count is None:
            count = _as_number(ordering.cell(row, 33).value)  # AG fallback
        entries.append({
            "sheet": sheet_name,
            "day": day,
            "base_count": _rounded(count or 0),
            "map_row": row,
        })
        seen.add(sheet_name)
    if not entries:
        raise ValueError(f"لا توجد وصفات فطار مرتبطة باليوم {day} في Ordering!AA:AC")
    return entries


def _selected_day_recipe_sheets(workbook_path, day_no):
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        return [entry["sheet"] for entry in _day_recipe_entries(wb, day_no)]
    finally:
        wb.close()


def _english_first_title(value):
    text = str(value or "").strip()
    if " - " not in text:
        return text
    left, right = [part.strip() for part in text.split(" - ", 1)]
    if any("A" <= ch <= "Z" or "a" <= ch <= "z" for ch in right):
        return f"{right} - {left}"
    return text


def _norm_text(value):
    return " ".join(str(value or "").replace("\u00a0", " ").split()).strip().lower()


def _norm_recipe_text(value):
    """Normalize Arabic/English recipe labels without depending on tab names."""
    text = unicodedata.normalize("NFKC", str(value or "")).lower()
    text = re.sub(r"[\u064b-\u065f\u0670\u0640]", "", text)
    text = text.translate(str.maketrans({
        "أ": "ا", "إ": "ا", "آ": "ا", "ٱ": "ا", "ى": "ي", "ة": "ه",
    }))
    text = re.sub(r"[^0-9a-z\u0600-\u06ff]+", " ", text)
    return " ".join(text.split()).strip()


def _breakfast_aliases(item):
    sheet = str(item.get("sheet") or "").strip()
    name = str(item.get("name") or "").strip()
    aliases = {sheet, name}
    for value in (sheet, name):
        aliases.update(part.strip() for part in re.split(r"\s+-\s+|/", value) if part.strip())

    # Known operational wording may differ from the recipe-tab wording while
    # still referring to the exact same breakfast recipe.
    operational_aliases = {
        "foul": {"فول"},
        "egg and cheese sandwich": {
            "ساندوتش بيض بالجبن", "ساندوتش بيض بالجبنة", "ساندوتش بيض وجبنة",
        },
        "croissant spinach and cheese": {
            "كرواسون جبنة بالسبانخ", "كرواسون جبنه بالسبانخ",
            "كروسوان سبانخ وجبنة", "كروسان سبانخ وجبنة",
        },
        "club sandwich": {"كلوب ساندوتش", "ساندوتش كلوب"},
    }
    aliases.update(operational_aliases.get(sheet.lower(), set()))
    return {_norm_recipe_text(value) for value in aliases if _norm_recipe_text(value)}


def _match_uploaded_recipe(recipe_name, breakfasts):
    wanted = _norm_recipe_text(recipe_name)
    if not wanted:
        return None

    best_item = None
    best_score = 0.0
    wanted_tokens = set(wanted.split())
    for item in breakfasts:
        for alias in _breakfast_aliases(item):
            if wanted == alias:
                return item
            alias_tokens = set(alias.split())
            union = wanted_tokens | alias_tokens
            token_score = len(wanted_tokens & alias_tokens) / len(union) if union else 0.0
            text_score = SequenceMatcher(None, wanted, alias).ratio()
            score = max(text_score, token_score)
            if score > best_score:
                best_score = score
                best_item = item
    return best_item if best_score >= 0.67 else None


def _header_index(values, *needles):
    for index, value in enumerate(values, 1):
        normalized = _norm_recipe_text(value)
        if all(needle in normalized for needle in needles):
            return index
    return None


def _extract_sheet1_shift_counts(upload_path):
    """Read final morning/evening counts from Sheet1 by semantic headers."""
    wb = load_workbook(upload_path, data_only=True, read_only=True)
    try:
        if "Sheet1" not in wb.sheetnames:
            raise ValueError("ملف يوم التشغيل لازم يحتوي على التاب Sheet1")
        ws = wb["Sheet1"]
        header_row = None
        recipe_col = morning_col = evening_col = None
        for row_no in range(1, min(ws.max_row, 20) + 1):
            values = [ws.cell(row_no, col).value for col in range(1, ws.max_column + 1)]
            recipe = _header_index(values, "recipe") or _header_index(values, "اسم", "وجبه")
            morning = (
                _header_index(values, "final", "morning", "count")
                or _header_index(values, "نهائي", "صباح", "عدد")
            )
            evening = (
                _header_index(values, "final", "evening", "count")
                or _header_index(values, "نهائي", "مساء", "عدد")
            )
            if recipe and morning and evening:
                header_row, recipe_col, morning_col, evening_col = row_no, recipe, morning, evening
                break
        if header_row is None:
            raise ValueError(
                "لم أجد أعمدة Recipe وFinal Morning Count وFinal Evening Count داخل Sheet1"
            )

        rows = []
        for row_no in range(header_row + 1, ws.max_row + 1):
            recipe_name = str(ws.cell(row_no, recipe_col).value or "").strip()
            if not recipe_name or recipe_name in {"-", "—"}:
                continue
            morning = _as_number(ws.cell(row_no, morning_col).value)
            evening = _as_number(ws.cell(row_no, evening_col).value)
            if morning is None and evening is None:
                continue
            rows.append({
                "row": row_no,
                "recipe": recipe_name,
                "morning": _rounded(morning or 0),
                "evening": _rounded(evening or 0),
            })
        if not rows:
            raise ValueError("Sheet1 لا يحتوي على أعداد صباحية أو مسائية قابلة للقراءة")
        return rows
    finally:
        wb.close()


def _template_day_breakfasts(template_path, day_no):
    wb = load_workbook(template_path, data_only=True, read_only=True)
    try:
        breakfasts = []
        for entry in _day_recipe_entries(wb, day_no):
            ws = wb[entry["sheet"]]
            breakfasts.append({
                "sheet": entry["sheet"],
                "row": entry["map_row"],
                "name": ws["B2"].value or entry["sheet"],
                "fill": _fill_rgb(ws["B2"]),
            })
        return breakfasts
    finally:
        wb.close()


def analyze_breakfast_shift_upload(file_storage, day_no=1, template_path=BREAKFAST_TEMPLATE_PATH):
    """Map Sheet1 final counts to the selected breakfast day and split shifts."""
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Tokyo_Breakfast.xlsm غير موجود في data")
    suffix = os.path.splitext(file_storage.filename or "")[1].lower()
    if suffix not in (".xlsx", ".xlsm"):
        raise ValueError("ملف يوم التشغيل لازم يكون Excel بصيغة XLSX أو XLSM")
    upload_path = tempfile.NamedTemporaryFile(suffix=suffix, delete=False).name
    file_storage.seek(0)
    file_storage.save(upload_path)

    day = max(1, int(_as_number(day_no) or 1))
    breakfasts = _template_day_breakfasts(template_path, day)
    source_rows = _extract_sheet1_shift_counts(upload_path)
    matched = {}
    unmatched = []
    for source in source_rows:
        item = _match_uploaded_recipe(source["recipe"], breakfasts)
        if not item:
            unmatched.append({"row": source["row"], "recipe": source["recipe"]})
            continue
        matched[item["sheet"]] = source

    if not matched:
        raise ValueError("لم أجد في Sheet1 وجبات فطار مطابقة لوصفات اليوم المختار")

    shifts = {"morning": [], "evening": []}
    missing = []
    for item in breakfasts:
        source = matched.get(item["sheet"])
        if source is None:
            missing.append(item["name"])
            continue
        for shift in ("morning", "evening"):
            base_count = _rounded(source[shift])
            shifts[shift].append({
                **item,
                "base_count": base_count,
                "required_count": base_count,
                "safety_count": 0,
                "final_count": base_count,
                "source_row": source["row"],
                "source_recipe": source["recipe"],
                "shift": shift,
            })

    return {
        "day": day,
        "source_file": os.path.basename(file_storage.filename or "day.xlsx"),
        "shifts": shifts,
        "matched_count": len(matched),
        "missing_recipes": missing,
        "unmatched_rows": unmatched,
    }


def _extract_uploaded_counts(upload_path, known_breakfasts):
    known = {}
    for item in known_breakfasts:
        for value in (item.get("name"), item.get("sheet")):
            key = _norm_text(value)
            if key:
                known[key] = item
    matched = {}
    if not known:
        return matched
    wb = load_workbook(upload_path, data_only=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                row_values = [cell.value for cell in row]
                for idx, value in enumerate(row_values):
                    key = _norm_text(value)
                    if key not in known or key in matched:
                        continue
                    candidates = row_values[idx + 1:idx + 5] + row_values[max(0, idx - 3):idx]
                    for candidate in candidates:
                        number = _as_number(candidate)
                        if number is not None:
                            matched[key] = number
                            break
    finally:
        wb.close()
    return matched


def extract_workbook_state(workbook_path):
    wb_formula = load_workbook(workbook_path, data_only=False, keep_vba=True)
    wb_values = load_workbook(workbook_path, data_only=True, keep_vba=True)
    sheets = []
    for sheet_name in wb_formula.sheetnames:
        if sheet_name == "NameNormalizationLog":
            continue
        ws_formula = wb_formula[sheet_name]
        ws_values = wb_values[sheet_name]
        rows = []
        for row in range(1, ws_formula.max_row + 1):
            rows.append([_cell_payload(ws_formula, ws_values, row, col) for col in range(1, ws_formula.max_column + 1)])
        sheets.append({
            "name": sheet_name,
            "max_row": ws_formula.max_row,
            "max_col": ws_formula.max_column,
            "columns": [get_column_letter(col) for col in range(1, ws_formula.max_column + 1)],
            "rows": rows,
        })
    wb_formula.close()
    wb_values.close()
    return {"sheets": sheets}


def extract_dashboard_state(workbook_path, day_no=None):
    wb = load_workbook(workbook_path, data_only=True)
    ordering = wb["Ordering"]
    day = max(1, int(_as_number(day_no) or _as_number(ordering["R1"].value) or 1))
    day_entries = _day_recipe_entries(wb, day)
    breakfast = []
    for entry in day_entries:
        sheet_name = entry["sheet"]
        ws = wb[sheet_name]
        name = ws["B2"].value or sheet_name
        base_count = _rounded(entry["base_count"])
        safety_count = _rounded(ws["S14"].value)
        final_count = _rounded(ws["S15"].value)
        total_cost = _as_number(ws["S16"].value) or 0
        unit_cost = _as_number(ws["Y8"].value)
        numeric_final = _as_number(final_count) or (_as_number(base_count) or 0)
        if unit_cost is None and numeric_final:
            unit_cost = total_cost / numeric_final
        breakfast.append({
            "sheet": sheet_name,
            "row": entry["map_row"],
            "name": name,
            "count": base_count,
            "base_count": base_count,
            "required_count": base_count,
            "safety_count": safety_count or 0,
            "final_count": final_count,
            "unit_cost": round(unit_cost or 0, 3),
            "total_cost": round(total_cost, 3),
            "fill": _fill_rgb(ws["B2"]),
        })

    ingredients = []
    for row in range(1, ordering.max_row + 1):
        item = ordering[f"A{row}"].value
        category = ordering[f"B{row}"].value
        if row != 1 and not item and not category:
            continue
        ingredients.append({
            "row": row,
            "item": item,
            "category": category,
            "unit": ordering[f"C{row}"].value,
            "daily_weight": _rounded(ordering[f"D{row}"].value),
            "weekly_weight": _rounded(ordering[f"E{row}"].value),
            "daily_order": _rounded(ordering[f"L{row}"].value),
            "fill": _fill_rgb(ordering[f"A{row}"]),
        })

    wb.close()
    return {"breakfast": breakfast, "ingredients": ingredients, "day": day}


def get_breakfast_template_state(template_path=BREAKFAST_TEMPLATE_PATH, day_no=None):
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Tokyo_Breakfast.xlsm غير موجود في data")
    recalculated = recalc_workbook_to_xlsx(template_path)
    state = extract_dashboard_state(recalculated, day_no=day_no)
    state["template_file"] = os.path.basename(template_path)
    state["template_updated_at"] = datetime.fromtimestamp(
        os.path.getmtime(template_path), tz=timezone.utc
    ).isoformat()
    return state


def recalculate_breakfast_with_edits(edits, template_path=BREAKFAST_TEMPLATE_PATH):
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Tokyo_Breakfast.xlsm غير موجود في data")
    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    _apply_edits(wb, edits)
    out_path = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
    wb.save(out_path)
    wb.close()
    recalculated = recalc_workbook_to_xlsx(out_path)
    state = extract_dashboard_state(recalculated)
    return state


def update_breakfast_counts_from_upload(file_storage, template_path=BREAKFAST_TEMPLATE_PATH):
    if not os.path.exists(template_path):
        raise FileNotFoundError("ملف Tokyo_Breakfast.xlsm غير موجود في data")
    suffix = os.path.splitext(file_storage.filename or "")[1].lower()
    if suffix not in (".xlsx", ".xlsm", ".xls"):
        raise ValueError("ملف الأعداد لازم يكون Excel")
    upload_path = tempfile.NamedTemporaryFile(suffix=suffix or ".xlsx", delete=False).name
    file_storage.seek(0)
    file_storage.save(upload_path)

    current_xlsx = recalc_workbook_to_xlsx(template_path)
    current_state = extract_dashboard_state(current_xlsx)
    matched = _extract_uploaded_counts(upload_path, current_state["breakfast"])
    if not matched:
        raise ValueError("ملف الأعداد مفيهوش أسماء فطار مطابقة للشيت الرئيسي")

    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    try:
        changed = 0
        for item in current_state["breakfast"]:
            keys = [_norm_text(item["name"]), _norm_text(item["sheet"])]
            value = next((matched[key] for key in keys if key in matched), None)
            if value is None or item["sheet"] not in wb.sheetnames:
                continue
            wb[item["sheet"]]["V1"] = value
            changed += 1
        out_path = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
        wb.save(out_path)
    finally:
        wb.close()

    recalculated = recalc_workbook_to_xlsx(out_path)
    state = extract_dashboard_state(recalculated)
    return state, {"matched_count": changed}


def _updated_workbook(edits, template_path=BREAKFAST_TEMPLATE_PATH):
    wb = load_workbook(template_path, data_only=False, keep_vba=True)
    _apply_edits(wb, edits)
    out_path = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
    day = int(_as_number(wb["Ordering"]["R1"].value) or 1) if "Ordering" in wb.sheetnames else 1
    wb.save(out_path)
    wb.close()
    return recalc_workbook_to_xlsx(out_path), int(day)


def export_breakfast_excel_with_edits(edits, template_path=BREAKFAST_TEMPLATE_PATH):
    xlsx, day = _updated_workbook(edits, template_path)
    return xlsx, {"day_no": day}


def export_breakfast_pdf_with_edits(edits, day_no=1, template_path=BREAKFAST_TEMPLATE_PATH):
    xlsx, day = _updated_workbook(edits, template_path)
    selected = _selected_day_recipe_sheets(xlsx, day_no)
    wb_values = load_workbook(xlsx, data_only=True)
    out_wb = Workbook()
    out_wb.remove(out_wb.active)
    dark_fill = PatternFill("solid", fgColor="303D4D")
    green_fill = PatternFill("solid", fgColor="C6E0B4")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(color="000000", bold=True, size=16)
    body_font = Font(color="000000", size=10)
    body_bold = Font(color="000000", bold=True, size=10)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    try:
        total_pages = len(selected)
        for page_index, sheet_name in enumerate(selected, 1):
            vals = wb_values[sheet_name]
            ws = out_wb.create_sheet(sheet_name[:31])
            ws.sheet_view.showGridLines = False

            ws.merge_cells("A1:H1")
            ws["A1"] = sheet_name
            ws["A1"].font = title_font
            ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
            ws["I1"] = f"Day {int(_as_number(day_no) or day)}"
            ws["I1"].font = title_font
            ws["I1"].alignment = Alignment(horizontal="right", vertical="center")
            ws.row_dimensions[1].height = 28

            breakfast_title = _english_first_title(vals["B2"].value or sheet_name)
            meal_title = vals["B3"].value or ""
            ws.merge_cells("B4:H4")
            ws["B4"] = breakfast_title
            ws["B4"].fill = dark_fill
            ws["B4"].font = white_font
            ws["B4"].alignment = Alignment(horizontal="center", vertical="center")
            ws["A4"].fill = dark_fill
            ws["A4"].border = border
            if meal_title:
                ws["I4"] = meal_title
                ws["I4"].font = body_bold
                ws["I4"].alignment = Alignment(horizontal="center", vertical="center")

            # Keep each recipe's own headings.  Some breakfast sheets have a
            # Cutting Method column while others intentionally leave it blank.
            headers = [vals.cell(4, col).value for col in range(1, 9)]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col)
                cell.value = header
                cell.fill = green_fill if col in (5, 6) else dark_fill
                cell.font = Font(color="000000" if col in (5, 6) else "FFFFFF", bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            out_row = 6
            for src_row in range(5, vals.max_row + 1):
                ingredient = vals.cell(src_row, 2).value
                if not ingredient:
                    continue
                values = [vals.cell(src_row, col).value for col in range(1, 9)]
                for col, value in enumerate(values, 1):
                    cell = ws.cell(out_row, col)
                    cell.value = value
                    cell.font = body_bold if col in (2, 8) else body_font
                    cell.alignment = Alignment(
                        horizontal="right" if col in (2, 8) else "center",
                        vertical="center",
                        wrap_text=False,
                    )
                    cell.border = border
                    if isinstance(value, (int, float)):
                        cell.number_format = "0%" if col == 6 else "#,##0.##"
                out_row += 1

            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 37
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 17
            ws.column_dimensions["F"].width = 15
            ws.column_dimensions["G"].width = 16
            ws.column_dimensions["H"].width = 34
            ws.column_dimensions["I"].width = 14
            ws.row_dimensions[4].height = 20
            ws.row_dimensions[5].height = 52
            for row in range(6, out_row):
                ws.row_dimensions[row].height = 18

            ws.page_setup.orientation = "landscape"
            ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            ws.print_area = f"A1:I{max(out_row + 18, 42)}"
            ws.page_margins.left = 0.25
            ws.page_margins.right = 0.25
            ws.page_margins.top = 0.45
            ws.page_margins.bottom = 0.35
            ws.oddFooter.center.text = "Page &P of &N"
            ws.oddFooter.center.size = 10
            ws.oddFooter.center.font = "Arial,Bold"
            footer_row = max(out_row + 18, 42)
            ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=9)
            footer = ws.cell(footer_row, 1)
            footer.value = f"Page {page_index} of {total_pages}"
            footer.font = Font(color="000000", bold=True, size=10)
            footer.alignment = Alignment(horizontal="center", vertical="center")
        if out_wb.sheetnames:
            out_wb.active = 0
        out_path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
        out_wb.save(out_path)
    finally:
        wb_values.close()
        out_wb.close()
    return export_workbook_to_pdf(out_path), {"day_no": int(_as_number(day_no) or day), "sheets": selected}


def replace_breakfast_template(file_storage, template_path=BREAKFAST_TEMPLATE_PATH):
    suffix = os.path.splitext(file_storage.filename or "")[1].lower()
    if suffix != ".xlsm":
        raise ValueError("الشيت الرئيسي لازم يكون .xlsm")
    upload_path = tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False).name
    file_storage.seek(0)
    file_storage.save(upload_path)
    wb = load_workbook(upload_path, data_only=False, keep_vba=True)
    try:
        if "Ordering" not in wb.sheetnames:
            raise ValueError("الشيت الجديد لازم يحتوي على Ordering")
        mapped_days = set()
        ordering = wb["Ordering"]
        for row in range(3, ordering.max_row + 1):
            day = int(_as_number(ordering.cell(row, 28).value) or 0)
            if day:
                mapped_days.add(day)
        if not mapped_days:
            raise ValueError("الشيت الجديد لا يحتوي على خريطة الأيام في Ordering!AA:AC")
        for day in mapped_days:
            _day_recipe_entries(wb, day)
    finally:
        wb.close()
    shutil.copyfile(upload_path, template_path)
    return get_breakfast_template_state(template_path), {"template_file": os.path.basename(template_path)}


def export_breakfast_cost_report_with_edits(edits, template_path=BREAKFAST_TEMPLATE_PATH):
    xlsx, day = _updated_workbook(edits, template_path)
    source = load_workbook(xlsx, data_only=True)
    state = extract_dashboard_state(xlsx)

    report = Workbook()
    summary = report.active
    summary.title = "Summary"
    summary.sheet_view.showGridLines = False
    summary.append(["Metric", "Value"])
    summary.append(["Day", day])
    summary.append(["Breakfast Count", len(state["breakfast"])])
    summary.append(["Total Cost", round(sum(s["total_cost"] for s in state["breakfast"]), 3)])

    salad_ws = report.create_sheet("Breakfast Costs")
    salad_ws.append(["Breakfast", "Count", "Extra Count", "Unit Cost", "Total Cost"])
    for item in state["breakfast"]:
        salad_ws.append([item["name"], item["count"], item.get("extra_count"), item["unit_cost"], item["total_cost"]])

    ing_ws = report.create_sheet("Ordering Map")
    ing_ws.append(["Item", "Category", "Unit", "Daily Weight", "Weekly Weight", "Daily Order"])
    for item in state["ingredients"][1:]:
        ing_ws.append([item["item"], item["category"], item["unit"], item["daily_weight"], item["weekly_weight"], item["daily_order"]])

    usage_ws = report.create_sheet("Recipe Details")
    usage_ws.append(["Breakfast", "Ingredient", "Unit", "Base Recipe", "Scaled Amount", "Ordering Qty", "Cost"])
    for breakfast in state["breakfast"]:
        ws = source[breakfast["sheet"]]
        for row in range(5, ws.max_row + 1):
            ingredient = ws[f"B{row}"].value
            if not ingredient:
                continue
            usage_ws.append([
                breakfast["name"],
                ingredient,
                ws[f"C{row}"].value,
                ws[f"D{row}"].value,
                ws[f"H{row}"].value,
                ws[f"K{row}"].value,
                ws[f"L{row}"].value,
            ])

    for ws in report.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="70306F")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.000"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_margins.left = 0.25
        ws.page_margins.right = 0.25

    if salad_ws.max_row > 1:
        chart = BarChart()
        chart.title = "Total Cost by Breakfast"
        chart.y_axis.title = "Cost"
        data = Reference(salad_ws, min_col=5, min_row=1, max_row=salad_ws.max_row)
        cats = Reference(salad_ws, min_col=1, min_row=2, max_row=salad_ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend = None
        chart.height = 8
        chart.width = 18
        summary.add_chart(chart, "D2")

    out_path = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False).name
    report.save(out_path)
    source.close()
    return out_path, {"day_no": day, "breakfast_count": len(state["breakfast"])}


def export_breakfast_cost_report_pdf_with_edits(edits, template_path=BREAKFAST_TEMPLATE_PATH):
    report, meta = export_breakfast_cost_report_with_edits(edits, template_path)
    return export_workbook_to_pdf(report), meta
